# -*- coding: utf-8 -*-
"""
Snowball Link11 - 정보보호공시 모듈

정보보호공시를 위한 질문-답변 시스템과 증빙 자료 관리 기능을 제공합니다.
ITGC 인터뷰 모듈과 유사한 구조로 설계되었습니다.
"""
import sys
import os
import json
import uuid
from datetime import datetime
from flask import Blueprint, render_template, jsonify, request, session, send_file
from werkzeug.utils import secure_filename

# Windows 콘솔 UTF-8 설정
if os.name == 'nt':
    os.system('chcp 65001 > nul')
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

bp_link11 = Blueprint('link11', __name__)

# 파일 업로드 설정
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'disclosure')
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'jpg', 'jpeg', 'png', 'gif', 'txt', 'hwp'}
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB

# 파일 시그니처 (매직 바이트) - 실제 파일 내용 검증용
FILE_SIGNATURES = {
    'pdf': [b'%PDF'],
    'doc': [b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'],  # OLE Compound Document
    'docx': [b'PK\x03\x04'],  # ZIP (Office Open XML)
    'xls': [b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'],  # OLE Compound Document
    'xlsx': [b'PK\x03\x04'],  # ZIP (Office Open XML)
    'ppt': [b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'],  # OLE Compound Document
    'pptx': [b'PK\x03\x04'],  # ZIP (Office Open XML)
    'jpg': [b'\xff\xd8\xff'],
    'jpeg': [b'\xff\xd8\xff'],
    'png': [b'\x89PNG\r\n\x1a\n'],
    'gif': [b'GIF87a', b'GIF89a'],
    'hwp': [b'HWP Document File', b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'],  # HWP 5.0+ 또는 OLE
    'txt': None  # 텍스트 파일은 시그니처 검증 생략
}

# 업로드 폴더 생성
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ============================================================================
# 질문 ID 상수 정의 (순차형 ID 체계)
# - ID는 순차적 (Q1, Q2, ...), display_number는 화면 표시용 계층형
# ============================================================================
class QID:
    """정보보호공시 질문 ID 상수"""
    # 카테고리 1: 정보보호 투자
    INV_HAS_INVESTMENT = "Q1"      # 정보보호 투자 발생 여부
    INV_IT_AMOUNT = "Q2"           # 정보기술부문 투자액 A
    INV_SEC_GROUP = "Q3"           # 정보보호부문 투자액 B Group
    INV_SEC_DEPRECIATION = "Q4"    # 감가상각비
    INV_SEC_SERVICE = "Q5"         # 서비스비용
    INV_SEC_LABOR = "Q6"           # 인건비
    INV_HAS_PLAN = "Q7"            # 향후 투자 계획 여부
    INV_PLAN_AMOUNT = "Q8"         # 예정 투자액
    INV_MAIN_ITEMS = "Q27"         # 주요 투자 항목 (신규)

    # 카테고리 2: 정보보호 인력
    PER_HAS_TEAM = "Q9"            # 전담 부서/인력 여부
    PER_TOTAL_EMPLOYEES = "Q10"    # 총 임직원 수
    PER_INTERNAL = "Q11"           # 내부 전담인력 수
    PER_EXTERNAL = "Q12"           # 외주 전담인력 수
    PER_HAS_CISO = "Q13"           # CISO/CPO 지정 여부
    PER_CISO_DETAIL = "Q14"        # CISO/CPO 상세 현황
    PER_IT_EMPLOYEES = "Q28"       # 정보기술인력(C) (신규)
    PER_CISO_ACTIVITY = "Q29"      # CISO 활동내역 (신규)

    # 카테고리 3: 정보보호 인증
    CERT_HAS_CERT = "Q15"          # 인증 보유 여부
    CERT_DETAIL = "Q16"            # 인증 보유 현황

    # 카테고리 4: 정보보호 활동
    ACT_HAS_ACTIVITY = "Q17"       # 이용자 보호 활동 여부
    ACT_IT_ASSET = "Q18"           # IT 자산 관리
    ACT_TRAINING = "Q19"           # 교육/훈련 실적
    ACT_PROCEDURE = "Q20"          # 지침/절차서
    ACT_VULN_ASSESS = "Q21"        # 취약점 분석
    ACT_ZERO_TRUST = "Q22"         # 제로트러스트
    ACT_SBOM = "Q23"               # SBOM
    ACT_CTAS = "Q24"               # C-TAS
    ACT_DRILL = "Q25"              # 모의훈련
    ACT_INSURANCE = "Q26"          # 배상책임보험


# ============================================================================
# Helper Functions
# ============================================================================

def is_logged_in():
    """로그인 상태 확인"""
    return 'user_id' in session


def get_user_info():
    """현재 로그인한 사용자 정보 반환 (세션 우선)"""
    if is_logged_in():
        if 'user_info' in session:
            return session['user_info']
        try:
            from auth import get_current_user
            return get_current_user()
        except ImportError:
            pass
    return None


def get_db():
    """데이터베이스 연결 가져오기"""
    try:
        from auth import get_db as auth_get_db
        return auth_get_db()
    except ImportError:
        import sqlite3
        # 항상 이 파일이 위치한 디렉토리(snowball)에 DB 파일 사용
        _db_dir = os.path.dirname(os.path.abspath(__file__))
        _db_path = os.path.join(_db_dir, 'snowball.db')
        conn = sqlite3.connect(_db_path)
        conn.row_factory = sqlite3.Row
        return conn


def allowed_file(filename):
    """허용된 파일 확장자 확인"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_file_signature(file_stream, file_ext: str) -> tuple[bool, str]:
    """
    파일 시그니처(매직 바이트)를 검증하여 실제 파일 타입 확인

    Args:
        file_stream: 파일 스트림 객체
        file_ext: 파일 확장자 (소문자)

    Returns:
        tuple: (검증 성공 여부, 오류 메시지)
    """
    # 텍스트 파일은 시그니처 검증 생략
    if file_ext == 'txt':
        return True, ""

    # 알 수 없는 확장자
    if file_ext not in FILE_SIGNATURES:
        return True, ""  # 시그니처 목록에 없으면 통과 (확장자 검사는 이미 완료)

    expected_signatures = FILE_SIGNATURES.get(file_ext)
    if expected_signatures is None:
        return True, ""

    # 파일 시작 부분 읽기 (최대 20바이트)
    current_pos = file_stream.tell()
    file_stream.seek(0)
    header = file_stream.read(20)
    file_stream.seek(current_pos)  # 원래 위치로 복원

    if not header:
        return False, "파일 내용을 읽을 수 없습니다."

    # 시그니처 매칭 확인
    for signature in expected_signatures:
        if header.startswith(signature):
            return True, ""

    return False, f"파일 내용이 {file_ext.upper()} 형식과 일치하지 않습니다. 파일이 손상되었거나 확장자가 변조되었을 수 있습니다."


def generate_uuid():
    """UUID 생성"""
    return str(uuid.uuid4())


def validate_company_ownership(user_info: dict, requested_company_id: str) -> tuple[bool, str]:
    """
    로그인한 사용자가 요청된 회사 데이터에 접근 권한이 있는지 검증

    Args:
        user_info: 로그인 사용자 정보 딕셔너리
        requested_company_id: 요청된 company_id

    Returns:
        tuple: (권한 있음 여부, 오류 메시지)
    """
    if not user_info:
        return False, "로그인 정보가 없습니다."

    user_company = user_info.get('company_name', '')

    # 관리자(admin)는 모든 회사 접근 가능
    if user_info.get('is_admin') or user_info.get('role') == 'admin':
        return True, ""

    # 본인 회사 데이터만 접근 가능
    if requested_company_id and requested_company_id != user_company:
        return False, f"다른 회사의 데이터에 접근할 수 없습니다."

    return True, ""


def get_validated_company_id(user_info: dict, requested_company_id: str = None) -> str:
    """
    검증된 company_id 반환 (요청값 무시하고 사용자 소속 회사 반환)

    보안 강화: 클라이언트에서 전달된 company_id를 신뢰하지 않고
    항상 로그인 사용자의 소속 회사를 사용

    Args:
        user_info: 로그인 사용자 정보
        requested_company_id: 클라이언트 요청 company_id (무시됨, 로깅용)

    Returns:
        str: 사용자 소속 company_id
    """
    user_company = user_info.get('company_name', 'default')

    # 요청된 company_id와 다르면 경고 로그
    if requested_company_id and requested_company_id != user_company:
        print(f"⚠️ [보안] company_id 불일치 감지: 요청={requested_company_id}, 실제={user_company}")

    return user_company


def get_company_name_by_user_id(user_id):
    """user_id로 company_name 조회"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT company_name FROM sb_user WHERE user_id = ?
            ''', (user_id,))
            result = cursor.fetchone()
            return result['company_name'] if result else 'default'
    except Exception as e:
        print(f"회사명 조회 오류: {e}")
        return 'default'


# 카테고리 ID ↔ 이름 매핑
CATEGORY_MAP = {
    1: '정보보호 투자 현황',
    2: '정보보호 인력 현황',
    3: '정보보호 관련 인증',
    4: '정보보호 관련 활동'
}

CATEGORY_REVERSE_MAP = {v: k for k, v in CATEGORY_MAP.items()}


def get_category_name(category_id):
    """카테고리 ID로 이름 조회"""
    return CATEGORY_MAP.get(category_id)


def get_category_id(category_name):
    """카테고리 이름으로 ID 조회"""
    return CATEGORY_REVERSE_MAP.get(category_name)


# ============================================================================
# Routes - 메인 페이지
# ============================================================================

@bp_link11.route('/link11')
def index():
    """정보보호공시 메인 페이지 (대시보드)"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info)


@bp_link11.route('/link11/category/<category_id>')
def category_view(category_id):
    """카테고리별 질문 응답 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info,
                           view='category',
                           category_id=category_id)


@bp_link11.route('/link11/progress')
def progress_view():
    """진행 상황 추적 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info,
                           view='progress')


@bp_link11.route('/link11/evidence')
def evidence_view():
    """증빙 자료 관리 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11_evidence.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info)


@bp_link11.route('/link11/report')
def report_view():
    """공시 자료 생성/검토 페이지"""
    user_logged_in = is_logged_in()
    user_info = get_user_info()
    return render_template('link11_report.jsp',
                           is_logged_in=user_logged_in,
                           user_info=user_info)


# ============================================================================
# API Endpoints - 질문 관리
# ============================================================================

@bp_link11.route('/link11/api/questions', methods=['GET'])
def get_questions():
    """모든 질문 조회 (카테고리별 필터링 가능)"""
    try:
        category_id = request.args.get('category', type=int)

        with get_db() as conn:
            if category_id:
                cursor = conn.execute('''
                    SELECT * FROM sb_disclosure_questions
                    WHERE category_id = ?
                    ORDER BY sort_order
                ''', (category_id,))
            else:
                cursor = conn.execute('''
                    SELECT * FROM sb_disclosure_questions
                    ORDER BY category_id, sort_order
                ''')

            questions = []
            for row in cursor.fetchall():
                q = dict(row)
                # JSON 필드 파싱
                if q.get('options'):
                    q['options'] = json.loads(q['options'])
                if q.get('dependent_question_ids'):
                    q['dependent_question_ids'] = json.loads(q['dependent_question_ids'])
                if q.get('evidence_list'):
                    q['evidence_list'] = json.loads(q['evidence_list'])
                questions.append(q)

            return jsonify({
                'success': True,
                'questions': questions,
                'count': len(questions)
            })

    except Exception as e:
        print(f"질문 조회 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/questions/<question_id>', methods=['GET'])
def get_question(question_id):
    """특정 질문 상세 조회"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT * FROM sb_disclosure_questions WHERE id = ?
            ''', (question_id,))
            row = cursor.fetchone()

            if not row:
                return jsonify({'success': False, 'message': '질문을 찾을 수 없습니다.'}), 404

            q = dict(row)
            # JSON 필드 파싱
            if q.get('options'):
                q['options'] = json.loads(q['options'])
            if q.get('dependent_question_ids'):
                q['dependent_question_ids'] = json.loads(q['dependent_question_ids'])
            if q.get('evidence_list'):
                q['evidence_list'] = json.loads(q['evidence_list'])

            return jsonify({'success': True, 'question': q})

    except Exception as e:
        print(f"질문 상세 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/categories', methods=['GET'])
def get_categories():
    """카테고리 목록 조회"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT category_id, category as name, MIN(sort_order) as min_order
                FROM sb_disclosure_questions
                GROUP BY category_id, category
                ORDER BY category_id
            ''')
            categories = [dict(row) for row in cursor.fetchall()]

            # 카테고리별 질문 수 계산
            category_stats = []
            for cat in categories:
                cursor = conn.execute('''
                    SELECT COUNT(*) as total,
                           SUM(CASE WHEN level = 1 THEN 1 ELSE 0 END) as level1_count
                    FROM sb_disclosure_questions 
                    WHERE category_id = ? AND type != 'group'
                ''', (cat['category_id'],))
                stats = cursor.fetchone()
                category_stats.append({
                    'id': cat['category_id'],
                    'name': cat['name'],
                    'total': stats['total'],
                    'level1_count': stats['level1_count']
                })

            return jsonify({
                'success': True,
                'categories': category_stats
            })

    except Exception as e:
        print(f"카테고리 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 답변 관리
# ============================================================================

@bp_link11.route('/link11/api/answers', methods=['POST'])
def save_answer():
    """답변 저장"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        data = request.get_json()
        user_info = get_user_info()

        question_id = data.get('question_id')
        value = data.get('value')
        # 보안: 클라이언트 요청 company_id 무시, 로그인 사용자 소속 회사 사용
        company_id = get_validated_company_id(user_info, data.get('company_id'))
        year = data.get('year', datetime.now().year)

        if not question_id:
            return jsonify({'success': False, 'message': '질문 ID가 필요합니다.'}), 400

        # 확정 상태 차단 (confirmed 세션은 수정 불가)
        with get_db() as conn:
            sess = conn.execute(
                'SELECT status FROM sb_disclosure_sessions WHERE company_id = ? AND year = ?',
                (company_id, year)
            ).fetchone()
            if sess and sess['status'] == 'confirmed':
                return jsonify({'success': False, 'message': '확정된 공시는 수정할 수 없습니다.'}), 403

        # 값이 리스트인 경우 JSON으로 직렬화
        if isinstance(value, list):
            value = json.dumps(value, ensure_ascii=False)

        # 숫자 필드 음수 방지 검증
        numeric_fields = [
            QID.INV_IT_AMOUNT, QID.INV_SEC_GROUP, QID.INV_SEC_DEPRECIATION,
            QID.INV_SEC_SERVICE, QID.INV_SEC_LABOR, QID.INV_PLAN_AMOUNT,
            QID.PER_TOTAL_EMPLOYEES, QID.PER_IT_EMPLOYEES, QID.PER_INTERNAL, QID.PER_EXTERNAL
        ]
        if question_id in numeric_fields and value is not None:
            try:
                num_val = float(str(value).replace(',', ''))
                if num_val < 0:
                    return jsonify({'success': False, 'message': '음수는 입력할 수 없습니다.'}), 400
            except ValueError:
                pass  # 숫자가 아닌 경우 무시

        with get_db() as conn:
            # 0. 인력 수 계층 검증 (총 임직원 >= IT 인력 >= 보안 인력)
            personnel_ids = [QID.PER_TOTAL_EMPLOYEES, QID.PER_IT_EMPLOYEES, QID.PER_INTERNAL, QID.PER_EXTERNAL]
            if question_id in personnel_ids:
                # 현재 저장된 값들과 새 값을 조합하여 검증
                cursor = conn.execute('''
                    SELECT question_id, value FROM sb_disclosure_answers
                    WHERE question_id IN (?, ?, ?, ?) AND company_id = ? AND year = ? AND deleted_at IS NULL
                ''', (*personnel_ids, company_id, year))
                existing_personnel = {row['question_id']: row['value'] for row in cursor.fetchall()}

                # 새 값으로 업데이트
                existing_personnel[question_id] = value

                try:
                    total_emp = float(str(existing_personnel.get(QID.PER_TOTAL_EMPLOYEES, 0) or 0).replace(',', ''))
                    it_emp = float(str(existing_personnel.get(QID.PER_IT_EMPLOYEES, 0) or 0).replace(',', ''))
                    internal = float(str(existing_personnel.get(QID.PER_INTERNAL, 0) or 0).replace(',', ''))
                    external = float(str(existing_personnel.get(QID.PER_EXTERNAL, 0) or 0).replace(',', ''))
                    security_total = internal + external

                    # 검증 1: IT 인력 <= 총 임직원
                    if total_emp > 0 and it_emp > total_emp:
                        return jsonify({
                            'success': False,
                            'message': f'정보기술부문 인력({int(it_emp)}명)은 총 임직원 수({int(total_emp)}명)를 초과할 수 없습니다.'
                        }), 400

                    # 검증 2: 보안 인력 <= IT 인력
                    if it_emp > 0 and security_total > it_emp:
                        return jsonify({
                            'success': False,
                            'message': f'정보보호 전담인력({int(security_total)}명)은 정보기술부문 인력({int(it_emp)}명)을 초과할 수 없습니다.'
                        }), 400
                except ValueError:
                    pass  # 숫자 변환 오류 시 무시

            # 1. 정보보호 투자액 검증 (B <= A)
            # B = Q4+Q5+Q6 합산 — Q3 저장값 미사용 (정규화 원칙)
            inv_b_ids = [QID.INV_SEC_DEPRECIATION, QID.INV_SEC_SERVICE, QID.INV_SEC_LABOR]
            if question_id in inv_b_ids or question_id == QID.INV_IT_AMOUNT:
                # 현재 저장된 Q4/Q5/Q6 조회
                cursor = conn.execute(f'''
                    SELECT question_id, value FROM sb_disclosure_answers
                    WHERE question_id IN (?, ?, ?) AND company_id = ? AND year = ? AND deleted_at IS NULL
                ''', (*inv_b_ids, company_id, year))
                b_vals = {row['question_id']: row['value'] for row in cursor.fetchall()}
                # 현재 저장 중인 값으로 override
                if question_id in inv_b_ids:
                    b_vals[question_id] = value
                # Q2(A) 조회
                cursor = conn.execute(
                    'SELECT value FROM sb_disclosure_answers WHERE question_id = ? AND company_id = ? AND year = ? AND deleted_at IS NULL',
                    (QID.INV_IT_AMOUNT, company_id, year)
                )
                q_a = cursor.fetchone()
                val_a = float(str(value).replace(',', '')) if question_id == QID.INV_IT_AMOUNT \
                    else float(str(q_a['value']).replace(',', '')) if q_a and q_a['value'] else 0
                if val_a > 0:
                    try:
                        val_b = sum(float(str(b_vals.get(qid, 0) or 0).replace(',', '')) for qid in inv_b_ids)
                        if val_b > val_a:
                            return jsonify({'success': False, 'message': f'정보보호 투자액(B) {int(val_b):,}원이 정보기술 투자액(A) {int(val_a):,}원을 초과합니다.'}), 400
                    except ValueError:
                        pass

            # 2. 기존 답변 확인
            cursor = conn.execute('''
                SELECT id, value FROM sb_disclosure_answers
                WHERE question_id = ? AND company_id = ? AND year = ?
            ''', (question_id, company_id, year))
            existing = cursor.fetchone()
            old_value = existing['value'] if existing else None

            if existing:
                # 기존 답변 업데이트 (소프트 삭제된 경우 복구 가능성 포함)
                conn.execute('''
                    UPDATE sb_disclosure_answers
                    SET value = ?, status = 'completed', updated_at = CURRENT_TIMESTAMP, deleted_at = NULL
                    WHERE id = ?
                ''', (value, existing['id']))
                answer_id = existing['id']
            else:
                # 새 답변 생성
                answer_id = generate_uuid()
                conn.execute('''
                    INSERT INTO sb_disclosure_answers
                    (id, question_id, company_id, user_id, year, value, status)
                    VALUES (?, ?, ?, ?, ?, ?, 'completed')
                ''', (answer_id, question_id, company_id, str(user_info.get('user_id', '')), year, value))

            # Audit Trail 기록 (K-SOX 대응)
            conn.execute('''
                INSERT INTO sb_disclosure_answer_history
                (company_id, year, question_id, old_value, new_value, changed_by)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (company_id, year, question_id, old_value, value, str(user_info.get('user_id', ''))))

            # 3. 재귀적 데이터 클렌징 (상위 질문 'NO' 또는 N/A성 답변 시)
            cursor = conn.execute('SELECT type FROM sb_disclosure_questions WHERE id = ?', (question_id,))
            q_info = cursor.fetchone()

            is_negative = False
            if q_info and q_info['type'] == 'yes_no':
                if str(value).strip().upper() in ('NO', 'N', 'FALSE', '0', '아니오', '아니요'):
                    is_negative = True

            if is_negative:
                # 하위 질문 전체 N/A 처리 (답변 없던 것도 포함)
                _mark_dependent_questions_na(conn, question_id, company_id, year, str(user_info.get('user_id', '')))
            else:
                # YES 복귀 시 N/A 행 삭제 → 사용자 재입력 (기존 답변 복구 없음)
                _clear_na_from_dependents(conn, question_id, company_id, year)

            conn.commit()

            # 진행률 업데이트
            _update_session_progress(conn, company_id, year)

            return jsonify({
                'success': True,
                'message': '답변이 저장되었습니다.',
                'answer_id': answer_id
            })

    except Exception as e:
        print(f"답변 저장 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/answers/<int:user_id>/<int:year>', methods=['GET'])
def get_answers(user_id, year):
    """특정 회사의 특정 연도 모든 답변 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT a.*, q.text as question_text, q.type as question_type
                FROM sb_disclosure_answers a
                JOIN sb_disclosure_questions q ON a.question_id = q.id
                WHERE a.company_id = ? AND a.year = ? AND a.deleted_at IS NULL
                ORDER BY q.category_id, q.sort_order
            ''', (company_id, year))

            answers = []
            for row in cursor.fetchall():
                a = dict(row)
                # JSON 값 파싱 시도
                if a.get('value'):
                    try:
                        a['value'] = json.loads(a['value'])
                    except (json.JSONDecodeError, TypeError):
                        pass  # 일반 문자열인 경우 그대로 유지
                answers.append(a)

            sess = conn.execute(
                'SELECT status FROM sb_disclosure_sessions WHERE company_id = ? AND year = ?',
                (company_id, year)
            ).fetchone()
            session_status = sess['status'] if sess else 'draft'

            return jsonify({
                'success': True,
                'answers': answers,
                'count': len(answers),
                'status': session_status
            })

    except Exception as e:
        print(f"답변 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/answers/<question_id>', methods=['DELETE'])
def delete_answer(question_id):
    """특정 질문의 답변 삭제 (하위 질문 포함)"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        data = request.get_json()
        year = data.get('year', datetime.now().year)

        user_info = get_user_info()
        company_id = get_company_name_by_user_id(user_info.get('user_id'))

        with get_db() as conn:
            # 해당 질문의 종속 질문 ID 조회
            cursor = conn.execute('''
                SELECT dependent_question_ids FROM sb_disclosure_questions
                WHERE id = ?
            ''', (question_id,))
            row = cursor.fetchone()

            # 삭제할 질문 ID 목록
            question_ids_to_delete = [question_id]

            # 종속 질문이 있으면 재귀적으로 모든 하위 질문 추가
            if row and row['dependent_question_ids']:
                try:
                    dependent_ids = json.loads(row['dependent_question_ids'])
                    question_ids_to_delete.extend(_get_all_dependent_question_ids(conn, dependent_ids))
                except (json.JSONDecodeError, TypeError):
                    pass

            # 모든 질문의 답변 삭제
            placeholders = ','.join(['?'] * len(question_ids_to_delete))
            cursor = conn.execute(f'''
                UPDATE sb_disclosure_answers
                SET deleted_at = CURRENT_TIMESTAMP
                WHERE question_id IN ({placeholders}) AND company_id = ? AND year = ?
            ''', (*question_ids_to_delete, company_id, year))

            deleted_count = cursor.rowcount
            conn.commit()

            # 진행률 업데이트
            _update_session_progress(conn, company_id, year)

            return jsonify({
                'success': True,
                'message': f'{deleted_count}개의 답변이 삭제되었습니다.',
                'deleted_count': deleted_count,
                'deleted_questions': question_ids_to_delete
            })

    except Exception as e:
        print(f"답변 삭제 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


def _get_all_dependent_question_ids(conn, question_ids):
    """재귀적으로 모든 하위 질문 ID를 수집"""
    all_ids = list(question_ids)

    for q_id in question_ids:
        cursor = conn.execute('''
            SELECT dependent_question_ids FROM sb_disclosure_questions WHERE id = ?
        ''', (q_id,))
        row = cursor.fetchone()

        if row and row['dependent_question_ids']:
            try:
                child_ids = json.loads(row['dependent_question_ids'])
                if child_ids:
                    all_ids.extend(_get_all_dependent_question_ids(conn, child_ids))
            except (json.JSONDecodeError, TypeError):
                pass

    return all_ids



def _calculate_ratios(conn, company_id, year):
    """투자 및 인력 비율 자동 계산"""
    ratios = {
        'investment_ratio': 0,
        'personnel_ratio': 0,
        'activity_score': 0
    }
    
    try:
        # 1. 투자 비율 (감가상각비 + 서비스비용 + 인건비 / 정보기술부문 투자액)
        # Q1(투자 발생 여부)이 'YES'인 경우에만 계산, 아니면 0%
        cursor = conn.execute('SELECT value FROM sb_disclosure_answers WHERE question_id = ? AND company_id = ? AND year = ? AND deleted_at IS NULL', (QID.INV_HAS_INVESTMENT, company_id, year))
        q1_ans = cursor.fetchone()
        
        has_investment = False
        if q1_ans and str(q1_ans['value']).strip().upper() in ('YES', 'Y', 'TRUE', '1', '예', '네'):
            has_investment = True

        if has_investment:
            inv_ids = [QID.INV_IT_AMOUNT, QID.INV_SEC_GROUP, QID.INV_SEC_DEPRECIATION, QID.INV_SEC_SERVICE, QID.INV_SEC_LABOR]
            cursor = conn.execute(f'''
                SELECT question_id, value FROM sb_disclosure_answers
                WHERE question_id IN ({','.join(['?']*len(inv_ids))})
                AND company_id = ? AND year = ? AND deleted_at IS NULL
            ''', (*inv_ids, company_id, year))
            ans = {row['question_id']: row['value'] for row in cursor.fetchall()}

            if QID.INV_IT_AMOUNT in ans:
                try:
                    a = float(str(ans[QID.INV_IT_AMOUNT]).replace(',', ''))
                    # 감가상각비 + 서비스비용 + 인건비 합계 계산
                    b1 = float(str(ans.get(QID.INV_SEC_DEPRECIATION, 0)).replace(',', ''))
                    b2 = float(str(ans.get(QID.INV_SEC_SERVICE, 0)).replace(',', ''))
                    b3 = float(str(ans.get(QID.INV_SEC_LABOR, 0)).replace(',', ''))
                    b_sum = b1 + b2 + b3

                    # 하위 항목 합계가 있으면 우선 사용, 없으면 Group 값 사용
                    b = b_sum if b_sum > 0 else float(str(ans.get(QID.INV_SEC_GROUP, 0)).replace(',', ''))

                    if a > 0:
                        ratios['investment_ratio'] = round((b / a) * 100, 2)
                except: pass
        else:
            ratios['investment_ratio'] = 0.0

        # 2. 인력 비율 (정보보호 전담인력 / 정보기술부문 인력 수 = D/C)
        # Q9(전담 부서/인력 여부)가 'YES'인 경우에만 계산
        cursor = conn.execute('SELECT value FROM sb_disclosure_answers WHERE question_id = ? AND company_id = ? AND year = ? AND deleted_at IS NULL', (QID.PER_HAS_TEAM, company_id, year))
        q9_ans = cursor.fetchone()
        
        has_personnel = False
        if q9_ans and str(q9_ans['value']).strip().upper() in ('YES', 'Y', 'TRUE', '1', '예', '네'):
            has_personnel = True

        if has_personnel:
            per_ids = [QID.PER_TOTAL_EMPLOYEES, QID.PER_INTERNAL, QID.PER_EXTERNAL, QID.PER_IT_EMPLOYEES]
            cursor = conn.execute(f'''
                SELECT question_id, value FROM sb_disclosure_answers
                WHERE question_id IN ({','.join(['?']*len(per_ids))}) AND company_id = ? AND year = ? AND deleted_at IS NULL
            ''', (*per_ids, company_id, year))
            ans = {row['question_id']: row['value'] for row in cursor.fetchall()}

            internal = float(str(ans.get(QID.PER_INTERNAL, 0)).replace(',', ''))
            external = float(str(ans.get(QID.PER_EXTERNAL, 0)).replace(',', ''))
            d_sum = internal + external

            if QID.PER_IT_EMPLOYEES in ans:
                try:
                    it_total = float(str(ans[QID.PER_IT_EMPLOYEES]).replace(',', ''))
                    if it_total > 0:
                        ratios['personnel_ratio'] = round((d_sum / it_total) * 100, 2)
                except: pass
            elif QID.PER_TOTAL_EMPLOYEES in ans:
                try:
                    total = float(str(ans[QID.PER_TOTAL_EMPLOYEES]).replace(',', ''))
                    if total > 0:
                        ratios['personnel_ratio'] = round((d_sum / total) * 100, 2)
                except: pass
        else:
            ratios['personnel_ratio'] = 0.0

        # 3. 활동 지수 (카테고리 4의 긍정 응답 비율)
        try:
            # 전체 활동 질문 수 (Q17 트리거 제외, Group 제외)
            cursor = conn.execute(f'SELECT COUNT(*) FROM sb_disclosure_questions WHERE category_id = 4 AND id != ? AND type != "group"', (QID.ACT_HAS_ACTIVITY,))
            total_act = cursor.fetchone()[0]

            # 활동 카테고리 질문 ID 범위 (Q18~Q26)
            act_ids = [QID.ACT_IT_ASSET, QID.ACT_TRAINING, QID.ACT_PROCEDURE, QID.ACT_VULN_ASSESS,
                       QID.ACT_ZERO_TRUST, QID.ACT_SBOM, QID.ACT_CTAS, QID.ACT_DRILL, QID.ACT_INSURANCE]
            cursor = conn.execute(f'''
                SELECT COUNT(*) FROM sb_disclosure_answers
                WHERE question_id IN ({','.join(['?']*len(act_ids))})
                AND company_id = ? AND year = ? AND deleted_at IS NULL
                AND value NOT IN ("NO", "no", "No", "0", "아니요", "아니오", "미수행", "도입전")
            ''', (*act_ids, company_id, year))
            active_act = cursor.fetchone()[0]

            if total_act > 0:
                ratios['activity_score'] = round((active_act / total_act) * 100)
        except: pass
            
    except Exception as e:
        print(f"비율 계산 오류: {e}")
        
    return ratios


def _is_question_active(q, questions_dict, answers):
    """질문이 현재 답변 상태에 따라 활성화(표시)되어야 하는지 확인 (재귀적)"""
    if q['level'] == 1:
        return True

    parent_id = q.get('parent_question_id')
    if not parent_id:
        return True

    parent_q = questions_dict.get(parent_id)
    if not parent_q:
        return False

    # 부모 질문이 활성 상태여야 함
    if not _is_question_active(parent_q, questions_dict, answers):
        return False

    # 부모 질문의 답변에 따라 트리거 여부 확인
    parent_answer = answers.get(parent_id)
    if parent_answer is None:
        return False

    parent_answer_str = str(parent_answer).strip().upper()

    # 1. Group 유형은 항상 자식 트리거
    if parent_q['type'] == 'group':
        return True

    # 2. YES/NO 유형
    if parent_q['type'] == 'yes_no':
        # 다양한 '예'/'YES' 형태 지원
        return parent_answer_str in ('YES', 'Y', 'TRUE', '1', '예', '네')

    return False


def _is_question_skipped(q, questions_dict, answers):
    """질문이 건너뛰기 상태인지 확인 (부모가 NO로 명시 답변된 경우만 True)

    - True : 상위 질문이 NO → 해당없음, 자동 완료 처리
    - False: 상위 질문이 미답변이거나 활성 상태 → 미완료로 처리
    """
    if q['level'] == 1:
        return False

    parent_id = q.get('parent_question_id')
    if not parent_id:
        return False

    parent_q = questions_dict.get(parent_id)
    if not parent_q:
        return False

    parent_answer = answers.get(parent_id)

    if parent_q['type'] == 'group':
        # Group은 답변 없이도 트리거하므로 조상을 재귀 확인
        return _is_question_skipped(parent_q, questions_dict, answers)

    if parent_q['type'] == 'yes_no':
        if parent_answer is None:
            # 부모 미답변 → 건너뛰기 아님 (아직 미결)
            return False
        parent_answer_str = str(parent_answer).strip().upper()
        if parent_answer_str not in ('YES', 'Y', 'TRUE', '1', '예', '네'):
            # 부모가 명시적으로 NO → 건너뛰기
            return True
        else:
            # 부모가 YES → 조상 체인을 재귀 확인
            return _is_question_skipped(parent_q, questions_dict, answers)

    return False


# ============================================================================
# API Endpoints - 증빙 자료 관리
# ============================================================================

@bp_link11.route('/link11/api/evidence', methods=['POST'])
def upload_evidence():
    """증빙 자료 업로드"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        user_info = get_user_info()

        # 파일 확인
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400

        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': '허용되지 않는 파일 형식입니다.'}), 400

        # 파일 확장자 추출
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''

        # 파일 시그니처(MIME 타입) 검증 - 확장자 변조 방지
        is_valid, error_msg = validate_file_signature(file.stream, file_ext)
        if not is_valid:
            return jsonify({'success': False, 'message': error_msg}), 400

        # 파일 크기 확인
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)

        if file_size > MAX_FILE_SIZE:
            return jsonify({'success': False, 'message': '파일 크기가 100MB를 초과합니다.'}), 400

        # 폼 데이터 가져오기
        question_id = request.form.get('question_id')
        answer_id = request.form.get('answer_id')
        evidence_type = request.form.get('evidence_type', '')
        company_id = user_info.get('company_name', 'default')
        user_id = user_info.get('user_id', 0)
        year = int(request.form.get('year', datetime.now().year))

        # 안전한 파일명 생성
        original_filename = secure_filename(file.filename)
        file_ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
        stored_filename = f"{generate_uuid()}.{file_ext}"

        # user_id/연도별 폴더 생성 (숫자 기반으로 안전한 경로)
        upload_dir = os.path.join(UPLOAD_FOLDER, str(user_id), str(year))
        os.makedirs(upload_dir, exist_ok=True)

        # 파일 저장
        file_path = os.path.join(upload_dir, stored_filename)
        file.save(file_path)

        # 데이터베이스에 기록
        evidence_id = generate_uuid()
        with get_db() as conn:
            conn.execute('''
                INSERT INTO sb_disclosure_evidence
                (id, answer_id, question_id, company_id, year, file_name, file_url,
                 file_size, file_type, evidence_type, uploaded_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                evidence_id, answer_id, question_id, company_id, year,
                original_filename, file_path, file_size, file_ext,
                evidence_type, str(user_info.get('user_id', ''))
            ))
            conn.commit()

        return jsonify({
            'success': True,
            'message': '파일이 업로드되었습니다.',
            'evidence': {
                'id': evidence_id,
                'file_name': original_filename,
                'file_size': file_size,
                'evidence_type': evidence_type
            }
        })

    except Exception as e:
        print(f"파일 업로드 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/<evidence_id>', methods=['DELETE'])
def delete_evidence(evidence_id):
    """증빙 자료 삭제"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        with get_db() as conn:
            # 증빙 자료 조회
            cursor = conn.execute('''
                SELECT * FROM sb_disclosure_evidence WHERE id = ?
            ''', (evidence_id,))
            evidence = cursor.fetchone()

            if not evidence:
                return jsonify({'success': False, 'message': '증빙 자료를 찾을 수 없습니다.'}), 404

            # 파일 삭제
            file_path = evidence['file_url']
            if os.path.exists(file_path):
                os.remove(file_path)

            # 데이터베이스에서 삭제
            conn.execute('DELETE FROM sb_disclosure_evidence WHERE id = ?', (evidence_id,))
            conn.commit()

        return jsonify({'success': True, 'message': '증빙 자료가 삭제되었습니다.'})

    except Exception as e:
        print(f"증빙 자료 삭제 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/<int:user_id>/<int:year>', methods=['GET'])
def get_evidence_list(user_id, year):
    """특정 회사/연도의 모든 증빙 자료 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT e.*, q.text as question_text
                FROM sb_disclosure_evidence e
                LEFT JOIN sb_disclosure_questions q ON e.question_id = q.id
                WHERE e.company_id = ? AND e.year = ?
                ORDER BY e.uploaded_at DESC
            ''', (company_id, year))

            evidence_list = [dict(row) for row in cursor.fetchall()]

            return jsonify({
                'success': True,
                'evidence': evidence_list,
                'count': len(evidence_list)
            })

    except Exception as e:
        print(f"증빙 자료 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/missing/<int:user_id>/<int:year>', methods=['GET'])
def get_missing_evidence(user_id, year):
    """누락된 증빙 자료 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            # 답변이 완료된 질문 중 증빙 자료가 필요하지만 없는 것 찾기
            cursor = conn.execute('''
                SELECT q.id, q.text, q.evidence_list
                FROM sb_disclosure_questions q
                JOIN sb_disclosure_answers a ON q.id = a.question_id
                    AND a.company_id = ? AND a.year = ? AND a.status = 'completed'
                WHERE q.evidence_list IS NOT NULL AND q.evidence_list != '[]'
                    AND NOT EXISTS (
                        SELECT 1 FROM sb_disclosure_evidence e
                        WHERE e.question_id = q.id
                            AND e.company_id = ? AND e.year = ?
                    )
                ORDER BY q.sort_order
            ''', (company_id, year, company_id, year))

            missing = []
            for row in cursor.fetchall():
                q = dict(row)
                if q.get('evidence_list'):
                    q['evidence_list'] = json.loads(q['evidence_list'])
                missing.append(q)

            return jsonify({
                'success': True,
                'missing': missing,
                'count': len(missing)
            })

    except Exception as e:
        print(f"누락 증빙 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/download/<evidence_id>', methods=['GET'])
def download_evidence(evidence_id):
    """증빙 자료 다운로드"""
    try:
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT * FROM sb_disclosure_evidence WHERE id = ?
            ''', (evidence_id,))
            evidence = cursor.fetchone()

            if not evidence:
                return jsonify({'success': False, 'message': '증빙 자료를 찾을 수 없습니다.'}), 404

            file_path = evidence['file_url']
            if not os.path.exists(file_path):
                return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404

            return send_file(
                file_path,
                as_attachment=True,
                download_name=evidence['file_name']
            )

    except Exception as e:
        print(f"파일 다운로드 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/evidence/stats/<int:user_id>/<int:year>', methods=['GET'])
def get_evidence_stats(user_id, year):
    """증빙 자료 통계 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            # 전체 파일 수
            cursor = conn.execute('''
                SELECT COUNT(*) as total FROM sb_disclosure_evidence
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            total = cursor.fetchone()['total']

            # 연결된 질문 수
            cursor = conn.execute('''
                SELECT COUNT(DISTINCT question_id) as questions FROM sb_disclosure_evidence
                WHERE company_id = ? AND year = ? AND question_id IS NOT NULL
            ''', (company_id, year))
            questions = cursor.fetchone()['questions']

            # 총 용량
            cursor = conn.execute('''
                SELECT COALESCE(SUM(file_size), 0) as total_size FROM sb_disclosure_evidence
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            total_size = cursor.fetchone()['total_size']

            # 최근 7일 업로드
            cursor = conn.execute('''
                SELECT COUNT(*) as recent FROM sb_disclosure_evidence
                WHERE company_id = ? AND year = ?
                AND uploaded_at >= datetime('now', '-7 days')
            ''', (company_id, year))
            recent = cursor.fetchone()['recent']

            return jsonify({
                'success': True,
                'stats': {
                    'total': total,
                    'questions': questions,
                    'total_size': total_size,
                    'recent': recent
                }
            })

    except Exception as e:
        print(f"증빙 통계 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 진행 상황
# ============================================================================

@bp_link11.route('/link11/api/progress/<int:user_id>/<int:year>', methods=['GET'])
def get_progress(user_id, year):
    """공시 진행 상황 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            # 모든 질문 로드
            cursor = conn.execute('SELECT * FROM sb_disclosure_questions ORDER BY category_id, sort_order')
            all_questions = [dict(row) for row in cursor.fetchall()]
            questions_dict = {q['id']: q for q in all_questions}

            cursor = conn.execute('''
                SELECT question_id, value FROM sb_disclosure_answers 
                WHERE company_id = ? AND year = ? AND deleted_at IS NULL
            ''', (company_id, year))
            answers = {row['question_id']: row['value'] for row in cursor.fetchall()}

            categories = {}
            total_questions = 0
            answered_questions = 0

            # 질문별 통계 합산 (전체 질문 수 기준, 비활성화 질문도 포함)
            for q in all_questions:
                cat_name = q['category']
                if cat_name not in categories:
                    categories[cat_name] = {
                        'id': q['category_id'],
                        'total': 0,
                        'completed': 0,
                        'rate': 0
                    }

                # Group 유형은 입력 필드가 없으므로 통계에서 제외 (안내 문구 역할)
                if q['type'] == 'group':
                    continue

                # 전체 질문 수에 포함
                categories[cat_name]['total'] += 1
                total_questions += 1

                # 활성화 상태 확인 (재귀적)
                is_active = _is_question_active(q, questions_dict, answers)

                if is_active:
                    # 활성화된 질문: 답변이 있으면 완료
                    if q['id'] in answers and answers[q['id']] not in (None, ''):
                        categories[cat_name]['completed'] += 1
                        answered_questions += 1
                else:
                    # 비활성화 이유 구분:
                    # - 상위 질문이 NO(건너뛰기) → 자동 완료
                    # - 상위 질문이 미답변 → 미완료 (카운트 제외)
                    if _is_question_skipped(q, questions_dict, answers):
                        categories[cat_name]['completed'] += 1
                        answered_questions += 1

            # 각 카테고리별 퍼센트 계산
            for cat_name in categories:
                cat = categories[cat_name]
                if cat['total'] > 0:
                    cat['rate'] = round((cat['completed'] / cat['total']) * 100)

            # 전체 진행률 계산
            completion_rate = round((answered_questions / total_questions) * 100) if total_questions > 0 else 0

            # 세션 조회
            cursor = conn.execute('''
                SELECT * FROM sb_disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            session_data = cursor.fetchone()

            session_dict = dict(session_data) if session_data else None

            # 비율 계산 (투자 비율, 인력 비율)
            ratios = _calculate_ratios(conn, company_id, year)

            return jsonify({
                'success': True,
                'session': session_dict,
                'progress': {
                    'total_questions': total_questions,
                    'answered_questions': answered_questions,
                    'completion_rate': completion_rate
                },
                'categories': categories,
                'ratios': ratios
            })

    except Exception as e:
        print(f"진행 상황 조회 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/session', methods=['POST'])
def create_or_update_session():
    """공시 세션 생성/업데이트"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        data = request.get_json()
        user_info = get_user_info()

        # 보안: 클라이언트 요청 company_id 무시, 로그인 사용자 소속 회사 사용
        company_id = get_validated_company_id(user_info, data.get('company_id'))
        year = data.get('year', datetime.now().year)
        status = data.get('status', 'draft')

        with get_db() as conn:
            # 기존 세션 확인
            cursor = conn.execute('''
                SELECT id FROM sb_disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            existing = cursor.fetchone()

            if existing:
                # 세션 업데이트
                conn.execute('''
                    UPDATE sb_disclosure_sessions
                    SET status = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (status, existing['id']))
                session_id = existing['id']
            else:
                # 새 세션 생성
                session_id = generate_uuid()
                conn.execute('''
                    INSERT INTO sb_disclosure_sessions
                    (id, company_id, user_id, year, status)
                    VALUES (?, ?, ?, ?, ?)
                ''', (session_id, company_id, str(user_info.get('user_id', '')), year, status))

            conn.commit()

            # 진행률 업데이트
            _update_session_progress(conn, company_id, year)

            return jsonify({
                'success': True,
                'message': '세션이 저장되었습니다.',
                'session_id': session_id
            })

    except Exception as e:
        print(f"세션 저장 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500





def _mark_dependent_questions_na(conn, parent_question_id, company_id, year, user_id):
    """부모 질문이 '아니요' 등일 때 하위 질문들을 DB에서 'N/A' 및 'completed' 상태로 처리 (재귀적)
    이렇게 함으로써 하위 질문이 화면에서 사라지더라도 진행률에는 '완료'로 계산되도록 함.
    """
    try:
        # 해당 질문의 dependent_question_ids 조회
        cursor = conn.execute('''
            SELECT dependent_question_ids FROM sb_disclosure_questions
            WHERE id = ?
        ''', (parent_question_id,))
        result = cursor.fetchone()

        if not result or not result['dependent_question_ids']:
            return

        # dependent_question_ids 파싱
        dependent_ids_str = result['dependent_question_ids']
        try:
            dependent_ids = json.loads(dependent_ids_str)
        except (json.JSONDecodeError, TypeError):
            if dependent_ids_str:
                dependent_ids = [id.strip() for id in dependent_ids_str.split(',') if id.strip()]
            else:
                dependent_ids = []

        if not dependent_ids:
            return

        # 각 하위 질문 처리
        for dep_id in dependent_ids:
            # 1. DB에서 기존 답변 확인
            cursor = conn.execute('''
                SELECT id FROM sb_disclosure_answers
                WHERE question_id = ? AND company_id = ? AND year = ?
            ''', (dep_id, company_id, year))
            existing = cursor.fetchone()

            if existing:
                # 기존 답변이 있으면 'N/A'로 업데이트하고 완료 처리
                conn.execute('''
                    UPDATE sb_disclosure_answers
                    SET value = 'N/A', status = 'completed', updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (existing['id'],))
            else:
                # 답변이 없으면 'N/A' 상태로 신규 생성하여 진행률에 반영
                ans_id = str(uuid.uuid4())
                conn.execute('''
                    INSERT INTO sb_disclosure_answers
                    (id, question_id, company_id, user_id, year, value, status)
                    VALUES (?, ?, ?, ?, ?, 'N/A', 'completed')
                ''', (ans_id, dep_id, company_id, user_id, year))
            
            # 2. 재귀적으로 해당 하위 질문의 하위 질문도 처리
            _mark_dependent_questions_na(conn, dep_id, company_id, year, user_id)

    except Exception as e:
        print(f"하위 질문 정리 오류: {e}")
        import traceback
        traceback.print_exc()


def _clear_na_from_dependents(conn, parent_question_id, company_id, year):
    """부모 질문이 다시 활성화('예' 등) 되었을 때, 기존에 자동 N/A 처리되었던 하위 질문 답변들만 삭제 (재귀적)"""
    try:
        cursor = conn.execute('SELECT dependent_question_ids FROM sb_disclosure_questions WHERE id = ?', (parent_question_id,))
        row = cursor.fetchone()
        if not row or not row['dependent_question_ids']:
            return

        dependent_ids = []
        try:
            dependent_ids = json.loads(row['dependent_question_ids'])
        except:
            return

        for dep_id in dependent_ids:
            # value가 'N/A'인 경우만 삭제 (사용자가 직접 입력했던 데이터는 보호)
            conn.execute('''
                DELETE FROM sb_disclosure_answers
                WHERE question_id = ? AND company_id = ? AND year = ? AND value = 'N/A'
            ''', (dep_id, company_id, year))
            
            _clear_na_from_dependents(conn, dep_id, company_id, year)

    except Exception as e:
        print(f"하위 질문 N/A 정리 오류: {e}")


def _update_session_progress(conn, company_id, year, user_id=None):
    """세션 진행률 자동 업데이트 (내부 함수)"""
    try:
        # 전체 질문 수 (Group 유형 제외)
        cursor = conn.execute("SELECT COUNT(*) as total FROM sb_disclosure_questions WHERE type != 'group'")
        total_result = cursor.fetchone()
        total = total_result['total'] if total_result else 0

        # 모든 질문 및 답변 조회
        cursor = conn.execute('SELECT * FROM sb_disclosure_questions ORDER BY category_id, sort_order')
        all_questions = [dict(row) for row in cursor.fetchall()]
        questions_dict = {q['id']: q for q in all_questions}

        cursor = conn.execute('''
            SELECT question_id, value FROM sb_disclosure_answers
            WHERE company_id = ? AND year = ? AND deleted_at IS NULL
        ''', (company_id, year))
        answers = {row['question_id']: row['value'] for row in cursor.fetchall()}

        # 완료된 질문 수 계산 (활성화된 질문의 답변 + 비활성화된 질문)
        answered = 0
        for q in all_questions:
            if q['type'] == 'group':
                continue

            is_active = _is_question_active(q, questions_dict, answers)
            if is_active:
                # 활성화된 질문: 답변이 있으면 완료
                if q['id'] in answers and answers[q['id']] not in (None, ''):
                    answered += 1
            else:
                # 비활성화 이유 구분:
                # - 상위 질문이 NO(건너뛰기) → 자동 완료
                # - 상위 질문이 미답변 → 미완료 (카운트 제외)
                if _is_question_skipped(q, questions_dict, answers):
                    answered += 1

        # 진행률 계산
        completion_rate = round((answered / total) * 100) if total > 0 else 0

        # 세션 존재 여부 확인
        cursor = conn.execute('''
            SELECT id FROM sb_disclosure_sessions
            WHERE company_id = ? AND year = ?
        ''', (company_id, year))
        existing = cursor.fetchone()

        if existing:
            # 세션 업데이트
            conn.execute('''
                UPDATE sb_disclosure_sessions
                SET total_questions = ?, answered_questions = ?, completion_rate = ?, updated_at = CURRENT_TIMESTAMP,
                    status = CASE WHEN ? = 100 THEN 'completed' ELSE 'in_progress' END
                WHERE company_id = ? AND year = ?
            ''', (total, answered, completion_rate, completion_rate, company_id, year))
        else:
            # 세션 자동 생성
            session_id = generate_uuid()
            status = 'completed' if completion_rate == 100 else ('in_progress' if answered > 0 else 'draft')
            conn.execute('''
                INSERT INTO sb_disclosure_sessions
                (id, company_id, user_id, year, status, total_questions, answered_questions, completion_rate)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (session_id, company_id, user_id or '', year, status, total, answered, completion_rate))

        conn.commit()

    except Exception as e:
        print(f"진행률 업데이트 오류: {e}")
        import traceback
        traceback.print_exc()


# ============================================================================
# API Endpoints - 보고서 생성
# ============================================================================

@bp_link11.route('/link11/api/report/generate', methods=['POST'])
def generate_report():
    """공시 자료 보고서 생성"""
    try:
        data = request.get_json()
        user_info = get_user_info()

        # 보안: 클라이언트 요청 company_id 무시, 로그인 사용자 소속 회사 사용
        company_id = get_validated_company_id(user_info, data.get('company_id'))
        year = data.get('year', datetime.now().year)
        format_type = data.get('format', 'json')  # json, pdf, excel

        with get_db() as conn:
            # 모든 답변 조회
            cursor = conn.execute('''
                SELECT q.*, a.value, a.status
                FROM sb_disclosure_questions q
                LEFT JOIN sb_disclosure_answers a ON q.id = a.question_id
                    AND a.company_id = ? AND a.year = ? AND a.deleted_at IS NULL
                ORDER BY q.category_id, q.sort_order
            ''', (company_id, year))

            report_data = {
                'company_id': company_id,
                'year': year,
                'generated_at': datetime.now().isoformat(),
                'categories': {}
            }

            # 활동성 체크를 위한 사전 준비
            all_rows = [dict(row) for row in cursor.fetchall()]
            questions_dict = {q['id']: q for q in all_rows}
            answers_dict = {q['id']: q['value'] for q in all_rows if q['value'] is not None}

            for row in all_rows:
                category = row['category']
                if category not in report_data['categories']:
                    report_data['categories'][category] = {
                        'id': row['category_id'],
                        'questions': [],
                        'completed': 0,
                        'total': 0
                    }

                # 활성화 상태 확인
                is_active = _is_question_active(row, questions_dict, answers_dict)

                q_data = {
                    'id': row['id'],
                    'text': row['text'],
                    'type': row['type'],
                    'value': row['value'],
                    'status': row['status'] or 'pending',
                    'is_active': is_active
                }

                if not is_active:
                    # 숫자 유형이거나 Q3(합계)인 경우 '0'으로 표시, 그 외엔 '해당 없음'
                    if row['type'] == 'number' or row['id'] == QID.INV_SEC_GROUP:
                        q_data['value'] = '0'
                    else:
                        q_data['value'] = '해당 없음'
                    q_data['status'] = 'completed'
                elif row['id'] == QID.INV_SEC_GROUP:
                    # Q3(group)는 저장값 없이 Q4+Q5+Q6 합산으로 표시
                    try:
                        b_sum = sum(
                            float(str(answers_dict.get(qid, 0) or 0).replace(',', ''))
                            for qid in [QID.INV_SEC_DEPRECIATION, QID.INV_SEC_SERVICE, QID.INV_SEC_LABOR]
                        )
                        q_data['value'] = int(b_sum) if b_sum == int(b_sum) else b_sum
                        q_data['status'] = 'completed' if b_sum > 0 else 'pending'
                    except (ValueError, TypeError):
                        pass
                else:
                    # JSON 값 파싱
                    if q_data['value']:
                        try:
                            q_data['value'] = json.loads(q_data['value'])
                        except (json.JSONDecodeError, TypeError):
                            pass

                report_data['categories'][category]['questions'].append(q_data)
                report_data['categories'][category]['total'] += 1
                if q_data['status'] == 'completed':
                    report_data['categories'][category]['completed'] += 1

            if format_type == 'json':
                return jsonify({
                    'success': True,
                    'report': report_data
                })
            else:
                # PDF/Excel 생성은 추후 구현
                return jsonify({
                    'success': False,
                    'message': f'{format_type} 형식은 아직 지원되지 않습니다.'
                }), 501

    except Exception as e:
        print(f"보고서 생성 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/report/download', methods=['POST'])
def download_report():
    """공시 자료 Excel 다운로드"""
    try:
        data = request.get_json()
        user_info = get_user_info()

        # 보안: 클라이언트 요청 company_id 무시, 로그인 사용자 소속 회사 사용
        company_id = get_validated_company_id(user_info, data.get('company_id'))
        year = data.get('year', datetime.now().year)
        format_type = data.get('format', 'excel')
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from io import BytesIO

        with get_db() as conn:
            # 모든 답변 조회 (parent_question_id 포함하여 모든 컬럼 조회)
            cursor = conn.execute('''
                SELECT q.*, a.value, a.status
                FROM sb_disclosure_questions q
                LEFT JOIN sb_disclosure_answers a ON q.id = a.question_id
                    AND a.company_id = ? AND a.year = ? AND a.deleted_at IS NULL
                ORDER BY q.category_id, q.sort_order
            ''', (company_id, year))

            questions_flat = [dict(row) for row in cursor.fetchall()]

        # 활동성 체크를 위한 사전 준비
        questions_dict = {q['id']: q for q in questions_flat}
        answers_dict = {q['id']: q['value'] for q in questions_flat if q['value'] is not None}

        # Excel 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "정보보호공시"

        # 스타일 정의
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        category_font = Font(bold=True, size=11)
        category_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 제목
        ws.merge_cells('A1:C1')
        ws['A1'] = f"정보보호공시 - {company_id} ({year}년)"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:C2')
        ws['A2'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # 헤더
        headers = ['ID', '질문(항목)', '공시 내용']
        ws.append([])  # 빈 줄
        ws.append(headers)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 데이터
        current_category = None
        row_num = 5

        for q in questions_flat:
            # 카테고리 헤더
            if q['category'] != current_category:
                current_category = q['category']
                ws.merge_cells(f'A{row_num}:C{row_num}')
                ws[f'A{row_num}'] = current_category
                ws[f'A{row_num}'].font = category_font
                ws[f'A{row_num}'].fill = category_fill
                ws[f'A{row_num}'].border = thin_border
                row_num += 1

            # 활성화 상태 확인
            is_active = _is_question_active(q, questions_dict, answers_dict)
            
            if not is_active:
                # 숫자 유형이거나 Q3(합계)인 경우 '0'으로 표시, 그 외엔 '해당 없음'
                if q.get('type') == 'number' or q['id'] == QID.INV_SEC_GROUP:
                    value = '0'
                else:
                    value = '해당 없음'
            elif q['id'] == QID.INV_SEC_GROUP:
                # Q3(group)는 저장값 없이 Q4+Q5+Q6 합산으로 표시
                try:
                    b_sum = sum(
                        float(str(answers_dict.get(qid, 0) or 0).replace(',', ''))
                        for qid in [QID.INV_SEC_DEPRECIATION, QID.INV_SEC_SERVICE, QID.INV_SEC_LABOR]
                    )
                    value = "{:,}".format(int(b_sum)) if b_sum == int(b_sum) else "{:,.2f}".format(b_sum)
                except (ValueError, TypeError):
                    value = ''
            else:
                # 질문 데이터 타입에 따른 포맷팅
                raw_value = q['value'] or ''
                q_type = q.get('type')
                
                if not raw_value:
                    value = ''
                elif q_type == 'number':
                    try:
                        # 숫자인 경우 천단위 구분기호 추가
                        num_val = float(str(raw_value).replace(',', ''))
                        if num_val == int(num_val):
                            value = "{:,}".format(int(num_val))
                        else:
                            value = "{:,.2f}".format(num_val)
                    except:
                        value = raw_value
                elif q_type == 'yes_no':
                    if str(raw_value).upper() == 'YES':
                        value = '예'
                    elif str(raw_value).upper() == 'NO':
                        value = '아니요'
                    else:
                        value = raw_value
                elif raw_value.startswith('[') and raw_value.endswith(']'):
                    try:
                        data = json.loads(raw_value)
                        if isinstance(data, list):
                            if all(isinstance(item, list) for item in data):
                                # 테이블 형태 (리스트의 리스트)
                                value = '\n'.join([', '.join(map(str, row)) for row in data])
                            elif all(isinstance(item, dict) for item in data):
                                # 테이블 형태 (리스트의 딕셔너리)
                                value = '\n'.join([', '.join([f"{v}" for v in item.values()]) for item in data])
                            else:
                                # 일반 리스트 (체크박스 등)
                                value = ', '.join(map(str, data))
                    except:
                        value = raw_value
                elif raw_value.startswith('{') and raw_value.endswith('}'):
                    try:
                        comp = json.loads(raw_value)
                        if isinstance(comp, dict):
                            val_parts = []
                            total = 0
                            for k, v in comp.items():
                                val_parts.append(f"{k}: {v}")
                                try: total += float(str(v).replace(',', ''))
                                except: pass
                            
                            # 인력 현황 등의 경우 총계 표시 (질문 텍스트에 따라 유동적 가능하나 여기선 일반화)
                            if '인원' in q['text'] or '인력' in q['text']:
                                value = f"{', '.join(val_parts)} (총 {total:.1f}명)" if total % 1 != 0 else f"{', '.join(val_parts)} (총 {int(total)}명)"
                            else:
                                value = '\n'.join(val_parts)
                    except:
                        value = raw_value
                else:
                    value = raw_value

            ws.append([q['id'], q['text'], value])

            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                
                # 수직 정렬 공통 적용
                align_params = {'vertical': 'center', 'wrap_text': True}
                
                # 3번째 열(공시 내용)이면서 숫자 유형인 경우 오른쪽 정렬
                if col == 3 and q.get('type') == 'number':
                    align_params['horizontal'] = 'right'
                elif col == 1:
                    align_params['horizontal'] = 'center'
                
                cell.alignment = Alignment(**align_params)

            row_num += 1

        # 열 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 45

        # ============================================
        # 증빙자료 시트 추가 (내부 관리용)
        # ============================================
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT e.id, e.question_id, e.file_name, e.file_size, e.file_type,
                       e.evidence_type, e.uploaded_at, e.uploaded_by,
                       q.category, q.text as question_text
                FROM sb_disclosure_evidence e
                LEFT JOIN sb_disclosure_questions q ON e.question_id = q.id
                WHERE e.company_id = ? AND e.year = ?
                ORDER BY q.category, e.uploaded_at
            ''', (company_id, year))
            evidence_list = [dict(row) for row in cursor.fetchall()]

        if evidence_list:
            ws_evidence = wb.create_sheet(title="증빙자료")

            # 제목
            ws_evidence.merge_cells('A1:F1')
            ws_evidence['A1'] = f"증빙자료 목록 - {company_id} ({year}년)"
            ws_evidence['A1'].font = Font(bold=True, size=16)
            ws_evidence['A1'].alignment = Alignment(horizontal='center')

            ws_evidence.merge_cells('A2:F2')
            ws_evidence['A2'] = f"총 {len(evidence_list)}개 파일"
            ws_evidence['A2'].alignment = Alignment(horizontal='center')

            # 헤더
            evidence_headers = ['카테고리', '관련 질문', '파일명', '파일크기', '유형', '업로드일']
            ws_evidence.append([])  # 빈 줄
            ws_evidence.append(evidence_headers)

            for col, header in enumerate(evidence_headers, 1):
                cell = ws_evidence.cell(row=4, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # 데이터
            row_num = 5
            for ev in evidence_list:
                # 파일 크기 포맷
                size = ev['file_size'] or 0
                if size < 1024:
                    size_str = f"{size} B"
                elif size < 1024 * 1024:
                    size_str = f"{size / 1024:.1f} KB"
                else:
                    size_str = f"{size / (1024 * 1024):.1f} MB"

                # 업로드일 포맷
                uploaded_at = ev['uploaded_at'] or ''
                if uploaded_at and len(uploaded_at) >= 10:
                    uploaded_at = uploaded_at[:10]  # YYYY-MM-DD만

                ws_evidence.append([
                    ev['category'] or '',
                    ev['question_text'] or '',
                    ev['file_name'] or '',
                    size_str,
                    ev['evidence_type'] or '',
                    uploaded_at
                ])

                for col in range(1, 7):
                    cell = ws_evidence.cell(row=row_num, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

                row_num += 1

            # 열 너비 조정
            ws_evidence.column_dimensions['A'].width = 20
            ws_evidence.column_dimensions['B'].width = 50
            ws_evidence.column_dimensions['C'].width = 35
            ws_evidence.column_dimensions['D'].width = 12
            ws_evidence.column_dimensions['E'].width = 15
            ws_evidence.column_dimensions['F'].width = 12

        # 파일 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'정보보호공시_{company_id}_{year}.xlsx'
        )

    except ImportError:
        return jsonify({'success': False, 'message': 'openpyxl 라이브러리가 필요합니다.'}), 500
    except Exception as e:
        print(f"보고서 다운로드 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 공시 제출
# ============================================================================

@bp_link11.route('/link11/api/submit/<int:user_id>/<int:year>', methods=['POST'])
def submit_disclosure(user_id, year):
    """공시 제출"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        data = request.get_json() or {}
        user_info = get_user_info()

        company_id = get_company_name_by_user_id(user_id)
        submission_details = data.get('details', '')

        with get_db() as conn:
            # 세션 조회
            cursor = conn.execute('''
                SELECT * FROM sb_disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))
            session_data = cursor.fetchone()

            if not session_data:
                return jsonify({'success': False, 'message': '공시 세션을 찾을 수 없습니다.'}), 404

            # 완료율 확인 (필수 질문 모두 완료 여부)
            if session_data['completion_rate'] < 100:
                return jsonify({
                    'success': False,
                    'message': f'모든 질문에 답변해야 제출할 수 있습니다. (현재 {session_data["completion_rate"]}%)'
                }), 400

            # 제출 기록 생성
            submission_id = generate_uuid()
            confirmation_number = f"DISC-{year}-{datetime.now().strftime('%Y%m%d%H%M%S')}"

            conn.execute('''
                INSERT INTO sb_disclosure_submissions
                (id, session_id, company_id, year, submitted_by, submission_details, confirmation_number, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'submitted')
            ''', (
                submission_id, session_data['id'], company_id, year,
                str(user_info.get('user_id', '')), submission_details, confirmation_number
            ))

            # 세션 상태 업데이트
            conn.execute('''
                UPDATE sb_disclosure_sessions
                SET status = 'submitted', submitted_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (session_data['id'],))

            conn.commit()

            return jsonify({
                'success': True,
                'message': '공시가 제출되었습니다.',
                'submission_id': submission_id,
                'confirmation_number': confirmation_number
            })

    except Exception as e:
        print(f"공시 제출 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/confirm/<int:user_id>/<int:year>', methods=['POST'])
def confirm_disclosure(user_id, year):
    """공시 최종 확정 (submitted → confirmed)"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401
    try:
        user_info = get_user_info()
        company_id = get_validated_company_id(user_info, None)
        with get_db() as conn:
            sess = conn.execute(
                'SELECT id, status FROM sb_disclosure_sessions WHERE company_id = ? AND year = ?',
                (company_id, year)
            ).fetchone()
            if not sess:
                return jsonify({'success': False, 'message': '세션이 존재하지 않습니다.'}), 404
            if sess['status'] != 'submitted':
                return jsonify({'success': False, 'message': f"제출 완료 상태에서만 확정 가능합니다. (현재: {sess['status']})"}), 400
            conn.execute(
                "UPDATE sb_disclosure_sessions SET status = 'confirmed' WHERE id = ?",
                (sess['id'],)
            )
            conn.commit()
        return jsonify({'success': True, 'message': '공시가 최종 확정되었습니다.'})
    except Exception as e:
        print(f"공시 확정 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/unconfirm/<int:user_id>/<int:year>', methods=['POST'])
def unconfirm_disclosure(user_id, year):
    """공시 확정 취소 (confirmed → submitted) — 관리자 전용"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401
    try:
        user_info = get_user_info()
        company_id = get_validated_company_id(user_info, None)
        with get_db() as conn:
            sess = conn.execute(
                'SELECT id, status FROM sb_disclosure_sessions WHERE company_id = ? AND year = ?',
                (company_id, year)
            ).fetchone()
            if not sess or sess['status'] != 'confirmed':
                return jsonify({'success': False, 'message': '확정 상태가 아닙니다.'}), 400
            conn.execute(
                "UPDATE sb_disclosure_sessions SET status = 'submitted' WHERE id = ?",
                (sess['id'],)
            )
            conn.commit()
        return jsonify({'success': True, 'message': '공시 확정이 취소되었습니다.'})
    except Exception as e:
        print(f"공시 확정 취소 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/submissions/<int:user_id>', methods=['GET'])
def get_submissions(user_id):
    """제출 이력 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        with get_db() as conn:
            cursor = conn.execute('''
                SELECT * FROM sb_disclosure_submissions
                WHERE company_id = ?
                ORDER BY submitted_at DESC
            ''', (company_id,))

            submissions = [dict(row) for row in cursor.fetchall()]

            return jsonify({
                'success': True,
                'submissions': submissions,
                'count': len(submissions)
            })

    except Exception as e:
        print(f"제출 이력 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API Endpoints - 데이터 관리 (새로하기, 이전 자료 불러오기)
# ============================================================================

@bp_link11.route('/link11/api/reset/<int:user_id>/<int:year>', methods=['POST'])
def reset_disclosure(user_id, year):
    """새로하기 - 현재 연도 데이터 초기화"""
    if not is_logged_in():
        return jsonify({'success': False, 'message': '로그인이 필요합니다.'}), 401

    try:
        company_id = get_company_name_by_user_id(user_id)

        with get_db() as conn:
            # 답변 삭제
            conn.execute('''
                DELETE FROM sb_disclosure_answers
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))

            # 세션 삭제
            conn.execute('''
                DELETE FROM sb_disclosure_sessions
                WHERE company_id = ? AND year = ?
            ''', (company_id, year))

            # 증빙자료는 유지 (필요시 별도 삭제)

            conn.commit()

            return jsonify({
                'success': True,
                'message': f'{year}년 데이터가 초기화되었습니다.'
            })

    except Exception as e:
        print(f"데이터 초기화 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/available-years/<int:user_id>', methods=['GET'])
def get_available_years(user_id):
    """이전 자료가 있는 연도 목록 조회"""
    try:
        company_id = get_company_name_by_user_id(user_id)
        current_year = datetime.now().year

        with get_db() as conn:
            cursor = conn.execute('''
                SELECT DISTINCT a.year, COUNT(*) as answer_count, s.status
                FROM sb_disclosure_answers a
                LEFT JOIN sb_disclosure_sessions s
                    ON a.company_id = s.company_id AND a.year = s.year
                WHERE a.company_id = ? AND a.year < ?
                GROUP BY a.year
                ORDER BY a.year DESC
            ''', (company_id, current_year))

            years = []
            for row in cursor.fetchall():
                years.append({
                    'year': row['year'],
                    'answer_count': row['answer_count'],
                    'status': row['status'] or 'draft'
                })

            return jsonify({
                'success': True,
                'years': years
            })

    except Exception as e:
        print(f"연도 목록 조회 오류: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link11.route('/link11/api/copy-from-year/<int:user_id>/<int:source_year>/<int:target_year>', methods=['POST'])
def copy_from_year(user_id, source_year, target_year):
    """이전 자료 불러오기 - 데이터 무결성을 위해 복사 기능을 비활성화하고 가이드 메시지 반환"""
    return jsonify({
        'success': False,
        'message': '데이터 무결성과 감사 품질을 위해 직접 복사 기능은 더 이상 지원되지 않습니다. 우측 상단의 [전년도 참고] 패널을 활용하여 내용을 직접 검토 후 입력해 주세요.'
    }), 403
