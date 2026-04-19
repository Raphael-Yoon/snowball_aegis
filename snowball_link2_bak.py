from datetime import datetime
import os
import re
import json
import tempfile
import uuid
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openai import OpenAI
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time
from requests.exceptions import RequestException
import socket
from flask import Blueprint, render_template, request, session, redirect, url_for, jsonify
from auth import log_user_activity, login_required, increment_ai_review_count
from snowball_mail import send_gmail_with_attachment

# Blueprint 생성
bp_link2 = Blueprint('link2', __name__)

# ================================
# Helper Functions
# ================================

def is_logged_in():
    """로그인 상태 확인"""
    return 'user_id' in session

def get_user_info():
    """현재 로그인한 사용자 정보 반환 (세션 우선)"""
    if is_logged_in():
        if 'user_info' in session:
            return session['user_info']
        try:
            from auth import get_current_user
            return get_current_user()
        except ImportError:
            pass
    return None

def reset_interview_session():
    """인터뷰 관련 세션만 초기화 (로그인 세션은 보존)"""
    # 인터뷰 관련 키만 제거
    interview_keys = ['question_index', 'answer', 'textarea_answer', 'System', 'Cloud', 'OS_Tool', 'DB_Tool', 'Batch_Tool']
    for key in interview_keys:
        session.pop(key, None)
    
    # 인터뷰 세션 재초기화
    user_info = get_user_info()
    if user_info and user_info.get('user_email'):
        user_email = user_info.get('user_email')
        # 로그인된 사용자: 첫 번째 질문에 이메일 자동 입력하고 두 번째 질문부터 시작
        session['question_index'] = 1
        session['answer'] = [''] * question_count
        session['textarea_answer'] = [''] * question_count
        session['answer'][0] = user_email  # 첫 번째 답변에 이메일 설정
    else:
        # 비로그인 사용자: 첫 번째 질문부터 시작
        session['question_index'] = START_QUESTION - 1 if 1 <= START_QUESTION <= question_count else 0
        session['answer'] = [''] * question_count
        session['textarea_answer'] = [''] * question_count
    

# 시작할 질문 번호 설정 (1부터 시작)
START_QUESTION = 0

# ================================
# AI 문장 다듬기 설정 (수기 조정 가능)
# ================================

# ===========================================
# 📝 AI 문장 다듬기 규칙 설정 (여기서 수정하세요!)
# ===========================================

# AI 프롬프트 템플릿 (토큰 절약을 위해 간소화)
AI_REFINEMENT_PROMPT = """문법교정만 하세요. 내용변경금지.

{answer_text}

오타수정, 문체통일, "답변:" 등 제거"""

# OpenAI 모델 설정
AI_MODEL_CONFIG = {
    'model': 'gpt-4o-mini',     # 사용할 모델
    'max_tokens': 800,          # 최대 토큰 수
    'temperature': 0.0          # 창의성 수준 (0.0=완전 일관적, 1.0=다양함)
}

# API Rate Limiting 설정 (안정성 우선)
API_RATE_LIMIT = {
    'requests_per_minute': 300,     # OpenAI 유료 티어 기준
    'delay_between_requests': 0.2   # 60초 / 300회 = 0.2초
}

# API 호출 추적용 전역 변수
_last_api_call_time = 0
_api_call_lock = threading.Lock()

# 텍스트 길이 제한 (토큰 절약)
TEXT_LENGTH_LIMITS = {
    'min_length': 20,        # 이보다 짧으면 AI 다듬기 건너뜀 (Summary 시트 포함을 위해 낮춤)
    'max_length': 2000,      # 이보다 길면 AI 다듬기 건너뜀
}

# 제거할 접두사 목록 (AI 응답에서 자동 제거)
PREFIXES_TO_REMOVE = [
    '답변:', '결과:', '개선된 답변:', '수정된 답변:', '다듬어진 답변:', '교정된 답변:',
    '중요한 규칙:', '지침:', '교정된 텍스트:', '수정된 내용:', '개선 결과:'
]

# 제거할 불필요한 문구 목록 (AI 응답 중간에 나타날 수 있는 메타 정보)
UNWANTED_PHRASES = [
    '중요한 규칙:', '지침:', '다음은 교정된 내용입니다:', 
    '교정된 텍스트는 다음과 같습니다:', '개선된 버전:'
]

# 자동 문단 나누기 설정
AUTO_PARAGRAPH_BREAK = {
    'enable_sentence_break': True,   # 마침표 뒤 자동 줄바꿈 (예: "있습니다. 새로운" → "있습니다.\n\n새로운")
    'enable_phrase_break': True,     # "아래와 같습니다" 뒤 자동 줄바꿈
}

# 추가 텍스트 처리 규칙
TEXT_PROCESSING_RULES = {
    'remove_double_spaces': True,    # 이중 공백 제거
    'unify_punctuation': True,       # 문장부호 통일 (예: "。" → ".")
    'normalize_line_breaks': True,   # 줄바꿈 정규화
}

# ================================
# 진행률 관리 설정 (인터뷰 처리용)
# ================================

# WSGI 환경에서 안전한 경로 사용
PROGRESS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'progress_data')
if not os.path.exists(PROGRESS_DIR):
    try:
        os.makedirs(PROGRESS_DIR, exist_ok=True)
    except Exception as e:
        # 시스템 임시 디렉토리를 대안으로 사용
        PROGRESS_DIR = os.path.join(tempfile.gettempdir(), 'snowball_progress')
        os.makedirs(PROGRESS_DIR, exist_ok=True)

def get_progress_status(task_id):
    """파일에서 진행률 상태 읽기"""
    progress_file = os.path.join(PROGRESS_DIR, f"{task_id}.progress")
    if os.path.exists(progress_file):
        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data
        except (IOError, json.JSONDecodeError) as e:
            pass
    return {
        'percentage': 0,
        'current_task': 'AI 검토를 준비하고 있습니다...',
        'is_processing': True # 클라이언트가 폴링을 계속하도록 설정
    }

def set_progress_status(task_id, status):
    """파일에 진행률 상태 저장"""
    progress_file = os.path.join(PROGRESS_DIR, f"{task_id}.progress")
    try:
        # 임시 파일에 먼저 쓰고 원자적으로 이동 (WSGI 환경에서 안전)
        temp_file = progress_file + '.tmp'
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(status, f, ensure_ascii=False, indent=2)
            f.flush()  # 버퍼를 즉시 플러시
            os.fsync(f.fileno())  # 디스크에 강제로 쓰기
        
        # 원자적으로 파일 이동
        os.replace(temp_file, progress_file)
    except IOError as e:
        # 임시 파일이 남아있으면 삭제
        try:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
        except:
            pass

def update_progress(task_id, percentage, task_description):
    """진행률 업데이트 함수 (파일 기반)"""
    if task_id is None:
        return
    
    status = {
        'percentage': percentage,
        'current_task': task_description,
        'is_processing': percentage < 100,
        'timestamp': datetime.now().isoformat()  # 타임스탬프 추가
    }
    set_progress_status(task_id, status)

def reset_progress(task_id):
    """진행률 파일 삭제 함수"""
    progress_file = os.path.join(PROGRESS_DIR, f"{task_id}.progress")
    try:
        if os.path.exists(progress_file):
            os.unlink(progress_file)
    except IOError as e:
        pass

# ITGC 통제 정의 (반복 코드 제거를 위한 데이터 구조)
ITGC_CONTROLS = {
    'APD01': {
        'title': '사용자 신규 권한 승인',
        'template': '사용자 권한 부여 이력이 시스템에 {history_status}\n\n{procedure_text}',
        'history_idx': 6,
        'procedure_idx': 10,
        'textarea_idx': 10,
        'history_yes': '기록되고 있어 모집단 확보가 가능합니다.',
        'history_no': '기록되지 않아 모집단 확보가 불가합니다.',
        'procedure_prefix': '새로운 권한 요청 시, 요청서를 작성하고 부서장의 승인을 득하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '새로운 권한 요청 시 승인 절차가 없습니다.',
        'default_msg': '권한 부여 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD02': {
        'title': '부서이동자 권한 회수',
        'template': '사용자 권한 회수 이력이 시스템에 {history_status}\n\n{procedure_text}',
        'history_idx': 7,
        'procedure_idx': 11,
        'textarea_idx': 11,
        'history_yes': '기록되고 있습니다.',
        'history_no': '기록되지 않습니다.',
        'procedure_prefix': '부서 이동 시 기존 권한을 회수하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '부서 이동 시 기존 권한 회수 절차가 없습니다.',
        'default_msg': '부서 이동 시 권한 회수 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD03': {
        'title': '퇴사자 접근권한 회수',
        'type': 'simple_procedure',
        'procedure_idx': 12,
        'textarea_idx': 12,
        'procedure_prefix': '퇴사자 발생 시 접근권한을 차단하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '퇴사자 발생 시 접근권한을 차단 절차가 없습니다.',
        'default_msg': '퇴사자 접근권한 차단 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD04': {
        'title': 'Application 관리자 권한 제한',
        'type': 'simple_list',
        'template': 'Application 관리자 권한을 보유한 인원은 아래와 같습니다.\n\n{content}',
        'answer_idx': 13,
        'default_msg': 'Application 관리자 권한을 보유한 인원에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD05': {
        'title': '사용자 권한 Monitoring',
        'type': 'simple_status',
        'template': '전체 사용자가 보유한 권한에 대한 적절성을 {status}',
        'answer_idx': 14,
        'status_yes': '모니터링하는 절차가 있습니다.',
        'status_no': '모니터링 절차가 존재하지 않습니다.'
    },
    'APD06': {
        'title': 'Application 패스워드',
        'type': 'simple_list',
        'template': '패스워드 설정 사항은 아래와 같습니다.\n\n{content}',
        'answer_idx': 15,
        'default_msg': '패스워드 설정 사항에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD07': {
        'title': '데이터 직접 변경',
        'template': '데이터 변경 이력이 시스템에 {history_status}\n\n{procedure_text}',
        'history_idx': 17,
        'procedure_idx': 18,
        'textarea_idx': 18,
        'history_yes': '기록되고 있어 모집단 확보가 가능합니다.',
        'history_no': '기록되지 않아 모집단 확보가 불가합니다.',
        'procedure_prefix': '데이터 변경이 필요한 경우 요청서를 작성하고 부서장의 승인을 득하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '데이터 변경 시 승인 절차가 없습니다.',
        'default_msg': '데이터 변경 승인 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD08': {
        'title': '데이터 변경 권한 제한',
        'type': 'simple_list',
        'template': '데이터 변경 권한을 보유한 인원은 아래와 같습니다.\n\n{content}',
        'answer_idx': 19,
        'default_msg': '데이터 변경 권한을 보유한 인원에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD09': {
        'title': 'DB 접근권한 승인',
        'type': 'complex_db',
        'db_type_idx': 20,
        'db_tool_idx': 21,
        'history_idx': 22,
        'procedure_idx': 23,
        'textarea_idx': 23,
        'history_yes': '기록되고 있어 모집단 확보가 가능합니다.',
        'history_no': '기록되지 않아 모집단 확보가 불가합니다.',
        'procedure_prefix': 'DB 접근권한이 필요한 경우 요청서를 작성하고 부서장의 승인을 득하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': 'DB 접근권한 요청 시 승인 절차가 없습니다.',
        'default_msg': 'DB 접근권한 승인 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD10': {
        'title': 'DB 관리자 권한 제한',
        'type': 'simple_list',
        'template': 'DB 관리자 권한을 보유한 인원은 아래와 같습니다.\n\n{content}',
        'answer_idx': 24,
        'default_msg': 'DB 관리자 권한을 보유한 인원에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD11': {
        'title': 'DB 패스워드',
        'type': 'simple_list',
        'template': 'DB 패스워드 설정사항은 아래와 같습니다.\n\n{content}',
        'answer_idx': 25,
        'default_msg': 'DB 패스워드 설정 사항에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD12': {
        'title': 'OS 접근권한 승인',
        'type': 'complex_os',
        'os_type_idx': 27,
        'os_tool_idx': 28,
        'history_idx': 29,
        'procedure_idx': 30,
        'textarea_idx': 30,
        'history_yes': '기록되고 있어 모집단 확보가 가능합니다.',
        'history_no': '기록되지 않아 모집단 확보가 불가합니다.',
        'procedure_prefix': 'OS 접근권한이 필요한 경우 요청서를 작성하고 부서장의 승인을 득하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': 'OS 접근권한 요청 시 승인 절차가 없습니다.',
        'default_msg': 'OS 접근권한 승인 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD13': {
        'title': 'OS 관리자 권한 제한',
        'type': 'simple_list',
        'template': 'OS 관리자 권한을 보유한 인원은 아래와 같습니다.\n\n{content}',
        'answer_idx': 31,
        'default_msg': 'OS 관리자 권한을 보유한 인원에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD14': {
        'title': 'OS 패스워드',
        'type': 'simple_list',
        'template': 'OS 패스워드 설정사항은 아래와 같습니다.\n\n{content}',
        'answer_idx': 32,
        'default_msg': 'OS 패스워드 설정 사항에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'APD15': {
        'title': '사용자 ID 개인별 발급 여부',
        'type': 'simple_status',
        'template': '시스템 사용자 ID는 {status}',
        'answer_idx': 8,
        'status_yes': '사용자 개인별로 발급되고 있습니다.',
        'status_no': '팀·부서 단위 공용으로 발급되어, 일부 사용자가 동일한 ID를 공동으로 사용하고 있습니다.',
    },
    'APD16': {
        'title': '외부 접속(VPN) 관리',
        'type': 'simple_yn_tool',
        'answer_idx': 33,
        'textarea_idx': 33,
        'template_yes': '외부 원격 접속 시 VPN 등의 접근 통제 수단을 사용하고 있습니다.\n\n사용 솔루션: {tool_name}',
        'template_no': '외부 원격 접속 시 별도의 VPN 등 접근 통제 수단을 사용하고 있지 않습니다.',
    },
    'PC01': {
        'title': '프로그램 변경 승인',
        'template': '프로그램 변경 이력이 시스템에 {history_status}\n\n{procedure_text}',
        'history_idx': 35,
        'procedure_idx': 36,
        'textarea_idx': 36,
        'history_yes': '기록되고 있어 모집단 확보가 가능합니다.',
        'history_no': '기록되지 않아 모집단 확보가 불가합니다.',
        'procedure_prefix': '프로그램 변경 시 요청서를 작성하고 부서장의 승인을 득하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '프로그램 변경 시 승인 절차가 없습니다.',
        'default_msg': '프로그램 변경 승인 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'PC02': {
        'title': '프로그램 변경 사용자 테스트',
        'type': 'simple_procedure',
        'procedure_idx': 37,
        'textarea_idx': 37,
        'procedure_prefix': '프로그램 변경 시 사용자 테스트를 수행하고 그 결과를 문서화하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '프로그램 변경 시 사용자 테스트를 수행하지 않습니다.',
        'default_msg': '프로그램 변경 사용자 테스트 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'PC03': {
        'title': '프로그램 변경 이관 승인',
        'type': 'simple_procedure',
        'procedure_idx': 38,
        'textarea_idx': 38,
        'procedure_prefix': '프로그램 변경 완료 후 이관(배포)을 위해 부서장 등의 승인을 득하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '프로그램 변경 완료 후 이관(배포) 절차가 없습니다.',
        'default_msg': '프로그램 변경 이관 승인 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'PC04': {
        'title': '이관(배포) 권한 제한',
        'type': 'simple_list',
        'template': '이관(배포) 권한을 보유한 인원은 아래와 같습니다.\n\n{content}',
        'answer_idx': 39,
        'default_msg': '이관(배포) 권한을 보유한 인원에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'PC05': {
        'title': '개발/운영 환경 분리',
        'type': 'simple_status',
        'template': '운영서버 외 별도의 개발 또는 테스트 서버를 {status}',
        'answer_idx': 40,
        'status_yes': '운용하고 있습니다.',
        'status_no': '운용하지 않습니다.'
    },
    'CO01': {
        'title': '배치 스케줄 등록/변경 승인',
        'template': '배치 스케줄 등록/변경 이력이 시스템에 {history_status}\n\n{procedure_text}',
        'history_idx': 43,
        'procedure_idx': 44,
        'textarea_idx': 44,
        'history_yes': '기록되고 있어 모집단 확보가 가능합니다.',
        'history_no': '기록되지 않아 모집단 확보가 불가합니다.',
        'procedure_prefix': '배치 스케줄 등록/변경 시 요청서를 작성하고 부서장 등의 승인을 득하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': '배치 스케줄 등록/변경 시 승인 절차가 없습니다.',
        'default_msg': '배치 스케줄 등록/변경 승인 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'CO02': {
        'title': '배치 스케줄 등록/변경 권한 제한',
        'type': 'simple_list',
        'template': '배치 스케줄 등록/변경 권한을 보유한 인원은 아래와 같습니다.\n\n{content}',
        'answer_idx': 45,
        'default_msg': '배치 스케줄 등록/변경 권한을 보유한 인원에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'CO03': {
        'title': '배치 실행 모니터링',
        'type': 'simple_list',
        'template': '배치 실행 오류 등에 대한 모니터링은 아래와 같이 수행되고 있습니다.\n\n{content}',
        'answer_idx': 46,
        'default_msg': '배치 실행 오류 등에 대한 모니터링 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'CO04': {
        'title': '장애 대응 절차',
        'type': 'simple_list',
        'template': '장애 발생시 이에 대응하고 조치하는 절차는 아래와 같습니다.\n\n{content}',
        'answer_idx': 47,
        'default_msg': '장애 발생시 이에 대응하고 조치하는 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'CO05': {
        'title': '백업 및 모니터링',
        'type': 'simple_list',
        'template': '백업 수행 및 모니터링 절차는 아래와 같습니다.\n\n{content}',
        'answer_idx': 48,
        'default_msg': '백업 수행 및 모니터링 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'CO06': {
        'title': '서버실 출입 절차',
        'type': 'simple_list',
        'template': '서버실 출입 절차는 아래와 같습니다.\n\n{content}',
        'answer_idx': 49,
        'default_msg': '서버실 출입 절차에 대한 상세 기술이 제공되지 않았습니다.'
    },
    'CO07': {
        'title': '보안 패치·취약점 관리',
        'type': 'simple_procedure',
        'procedure_idx': 50,
        'textarea_idx': 50,
        'procedure_prefix': 'OS, DB 등 시스템 소프트웨어에 대한 보안 패치를 주기적으로 적용하는 절차가 있으며 그 절차는 아래와 같습니다.',
        'procedure_no': 'OS, DB 등 시스템 소프트웨어에 대한 보안 패치를 주기적으로 적용하는 절차가 없습니다.',
        'default_msg': '보안 패치 적용 절차에 대한 상세 기술이 제공되지 않았습니다.',
    }
}
# 인터뷰 질문 리스트 (생략 없이 전체 복사)
s_questions = [
    {"index": 0, "text": "산출물을 전달받을 이메일 주소를 입력해 주세요.", "category": "Complete", "help": "인터뷰 완료 후 최종 평가 결과 및 관련 산출물을 받을 이메일 주소를 입력하세요.<br>예: admin@company.com", "answer_type": "2", "text_help": ""},
    {"index": 1, "text": "시스템 이름을 적어 주세요.", "category": "IT PwC", "help": "평가 대상이 되는 정보시스템의 이름을 입력하세요.<br>예: SAP ERP, 인사관리시스템, 회계시스템, 고객관리시스템 등", "answer_type": "2", "text_help": ""},
    {"index": 2, "text": "사용하고 있는 시스템은 상용소프트웨어입니까?", "category": "IT PwC", "help": "", "answer_type": "3", "text_help": "SAP ERP, Oracle ERP, 더존ERP 등"},
    {"index": 3, "text": "Cloud 서비스를 사용하고 있습니까?", "category": "IT PwC", "help": "해당 시스템이 클라우드 환경에서 운영되고 있는지 확인하는 질문입니다.<br>- 예: AWS, Azure, GCP 등의 클라우드 서비스를 이용<br>- 아니오: 자체 데이터센터나 온프레미스 환경에서 운영", "answer_type": "1", "text_help": ""},
    {"index": 4, "text": "어떤 종류의 Cloud입니까?", "category": "IT PwC", "help": "SaaS (Software as a Service): 사용자가 직접 설치 및 관리할 필요 없이, 클라우드에서 제공되는 ERP 소프트웨어를 사용하는 방식.<br>예: SAP S/4HANA Cloud, Oracle NetSuite, Microsoft Dynamics 365<br><br>PaaS (Platform as a Service): 애플리케이션 개발 및 배포를 위한 플랫폼을 클라우드에서 제공하는 방식.<br>예: Microsoft Azure App Service, Google App Engine<br><br>IaaS (Infrastructure as a Service): 기업이 자체적으로 ERP 시스템을 구축하고 운영할 수 있도록 서버, 스토리지, 네트워크 등의 인프라를 제공하는 방식.<br>예: AWS EC2, Microsoft Azure Virtual Machines, Google Cloud Compute Engine", "answer_type": "6", "text_help": "SaaS|PaaS|IaaS"},
    {"index": 5, "text": "클라우드 서비스 업체의 SOC1 Report를 내부에서 검토하는 절차가 있습니까?", "category": "IT PwC", "help": "예: 클라우드 업체로부터 연간 SOC1 Report를 수령하여 내부 감사팀 또는 IT 담당자가 주요 통제 항목을 검토하고 그 결과를 경영진에게 보고함", "answer_type": "1", "text_help": ""},
    {"index": 6, "text": "시스템 사용자 ID는 사용자 개인별로 발급되고 있습니까?", "category": "APD", "help": "사용자가 시스템에 접속할 때 사용하는 ID의 발급 방식을 확인하는 질문입니다.<br><br>예(개인별 발급):<br>- 홍길동 → hong123, 김철수 → kim456 등 각자의 ID로 로그인<br><br>아니오(공용 발급):<br>- 재무팀 팀원 전체가 'finance01' 등 하나의 ID를 공동으로 사용<br>- 구매팀 전원이 팀 공용 ID로 발주·승인 업무 처리", "answer_type": "1", "text_help": ""},
    {"index": 7, "text": "사용자 권한부여 이력이 시스템에 기록되고 있습니까?", "category": "APD", "help": "사용자A가 재무권한을 가지고 있었는데 당기에 구매권한을 추가로 받았을 경우 언제(날짜 등) 구매권한을 받았는지 시스템에서 관리되는 경우를 의미합니다.", "answer_type": "1", "text_help": ""},
    {"index": 8, "text": "사용자 권한회수 이력이 시스템에 기록되고 있습니까?", "category": "APD", "help": "기존 권한 회수 시 변경 이력을 관리하고 있는지 확인합니다.<br>Standard 기능을 기준으로 SAP ERP의 경우 권한회수이력을 별도로 저장하며 Oracle ERP의 경우 권한 데이터를 삭제하지 않고 Effective Date로 관리합니다", "answer_type": "1", "text_help": ""},
    {"index": 9, "text": "공용 계정 환경에서 권한 회수 및 퇴사자 접근 차단은 어떻게 관리하고 있습니까?", "category": "APD", "help": "공용 계정을 사용하는 경우 개인별 권한 회수·퇴사자 접근 차단 적용이 어려울 수 있습니다. 현재 시스템에서 이를 대신하는 관리 방식이나 대안적 통제가 있다면 자유롭게 기술해 주세요.<br><br>예:<br>- 퇴사 발생 시 팀장이 IT팀에 연락하여 공용 ID의 비밀번호를 변경<br>- 공용 ID 사용 가능 인원 목록을 주기적으로 갱신하고 퇴사자를 제외<br>- 별도 접근 로그로 사용자별 행위 추적<br>- 특별한 관리 방식이 없는 경우 그대로 기술해 주세요.", "answer_type": "5", "text_help": ""},
    {"index": 10, "text": "사용자가 새로운 권한이 필요한 경우 요청서를 작성하고 부서장 등의 승인을 받는 절차가 있습니까?", "category": "APD", "help": "예) 새로운 권한이 필요한 경우 ITSM을 통해 요청서를 작성하고 팀장의 승인을 받은 후 IT팀에서 해당 권한을 부여함", "answer_type": "4", "text_help": "예) 새로운 권한이 필요한 경우 ITSM을 통해 요청서를 작성하고 팀장의 승인을 받은 후 IT팀에서 해당 권한을 부여함"},
    {"index": 11, "text": "부서이동 등 기존 권한의 회수가 필요한 경우 기존 권한을 회수하는 절차가 있습니까?", "category": "APD", "help": "예1) 인사팀에서 인사시스템에 인사명령을 입력하면 시스템에서 자동으로 기존 권한을 회수함<br>예2) 인사팀에서 인사명령을 IT팀으로 전달하면 IT팀에서 해당 인원의 기존 권한을 회수함", "answer_type": "4", "text_help": "예) 인사팀에서 인사시스템에 인사명령을 입력하면 시스템에서 자동으로 기존 권한을 회수함"},
    {"index": 12, "text": "퇴사자 발생 시 접근권한을 차단하는 절차가 있습니까?", "category": "APD", "help": "예1) 인사팀에서 인사시스템에 인사명령을 입력하면 시스템에서 자동으로 접근권한을 차단함<br>예2) 인사팀에서 인사명령을 IT팀으로 전달하면 IT팀에서 해당 인원의 접근권한을 차단함", "answer_type": "4", "text_help": "예) 인사팀에서 인사시스템에 인사명령을 입력하면 시스템에서 자동으로 접근권한을 차단함"},
    {"index": 13, "text": "Application 관리자(Superuser) 권한을 보유한 인원에 대해 기술해 주세요.", "category": "APD", "help": "예1) IT운영팀 김xx 책임", "answer_type": "5", "text_help": "권한 보유 인원의 부서, 직급, 직무 등"},
    {"index": 14, "text": "전체 사용자가 보유한 권한에 대한 적절성을 모니터링하는 절차가 있습니까?", "category": "APD", "help": "사용자에게 부여된 권한이 현재 업무에 적합한지를 주기적으로 검토하는 절차를 의미합니다.<br>예: 매분기 사용자 권한 적정성 검토 수행, 부서이동 시 기존 권한 회수 및 새로운 권한 부여 모니터링", "answer_type": "4", "text_help": "예) 매분기 사용자 권한 적정성 검토 수행, 부서이동 시 기존 권한 회수 및 새로운 권한 부여 모니터링"},
    {"index": 15, "text": "패스워드 설정사항을 기술해 주세요.", "category": "APD", "help": "예) 최소자리: 8, 복잡성: 영문/숫자/특수문자, 변경주기: 90일 등", "answer_type": "5", "text_help": ""},
    {"index": 16, "text": "회사에서 DB에 접속하여 필요한 작업을 수행하는 것이 가능합니까?", "category": "APD", "help": "회사에서 데이터가 저장된 곳에 직접 들어가서 데이터를 확인하거나 수정할 수 있는지를 묻는 질문입니다.<br><br>예시:<br>- 가능한 경우: IT 관리자가 데이터베이스에 직접 접속해서 고객 정보를 수정하거나 조회할 수 있음<br>- 불가능한 경우: 모든 데이터 작업은 시스템 화면을 통해서만 가능하고, 데이터베이스에 직접 들어갈 수 없음", "answer_type": "1", "text_help": ""},
    {"index": 17, "text": "데이터 변경 이력이 시스템에 기록되고 있습니까?", "category": "APD", "help": "시스템의 기능을 이용하여 데이터를 변경한 것이 아닌 관리자 등이 DB에 접속하여 쿼리를 통해 데이터를 변경한 건이 대상이며 해당 변경건만 추출이 가능해야 합니다", "answer_type": "1", "text_help": ""},
    {"index": 18, "text": "데이터 변경이 필요한 경우 요청서를 작성하고 부서장 등의 승인을 받는 절차가 있습니까?", "category": "APD", "help": "예) 데이터 변경 필요 시 담당자는 ITSM을 통해 요청서를 작성하고 책임자의 승인을 받은 후 IT담당자가 데이터를 변경함", "answer_type": "4", "text_help": "예) 데이터 변경 필요 시 담당자는 ITSM을 통해 요청서를 작성하고 책임자의 승인을 받은 후 IT담당자가 데이터를 변경함"},
    {"index": 19, "text": "데이터 변경 권한을 보유한 인원에 대해 기술해 주세요.", "category": "APD", "help": "예1) IT운영팀 최xx 책임", "answer_type": "5", "text_help": "권한 보유 인원의 부서, 직급, 직무 등"},
    {"index": 20, "text": "DB 종류와 버전을 작성해 주세요.", "category": "IT PwC", "help": "예: Oracle R12, MS SQL Server 2008 등", "answer_type": "2", "text_help": ""},
    {"index": 21, "text": "DB 접근제어 Tool을 사용하고 있습니까?", "category": "IT PwC", "help": "예: DBi, DB Safer 등", "answer_type": "3", "text_help": "제품명을 입력하세요"},
    {"index": 22, "text": "DB 접근권한 부여 이력이 시스템에 기록되고 있습니까?", "category": "APD", "help": "데이터베이스 접근 권한을 부여한 날짜, 사용자, 부여자, 권한 종류 등의 이력이 시스템에 기록되고 있는지 확인하는 질문입니다.<br>예: 권한 부여 시 로그 기록, DB 내장 관리 시스템에의 이력 저장", "answer_type": "1", "text_help": ""},
    {"index": 23, "text": "DB 접근권한이 필요한 경우 요청서를 작성하고 부서장 등의 승인을 받는 절차가 있습니까?", "category": "APD", "help": "예) DB 접근권한 필요시 담당자는 ITSM을 통해 요청서를 작성하고 서버 책임자에게 승인을 받은 후 서버 관리자가 접근 권한을 부여함", "answer_type": "4", "text_help": "예) DB 접근권한 필요시 담당자는 ITSM을 통해 요청서를 작성하고 서버 책임자에게 승인을 받은 후 서버 관리자가 접근 권한을 부여함"},
    {"index": 24, "text": "DB 관리자 권한을 보유한 인원에 대해 기술해 주세요.", "category": "APD", "help": "예) 인프라관리팀 김xx 과장, DBA", "answer_type": "5", "text_help": "권한 보유 인원의 부서, 직급, 직무 등"},
    {"index": 25, "text": "DB 패스워드 설정사항을 기술해 주세요.", "category": "APD", "help": "예) 최소자리: 8, 복잡성: 영문/숫자/특수문자, 변경주기: 90일 등", "answer_type": "5", "text_help": ""},
    {"index": 26, "text": "회사에서 OS서버에 접속하여 필요한 작업을 수행하는 것이 가능합니까?", "category": "APD", "help": "운영체제(OS) 수준에서 서버에 직접 접속하여 시스템 관리, 파일 수정, 로그 확인 등의 작업을 수행할 수 있는지를 확인하는 질문입니다.<br>예: SSH, Telnet, RDP 등을 통한 서버 접속 가능 여부<br>이 질문의 답변이 '아니오'일 경우, OS 관련 통제(APD12, APD13, APD14)는 해당사항이 없는 것으로 처리됩니다.", "answer_type": "1", "text_help": ""},
    {"index": 27, "text": "OS 종류와 버전을 작성해 주세요.", "category": "IT PwC", "help": "예: 윈도우즈 서버 2012, Unix AIX, Linux Redhat 등", "answer_type": "2", "text_help": ""},
    {"index": 28, "text": "OS 접근제어 Tool을 사용하고 있습니까?", "category": "IT PwC", "help": "예: Hiware, CyberArk 등", "answer_type": "3", "text_help": "제품명을 입력하세요"},
    {"index": 29, "text": "OS 접근권한 부여 이력이 시스템에 기록되고 있습니까?", "category": "APD", "help": "운영체제 수준의 접근 권한을 부여한 날짜, 사용자, 부여자, 권한 내용 등의 이력이 시스템에 기록되고 있는지 확인하는 질문입니다.<br>예: Active Directory 로그, 리눅스 audit 로그, 보안 관리 솔루션 내 이력 관리", "answer_type": "1", "text_help": ""},
    {"index": 30, "text": "OS 접근권한이 필요한 경우 요청서를 작성하고 부서장 등의 승인을 받는 절차가 있습니까?", "category": "APD", "help": "예) OS 접근권한 필요시 담당자는 ITSM을 통해 요청서를 작성하고 서버 책임자에게 승인을 받은 후 서버 관리자가 접근 권한을 부여함", "answer_type": "4", "text_help": "예) OS 접근권한 필요시 담당자는 ITSM을 통해 요청서를 작성하고 서버 책임자에게 승인을 받은 후 서버 관리자가 접근 권한을 부여함"},
    {"index": 31, "text": "OS 관리자 권한을 보유한 인원에 대해 기술해 주세요.", "category": "APD", "help": "예) 인프라관리팀 이xx 책임, 보안관리자", "answer_type": "5", "text_help": "권한 보유 인원의 부서, 직급, 직무 등"},
    {"index": 32, "text": "OS 패스워드 설정사항을 기술해 주세요.", "category": "APD", "help": "예) 최소자리: 8, 복잡성: 영문/숫자/특수문자, 변경주기: 90일 등", "answer_type": "5", "text_help": ""},
    {"index": 33, "text": "외부에서 시스템에 원격 접속하는 경우 VPN 등의 접근 통제 수단을 사용하고 있습니까?", "category": "APD", "help": "재택근무, 출장, 외부 업체 접속 등 외부 네트워크에서 시스템에 접근할 때 사용하는 보안 수단을 확인하는 질문입니다.<br>예:<br>- VPN(Virtual Private Network) 사용<br>- 2차 인증(MFA) 사용<br>- 아니오: 별도의 접근 통제 없이 인터넷으로 직접 접속", "answer_type": "3", "text_help": "VPN 솔루션명을 입력하세요"},
    {"index": 34, "text": "주요 로직을 회사 내부에서 수정하여 사용할 수 있습니까?", "category": "IT PwC", "help": "시스템의 핵심 기능을 회사에서 직접 변경할 수 있는지를 의미합니다.<br>예:<br>- 가능: 회사 개발팀이 계산 방식이나 업무 절차를 수정할 수 있음<br>- 불가능: 외부에서 만든 패키지 소프트웨어로 수정 불가", "answer_type": "1", "text_help": ""},
    {"index": 35, "text": "프로그램 변경 이력이 시스템에 기록되고 있습니까?", "category": "PC", "help": "변경에 대한 History가 시스템에 의해 기록되어야 합니다. A화면을 1, 3, 5월에 요청서를 받아 변경했다면 각각의 이관(배포)이력이 기록되어야 하며 자체기능, 배포툴, 형상관리툴 등을 사용할 수 있습니다.", "answer_type": "1", "text_help": ""},
    {"index": 36, "text": "프로그램 변경이 필요한 경우 요청서를 작성하고 부서장의 승인을 받는 절차가 있습니까?", "category": "PC", "help": "예) 프로그램 기능 변경 필요 시 ITSM을 통해 요청서를 작성하고 부서장의 승인을 받음", "answer_type": "4", "text_help": "예) 프로그램 기능 변경 필요 시 ITSM을 통해 요청서를 작성하고 부서장의 승인을 받음"},
    {"index": 37, "text": "프로그램 변경 시 사용자 테스트를 수행하고 그 결과를 문서화하는 절차가 있습니까?", "category": "PC", "help": "예) 프로그램 기능 변경 완료 후 요청자에 의해 사용자 테스트가 수행되고 그 결과가 문서화됨", "answer_type": "4", "text_help": "예) 프로그램 기능 변경 완료 후 요청자에 의해 사용자 테스트가 수행되고 그 결과가 문서화됨"},
    {"index": 38, "text": "프로그램 변경 완료 후 이관(배포)을 위해 부서장 등의 승인을 받는 절차가 있습니까?", "category": "PC", "help": "예) 프로그램 기능 변경 및 사용자 테스트 완료 후 변경담당자로부터 이관 요청서가 작성되고 부서장의 승인을 받음", "answer_type": "4", "text_help": "예) 프로그램 기능 변경 및 사용자 테스트 완료 후 변경담당자로부터 이관 요청서가 작성되고 부서장의 승인을 받음"},
    {"index": 39, "text": "이관(배포)권한을 보유한 인원에 대해 기술해 주세요.", "category": "PC", "help": "예) 인프라관리팀 박xx 수석, 서버관리자", "answer_type": "5", "text_help": "권한 보유 인원의 부서, 직급, 직무 등"},
    {"index": 40, "text": "운영서버 외 별도의 개발 또는 테스트 서버를 운용하고 있습니까?", "category": "PC", "help": "JSP, ASP 등으로 개발된 웹시스템의 경우 localhost 또는 127.0.0.1을 개발서버로도 볼 수 있습니다", "answer_type": "1", "text_help": ""},
    {"index": 41, "text": "현재 실행 중인 배치 스케줄이 있습니까?", "category": "CO", "help": "시스템이 정해진 시간에 자동으로 수행하는 작업이 있는지 확인하는 질문입니다.<br>예:<br>- 매일 밤 12시에 주문 데이터 집계 작업<br>- 매월 말 월천결산 데이터 백업<br>- 매주 월요일 사용자 권한 점검", "answer_type": "1", "text_help": ""},
    {"index": 42, "text": "별도의 Batch Schedule Tool을 사용하고 있습니까?", "category": "IT PwC", "help": "예: Waggle, JobScheduler 등", "answer_type": "3", "text_help": "제품명을 입력하세요"},
    {"index": 43, "text": "배치 스케줄 등록/변경 이력이 시스템에 기록되고 있습니까?", "category": "CO", "help": "개발되어 등록된 배치 프로그램(Background Job)을 스케줄로 등록 또는 변경한 경우로 한정합니다. 배치 프로그램을 개발하여 운영서버에 반영하는 것은 이 경우에 포함되지 않습니다", "answer_type": "1", "text_help": ""},
    {"index": 44, "text": "배치 스케줄 등록/변경이 필요한 경우 요청서를 작성하고 부서장 등의 승인을 받는 절차가 있습니까?", "category": "CO", "help": "예) 배치 스케줄이 필요한 경우 ITSM을 통해 요청서를 작성하고 승인권자의 승인을 받은 후 적절한 담당자에 의해 스케줄이 등록됨", "answer_type": "4", "text_help": "예) 배치 스케줄이 필요한 경우 ITSM을 통해 요청서를 작성하고 승인권자의 승인을 받은 후 적절한 담당자에 의해 스케줄이 등록됨"},
    {"index": 45, "text": "배치 스케줄을 등록/변경할 수 있는 인원에 대해 기술해 주세요.", "category": "CO", "help": "예) 시스템 운영팀 최xx 과장, 시스템운영자", "answer_type": "5", "text_help": "권한 보유 인원의 부서, 직급, 직무 등"},
    {"index": 46, "text": "배치 실행 오류 등에 대한 모니터링은 어떻게 수행되고 있는지 기술해 주세요.", "category": "CO", "help": "예1) 매일 아침 배치 수행 결과를 확인하며 문서화하며 오류 발생 시 원인 파악 및 조치 현황 등을 함께 기록함<br>예2) 오류 발생 시에만 점검결과를 작성하며 오류 발생 기록은 삭제하지 않고 유지됨", "answer_type": "5", "text_help": ""},
    {"index": 47, "text": "장애 발생 시 이에 대응하고 조치하는 절차에 대해 기술해 주세요.", "category": "CO", "help": "시스템에 문제가 생겼을 때 어떻게 대처하고 해결하는지에 대한 절차를 설명해 주세요.<br>예:<br>- 시스템 중단 시 연락체계 및 복구 절차<br>- 데이터 오류 발견 시 수정 및 보고 절차<br>- 외부 공격이나 보안 사고 대응 방법", "answer_type": "5", "text_help": ""},
    {"index": 48, "text": "백업은 어떻게 수행되고 또 어떻게 모니터링되고 있는지 기술해 주세요.", "category": "CO", "help": "시스템의 데이터를 안전하게 복사해서 보조하는 방법과 이것이 제대로 되고 있는지 확인하는 방법을 설명해 주세요.<br>예:<br>- 매일 밤에 전체 데이터 복사 후 외부 저장장치에 저장<br>- 백업 완료 시 알림 메일 발송<br>- 주기적으로 백업 데이터 복구 테스트 수행", "answer_type": "5", "text_help": ""},
    {"index": 49, "text": "서버실 출입 시 절차에 대해 기술해 주세요.", "category": "CO", "help": "서버나 주요 IT 장비가 있는 방에 들어갈 때의 보안 절차를 설명해 주세요.<br>예:<br>- 신분증 확인 및 출입자 명부 작성<br>- 보안카드나 비밀번호로 출입문 통과<br>- 출입 시간 기록 및 CCTV 모니터링<br>- 외부 인원의 경우 동반자 필요 여부", "answer_type": "5", "text_help": ""},
    {"index": 50, "text": "OS, DB 등 시스템 소프트웨어에 대한 보안 패치를 주기적으로 적용하는 절차가 있습니까?", "category": "CO", "help": "운영체제(OS), 데이터베이스(DB), 미들웨어 등 시스템 구성요소에 대한 보안 취약점 패치를 정기적으로 적용하는 절차가 있는지 확인합니다.<br>예:<br>- 월 1회 보안 패치 현황 점검 및 적용<br>- 긴급 보안 패치의 경우 발표 후 1주일 이내 적용<br>- 패치 적용 전 테스트 환경에서 검증 절차 수행", "answer_type": "4", "text_help": "예) 월 1회 보안 패치 현황 점검 및 적용, 긴급 패치의 경우 발표 후 1주일 이내 적용"},
    {"index": 51, "text": "SOC1 Report를 검토/승인하는 절차 및 통제가 있습니까?", "category": "IT PwC", "help": "클라우드 서비스 제공업체에서 발행하는 SOC1 보고서를 회사에서 검토하고 승인하는 공식적인 절차가 있는지 확인하는 질문입니다.<br>예:<br>- 연간 또는 정기적으로 SOC1 보고서 수령 및 검토<br>- IT 보안팀 또는 감사팀의 검토 및 승인 절차<br>- SOC1 보고서상의 예외사항 또는 개선권고사항에 대한 대응 계획 수립<br>- 검토 결과의 문서화 및 보고", "answer_type": "1", "text_help": ""}
]

question_count = len(s_questions)

# ================================
# 질문 인덱스 명칭 매핑 (Q_ID)
# 질문 추가·삭제 시 이 딕셔너리의 숫자값만 수정하면
# 하위 함수 코드를 건드리지 않아도 자동 반영됩니다.
# ================================
Q_ID = {
    'email':                0,   # 이메일 주소
    'system_name':          1,   # 시스템 이름
    'commercial_sw':        2,   # 상용소프트웨어 여부
    'use_cloud':            3,   # Cloud 사용 여부
    'cloud_type':           4,   # Cloud 종류 (SaaS/PaaS/IaaS)
    'soc1_report':          5,   # SOC1 Report 내부 검토 절차 여부
    'apd15_shared_account': 6,   # 공유 계정 존재 여부 (분기점, 구 index 8)
    'apd01_auth_history':   7,   # 사용자 권한부여 이력
    'apd02_revoke_history': 8,   # 사용자 권한회수 이력
    'apd15_shared_mgmt':    9,   # 공용 계정 환경 권한 관리 방식 (apd15=N 시 표시)
    'apd01_procedure':      10,  # 신규 권한 승인 절차
    'apd02_procedure':      11,  # 권한 회수 절차
    'apd03_procedure':      12,  # 퇴사자 접근권한 차단
    'apd04_admin':          13,  # Application 관리자 권한자
    'apd05_monitoring':     14,  # 사용자 권한 모니터링
    'apd06_password':       15,  # Application 패스워드 설정
    'apd07_db_access':      16,  # DB 접속 가능 여부 (분기점)
    'apd07_data_history':   17,  # 데이터 변경 이력 기록
    'apd07_procedure':      18,  # 데이터 변경 승인 절차
    'apd08_data_auth':      19,  # 데이터 변경 권한자
    'apd09_db_type':        20,  # DB 종류와 버전
    'apd09_db_tool':        21,  # DB 접근제어 Tool
    'apd09_db_history':     22,  # DB 접근권한 부여 이력
    'apd09_procedure':      23,  # DB 접근권한 승인 절차
    'apd10_db_admin':       24,  # DB 관리자 권한자
    'apd11_db_password':    25,  # DB 패스워드 설정
    'apd12_os_access':      26,  # OS 접속 가능 여부 (분기점)
    'apd12_os_type':        27,  # OS 종류와 버전
    'apd12_os_tool':        28,  # OS 접근제어 Tool
    'apd12_os_history':     29,  # OS 접근권한 부여 이력
    'apd12_procedure':      30,  # OS 접근권한 승인 절차
    'apd13_os_admin':       31,  # OS 관리자 권한자
    'apd14_os_password':    32,  # OS 패스워드 설정
    'apd16_vpn':            33,  # 외부 접속(VPN) 사용 여부
    'pc_can_modify':        34,  # 주요 로직 수정 가능 여부 (컨텍스트 파악용)
    'pc01_change_history':  35,  # 프로그램 변경 이력
    'pc01_procedure':       36,  # 프로그램 변경 승인 절차
    'pc02_procedure':       37,  # 사용자 테스트 절차
    'pc03_procedure':       38,  # 이관 승인 절차
    'pc04_deploy_auth':     39,  # 이관 권한자
    'pc05_dev_env':         40,  # 개발/운영 환경 분리
    'co_has_batch':         41,  # 배치 스케줄 유무 (분기점)
    'co01_batch_tool':      42,  # Batch Schedule Tool
    'co01_batch_history':   43,  # 배치 스케줄 변경 이력
    'co01_procedure':       44,  # 배치 스케줄 승인 절차
    'co02_batch_auth':      45,  # 배치 스케줄 등록/변경 권한자
    'co03_monitoring':      46,  # 배치 실행 모니터링
    'co04_incident':        47,  # 장애 대응 절차
    'co05_backup':          48,  # 백업 수행 및 모니터링
    'co06_server_room':     49,  # 서버실 출입 절차
    'co07_security_patch':  50,  # 보안 패치 적용 절차
    'soc1_review':          51,  # SOC1 Report 검토/승인 절차
}

# --- 조건부 질문 생략 로직 ---
def get_conditional_questions(answers):
    """
    답변에 따라 조건부로 질문을 필터링하는 함수
    """
    
    if not answers or len(answers) < 4:
        return s_questions

    skip_ranges = []

    # Cloud 미사용 시 4~5번, soc1_review 질문 생략
    if len(answers) > Q_ID['use_cloud'] and answers[Q_ID['use_cloud']] and str(answers[Q_ID['use_cloud']]).upper() == 'N':
        skip_ranges.append((4, 5))
        skip_ranges.append((Q_ID['soc1_review'], Q_ID['soc1_review']))

    # Cloud 타입별 스킵 (SOC1 여부 무관)
    cloud_type = str(answers[Q_ID['cloud_type']]) if len(answers) > Q_ID['cloud_type'] and answers[Q_ID['cloud_type']] else ''

    # SaaS: DB+OS그룹(Q16~32), PC전체(Q34~40), CO백업·서버실·보안패치(Q48~50) 생략
    # apd04_admin(Q13): App레벨 SuperUser 자사관리 → 유지
    # apd16_vpn(Q33): Cloud 무관 → 유지
    # co04_incident(Q47): 업체 장애 시 자사 내부 에스컬레이션 절차 필요 → 유지
    if cloud_type == 'SaaS':
        skip_ranges.append((Q_ID['apd07_db_access'], Q_ID['apd14_os_password']))   # Q16~32 (DB+OS그룹)
        skip_ranges.append((Q_ID['pc_can_modify'], Q_ID['pc05_dev_env']))           # Q34~40 (PC전체)
        skip_ranges.append((Q_ID['co05_backup'], Q_ID['co07_security_patch']))      # Q48~50 (백업·서버실·보안패치)

    # IaaS: apd10_db_admin, apd13_os_admin, co06_server_room만 생략
    # co04_incident(Q47)·co05_backup(Q48)·co07_security_patch(Q50): 자사 관리 → 유지
    # co06_server_room(Q49): 물리 서버실 없음 → 스킵
    elif cloud_type == 'IaaS':
        skip_ranges.append((Q_ID['apd10_db_admin'], Q_ID['apd10_db_admin']))
        skip_ranges.append((Q_ID['apd13_os_admin'], Q_ID['apd13_os_admin']))
        skip_ranges.append((Q_ID['co06_server_room'], Q_ID['co06_server_room']))    # Q49만 스킵

    # PaaS: DB+OS그룹(Q16~32), CO백업·서버실·보안패치(Q48~50) 생략
    # co04_incident(Q47): 앱 레벨 장애 자사 대응 → 유지
    elif cloud_type == 'PaaS':
        skip_ranges.append((Q_ID['apd07_db_access'], Q_ID['apd14_os_password']))   # Q16~32 (DB+OS그룹)
        skip_ranges.append((Q_ID['co05_backup'], Q_ID['co07_security_patch']))      # Q48~50 (백업·서버실·보안패치)

    # DB 접속 불가 시 apd07_data_history~apd11_db_password(16~24) 생략
    if len(answers) > Q_ID['apd07_db_access'] and answers[Q_ID['apd07_db_access']] and str(answers[Q_ID['apd07_db_access']]).upper() == 'N':
        skip_ranges.append((Q_ID['apd07_data_history'], Q_ID['apd11_db_password']))

    # OS 접속 불가 시 apd12_os_type~apd14_os_password(26~31) 생략
    if len(answers) > Q_ID['apd12_os_access'] and answers[Q_ID['apd12_os_access']] and str(answers[Q_ID['apd12_os_access']]).upper() == 'N':
        skip_ranges.append((Q_ID['apd12_os_type'], Q_ID['apd14_os_password']))

    # 배치 스케줄 없을 시 co01_batch_tool~co03_monitoring(41~45) 생략
    if len(answers) > Q_ID['co_has_batch'] and answers[Q_ID['co_has_batch']] and str(answers[Q_ID['co_has_batch']]).upper() == 'N':
        skip_ranges.append((Q_ID['co01_batch_tool'], Q_ID['co03_monitoring']))

    # 공용 ID 발급(apd15=N): 공용 계정 관리 방식 질문(Q9) 표시, apd01~apd03_procedure 스킵
    # 개인별 발급(apd15=Y): 공용 계정 관리 방식 질문(Q9) 스킵, apd01~apd03_procedure 표시
    if len(answers) > Q_ID['apd15_shared_account'] and answers[Q_ID['apd15_shared_account']]:
        if str(answers[Q_ID['apd15_shared_account']]).upper() == 'N':
            skip_ranges.append((Q_ID['apd01_procedure'], Q_ID['apd03_procedure']))
        elif str(answers[Q_ID['apd15_shared_account']]).upper() == 'Y':
            skip_ranges.append((Q_ID['apd15_shared_mgmt'], Q_ID['apd15_shared_mgmt']))
    
    
    if not skip_ranges:
        return s_questions
    
    filtered_questions = []
    skipped_count = 0
    for question in s_questions:
        question_index = question['index']
        should_skip = False
        
        for start, end in skip_ranges:
            if start <= question_index <= end:
                should_skip = True
                skipped_count += 1
                break
        
        if not should_skip:
            filtered_questions.append(question)
    
    return filtered_questions

def get_conditional_question_count(answers):
    """
    조건부 질문 생략을 고려한 총 질문 수 반환
    """
    conditional_questions = get_conditional_questions(answers)
    return len(conditional_questions)

def get_skipped_controls(answers):
    """
    스킵된 질문과 관련된 통제 목록을 반환
    """
    skipped_controls = set()
    
    if not answers or len(answers) < 4:
        return skipped_controls

    # Cloud 서비스 관련 스킵 로직
    if len(answers) > Q_ID['soc1_report']:
        cloud_type = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else None
        has_soc1_report = answers[Q_ID['soc1_report']] and str(answers[Q_ID['soc1_report']]).upper() == 'Y'

        if cloud_type and has_soc1_report:
            if cloud_type == 'SaaS':
                # CO04(장애대응): 자사 내부 에스컬레이션 절차 필요 → 유지
                skipped_controls.update([
                    'APD07', 'APD08', 'APD09', 'APD10', 'APD11',
                    'APD12', 'APD13', 'APD14',
                    'PC01', 'PC02', 'PC03', 'PC04', 'PC05',
                    'CO05', 'CO06', 'CO07'
                ])
            elif cloud_type == 'PaaS':
                # CO04(장애대응): 앱 레벨 장애 자사 대응 → 유지
                skipped_controls.update([
                    'APD07', 'APD08', 'APD09', 'APD10', 'APD11',
                    'APD12', 'APD13', 'APD14',
                    'CO05', 'CO06', 'CO07'
                ])
            elif cloud_type == 'IaaS':
                # CO04(장애대응)·CO05(백업)·CO07(보안패치): 자사 관리 → 유지
                # CO06(서버실): 물리 서버실 없음 → 스킵
                skipped_controls.update([
                    'APD11',
                    'APD13',
                    'CO06'
                ])

    # DB 접속 불가 시 DB 관련 통제 해당없음
    if len(answers) > Q_ID['apd07_db_access'] and answers[Q_ID['apd07_db_access']] and str(answers[Q_ID['apd07_db_access']]).upper() == 'N':
        skipped_controls.update(['APD07', 'APD08', 'APD09', 'APD10', 'APD11'])

    # OS 접속 불가 시 OS 관련 통제 해당없음
    if len(answers) > Q_ID['apd12_os_access'] and answers[Q_ID['apd12_os_access']] and str(answers[Q_ID['apd12_os_access']]).upper() == 'N':
        skipped_controls.update(['APD12', 'APD13', 'APD14'])

    # 배치 스케줄 없을 시 CO01~CO03 해당없음
    if len(answers) > Q_ID['co_has_batch'] and answers[Q_ID['co_has_batch']] and str(answers[Q_ID['co_has_batch']]).upper() == 'N':
        skipped_controls.update(['CO01', 'CO02', 'CO03'])

    # 공용 ID 발급(apd15=N) 시 APD01·APD02·APD03 해당없음 (개인 단위 권한 부여·회수·퇴사자 차단 불가)
    if len(answers) > Q_ID['apd15_shared_account'] and answers[Q_ID['apd15_shared_account']] and str(answers[Q_ID['apd15_shared_account']]).upper() == 'N':
        skipped_controls.update(['APD01', 'APD02', 'APD03'])
    
    return skipped_controls

def clear_skipped_answers(answers, textarea_answers):
    """조건부 질문 스킵 시 해당 범위의 답변을 공란으로 처리"""
    if not answers or len(answers) < 4:
        return

    skip_ranges = []

    # Cloud 미사용 시 4~5번 질문 생략
    if len(answers) > Q_ID['use_cloud'] and answers[Q_ID['use_cloud']] and str(answers[Q_ID['use_cloud']]).upper() == 'N':
        skip_ranges.append((4, 5))

    # DB 접속 불가 시 apd07_data_history~apd11_db_password 생략
    if len(answers) > Q_ID['apd07_db_access'] and answers[Q_ID['apd07_db_access']] and str(answers[Q_ID['apd07_db_access']]).upper() == 'N':
        skip_ranges.append((Q_ID['apd07_data_history'], Q_ID['apd11_db_password']))

    # OS 접속 불가 시 apd12_os_type~apd14_os_password 생략
    if len(answers) > Q_ID['apd12_os_access'] and answers[Q_ID['apd12_os_access']] and str(answers[Q_ID['apd12_os_access']]).upper() == 'N':
        skip_ranges.append((Q_ID['apd12_os_type'], Q_ID['apd14_os_password']))

    # 배치 스케줄 없을 시 co01_batch_tool~co03_monitoring 생략
    if len(answers) > Q_ID['co_has_batch'] and answers[Q_ID['co_has_batch']] and str(answers[Q_ID['co_has_batch']]).upper() == 'N':
        skip_ranges.append((Q_ID['co01_batch_tool'], Q_ID['co03_monitoring']))

    # 공용 ID 발급(apd15=N): apd01~apd03_procedure 초기화 / 개인별 발급(apd15=Y): apd15_shared_mgmt 초기화
    if len(answers) > Q_ID['apd15_shared_account'] and answers[Q_ID['apd15_shared_account']]:
        if str(answers[Q_ID['apd15_shared_account']]).upper() == 'N':
            skip_ranges.append((Q_ID['apd01_procedure'], Q_ID['apd03_procedure']))
        elif str(answers[Q_ID['apd15_shared_account']]).upper() == 'Y':
            skip_ranges.append((Q_ID['apd15_shared_mgmt'], Q_ID['apd15_shared_mgmt']))
    
    # 스킵된 범위의 답변을 공란으로 설정
    for start, end in skip_ranges:
        for i in range(start, end + 1):
            if i < len(answers):
                answers[i] = ''
            if i < len(textarea_answers):
                textarea_answers[i] = ''
    
    for start, end in skip_ranges:
        for i in range(start, end + 1):
            if i < len(answers):
                answers[i] = ''
            if i < len(textarea_answers):
                textarea_answers[i] = ''

def set_sheet_tab_color_for_skipped_controls(wb, answers):
    """
    스킵된 통제에 해당하는 시트 탭을 회색으로 설정
    """
    skipped_controls = get_skipped_controls(answers)
    
    for sheet_name in wb.sheetnames:
        if sheet_name in skipped_controls:
            try:
                ws = wb[sheet_name]
                # 시트 탭을 회색으로 설정 (RGB: 808080)
                ws.sheet_properties.tabColor = "808080"
            except Exception as e:
                pass


# --- 통제별 검토 기준 정의 (수기 수정 가능) ---
# 공통 감사기준 (엄격한 전문 접근)
COMMON_AUDIT_CRITERIA = [
    "명시적으로 '모집단 확보가 불가'라고 기술된 경우는 Ineffective",
    "명시적으로 '절차가 없다'고 기술된 경우는 Ineffective",
    "이력이 기록되고 적절한 절차가 있으면 Effective",
    "통제 목적이 달성되고 있으면 Effective",
    "소규모 회사의 현실적 제약을 고려하되 기본 요구사항은 충족되어야 함",
    "Package S/W 등 기술적 제약이 있는 경우 N/A 처리"
]

# 엄격한 전문 감사를 위한 핵심 고려사항
SPECIAL_CONSIDERATIONS = {
    "STRICT_APPROACH": "핵심 통제 요구사항이 충족되지 않으면 Ineffective",
    "PACKAGE_SW": "상용소프트웨어 사용시에만 해당 통제를 N/A로 처리",
    "AUDIT_TRAIL": "감사증적이 없으면 통제 검증이 불가능하므로 Ineffective",
    "CLOUD_SERVICE": "클라우드 서비스 이용시에도 기본적인 통제는 필요",
    "DOCUMENTATION": "절차 문서화는 통제의 일관성과 연속성 확보를 위해 필수",
    "SEGREGATION": "직무분리 원칙을 위반하면 통제 효과성이 현저히 저하"
}

CONTROL_SPECIFIC_CRITERIA = {
    'APD01': [
        "사용자 권한 부여 이력이 시스템에 자동으로 기록되는지 확인",
        "권한 요청 시 적절한 승인권자(부서장 등)의 사전 승인을 받는지 확인",
        "권한 부여 절차가 문서화되어 있고 준수되고 있는지 확인",
        "임시 권한의 경우 만료일 설정 및 자동 회수 여부 확인"
    ],
    'APD02': [
        "부서이동 시 기존 권한의 자동 또는 수동 회수 절차가 있는지 확인",
        "부서이동자가 직접 권한 회수를 요청하는 경우는 통제로 인정하지 않음",
        "인사시스템과 연계하여 자동 권한 회수가 되는지 확인",
        "권한 회수까지의 기간이 합리적인지 확인 (통상 1주일 이내)"
    ],
    'APD03': [
        "퇴사자 발생 시 즉시 접근 권한을 차단하는 절차가 있는지 확인",
        "퇴사일 당일 또는 익일까지 권한 회수가 완료되는지 확인",
        "인사팀과 IT팀 간의 퇴사자 정보 전달 체계가 있는지 확인",
        "퇴사자 계정의 완전 삭제 또는 비활성화 절차가 있는지 확인"
    ],
    'APD04': [
        "Application 관리자(Superuser) 권한 보유 인원이 명확히 식별되는지 확인",
        "관리자 권한 보유자의 소속, 직급, 담당 업무가 구체적으로 기술되어 있는지 확인",
        "관리자 권한 보유자가 업무상 필요한 적정 인원인지 판단",
        "관리자 권한의 정기적 검토 및 재승인 절차가 있는지 확인",
        "권한 보유자 수의 적정성은 회사 규모를 고려하여 판단"
    ],
    'APD05': [
        "사용자 권한의 적절성을 정기적으로 검토하는 절차가 있는지 확인",
        "전체 사용자 권한을 대상으로 검토하는지 확인 (일부만 검토 시 합리적 이유 필요)",
        "권한 검토 결과 발견된 이슈에 대한 조치 절차가 있는지 확인",
        "권한 검토 결과가 문서화되고 적절한 승인을 받는지 확인",
        "검토 주기의 적정성보다는 정기적 수행 여부에 중점",
        "권한 검토 시 직무분리 원칙 준수 여부도 확인"
    ],
    'APD06': [
        "Application 패스워드 최소 길이가 8자 이상으로 설정되어 있는지 확인",
        "패스워드 복잡성 요구사항(영문, 숫자, 특수문자 조합)이 있는지 확인",
        "패스워드 변경 주기 정책이 있는지 확인 (90일 이내 권장)",
        "최소 요구사항을 일부 충족하지 않더라도 다른 보안 조치가 있으면 경감요소로 고려",
        "패스워드 재사용 방지 정책이 있는지 확인"
    ],
    'APD07': [
        "데이터 변경 이력이 시스템에 자동으로 기록되는지 확인 (누가, 언제, 무엇을 변경했는지)",
        "데이터 변경 요청 시 적절한 승인권자의 사전 승인을 받는지 확인",
        "회사에서 직접 데이터 변경을 하지 않는 경우 이력 기록 불필요",
        "응급 상황 시 데이터 변경에 대한 사후 승인 절차가 있는지 확인",
        "데이터 변경 사유 및 영향범위가 기록되는지 확인"
    ],
    'APD08': [
        "데이터 변경 권한 보유 인원이 최소한으로 제한되어 있는지 확인",
        "데이터 변경 권한 보유자의 소속, 직급, 담당 업무가 명확히 기술되어 있는지 확인",
        "권한 보유자가 업무상 필요한 적정 인원인지 판단 (회사 규모 고려)",
        "데이터 변경 권한의 정기적 검토 절차가 있는지 확인",
        "일반 사용자와 데이터 변경 권한자의 명확한 분리가 되어 있는지 확인"
    ],
    'APD09': [
        "DB 접근권한 부여 이력이 시스템에 기록되는지 확인",
        "DB 접근 필요시 승인권자의 승인 여부"
    ],
    'APD10': [
        "DB 관리자 권한 보유 인원이 명확히 식별되는지 확인",
        "DB 관리자 권한 보유자의 부서, 직급, 직무가 구체적으로 기술되어 있는지 확인",
        "DB 관리자가 적정한 인원인지 판단",
        "권한 보유자가 몇 명인지는 판단하지 않음"

    ],
    'APD11': [
        "DB 패스워드 최소 길이가 8자 이상인지 확인",
        "DB 패스워드 복잡성(영문/숫자/특수문자) 요구사항이 있는지 확인"
    ],
    'APD12': [
        "OS 접근권한 부여 이력이 시스템에 기록되는지 확인",
        "OS 접근 필요시 승인권자의 승인 여부"
    ],
    'APD13': [
        "OS 관리자 권한 보유 인원이 명확히 식별되는지 확인",
        "OS 관리자 권한 보유자의 부서, 직급, 직무가 구체적으로 기술되어 있는지 확인",
        "OS 관리자가 적정한 인원인지 판단",
        "권한 보유자가 몇 명인지는 판단하지 않음"
    ],
    'APD14': [
        "OS 패스워드 최소 길이가 8자 이상인지 확인",
        "OS 패스워드 복잡성(영문/숫자/특수문자) 요구사항이 있는지 확인"
    ],
    'PC01': [
        "프로그램 변경 이력이 시스템에 자동으로 기록되는지 확인 (변경일시, 변경자, 변경내용)",
        "프로그램 변경 요청 시 적절한 승인권자의 사전 승인을 받는지 확인",
        "변경 요청서에 변경 사유, 영향범위, 테스트 계획이 포함되어 있는지 확인",
        "Package S/W 사용 시 본 통제는 적용되지 않음 (N/A 처리)",
        "형상관리 도구나 배포 도구를 통한 이력 관리도 인정"
    ],
    'PC02': [
        "프로그램 변경 후 사용자 테스트(UAT) 수행 절차가 있는지 확인",
        "사용자 테스트는 변경을 요청한 업무 담당자가 수행하는 것이 바람직",
        "테스트 결과가 문서화되고 승인되는 절차가 있는지 확인",
        "Package S/W 사용 시 본 통제는 적용되지 않음 (N/A 처리)",
        "테스트 실패 시 운영 반영을 중단하는 절차가 있는지 확인"
    ],
    'PC03': [
        "프로그램 변경 완료 후 운영 이관(배포) 전 최종 승인 절차가 있는지 확인",
        "이관 승인권자가 변경 요청 승인권자와 동일해도 무방",
        "Package S/W 사용 시 본 통제는 적용되지 않음 (N/A 처리)",
        "응급 변경 시에도 사후 승인 절차가 있는지 확인"
    ],
    'PC04': [
        "운영 서버 이관(배포) 권한 보유 인원이 명확히 제한되어 있는지 확인",
        "이관 권한 보유자의 소속, 직급, 담당 업무가 구체적으로 기술되어 있는지 확인",
        "이관 권한 보유자가 업무상 필요한 최소 인원인지 판단 (회사 규모 고려)",
        "Package S/W 사용 시 본 통제는 적용되지 않음 (N/A 처리)",
        "개발자와 이관 담당자의 분리가 되어 있는지 확인 (소규모는 예외)"
    ],
    'PC05': [
        "운영환경과 개발/테스트 환경이 물리적 또는 논리적으로 분리되어 있는지 확인",
        "개발 환경에서 운영 데이터에 직접 접근할 수 없도록 차단되어 있는지 확인",
        "localhost나 127.0.0.1을 개발서버로 사용하는 것도 분리로 인정",
        "Package S/W 사용 시에도 개발/운영 환경 분리는 적용 가능",
        "클라우드 환경의 경우 논리적 분리(VPC, 계정 분리 등)도 인정"
    ],
    'CO01': [
        "배치 스케줄 등록/변경 이력이 시스템에 기록되는지 확인",
        "배치 스케줄 등록/변경시 승인권자의 승인 여부"
    ],
    'CO02': [
        "배치 스케줄 등록/변경 권한 보유 인원이 명확히 식별되는지 확인",
        "배치 스케줄 등록 권한 보유자의 부서, 직급, 직무가 구체적으로 기술되어 있는지 확인",
        "배치 스케줄 등록 권한 보유자가 적정한 인원인지 판단",
        "권한 보유자가 몇 명인지는 판단하지 않음"
    ],
    'CO03': [
        "배치 실행 결과 모니터링 절차가 있는지 확인",
        "모니터링 결과 문서화 및 보관 절차가 있는지 확인"
    ],
    'CO04': [
        "장애 발생 시 대응 절차가 명확하게 정의되어 있는지 확인",
        "장애 처리 결과 문서화 및 사후 분석 절차가 있는지 확인"
    ],
    'CO05': [
        "백업 수행 절차가 명확하게 정의되어 있는지 확인",
        "백업 결과 모니터링 및 확인 절차가 있는지 확인",
        "백업 데이터 복구 테스트가 정기적으로 수행되는지 확인"
    ],
    'CO06': [
        "서버실 출입 승인 절차가 있는지 확인",
        "서버실 출입 기록이 관리되는지 확인",
        "서버실 물리적 보안 체계가 적절한지 확인"
    ],
    'APD15': [
        "시스템 내 공유 계정(Generic/Shared Account) 존재 여부 확인",
        "공유 계정이 존재하는 경우 개인 식별 불가로 인한 감사 증적 미확보 위험 평가",
        "공유 계정이 없으면 Effective, 존재하면 Ineffective로 판정",
        "불가피하게 공유 계정을 사용하는 경우 대체 통제(세션 로그, 별도 모니터링 등) 여부 확인"
    ],
    'APD16': [
        "외부 원격 접속 시 VPN 등 접근 통제 수단 사용 여부 확인",
        "재택근무, 출장, 외부 업체 등 다양한 외부 접속 시나리오에서 통제 적용 여부 확인",
        "VPN 미사용 시 대체 접근 통제 수단(MFA, IP 화이트리스트 등) 여부 확인",
        "접근 통제 수단 없이 인터넷으로 직접 접속하는 경우 Ineffective 판정"
    ],
    'CO07': [
        "OS, DB, 미들웨어 등 시스템 소프트웨어에 대한 보안 패치 절차 존재 여부 확인",
        "보안 패치 적용 주기의 적정성 확인 (권장: 월 1회 이상 정기 점검)",
        "긴급 보안 취약점 발견 시 신속 패치 적용 절차가 있는지 확인",
        "패치 적용 전 테스트 환경 검증 절차 여부 확인"
    ]
}

# --- 리팩토링: Ineffective 조건 체크 함수 ---
def is_ineffective(control, answers):
    # 조건부 질문 생략 로직 적용
    if not answers or len(answers) < 8:
        return False  # 답변이 부족하면 기본값 반환
    
    # 조건부 N/A 처리
    if control in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11'] and answers[Q_ID['apd07_db_access']] == 'N':
        return False  # DB 접속 불가 시 DB 관련 통제 N/A

    if control in ['APD12', 'APD13', 'APD14'] and answers[Q_ID['apd12_os_access']] == 'N':
        return False  # OS 접속 불가 시 OS 관련 통제 N/A

    if control in ['CO01', 'CO02', 'CO03'] and answers[Q_ID['co_has_batch']] == 'N':
        return False  # 배치 스케줄 없을 시 CO01-03 N/A

    # apd02_revoke_history가 N이면 관련 통제 N/A
    if control in ['APD07', 'APD08'] and answers[Q_ID['apd02_revoke_history']] == 'N':
        return False
    
    conditions = {
        'APD01': len(answers) > 14 and (answers[12] == 'N' or answers[14] == 'N'),
        'APD02': len(answers) > 15 and answers[15] == 'N',
        'APD03': len(answers) > 16 and answers[16] == 'N',
        'APD04': len(answers) > 17 and answers[17] == 'N',
        'APD05': False,  # 항상 적용되는 통제 (미비점 여부와 무관)
        'APD06': False,  # 항상 적용되는 통제 (미비점 여부와 무관)
        'APD07': len(answers) > 22 and answers[14] == 'Y' and (answers[21] == 'N' or answers[22] == 'N'),  # DB 접근 가능할 때만 체크
        'APD08': len(answers) > 22 and answers[14] == 'Y' and answers[22] == 'N',  # DB 접근 가능할 때만 체크
        'APD09': len(answers) > 26 and answers[14] == 'Y' and (answers[25] == 'N' or answers[26] == 'N'),  # DB 접근 가능할 때만 체크
        'APD10': len(answers) > 28 and answers[14] == 'Y' and (answers[27] == 'N' or answers[28] == 'N'),  # DB 접근 가능할 때만 체크
        'APD12': len(answers) > 32 and (answers[31] == 'N' or answers[32] == 'N'),  # OS 접근권한 부여 이력 + 승인절차
        'APD13': len(answers) > 33 and answers[33] == 'N',  # OS 관리자 권한 제한
        'PC01': (len(answers) > 36 and answers[24] == 'Y') and (answers[35] == 'N' or answers[36] == 'N'),
        'PC02': (len(answers) > 37 and answers[24] == 'Y') and (answers[37] == 'N'),
        'PC03': (len(answers) > 38 and answers[24] == 'Y') and (answers[38] == 'N'),
        'PC05': len(answers) > 40 and answers[40] == 'N',
        'CO01': len(answers) > 42 and answers[38] == 'Y' and (answers[41] == 'N' or answers[42] == 'N'),  # 배치 스케줄이 있을 때만 체크
        'CO02': len(answers) > 43 and answers[38] == 'Y' and answers[43] == 'N',  # 배치 스케줄이 있을 때만 체크
        'CO03': len(answers) > 44 and answers[38] == 'Y' and answers[44] == 'N',  # 배치 스케줄이 있을 때만 체크
    }
    return conditions.get(control, False)

def _wait_for_rate_limit():
    """API Rate Limiting을 위한 대기 함수"""
    global _last_api_call_time
    with _api_call_lock:
        current_time = time.time()
        time_since_last_call = current_time - _last_api_call_time

        if time_since_last_call < API_RATE_LIMIT['delay_between_requests']:
            wait_time = API_RATE_LIMIT['delay_between_requests'] - time_since_last_call
            if wait_time > 1:  # 1초 이상 대기 시 메시지 출력
                print(f"[INFO] API Rate Limit 준수: {wait_time:.1f}초 대기 중...")
            time.sleep(wait_time)

        _last_api_call_time = time.time()

def _call_openai_api_with_retry(client, messages, model, max_tokens, temperature, max_retries=5):
    """OpenAI API 호출 with retry 로직 (안정성 우선)"""
    for attempt in range(max_retries):
        try:
            _wait_for_rate_limit()

            response = client.chat.completions.create(
                model=model,
                messages=messages,
                max_tokens=max_tokens,
                temperature=temperature
            )

            return response.choices[0].message.content

        # 네트워크 예외를 명시적으로 처리
        except (RequestException, socket.timeout, socket.error, OSError) as network_err:
            error_name = type(network_err).__name__
            if attempt < max_retries - 1:
                wait_time = 10 * (attempt + 1)
                print(f"[INFO] 네트워크 연결 끊김 감지: {error_name}")
                print(f"[INFO] {wait_time}초 후 자동 재시도... ({attempt + 1}/{max_retries})")
                time.sleep(wait_time)
                continue
            else:
                print(f"[ERROR] 네트워크 연결 문제 지속 ({error_name}). 원본 텍스트 사용하여 계속 진행.")
                return None

        except Exception as e:
            error_str = str(e)

            # 할당량 초과 에러 처리
            if '429' in error_str or 'quota' in error_str.lower() or 'rate' in error_str.lower():
                if attempt < max_retries - 1:
                    wait_time = 20 * (attempt + 1)  # 20초, 40초, 60초, 80초
                    print(f"[INFO] API 할당량 초과. {wait_time}초 후 재시도... ({attempt + 1}/{max_retries})")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"[ERROR] API 할당량 초과. 최대 재시도 {max_retries}회 도달. 원본 사용.")
                    return None

            # 네트워크 오류 등 일시적 오류도 재시도
            elif any(keyword in error_str.lower() for keyword in [
                'connection', 'timeout', 'network', 'timed out',
                'connection reset', 'connection refused', 'connection aborted',
                'dns', 'socket', 'unreachable', 'unavailable',
                'temporary failure', 'service unavailable', '503', '502', '504'
            ]):
                if attempt < max_retries - 1:
                    wait_time = 10 * (attempt + 1)
                    print(f"[INFO] 네트워크 연결 문제 감지. {wait_time}초 후 재시도... ({attempt + 1}/{max_retries})")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"[ERROR] 네트워크 오류 지속. 최대 재시도 {max_retries}회 도달. 원본 텍스트 사용하여 계속 진행.")
                    return None
            else:
                # 그 외 예상치 못한 에러는 로그 출력 후 None 반환
                print(f"[ERROR] 예상치 못한 API 오류: {str(e)[:200]}")
                return None

    return None

def ai_improve_interview_answer(question_text, answer_text):
    """
    AI를 사용하여 인터뷰 답변을 문법적으로 다듬고 일관성 있게 개선합니다.
    """
    try:
        api_key = os.getenv('OPENAI_API_KEY')
        if api_key:
            # API 키가 있으면 AI로 개선
            client = OpenAI(api_key=api_key)
            model_name = os.getenv('OPENAI_MODEL', AI_MODEL_CONFIG['model'])

            prompt = AI_REFINEMENT_PROMPT.format(answer_text=answer_text)

            messages = [
                {"role": "system", "content": "당신은 한국어 문서 교정 전문가입니다. 문법 오류만 수정하고, 원본 의미와 내용은 절대 변경하지 마세요."},
                {"role": "user", "content": prompt}
            ]

            # OpenAI API 호출 (retry 및 rate limiting 포함)
            result = _call_openai_api_with_retry(
                client,
                messages,
                model_name,
                AI_MODEL_CONFIG['max_tokens'],
                AI_MODEL_CONFIG['temperature']
            )

            # None이 반환되면 원본 사용
            if result is None:
                result = answer_text
            else:
                result = result.strip()
        else:
            # API 키가 없으면 원본 텍스트 사용
            result = answer_text
        
        # 불필요한 접두사 및 문구 제거
        for prefix in PREFIXES_TO_REMOVE:
            if result.startswith(prefix):
                result = result[len(prefix):].strip()
                break
        
        # 중간에 나타날 수 있는 불필요한 문구들 제거
        for phrase in UNWANTED_PHRASES:
            if phrase in result:
                # 해당 문구가 포함된 줄을 제거
                lines = result.split('\n')
                result = '\n'.join([line for line in lines if phrase not in line])
        
        # 텍스트 형식 개선 (설정에 따라 적용)
        import re
        
        # 마침표 뒤 엔터값 추가 (마침표+공백 패턴만)
        if AUTO_PARAGRAPH_BREAK['enable_sentence_break']:
            result = re.sub(r'[.] ', '.\n\n', result)
        
        # 추가 텍스트 처리 규칙 적용
        if TEXT_PROCESSING_RULES['remove_double_spaces']:
            # 이중 공백 제거 (줄바꿈은 제외)
            result = re.sub(r' {2,}', ' ', result)
        
        if TEXT_PROCESSING_RULES['unify_punctuation']:
            # 문장부호 통일
            result = result.replace('。', '.')
            result = result.replace('、', ',')
        
        if TEXT_PROCESSING_RULES['normalize_line_breaks']:
            # 줄바꿈 정규화 (3개 이상의 연속 줄바꿈을 2개로)
            result = re.sub(r'\n{3,}', '\n\n', result)
        
        return {
            'improved_answer': result or answer_text,
            'suggestions': ""
        }
        
    except Exception as e:
        return {
            'improved_answer': answer_text,
            'suggestions': ""
        }

def check_answer_consistency(answers, textarea_answers):
    """
    답변들 간의 일관성을 체크합니다.
    """
    consistency_issues = []
    
    # 시스템 관련 일관성 체크
    if len(answers) > Q_ID['commercial_sw']:
        # 상용소프트웨어 사용 여부와 내부 수정 가능성 체크
        is_commercial = answers[Q_ID['commercial_sw']] == 'Y'
        can_modify = answers[Q_ID['pc_can_modify']] == 'Y' if len(answers) > Q_ID['pc_can_modify'] else False

        if is_commercial and can_modify:
            consistency_issues.append("상용소프트웨어를 사용하면서 내부에서 주요 로직을 수정할 수 있다고 답변하셨습니다. 일반적으로 상용소프트웨어는 내부 수정이 제한적입니다.")

    # 클라우드 관련 일관성 체크
    if len(answers) > Q_ID['cloud_type']:
        uses_cloud = answers[Q_ID['use_cloud']] == 'Y'
        cloud_type = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else ""

        if uses_cloud and not cloud_type.strip():
            consistency_issues.append("클라우드 서비스를 사용한다고 답변하셨지만 클라우드 종류가 선택되지 않았습니다.")

    # 권한 관리 일관성 체크
    if len(answers) > Q_ID['apd02_revoke_history']:
        has_auth_history = answers[Q_ID['apd01_auth_history']] == 'Y'
        has_revoke_history = answers[Q_ID['apd02_revoke_history']] == 'Y'

        if has_auth_history and not has_revoke_history:
            consistency_issues.append("권한 부여 이력은 기록되지만 권한 회수 이력은 기록되지 않는다고 답변하셨습니다.")

    # 데이터베이스 관련 일관성 체크
    if len(answers) > Q_ID['apd09_db_history']:
        has_data_change_history = answers[Q_ID['apd07_data_history']] == 'Y' if len(answers) > Q_ID['apd07_data_history'] else False
        has_db_auth_history = answers[Q_ID['apd09_db_history']] == 'Y'

        if has_db_auth_history and not has_data_change_history:
            consistency_issues.append("DB 접근권한 이력은 기록되지만 데이터 변경 이력은 기록되지 않는다고 답변하셨습니다.")
    
    return consistency_issues

def ai_improve_answer_consistency(answers, textarea_answers):
    """
    AI를 사용하여 답변들의 일관성을 체크하고 개선 제안을 제공합니다.
    """
    try:
        api_key = os.getenv('OPENAI_API_KEY')
        if not api_key:
            return {
                'consistency_check': "OpenAI API 키가 설정되지 않았습니다.",
                'suggestions': []
            }

        client = OpenAI(api_key=api_key)
        model_name = os.getenv('OPENAI_MODEL', AI_MODEL_CONFIG['model'])

        # 조건부 질문 필터링 적용
        filtered_questions = get_conditional_questions(answers)

        # 중요한 답변들만 선별하여 컨텍스트 구성
        key_answers = []
        for question in filtered_questions:
            question_index = question['index']
            if question_index < len(answers) and answers[question_index]:
                answer_text = answers[question_index]
                if question_index < len(textarea_answers) and textarea_answers[question_index]:
                    answer_text += f" ({textarea_answers[question_index]})"
                key_answers.append(f"Q{question_index+1}: {question['text']} -> A: {answer_text}")

        context = "\n".join(key_answers[:20])  # 처음 20개 질문만 사용

        prompt = f"""다음은 ITGC 인터뷰의 질문과 답변들입니다. 답변들 간의 논리적 일관성을 검토하고 개선 제안을 해주세요.

{context}

검토 기준:
1. 시스템 아키텍처 관련 답변들의 일관성
2. 보안 정책 및 권한 관리의 일관성
3. 프로세스 및 절차의 일관성
4. 기술적 구성요소들 간의 호환성

응답 형식:
일관성 검토: [전체적인 일관성 평가]
개선 제안: [구체적인 개선 사항들을 번호로 나열]"""

        messages = [
            {"role": "system", "content": "당신은 ITGC 전문가입니다. 답변의 논리적 일관성을 검토하고 개선 제안을 제공합니다."},
            {"role": "user", "content": prompt}
        ]

        # OpenAI API 호출 (retry 및 rate limiting 포함)
        result = _call_openai_api_with_retry(
            client,
            messages,
            model_name,
            600,
            0.0  # 완전히 일관된 결과
        )

        # None이 반환되면 기본 메시지
        if result is None:
            result = "AI 일관성 검토를 완료할 수 없습니다. (API 오류)"
        
        # 기본 일관성 체크도 추가
        basic_issues = check_answer_consistency(answers, textarea_answers)
        
        return {
            'ai_consistency_check': result,
            'basic_consistency_issues': basic_issues
        }
        
    except Exception as e:
        basic_issues = check_answer_consistency(answers, textarea_answers)
        return {
            'ai_consistency_check': f"AI 일관성 체크 중 오류가 발생했습니다: {str(e)}",
            'basic_consistency_issues': basic_issues
        }

def get_ai_review(content, control_number=None, answers=None):
    """
    AI를 사용하여 ITGC 내용을 검토하고 개선 제안을 반환합니다.
    Summary 시트의 C열(검토결과), D열(결론), E열(개선필요사항)에 맞는 구조화된 결과를 반환합니다.
    answers: 인터뷰 답변 리스트 (Package S/W 상황 판단용)
    """
    try:
        # 모집단 확보 불가 시 무조건 Ineffective 판정하는 통제 목록
        mandatory_ineffective_controls = ['APD01', 'APD07', 'APD09', 'APD12', 'PC01', 'PC02', 'PC03', 'CO01']
        
        # 모집단 확보 불가 키워드 체크
        population_unavailable_keywords = [
            '모집단 확보가 불가', '모집단 확보 불가', '모집단이 없', '모집단 부족',
            '이력이 없', '기록이 없', '데이터가 없', '자료가 없',
            '수행하지 않', '운영하지 않', '절차가 없'
        ]
        
        # 해당 통제이고 모집단 확보 불가 키워드가 포함된 경우 자동 Ineffective 판정
        if control_number in mandatory_ineffective_controls:
            content_lower = content.lower()
            for keyword in population_unavailable_keywords:
                if keyword in content:
                    return {
                        'review_result': f"모집단 확보가 불가능한 상황입니다. {control_number} 통제의 경우 모집단이 확보되지 않으면 통제 효과성을 평가할 수 없으므로 미비점으로 판정됩니다.",
                        'conclusion': "Ineffective",
                        'improvements': f"모집단 확보를 위한 시스템 구축 또는 절차 수립이 필요합니다. {control_number} 통제가 효과적으로 운영되기 위해서는 관련 데이터의 완전성과 정확성이 보장되어야 합니다."
                    }

        # 통제별 특정 기준만 가져오기 (토큰 절약)
        specific_criteria = CONTROL_SPECIFIC_CRITERIA.get(control_number, [])
        
        # 특별 상황 확인 및 컨텍스트 생성
        special_context = ""
        if answers and len(answers) > 4:
            is_package_sw = answers[Q_ID['commercial_sw']] == 'Y'
            cloud_type = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else None
            cannot_modify = answers[Q_ID['pc_can_modify']] == 'N' if len(answers) > Q_ID['pc_can_modify'] else False
            
            # Cloud 서비스 관련 컨텍스트 - 실제 스킵된 통제들만 N/A 처리
            has_soc1_report = len(answers) > 5 and answers[5] and str(answers[5]).upper() == 'Y'
            
            if cloud_type and has_soc1_report:
                # 스킵된 통제 목록 확인
                skipped_controls = get_skipped_controls(answers)
                
                # 실제로 스킵된 통제인 경우에만 N/A 컨텍스트 추가
                if control_number in skipped_controls:
                    if cloud_type == 'SaaS':
                        if control_number in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11']:
                            special_context += "\n중요: SaaS 환경에서는 데이터베이스 관리가 서비스 제공업체에서 담당합니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                        elif control_number in ['APD12', 'APD13']:
                            special_context += "\n중요: SaaS 환경에서는 운영체제 관리가 서비스 제공업체에서 담당합니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                        elif control_number and control_number.startswith('PC'):
                            special_context += "\n중요: SaaS 환경에서는 프로그램 변경이 서비스 제공업체에서 담당합니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                        elif control_number and control_number.startswith('CO') and control_number in ['CO04', 'CO05', 'CO06']:
                            special_context += "\n중요: SaaS 환경에서는 장애 대응 및 백업이 서비스 제공업체에서 담당합니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                    
                    elif cloud_type == 'PaaS':
                        if control_number in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11']:
                            special_context += "\n중요: PaaS 환경에서는 데이터베이스 관리가 플랫폼에서 담당됩니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                        elif control_number in ['APD12', 'APD13']:
                            special_context += "\n중요: PaaS 환경에서는 운영체제 관리가 플랫폼에서 담당됩니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                        elif control_number and control_number.startswith('CO') and control_number in ['CO04', 'CO05', 'CO06']:
                            special_context += "\n중요: PaaS 환경에서는 인프라 수준의 장애 대응 및 백업이 플랫폼에서 담당됩니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                    
                    elif cloud_type == 'IaaS':
                        if control_number in ['APD11', 'APD13']:
                            special_context += "\n중요: IaaS 환경에서는 관리자 권한이 클라우드 제공업체에서 관리됩니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
                        elif control_number and control_number.startswith('CO') and control_number in ['CO04', 'CO05', 'CO06']:
                            special_context += "\n중요: IaaS 환경에서는 인프라 수준의 장애 대응 및 백업이 클라우드 제공업체에서 담당됩니다. 이 통제는 질문이 스킵되어 적용되지 않으므로(N/A) 미비점이 아닙니다."
            
            # Package S/W 관련 컨텍스트
            if is_package_sw and cannot_modify:
                if control_number and control_number.startswith('PC'):
                    special_context += "\n중요: 이 시스템은 Package S/W이며 회사내에서 수정이 불가능합니다. 프로그램 변경(PC) 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
        
        # 배치 스케줄 상황 확인
        if answers and len(answers) > 6:
            has_batch_schedule = answers[40] == 'Y'  # 질문40: 배치 스케줄 유무
            
            if not has_batch_schedule:
                if control_number and control_number in ['CO01', 'CO02', 'CO03']:
                    special_context += "\n중요: 현재 실행중인 배치 스케줄이 없습니다. 배치 관련 통제(CO01-CO03)는 적용되지 않으므로(N/A) 미비점이 아닙니다."
        
        # 공통 감사기준 포함
        common_criteria_text = "공통기준: " + " | ".join(COMMON_AUDIT_CRITERIA)
        
        # 통제별 특정 기준 텍스트 생성
        if specific_criteria:
            specific_criteria_text = f"특정기준: " + " | ".join(specific_criteria[:3])  # 토큰 절약을 위해 최대 3개만
        else:
            specific_criteria_text = "특정기준: 표준 ITGC 검토 기준 적용"

        # Cloud 환경 정보 추가
        cloud_info = ""
        if cloud_type:
            cloud_info = f"\n**시스템 환경: {cloud_type} Cloud 서비스**"
        
        prompt = f"""ITGC {control_number} 검토:
{content}

{common_criteria_text}
{specific_criteria_text}{cloud_info}{special_context}

**검토 기준: 전문적이고 균형잡힌 감사 접근**
- 실제 사용자가 작성한 답변 내용을 기반으로 평가 (시스템 자동 생성 문구는 제외)
- 통제의 설계와 운영 상태를 실질적으로 분석
- 이력 기록 여부, 승인 절차 존재 여부, 권한 관리 적절성 등을 종합적으로 판단
- Cloud 서비스 환경에서는 제공업체가 담당하는 영역은 N/A 처리하여 미비점으로 판정하지 않음
- 소규모 조직의 현실적 제약을 고려하되 핵심 통제 요구사항은 충족되어야 함
- 명확한 통제 미비점이 확인되는 경우에만 Ineffective 판정
- 적절한 통제가 운영되고 있다면 Effective 판정

응답형식 (정확히 이 형식으로만 답변):
검토결과: [전문적 감사 관점에서 엄격하게 분석하되 정확한 서술형으로 작성]
결론: Effective (또는 결론: Ineffective - 반드시 이 두 단어 중 하나만 단독으로 사용)
개선필요사항: [Ineffective시 구체적이고 실행가능한 개선방안을 서술형으로 제시, Effective시 빈칸]

중요 작성 지침: 
1. 결론 부분에는 반드시 'Effective' 또는 'Ineffective' 단어만 사용할 것
2. 시스템이 자동으로 생성한 문구는 평가에서 제외하고, 실제 사용자 답변 내용만 분석할 것
3. 검토결과와 개선필요사항은 음슴체가 아닌 정확한 서술형으로 작성할 것 (예: "~입니다", "~되어야 합니다", "~것으로 판단됩니다")
4. 예시: "ITSM을 통해 요청서 작성 후 팀장 승인을 받아 권한 부여" → 사용자가 직접 작성한 절차 설명이므로 평가 대상
"""

        api_key = os.getenv('OPENAI_API_KEY')
        if not api_key:
            return {
                'review_result': "OpenAI API 키가 설정되지 않았습니다.",
                'conclusion': "검토 불가",
                'improvements': "OPENAI_API_KEY 환경변수를 설정해주세요."
            }

        client = OpenAI(api_key=api_key)
        model_name = os.getenv('OPENAI_MODEL', AI_MODEL_CONFIG['model'])

        messages = [
            {"role": "system", "content": "당신은 전문적이고 균형잡힌 ITGC 감사 전문가입니다. 명백한 미비점은 정확히 식별하되, 적절히 운영되는 통제는 Effective로 인정하는 합리적 판단을 수행합니다. 모든 검토 의견은 정확한 서술형으로 작성하며, 음슴체는 사용하지 않습니다."},
            {"role": "user", "content": prompt}
        ]

        # OpenAI API 호출 (retry 및 rate limiting 포함)
        ai_response = _call_openai_api_with_retry(
            client,
            messages,
            model_name,
            400,
            0.0  # 완전히 일관된 결과 (동일 입력 = 동일 출력)
        )

        # None이 반환되면 기본 응답
        if ai_response is None:
            return {
                'review_result': "AI API 호출 중 오류가 발생했습니다.",
                'conclusion': "검토 불가",
                'improvements': "API 연결을 확인하고 재시도하세요."
            }

        # AI 응답을 파싱하여 각 컬럼별로 분리
        result = {
            'review_result': '',  # C열: 검토결과
            'conclusion': '',     # D열: 결론
            'improvements': ''    # E열: 개선필요사항
        }

        lines = ai_response.split('\n')
        current_section = None

        for line in lines:
            line = line.strip()
            if line.startswith('검토결과:'):
                result['review_result'] = line.replace('검토결과:', '').strip()
                current_section = 'review_result'
            elif line.startswith('결론:'):
                result['conclusion'] = line.replace('결론:', '').strip()
                current_section = 'conclusion'
            elif line.startswith('개선필요사항:'):
                result['improvements'] = line.replace('개선필요사항:', '').strip()
                current_section = 'improvements'
            elif line and current_section:
                # 다음 섹션이 시작되기 전까지 내용을 이어서 추가
                if not any(line.startswith(prefix) for prefix in ['검토결과:', '결론:', '개선필요사항:']):
                    if result[current_section]:
                        result[current_section] += '\n' + line
                    else:
                        result[current_section] = line

        return result

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return {
            'review_result': f"AI 검토 중 오류 발생: {str(e)}\n\n상세 에러:\n{error_details[:500]}",
            'conclusion': "검토 불가",
            'improvements': "AI 검토 시스템 점검 필요"
        }

# --- 리팩토링: 시트 값 입력 함수 ---
def fill_sheet(ws, text_data, answers):
    if 'A1' in text_data:
        ws['C7'] = text_data['A1']
    if 'B1' in text_data:
        ws['C8'] = text_data['B1']
    
    # C12: AI로 다듬어진 내용 또는 원본 내용
    if 'C2' in text_data:
        ws['C12'] = text_data['C2']  # AI로 다듬어진 내용
        # 행 높이 조정 (AI 다듬어진 내용 기준)
        value = str(text_data['C2'])
        num_lines = value.count('\n') + 1
        approx_lines = num_lines + (len(value) // 50)
        ws.row_dimensions[12].height = 15 * approx_lines
        
    # C2가 없는 경우 (AI 비활성화 상태)
    elif 'B2' in text_data:
        ws['C12'] = text_data['B2']
        # 행 높이 조정
        value = str(text_data['B2'])
        num_lines = value.count('\n') + 1
        approx_lines = num_lines + (len(value) // 50)
        ws.row_dimensions[12].height = 15 * approx_lines
        
    # B3: 시스템명, B5: user_email
    if len(answers) > 1 and answers[1]:
        system_name = answers[1]  # 1번: 시스템 이름
        ws['B3'] = system_name
    
    if len(answers) > 0 and answers[0]:
        user_email = answers[0]  # 0번: 이메일 주소
        ws['B5'] = user_email

# 텍스트 길이 체크 (토큰 사용량 최적화)
def should_apply_ai_refinement(text):
    """AI 다듬기를 적용할지 결정 (설정 기반)"""
    text_len = len(text)
    return TEXT_LENGTH_LIMITS['min_length'] <= text_len <= TEXT_LENGTH_LIMITS['max_length']

# 공통 템플릿 처리 함수
def build_control_text(control_config, answers, textarea_answers):
    """ITGC 통제 텍스트를 공통 로직으로 생성"""
    control_type = control_config.get('type', 'history_procedure')
    
    if control_type == 'complex_db':
        # APD09 방식 (DB 복합)
        db_info = f"DB 종류와 버전: {answers[control_config['db_type_idx']]}"
        db_tool = f"DB 접근제어 Tool 사용 여부: {'사용' if answers[control_config['db_tool_idx']] == 'Y' else '미사용'}"
        
        history_status = (control_config['history_yes'] if answers[control_config['history_idx']] == 'Y' 
                         else control_config['history_no'])
        
        if answers[control_config['procedure_idx']] == 'Y':
            textarea_content = textarea_answers[control_config['textarea_idx']] if textarea_answers[control_config['textarea_idx']] else control_config['default_msg']
            procedure_text = f"{control_config['procedure_prefix']}\n\n{textarea_content}"
        else:
            procedure_text = control_config['procedure_no']
            
        return f"{db_info}\n\n{db_tool}\n\nDB 접근권한 부여 이력이 시스템에 {history_status}\n\n{procedure_text}"
    
    elif control_type == 'complex_os':
        # APD12 방식 (OS 복합)
        os_info = f"OS 종류와 버전: {answers[control_config['os_type_idx']]}"
        os_tool = f"OS 접근제어 Tool 사용 여부: {'사용' if answers[control_config['os_tool_idx']] == 'Y' else '미사용'}"
        
        history_status = (control_config['history_yes'] if answers[control_config['history_idx']] == 'Y' 
                         else control_config['history_no'])
        
        if answers[control_config['procedure_idx']] == 'Y':
            textarea_content = textarea_answers[control_config['textarea_idx']] if textarea_answers[control_config['textarea_idx']] else control_config['default_msg']
            procedure_text = f"{control_config['procedure_prefix']}\n\n{textarea_content}"
        else:
            procedure_text = control_config['procedure_no']
            
        return f"{os_info}\n\n{os_tool}\n\nOS 접근권한 부여 이력이 시스템에 {history_status}\n\n{procedure_text}"
    
    elif control_type == 'history_procedure' or 'history_idx' in control_config:
        # 기존 APD01, APD02 방식 (이력 + 절차)
        history_status = (control_config['history_yes'] if answers[control_config['history_idx']] == 'Y' 
                         else control_config['history_no'])
        
        if 'procedure_idx' in control_config:
            if answers[control_config['procedure_idx']] == 'Y':
                textarea_content = textarea_answers[control_config['textarea_idx']] if textarea_answers[control_config['textarea_idx']] else control_config['default_msg']
                procedure_text = f"{control_config['procedure_prefix']}\n\n{textarea_content}"
            else:
                procedure_text = control_config['procedure_no']
        else:
            procedure_text = ""
            
        return control_config['template'].format(
            history_status=history_status,
            procedure_text=procedure_text
        )
    
    elif control_type == 'simple_procedure':
        # APD03 방식 (단순 절차)
        if answers[control_config['procedure_idx']] == 'Y':
            textarea_content = textarea_answers[control_config['textarea_idx']] if textarea_answers[control_config['textarea_idx']] else control_config['default_msg']
            return f"{control_config['procedure_prefix']}\n\n{textarea_content}"
        else:
            return control_config['procedure_no']
    
    elif control_type == 'simple_list':
        # APD04, APD06 방식 (단순 리스트)
        content = answers[control_config['answer_idx']] if answers[control_config['answer_idx']] else control_config['default_msg']
        return control_config['template'].format(content=content)
    
    elif control_type == 'simple_status':
        # APD05, PC05 방식 (단순 상태)
        status = (control_config['status_yes'] if answers[control_config['answer_idx']] == 'Y'
                 else control_config['status_no'])
        return control_config['template'].format(status=status)

    elif control_type == 'simple_yn_tool':
        # APD16 방식 (Y이면 제품명 포함, N이면 미사용)
        if answers[control_config['answer_idx']] == 'Y':
            tool_name = textarea_answers[control_config['textarea_idx']] if textarea_answers[control_config['textarea_idx']] else '솔루션명 미기재'
            return control_config['template_yes'].format(tool_name=tool_name)
        else:
            return control_config['template_no']


    return control_config.get('default_msg', '상세 기술이 제공되지 않았습니다.')

def get_text_itgc(answers, control_number, textarea_answers=None, enable_ai_review=False):
    result = {}
    if textarea_answers is None:
        textarea_answers = [''] * len(answers)

    # 조건부 질문 생략 로직 적용
    if answers and len(answers) >= 8:
        # Cloud 환경에 따른 스킵 처리 추가
        cloud_type = answers[4] if len(answers) > 4 else None
        has_soc1_report = len(answers) > 5 and answers[5] and str(answers[5]).upper() == 'Y'
        
        # Cloud 스킵 조건 확인
        if cloud_type and has_soc1_report:
            cloud_skip_message = ""
            if cloud_type == 'SaaS':
                if control_number in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11',
                                    'APD12', 'APD13', 'APD14', 'PC01', 'PC02', 'PC03', 'PC04', 'PC05',
                                    'CO05', 'CO06', 'CO07']:
                    cloud_skip_message = f"SaaS 환경에서는 서비스 제공업체가 담당합니다. {control_number} 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
            elif cloud_type == 'PaaS':
                if control_number in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11',
                                    'APD12', 'APD13', 'APD14', 'CO05', 'CO06', 'CO07']:
                    cloud_skip_message = f"PaaS 환경에서는 플랫폼 제공업체가 담당합니다. {control_number} 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
            elif cloud_type == 'IaaS':
                if control_number in ['APD11', 'APD13', 'CO06']:
                    cloud_skip_message = f"IaaS 환경에서는 인프라 제공업체가 담당합니다. {control_number} 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
            
            if cloud_skip_message:
                result['A1'] = control_number
                result['B1'] = ITGC_CONTROLS.get(control_number, {}).get('title', control_number)
                result['B2'] = cloud_skip_message
                result['C2'] = result['B2']
                return result
        
        # 조건부 N/A 처리

        if control_number in ['APD01', 'APD02', 'APD03'] and answers[Q_ID['apd15_shared_account']] == 'N':
            result['A1'] = control_number
            result['B1'] = ITGC_CONTROLS.get(control_number, {}).get('title', control_number)
            shared_mgmt = textarea_answers[Q_ID['apd15_shared_mgmt']] if textarea_answers and len(textarea_answers) > Q_ID['apd15_shared_mgmt'] and textarea_answers[Q_ID['apd15_shared_mgmt']] else ''
            na_base = f"사용자 ID가 팀·부서 단위 공용으로 발급되어 개인 단위 권한 부여·회수 절차를 적용하기 어렵습니다. {control_number} 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
            if shared_mgmt:
                result['B2'] = f"{na_base}\n\n[공용 계정 환경 관리 방식]\n{shared_mgmt}"
            else:
                result['B2'] = na_base
            result['C2'] = result['B2']
            return result

        if control_number in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11'] and answers[Q_ID['apd07_db_access']] == 'N':
            result['A1'] = control_number
            result['B1'] = ITGC_CONTROLS.get(control_number, {}).get('title', control_number)
            result['B2'] = f"회사에서 DB에 접속하여 필요한 작업을 수행할 수 없습니다. {control_number} 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
            result['C2'] = result['B2']
            return result

        if control_number in ['APD12', 'APD13', 'APD14'] and answers[Q_ID['apd12_os_access']] == 'N':
            result['A1'] = control_number
            result['B1'] = ITGC_CONTROLS.get(control_number, {}).get('title', control_number)
            result['B2'] = f"회사에서 OS서버에 접속하여 필요한 작업을 수행할 수 없습니다. {control_number} 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
            result['C2'] = result['B2']
            return result

        if control_number in ['CO01', 'CO02', 'CO03'] and answers[Q_ID['co_has_batch']] == 'N':
            result['A1'] = control_number
            result['B1'] = ITGC_CONTROLS.get(control_number, {}).get('title', control_number)
            result['B2'] = f"현재 실행중인 배치 스케줄이 없습니다. {control_number} 통제는 적용되지 않으므로(N/A) 미비점이 아닙니다."
            result['C2'] = result['B2']
            return result
        

    # 공통 로직으로 처리
    result['A1'] = control_number
    
    if control_number in ITGC_CONTROLS:
        config = ITGC_CONTROLS[control_number]
        result['B1'] = config['title']
        result['B2'] = build_control_text(config, answers, textarea_answers)

    # APD03-CO06은 이제 모두 ITGC_CONTROLS에서 처리됨
    
    else:
        # 알 수 없는 통제 번호 처리
        result['A1'] = f"Unknown control number: {control_number}"
        result['B1'] = ""
        result['B2'] = "알 수 없는 통제 번호입니다."

    # AI 기능 적용 (토큰 사용량 최적화)
    if enable_ai_review and 'B2' in result and result['B2']:
        # 스킵된 통제인지 확인
        skipped_controls = get_skipped_controls(answers)
        is_skipped = control_number in skipped_controls
        
        if is_skipped:
            result['C2'] = result['B2']
        else:
            # 스킵되지 않은 통제만 AI 검토 수행
            text_length = len(result['B2'])
            should_apply_ai = should_apply_ai_refinement(result['B2'])
            
            if should_apply_ai:
                
                improved_text = ai_improve_interview_answer("", result['B2'])  # 질문 텍스트 제거로 토큰 절약
                if improved_text and improved_text.get('improved_answer'):
                    result['C2'] = improved_text['improved_answer']
                else:
                    result['C2'] = result['B2']
                
                # AI 검토 수행
                ai_review = get_ai_review(result['C2'], control_number, answers)
                result['AI_Review'] = ai_review
                result['AI_Summary'] = ai_review
            else:
                result['C2'] = result['B2']
                
                # 길이 제한 때문에 AI 다듬기는 건너뛰지만, 스킵되지 않은 통제는 Summary용 AI 검토 수행
                ai_review = get_ai_review(result['B2'], control_number, answers)
                result['AI_Review'] = ai_review
                result['AI_Summary'] = ai_review
    else:
        result['C2'] = result['B2']

    return result

# link2_prev의 핵심 로직만 분리 (세션 객체는 snowball.py에서 전달)
def link2_prev_logic(session):
    question_index = session.get('question_index', 0)
    
    if question_index > 0:
        # 조건부 질문 생략 로직 (역방향)
        answers = session.get('answer', [])

        if question_index == Q_ID['apd12_os_access']:  # OS서버 접속 가능 여부
            answer_db = answers[Q_ID['apd07_db_access']] if len(answers) > Q_ID['apd07_db_access'] else ''
            if answer_db == 'N':
                session['question_index'] = Q_ID['apd07_db_access']
            else:
                session['question_index'] = Q_ID['apd11_db_password']

        elif question_index == Q_ID['pc_can_modify']:  # 주요 로직 수정 가능 여부
            answer_os = answers[Q_ID['apd12_os_access']] if len(answers) > Q_ID['apd12_os_access'] else ''
            if answer_os == 'N':
                session['question_index'] = Q_ID['apd12_os_access']
            else:
                session['question_index'] = Q_ID['apd16_vpn']

        elif question_index == Q_ID['co_has_batch']:  # 배치 스케줄 유무
            answer_pc = answers[Q_ID['pc_can_modify']] if len(answers) > Q_ID['pc_can_modify'] else ''
            if answer_pc == 'N':
                session['question_index'] = Q_ID['pc_can_modify']
            else:
                session['question_index'] = Q_ID['pc05_dev_env']

        elif question_index == Q_ID['apd01_auth_history']:  # 사용자 권한부여 이력
            answer_cloud_type = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else ''
            answer_soc1 = answers[Q_ID['soc1_report']] if len(answers) > Q_ID['soc1_report'] else ''
            answer_use_cloud = answers[Q_ID['use_cloud']] if len(answers) > Q_ID['use_cloud'] else ''

            if answer_cloud_type == 'SaaS' and answer_soc1 == 'Y':
                session['question_index'] = Q_ID['soc1_report']
            elif answer_cloud_type == 'PaaS' and answer_soc1 == 'Y':
                session['question_index'] = Q_ID['apd15_shared_account']
            elif answer_use_cloud == 'N':
                session['question_index'] = Q_ID['use_cloud']
            else:
                session['question_index'] = Q_ID['soc1_report']

        elif question_index == Q_ID['apd15_shared_mgmt']:  # 공용 계정 관리 방식 (Q9, apd15=N 시만 표시)
            session['question_index'] = Q_ID['apd15_shared_account']

        elif question_index == Q_ID['apd01_procedure']:  # 신규 권한 승인 절차 (Q10)
            # apd15=Y 시 Q9(shared_mgmt)는 스킵 → Q8로
            session['question_index'] = Q_ID['apd15_shared_account']

        elif question_index == Q_ID['apd02_procedure']:  # 권한 회수 절차 (Q11)
            session['question_index'] = Q_ID['apd01_procedure']

        elif question_index == Q_ID['apd03_procedure']:  # 퇴사자 접근권한 차단 (Q12)
            session['question_index'] = Q_ID['apd02_procedure']

        elif question_index == Q_ID['apd04_admin']:  # Application 관리자 권한자 (Q13)
            answer_apd15 = answers[Q_ID['apd15_shared_account']] if len(answers) > Q_ID['apd15_shared_account'] else ''
            if answer_apd15 == 'N':
                # Q10·Q11·Q12 모두 스킵 → Q9(공용 계정 관리 방식)로
                session['question_index'] = Q_ID['apd15_shared_mgmt']
            else:
                session['question_index'] = Q_ID['apd03_procedure']

        elif question_index == Q_ID['cloud_type']:  # Cloud 종류: 이전 → use_cloud
            session['question_index'] = Q_ID['use_cloud']

        elif question_index == Q_ID['soc1_report']:  # SOC1 Report: 이전 → cloud_type
            session['question_index'] = Q_ID['cloud_type']

        elif question_index in [Q_ID['apd09_db_type'], Q_ID['apd09_db_tool'], Q_ID['apd09_db_history'],
                                 Q_ID['apd09_procedure'], Q_ID['apd10_db_admin'], Q_ID['apd11_db_password']]:  # DB 관련
            if question_index == Q_ID['apd11_db_password']:
                answer_cloud = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else ''
                answer_soc1 = answers[Q_ID['soc1_report']] if len(answers) > Q_ID['soc1_report'] else ''
                if answer_cloud == 'IaaS' and answer_soc1 == 'Y':
                    # IaaS+SOC1: apd10_db_admin 스킵됨 → apd09_procedure로
                    session['question_index'] = Q_ID['apd09_procedure']
                else:
                    session['question_index'] = question_index - 1  # apd10_db_admin
            else:
                session['question_index'] = question_index - 1

        elif question_index in [Q_ID['apd12_os_type'], Q_ID['apd12_os_tool'], Q_ID['apd12_os_history'],
                                 Q_ID['apd12_procedure'], Q_ID['apd13_os_admin'], Q_ID['apd14_os_password']]:  # OS 관련
            if question_index == Q_ID['apd14_os_password']:
                answer_cloud = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else ''
                answer_soc1 = answers[Q_ID['soc1_report']] if len(answers) > Q_ID['soc1_report'] else ''
                if answer_cloud == 'IaaS' and answer_soc1 == 'Y':
                    # IaaS+SOC1: apd13_os_admin 스킵됨 → apd12_procedure로
                    session['question_index'] = Q_ID['apd12_procedure']
                else:
                    session['question_index'] = question_index - 1  # apd13_os_admin
            else:
                session['question_index'] = question_index - 1

        elif question_index == Q_ID['pc01_change_history']:  # 프로그램 변경 이력
            answer_cloud = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else ''
            answer_soc1 = answers[Q_ID['soc1_report']] if len(answers) > Q_ID['soc1_report'] else ''
            if answer_cloud == 'PaaS' and answer_soc1 == 'Y':
                session['question_index'] = Q_ID['apd15_shared_account']
            else:
                session['question_index'] = Q_ID['pc_can_modify']

        elif question_index in [Q_ID['pc01_procedure'], Q_ID['pc02_procedure'], Q_ID['pc03_procedure'],
                                 Q_ID['pc04_deploy_auth'], Q_ID['pc05_dev_env']]:  # PC 관련
            session['question_index'] = Q_ID['pc_can_modify']

        elif question_index in [Q_ID['co01_batch_tool'], Q_ID['co01_batch_history'], Q_ID['co01_procedure'],
                                 Q_ID['co02_batch_auth'], Q_ID['co03_monitoring']]:  # 배치 관련
            session['question_index'] = Q_ID['co_has_batch']

        elif question_index in [Q_ID['co04_incident'], Q_ID['co05_backup'],
                                 Q_ID['co06_server_room'], Q_ID['co07_security_patch']]:  # CO04~CO07
            answer_cloud = answers[Q_ID['cloud_type']] if len(answers) > Q_ID['cloud_type'] else ''
            answer_soc1 = answers[Q_ID['soc1_report']] if len(answers) > Q_ID['soc1_report'] else ''
            answer_batch = answers[Q_ID['co_has_batch']] if len(answers) > Q_ID['co_has_batch'] else ''

            if answer_cloud == 'SaaS' and answer_soc1 == 'Y':
                session['question_index'] = Q_ID['apd15_shared_account']
            elif answer_cloud in ['PaaS', 'IaaS'] and answer_soc1 == 'Y':
                session['question_index'] = Q_ID['co03_monitoring']
            elif answer_batch == 'N':
                session['question_index'] = Q_ID['co_has_batch']
            else:
                if question_index == Q_ID['co04_incident']:
                    session['question_index'] = Q_ID['co03_monitoring']
                else:
                    session['question_index'] = question_index - 1
        else:
            session['question_index'] = question_index - 1
    
    return session

def export_interview_excel_and_send(answers, textarea_answers, get_text_itgc, fill_sheet, is_ineffective, send_gmail_with_attachment, enable_ai_review=False, progress_callback=None):
    """
    인터뷰 답변을 받아 엑셀 파일을 생성하고 메일로 전송합니다.
    진행률 처리 개선 및 서버 환경 호환성 강화
    answers: list (사용자 답변)
    textarea_answers: list (텍스트에어리어 답변)
    get_text_itgc: 텍스트 생성 함수
    fill_sheet: 시트 채우기 함수
    is_ineffective: 비효과적 통제 체크 함수
    send_gmail_with_attachment: 메일 전송 함수
    progress_callback: 진행률 업데이트 콜백 함수
    """
    today = datetime.today().strftime('%Y%m%d')

    # 파일명 생성 (한글 파일명 지원)
    if len(answers) > 1 and answers[1]:
        system_name = answers[1].strip()
        file_name = f"{system_name}_ITGC_{today}.xlsx"
    else:
        file_name = f"ITGC_System_{today}.xlsx"

    # 1. 템플릿 파일 불러오기
    template_path = os.path.join("static", "Design_Template.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
    wb = load_workbook(template_path)

    # 2. Summary sheet에 작성할 AI 검토 결과 수집
    summary_ai_reviews = {}

    # 3. 각 통제별 시트에 값 입력
    control_list = [
        'APD01', 'APD02', 'APD03', 'APD04', 'APD05', 'APD06', 'APD07', 'APD08', 'APD09', 'APD10', 'APD11', 'APD12', 'APD13', 'APD14',
        'APD15', 'APD16',
        'PC01', 'PC02', 'PC03', 'PC04', 'PC05',
        'CO01', 'CO02', 'CO03', 'CO04', 'CO05', 'CO06', 'CO07'
    ]
    
    # 스킵된 통제 목록 미리 가져오기
    skipped_controls = get_skipped_controls(answers)
    
    # AI 검토 대상 통제 결정 (로그인 상태에 따라)
    ai_review_controls = set()
    if enable_ai_review:
        from snowball import is_logged_in  # 로그인 상태 확인 함수 import
        
        if is_logged_in():
            # 로그인 상태: 스킵되지 않은 모든 통제 검토
            ai_review_controls = set(c for c in control_list if c not in skipped_controls)
        else:
            # 비로그인 상태: APD01, APD02, APD03만 검토 (스킵되지 않은 것만)
            limited_controls = ['APD01', 'APD02', 'APD03']
            ai_review_controls = set(c for c in limited_controls if c not in skipped_controls)
    
    total_controls = len(control_list)
    processed_controls = []
    failed_controls = []
    
    for idx, control in enumerate(control_list):
        # 진행률 계산 (20%에서 80% 사이에서 진행)
        progress_percent = 20 + int((idx / total_controls) * 60)
        
        if progress_callback:
            if enable_ai_review:
                progress_callback(progress_percent, f"AI가 {control} 통제를 검토하고 있습니다... ({idx+1}/{total_controls})")
            else:
                progress_callback(progress_percent, f"{control} 통제 문서를 생성하고 있습니다... ({idx+1}/{total_controls})")
        
        try:
            # 해당 통제가 AI 검토 대상인지 확인
            control_needs_ai_review = enable_ai_review and control in ai_review_controls
            text_data = get_text_itgc(answers, control, textarea_answers, control_needs_ai_review)
            ws = wb[control]
            fill_sheet(ws, text_data, answers)
            
            # 스킵된 통제인 경우 C12셀에 스킵 이유 표시
            if control in skipped_controls:
                # text_data에서 스킵 이유 가져오기 (B2 또는 C2에 있음)
                skip_reason = text_data.get('C2') or text_data.get('B2') or 'N/A (해당없음)'
                ws['C12'] = skip_reason
                # 행 높이도 조정
                value = str(skip_reason)
                num_lines = value.count('\n') + 1
                approx_lines = num_lines + (len(value) // 50)
                ws.row_dimensions[12].height = 15 * approx_lines
            
            processed_controls.append(control)
        except Exception as e:
            failed_controls.append((control, str(e)))

        # AI 검토 결과가 있는 경우 Summary 시트용 데이터 수집 (선택된 통제만)
        if control_needs_ai_review and 'AI_Review' in text_data and isinstance(text_data['AI_Review'], dict):
            summary_ai_reviews[control] = text_data['AI_Review']
        elif control_needs_ai_review:
            pass

        # AI 검토 결과가 있는 경우와 없는 경우에 따라 C14 처리 (선택된 통제만)
        ai_review_processed = False
        if control_needs_ai_review and 'AI_Review' in text_data:
            if isinstance(text_data['AI_Review'], dict):
                # 결론만 C14에 기록
                conclusion = text_data['AI_Review'].get('conclusion', '')
                ws['C14'] = conclusion
                ai_review_processed = True

                # AI 결론이 Ineffective인 경우 시트 탭 색상을 빨간색으로 변경
                if 'Ineffective' in conclusion:
                    ws.sheet_properties.tabColor = "FF0000"

                # 전체 AI 검토 결과를 C15 셀에 추가 (기존 기능 유지)
                # review_text = f"검토결과: {text_data['AI_Review'].get('review_result', '')}\n결론: {text_data['AI_Review'].get('conclusion', '')}\n개선필요사항: {text_data['AI_Review'].get('improvements', '')}"
                # ws['C15'] = f"=== AI 검토 결과 ===\n{review_text}"
                # # AI 검토 결과 셀의 높이 조정
                # ai_review_lines = review_text.count('\n') + 1
                # ws.row_dimensions[15].height = 15 * max(5, ai_review_lines // 3)
            else:
                # 기존 형식인 경우 그대로 C14에 기록
                # ws['C14'] = str(text_data['AI_Review'])
                ai_review_processed = True

                # 기존 형식에서도 Ineffective 체크
                if 'Ineffective' in str(text_data['AI_Review']):
                    ws.sheet_properties.tabColor = "FF0000"

        # AI 검토 결과가 없는 경우에만 기존 로직 적용
        if not ai_review_processed and is_ineffective(control, answers):
            ws['C13'] = ''
            ws['C14'] = 'Ineffective'
            ws.sheet_properties.tabColor = "FF0000"
            # ws.sheet_properties.tabColor = "FF0000"
        #else:
        #    ws['C13'] = '화면 증빙을 첨부해주세요'

    # 처리 결과 요약 출력
    if failed_controls:
        for control, error in failed_controls:
            pass

    # 4. Summary 시트 처리
    for control in summary_ai_reviews.keys():
        pass
    
    if enable_ai_review and summary_ai_reviews:
        # AI 검토가 활성화된 경우 Summary 시트 생성
        # 시스템 환경 정보 미리 정의 (try 블록 외부에서)
        cloud_type = answers[4] if len(answers) > 4 and answers[4] else '온프레미스'
        system_name = answers[1] if len(answers) > 1 else '시스템'
        
        # 스타일링 관련 import (try 블록 외부에서)
        from openpyxl.styles import Font, Alignment, PatternFill
        
        try:
            # Summary 시트 확인 (템플릿에 이미 있어야 함)
            if 'Summary' not in wb.sheetnames:
                return wb, processed_controls, failed_controls
            
            summary_ws = wb['Summary']
            
            # 스킵된 통제 목록 가져오기
            skipped_controls = get_skipped_controls(answers)
            
            # 전체 통제 목록 정의
            all_controls = [
                'APD01', 'APD02', 'APD03', 'APD04', 'APD05', 'APD06', 'APD07', 'APD08', 'APD09', 'APD10', 
                'APD11', 'APD12', 'APD13', 'APD14', 'PC01', 'PC02', 'PC03', 'PC04', 'PC05',
                'CO01', 'CO02', 'CO03', 'CO04', 'CO05', 'CO06'
            ]
            
            # 스킵 사유별 메시지를 동적으로 생성하는 함수
            def get_skip_reason_message(control, answers):
                """통제별로 스킵 사유에 맞는 메시지를 반환"""
                
                # Cloud 환경 체크
                cloud_type = answers[4] if len(answers) > 4 and answers[4] else None
                has_soc1_report = len(answers) > 5 and answers[5] and str(answers[5]).upper() == 'Y'
                
                if cloud_type and has_soc1_report:
                    if cloud_type == 'SaaS' and control in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11', 'APD12', 'APD13', 'APD14', 'PC01', 'PC02', 'PC03', 'PC04', 'PC05', 'CO05', 'CO06', 'CO07']:
                        return 'SaaS 환경에서는 서비스 제공업체가 해당 영역을 담당하므로 이 통제는 적용되지 않습니다.'
                    elif cloud_type == 'PaaS' and control in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11', 'APD12', 'APD13', 'APD14', 'CO05', 'CO06', 'CO07']:
                        return 'PaaS 환경에서는 플랫폼 제공업체가 해당 영역을 담당하므로 이 통제는 적용되지 않습니다.'
                    elif cloud_type == 'IaaS' and control in ['APD11', 'APD13', 'CO06']:
                        return 'IaaS 환경에서는 인프라 제공업체가 해당 영역을 담당하므로 이 통제는 적용되지 않습니다.'
                
                # Package S/W 체크
                if control.startswith('PC') and len(answers) > 31 and answers[31] == 'N':
                    return '상용소프트웨어 환경에서는 프로그램 변경이 제한되므로 이 통제는 적용되지 않습니다.'
                
                # DB 접근 불가 체크 (14번 답변 N -> 15~23번 스킵)
                if control in ['APD07', 'APD08', 'APD09', 'APD10', 'APD11'] and len(answers) > 14 and answers[14] == 'N':
                    return 'DB 직접 접근이 불가능한 환경이므로 이 통제는 적용되지 않습니다.'
                
                # OS 접근 불가 체크
                if control in ['APD12', 'APD13', 'APD14'] and len(answers) > 24 and answers[24] == 'N':
                    return 'OS 서버 직접 접근이 불가능한 환경이므로 이 통제는 적용되지 않습니다.'
                
                # 배치 스케줄 없음 체크
                if control in ['CO01', 'CO02', 'CO03'] and len(answers) > 38 and answers[38] == 'N':
                    return '배치 스케줄이 존재하지 않으므로 이 통제는 적용되지 않습니다.'
                
                # 기본 메시지
                return '현재 시스템 환경에서는 이 통제가 적용되지 않습니다.'
            
            # 전체 통제에 대해 Summary 시트 업데이트
            for control in all_controls:
                # Summary 시트에서 해당 통제번호가 있는 행 찾기
                found_row = None
                for row in range(1, summary_ws.max_row + 1):
                    cell_value = summary_ws[f'A{row}'].value
                    if cell_value and str(cell_value).strip() == control:
                        found_row = row
                        break
                
                if not found_row:
                    continue
                
                if control in summary_ai_reviews and isinstance(summary_ai_reviews[control], dict):
                    # AI 검토를 받은 통제: 검토 결과 작성
                    ai_review = summary_ai_reviews[control]
                    
                    # C열: 검토결과
                    review_result = ai_review.get('review_result', '')
                    if len(review_result) > 32767:  # 엑셀 셀 최대 문자 수 제한
                        review_result = review_result[:32760] + "..."
                    summary_ws[f'C{found_row}'] = review_result
                    
                    # D열: 결론
                    summary_ws[f'D{found_row}'] = ai_review.get('conclusion', '')
                    
                    # E열: 개선필요사항
                    improvements = ai_review.get('improvements', '')
                    if len(improvements) > 32767:  # 엑셀 셀 최대 문자 수 제한
                        improvements = improvements[:32760] + "..."
                    summary_ws[f'E{found_row}'] = improvements
                    
                
                elif control in skipped_controls:
                    # 스킵된 통제: 스킵 사유에 맞는 메시지 작성
                    
                    # 스킵 사유에 맞는 메시지 동적 생성
                    skip_message = get_skip_reason_message(control, answers)
                    
                    # C열: 스킵 사유
                    summary_ws[f'C{found_row}'] = skip_message
                    
                    # D열: N/A
                    summary_ws[f'D{found_row}'] = 'N/A'
                    
                    # E열: Cloud 스킵된 통제 + 47번 N인 경우 개선사항 작성
                    cloud_type = answers[4] if len(answers) > 4 and answers[4] else None
                    has_soc1_report = len(answers) > 5 and answers[5] and str(answers[5]).upper() == 'Y'
                    soc1_review_answer = answers[47] if len(answers) > 47 else None
                    
                    # Cloud로 인해 스킵된 통제인지 확인 (스킵 메시지에 Cloud 관련 키워드 포함)
                    is_cloud_skipped = any(keyword in skip_message for keyword in ['SaaS', 'PaaS', 'IaaS', '서비스 제공업체', '플랫폼 제공업체', '인프라 제공업체'])
                    
                    # Cloud 스킵된 통제이고 47번 답변이 N인 경우
                    if (is_cloud_skipped and soc1_review_answer and str(soc1_review_answer).upper() == 'N'):
                        summary_ws[f'E{found_row}'] = 'SOC1 Report를 검토 및 승인하는 절차와 통제가 필요합니다.'
                    else:
                        summary_ws[f'E{found_row}'] = ''
                    
        except Exception as e:
            # Summary 시트 오류가 발생해도 전체 프로세스는 계속 진행
            pass
    else:
        # AI 검토가 비활성화된 경우 기존 Summary 시트 삭제
        if 'Summary' in wb.sheetnames:
            try:
                wb.remove(wb['Summary'])
            except Exception as e:
                pass

    # 메모리 버퍼에 저장 (안전한 방식) - 한글 처리 개선
    excel_stream = BytesIO()
    excel_stream_copy = None
    try:
        # 엑셀 파일 저장 전 검증
        if not wb.worksheets:
            raise Exception("워크북에 시트가 없습니다.")
        
        # 한글 처리를 위한 엑셀 저장 옵션 설정
        from openpyxl.workbook.workbook import Workbook
        from openpyxl.writer.excel import ExcelWriter
        
        # 스킵된 통제의 시트 탭을 회색으로 설정
        set_sheet_tab_color_for_skipped_controls(wb, answers)
        
        # 엑셀 파일을 메모리에 저장 (한글 인코딩 처리)
        # MIME 타입을 명시적으로 설정하여 한글 처리 개선
        wb.save(excel_stream)
        excel_stream.seek(0)

        # 전송용 복사본 생성
        excel_data = excel_stream.getvalue()
        
        # 파일 크기 검증 (최소 크기 체크)
        if len(excel_data) < 1000:  # 1KB 미만이면 문제가 있을 가능성
            raise Exception("생성된 엑셀 파일이 너무 작습니다. 파일 생성에 문제가 있을 수 있습니다.")
        
        excel_stream_copy = BytesIO(excel_data)
        
    except Exception as e:
        # 오류 발생 시 리소스 정리
        if excel_stream:
            excel_stream.close()
        if excel_stream_copy:
            excel_stream_copy.close()
        wb.close()
        raise Exception(f"엑셀 파일 생성 중 오류 발생: {str(e)}")
    finally:
        # 원본 스트림은 항상 닫기
        if excel_stream:
            excel_stream.close()

    user_email = ''
    if answers and answers[0]:
        user_email = answers[0]

    if user_email:
        if progress_callback:
            progress_callback(85, "엑셀 파일을 메일로 전송하고 있습니다...")
            
        subject = '인터뷰 결과 파일'
        body = '인터뷰 내용에 따라 ITGC 설계평가 문서를 첨부합니다.'
        try:
            # 파일 스트림 검증
            if not excel_stream_copy:
                raise Exception("엑셀 파일 스트림이 생성되지 않았습니다.")
            
            # 파일 스트림 위치 확인 및 리셋
            excel_stream_copy.seek(0)
            
            if progress_callback:
                progress_callback(90, "메일 전송 중...")
            
            # 한글 파일명을 안전하게 처리하여 메일 첨부
            # 이메일 전송 (한글 파일명 지원)
            send_gmail_with_attachment(
                to=user_email,
                subject=subject,
                body=body,
                file_stream=excel_stream_copy,
                file_name=file_name
            )
            
            if progress_callback:
                progress_callback(95, "메일 전송이 완료되었습니다!")
                
            return True, user_email, None
        except Exception as e:
            return False, user_email, str(e)
        finally:
            # 전송 완료 후 스트림 정리
            if excel_stream_copy:
                excel_stream_copy.close()
            wb.close()
    else:
        # 이메일이 없는 경우에도 리소스 정리
        if excel_stream_copy:
            excel_stream_copy.close()
        wb.close()
        return False, None, '메일 주소가 입력되지 않았습니다. 1번 질문에 메일 주소를 입력해 주세요.'

def test_conditional_questions():
    """
    조건부 질문 생략 로직을 테스트하는 함수
    """
    
    # 테스트 케이스들
    test_cases = [
        {
            'name': '모든 질문 표시 (기본)',
            'answers': ['test@example.com', 'Test System', 'Y', 'Y', 'Y', 'Y', 'Y', 'Y'] + ['N'] * 40,
            'expected_count': 47
        },
        {
            'name': '3번 답변이 N (프로그램 변경 질문 생략)',
            'answers': ['test@example.com', 'Test System', 'Y', 'N', 'Y', 'Y', 'Y', 'Y'] + ['N'] * 40,
            'expected_count': 41  # 34-39번 질문 6개 생략
        },
        {
            'name': '4번 SaaS + 5번 SOC1 Report 발행 (11, 14~46번 질문 생략)',
            'answers': ['test@example.com', 'Test System', 'Y', 'Y', 'SaaS', 'Y', 'Y', 'Y'] + ['N'] * 40,
            'expected_count': 13  # 11번 1개 + 14~46번 33개 = 총 34개 생략, 13개 질문만 표시
        },
        {
            'name': '4번 IaaS + 5번 SOC1 Report 발행 (22, 29, 44~46번 질문 생략)',
            'answers': ['test@example.com', 'Test System', 'Y', 'Y', 'IaaS', 'Y', 'Y', 'Y'] + ['N'] * 40,
            'expected_count': 42  # 22번 1개 + 29번 1개 + 44~46번 3개 = 총 5개 생략
        },
        {
            'name': '4번 PaaS + 5번 SOC1 Report 발행 (14~31, 44~46번 질문 생략)',
            'answers': ['test@example.com', 'Test System', 'Y', 'Y', 'PaaS', 'Y', 'Y', 'Y'] + ['N'] * 40,
            'expected_count': 26  # 14~31번 18개 + 44~46번 3개 = 총 21개 생략
        },
        {
            'name': '4번 SaaS + 5번 SOC1 Report 미발행 (Cloud 스킵 없음)',
            'answers': ['test@example.com', 'Test System', 'Y', 'Y', 'SaaS', 'N', 'Y', 'Y'] + ['N'] * 40,
            'expected_count': 47  # 스킵 없음, 전체 질문 진행
        },
        {
            'name': '5번 답변이 N (OS 관련 질문 생략)',
            'answers': ['test@example.com', 'Test System', 'Y', 'Y', 'Y', 'N', 'Y', 'Y'] + ['N'] * 40,
            'expected_count': 41  # 28-33번 질문 6개 생략
        },
        {
            'name': '6번 답변이 N (배치 스케줄 질문 생략)',
            'answers': ['test@example.com', 'Test System', 'Y', 'Y', 'Y', 'Y', 'N', 'Y'] + ['N'] * 40,
            'expected_count': 43  # 41-43번 질문 3개 생략
        },
        {
            'name': '모든 조건부 질문 생략',
            'answers': ['test@example.com', 'Test System', 'Y', 'N', 'N', 'N', 'N', 'N'] + ['N'] * 40,
            'expected_count': 25  # 4-5, 15-23, 25-30, 32-37, 39-43번 질문들 생략
        }
    ]
    
    for test_case in test_cases:
        conditional_questions = get_conditional_questions(test_case['answers'])
        actual_count = len(conditional_questions)
        expected_count = test_case['expected_count']
        
        
        # 생략된 질문들 확인
        all_indices = [q['index'] for q in s_questions]
        conditional_indices = [q['index'] for q in conditional_questions]
        skipped_indices = [idx for idx in all_indices if idx not in conditional_indices]
        
        if skipped_indices:
            pass

    # 통제별 N/A 처리 테스트
    
    # 3번 답변이 N인 경우 PC 통제들 테스트
    test_answers = ['test@example.com', 'Test System', 'Y', 'N', 'Y', 'Y', 'Y', 'Y'] + ['N'] * 40
    pc_controls = ['PC01', 'PC02', 'PC03', 'PC04', 'PC05']
    
    for control in pc_controls:
        result = get_text_itgc(test_answers, control)
    
    # 4번 답변이 N인 경우 DB 관련 통제들 테스트
    test_answers = ['test@example.com', 'Test System', 'Y', 'Y', 'N', 'Y', 'Y', 'Y'] + ['N'] * 40
    db_controls = ['APD09', 'APD10', 'APD11']
    
    for control in db_controls:
        result = get_text_itgc(test_answers, control)
    
    # 7번 답변이 N인 경우 Cloud 관련 통제들 테스트
    test_answers = ['test@example.com', 'Test System', 'Y', 'Y', 'Y', 'Y', 'Y', 'N'] + ['N'] * 40
    cloud_controls = ['APD07', 'APD08']  # Cloud 관련 통제들
    
    for control in cloud_controls:
        result = get_text_itgc(test_answers, control)
    
    return True

def test_ai_review_feature():
    """
    AI 검토 기능을 테스트하는 함수 (Summary 시트 기능 포함)
    """
    # 테스트용 답변 데이터
    test_answers = ['test@example.com', 'Test System', 'Y'] + ['N'] * 40
    test_textarea_answers = [''] * 43

    # AI 검토 기능 활성화로 테스트
    result = get_text_itgc(test_answers, 'APD01', test_textarea_answers, enable_ai_review=True)


    if 'AI_Review' in result:
        ai_review = result['AI_Review']
        if isinstance(ai_review, dict):
            pass
        else:
            pass

    # 직접 AI 검토 함수 테스트
    test_content = "사용자 권한 부여 이력이 시스템에 기록되지 않아 모집단 확보가 불가합니다. 새로운 권한 요청 시 승인 절차가 없습니다."
    direct_ai_result = get_ai_review(test_content, 'APD01')

    if isinstance(direct_ai_result, dict):
        pass
    else:
        pass

    return result


# ================================
# Link2 Route Handlers
# ================================

@bp_link2.route('/link2', methods=['GET', 'POST'])
def link2():
    user_info = get_user_info()
    # Interview 기능 시작 시에만 로그 기록 (GET 요청이고 reset=1 파라미터가 있거나 최초 진입 시)
    if is_logged_in() and request.method == 'GET':
        if request.args.get('reset') == '1' or 'question_index' not in session:
            log_user_activity(user_info, 'FEATURE_START', 'Interview 기능 시작', '/link2',
                             request.remote_addr, request.headers.get('User-Agent'))

    if request.method == 'GET':
        # 쿼리 파라미터로 reset이 있을 때만 인터뷰 세션 초기화 (로그인 세션은 보존)
        if request.args.get('reset') == '1':
            reset_interview_session()
        # 세션에 값이 없으면(최초 진입)만 초기화
        elif 'question_index' not in session:
            user_info = get_user_info()
            if user_info and user_info.get('user_email'):
                user_email = user_info.get('user_email')
                # 로그인된 사용자: 첫 번째 질문에 이메일 자동 입력하고 두 번째 질문부터 시작
                session['question_index'] = 1
                session['answer'] = [''] * question_count
                session['textarea_answer'] = [''] * question_count
                session['answer'][0] = user_email  # 첫 번째 답변에 이메일 설정
            else:
                # 비로그인 사용자: 첫 번째 질문부터 시작
                session['question_index'] = 0
                session['answer'] = [''] * question_count
                session['textarea_answer'] = [''] * question_count

        user_info = get_user_info()
        users = user_info.get('user_name', 'Guest') if user_info else "Guest"
        # 현재 답변을 기반으로 필터링된 질문 목록 가져오기
        filtered_questions = get_conditional_questions(session.get('answer', []))
        current_question_index = session['question_index']

        # 현재 질문을 필터링된 목록에서 찾기
        current_question = None
        current_filtered_index = 0

        for idx, q in enumerate(filtered_questions):
            if q['index'] == current_question_index:
                current_question = q
                current_filtered_index = idx
                break

        if current_question is None:
            # 필터링된 목록에 없으면 원본에서 가져오기 (혹시 모를 상황 대비)
            current_question = s_questions[current_question_index] if current_question_index < len(s_questions) else s_questions[0]

        return render_template('link2_bak.jsp',
                             question=current_question,
                             question_count=len(filtered_questions),
                             current_index=current_filtered_index,  # 필터링된 목록에서의 인덱스
                             actual_question_number=current_question.get('index', 0) + 1,  # 실제 질문 번호
                             remote_addr=request.remote_addr,
                             users=users,
                             is_logged_in=is_logged_in(),
                             user_info=user_info,
                             answer=session['answer'],
                             textarea_answer=session['textarea_answer'])

    question_index = session.get('question_index', 0)

    if request.method == 'POST':
        form_data = request.form

        # 현재 질문의 filtered index를 구하기
        filtered_questions_before = get_conditional_questions(session.get('answer', []))
        current_filtered_index = 0
        for i, q in enumerate(filtered_questions_before):
            if q['index'] == question_index:
                current_filtered_index = i
                break

        # form 데이터는 current_filtered_index 기반으로 저장됨
        session['answer'][question_index] = form_data.get(f"a{current_filtered_index}", '')
        session['textarea_answer'][question_index] = form_data.get(f"a{current_filtered_index}_1", '')
        if form_data.get('a1_1'):
            session['System'] = form_data.get('a1_1')
        if form_data.get('a4_1'):
            session['Cloud'] = form_data.get('a4_1')
        if form_data.get('a6_1'):
            session['OS_Tool'] = form_data.get('a6_1')
        if form_data.get('a7_1'):
            session['DB_Tool'] = form_data.get('a7_1')
        if form_data.get('a8_1'):
            session['Batch_Tool'] = form_data.get('a8_1')

        # 현재 답변을 기반으로 필터링된 질문 목록 가져오기
        filtered_questions = get_conditional_questions(session.get('answer', []))

        # 다음 질문 인덱스를 결정하는 로직 (필터링된 질문 기준)
        current_filtered_index = 0
        for i, q in enumerate(filtered_questions):
            if q['index'] == question_index:
                current_filtered_index = i
                break

        # 다음 필터링된 질문으로 이동
        next_filtered_index = current_filtered_index + 1
        if next_filtered_index < len(filtered_questions):
            next_question_index = filtered_questions[next_filtered_index]['index']
        else:
            next_question_index = question_count  # 마지막 질문 이후

        # 마지막 질문 제출 시 AI 검토 선택 페이지로 이동
        if next_filtered_index >= len(filtered_questions):
            # AI 검토 선택 페이지로 리디렉션
            return redirect(url_for('link2.ai_review_selection'))

        session['question_index'] = next_question_index

        if next_question_index >= question_count:
            return redirect(url_for('index'))

    # 현재 질문을 렌더링
    filtered_questions = get_conditional_questions(session.get('answer', []))
    current_question_index = session['question_index']

    # 현재 질문을 필터링된 목록에서 찾기
    current_question = None
    current_filtered_index = 0

    for idx, q in enumerate(filtered_questions):
        if q['index'] == current_question_index:
            current_question = q
            current_filtered_index = idx
            break

    if current_question is None:
        current_question = s_questions[current_question_index] if current_question_index < len(s_questions) else s_questions[0]


    user_info = get_user_info()
    users = user_info.get('user_name', 'Guest') if user_info else "User List"
    return render_template(
        'link2_bak.jsp',
        question=current_question,
        question_count=len(filtered_questions),
        current_index=current_filtered_index,
        actual_question_number=current_question['index'] + 1,
        remote_addr=request.remote_addr,
        users=users,
        is_logged_in=is_logged_in(),
        user_info=user_info,
        answer=session['answer'],
        textarea_answer=session['textarea_answer']
    )

@bp_link2.route('/link2/prev')
def link2_prev():
    # 세션에서 현재 인덱스 가져오기 및 이전으로 이동 (로직 분리)
    link2_prev_logic(session)
    # 다시 질문 페이지로 이동
    return redirect(url_for('link2.link2'))

@bp_link2.route('/paper_request', methods=['POST'])
def paper_request():
    """
    Deprecated: 이 라우트는 더 이상 사용되지 않습니다.
    기존 코드와의 호환성을 위해 유지되고 있습니다.
    """
    # Link2로 리다이렉트
    return redirect(url_for('link2.link2'))

@bp_link2.route('/ai_review_selection')
def ai_review_selection():
    """AI 검토 옵션 선택 화면"""
    user_email = session.get('answer', [''])[0] if session.get('answer') else ''
    if not user_email:
        return redirect(url_for('link2.link2', reset=1))  # 세션이 없으면 인터뷰 처음으로
    
    user_info = get_user_info()

    return render_template('link2_ai_review.jsp',
                         user_email=user_email,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@bp_link2.route('/update_session_email', methods=['POST'])
@login_required
def update_session_email():
    """세션의 이메일 주소 업데이트 (인증 필요)"""
    try:
        data = request.get_json()
        new_email = data.get('email', '').strip()
        
        if not new_email:
            return jsonify({'success': False, 'message': '이메일 주소가 비어있습니다.'})
        
        # 이메일 유효성 검사 (서버 측)
        email_pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_pattern, new_email):
            return jsonify({'success': False, 'message': '올바른 이메일 형식이 아닙니다.'})
        
        # 세션의 answer 배열 첫 번째 요소(이메일) 업데이트
        if 'answer' in session:
            session['answer'][0] = new_email
            session.modified = True  # 세션이 수정되었음을 명시
        else:
            return jsonify({'success': False, 'message': '세션 정보를 찾을 수 없습니다.'})
        
        return jsonify({'success': True, 'message': '이메일이 성공적으로 변경되었습니다.'})

    except Exception as e:
        return jsonify({'success': False, 'message': '서버 오류가 발생했습니다.'})

@bp_link2.route('/process_with_ai_option', methods=['POST'])
def process_with_ai_option():
    """AI 검토 옵션에 따라 처리 페이지로 이동"""
    enable_ai_review = request.form.get('enable_ai_review', 'false').lower() == 'true'
    user_email = session.get('answer', [''])[0] if session.get('answer') else ''
    
    if enable_ai_review and is_logged_in():
        # 로그인한 사용자만 AI 검토 횟수 기록
        increment_ai_review_count(user_email)
    
    # 세션에 AI 검토 옵션 저장
    session['enable_ai_review'] = enable_ai_review

    # processing 페이지로 리디렉션
    return redirect(url_for('link2.processing'))

@bp_link2.route('/processing')
def processing():
    """인터뷰 완료 후 처리 중 화면"""
    user_email = session.get('answer', [''])[0] if session.get('answer') else ''
    enable_ai_review = session.get('enable_ai_review', False)
    
    # 고유한 작업 ID 생성 및 세션에 저장
    task_id = str(uuid.uuid4())
    session['processing_task_id'] = task_id
    
    # 초기 진행률 상태 파일 생성
    reset_progress(task_id) # 기존 파일이 있다면 삭제
    initial_status = {
        'percentage': 0,
        'current_task': 'AI 검토를 준비하고 있습니다...',
        'is_processing': True
    }
    set_progress_status(task_id, initial_status)
    
    return render_template('processing.jsp', user_email=user_email, enable_ai_review=enable_ai_review, task_id=task_id)

@bp_link2.route('/get_progress')
def get_progress():
    """진행률 상태 조회 엔드포인트"""
    task_id = request.args.get('task_id')
    if not task_id:
        return jsonify({'error': 'No task_id provided'}), 400
    status = get_progress_status(task_id)
    return jsonify(status)

@bp_link2.route('/process_interview', methods=['POST'])
def process_interview():
    """실제 인터뷰 처리 및 메일 발송"""
    data = request.get_json() or {}
    task_id = data.get('task_id')

    if not task_id:
        return jsonify({'success': False, 'error': 'No task_id provided'})

    try:
        # task_id를 포함하는 콜백 함수 생성
        progress_callback = lambda p, t: update_progress(task_id, p, t)
        
        progress_callback(5, "인터뷰 데이터를 확인하고 있습니다...")
        
        answers = session.get('answer', [])
        textarea_answers = session.get('textarea_answer', [])
        
        if not answers:
            reset_progress(task_id)
            return jsonify({'success': False, 'error': '인터뷰 데이터가 없습니다.'})
        
        user_email = answers[0] if answers else ''
        if not user_email:
            reset_progress(task_id)
            return jsonify({'success': False, 'error': '메일 주소가 입력되지 않았습니다.'})
        
        progress_callback(10, "AI 검토 설정을 확인하고 있습니다...")
        
        enable_ai_review = session.get('enable_ai_review', False)
        
        progress_callback(15, "ITGC 설계평가 문서 생성을 시작합니다...")
        
        # 스킵된 질문들의 답변을 최종적으로 공란으로 처리
        clear_skipped_answers(answers, textarea_answers)
        
        success, returned_email, error = export_interview_excel_and_send(
            answers,
            textarea_answers,
            get_text_itgc,
            fill_sheet,
            is_ineffective,
            send_gmail_with_attachment,
            enable_ai_review,
            progress_callback
        )
        
        if success:
            status = get_progress_status(task_id)
            status['percentage'] = 100
            status['current_task'] = "처리가 완료되었습니다!"
            status['is_processing'] = False
            set_progress_status(task_id, status)
            return jsonify({'success': True, 'email': returned_email})
        else:
            reset_progress(task_id)
            return jsonify({'success': False, 'error': error})

    except Exception as e:
        reset_progress(task_id)
        return jsonify({'success': False, 'error': str(e)})

if __name__ == "__main__":
    # 조건부 질문 생략 로직 테스트
    test_conditional_questions()
    
    # AI 검토 기능 테스트
    test_ai_review_feature()
    # 강제 수정1