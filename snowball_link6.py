from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, session
from auth import login_required, get_current_user, get_user_rcms, get_rcm_details, save_design_evaluation, get_design_evaluations, get_design_evaluations_by_header_id, get_user_evaluation_sessions, delete_evaluation_session, create_evaluation_structure, log_user_activity, get_db, get_or_create_evaluation_header, get_rcm_detail_mappings, archive_design_evaluation_session, unarchive_design_evaluation_session, get_completed_design_evaluation_sessions, get_key_rcm_details, count_completed_operation_evaluations
from snowball_link5 import get_user_info, is_logged_in
import sys
import os
import json
from werkzeug.utils import secure_filename
from flask import send_file
import tempfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles.colors import Color

bp_link6 = Blueprint('link6', __name__)

# 설계평가 관련 기능들

@bp_link6.route('/design-evaluation')
@login_required
def user_design_evaluation():
    """레거시 설계평가 페이지 - ITGC 통합 평가로 리디렉트"""
    return redirect(url_for('link6.itgc_evaluation'))

@bp_link6.route('/design-evaluation/rcm', methods=['GET', 'POST'])
@login_required
def user_design_evaluation_rcm():
    """사용자 디자인 평가 페이지 (RCM별)"""
    user_info = get_user_info()

    # POST로 전달된 RCM ID 받기 또는 세션에서 가져오기
    if request.method == 'POST':
        rcm_id = request.form.get('rcm_id')
        evaluation_type = request.form.get('evaluation_type', 'ITGC')  # 기본값 ITGC
        evaluation_session = request.form.get('session')  # 평가 세션

        if not rcm_id:
            flash('RCM 정보가 없습니다.', 'error')
            return redirect(url_for('link6.user_design_evaluation'))

        # RCM의 실제 control_category를 조회하여 evaluation_type 설정
        # (클라이언트에서 전달한 값이 잘못될 수 있으므로 서버에서 재확인)
        with get_db() as conn:
            rcm_category = conn.execute('''
                SELECT control_category FROM sb_rcm WHERE rcm_id = %s
            ''', (int(rcm_id),)).fetchone()

            if rcm_category and rcm_category['control_category']:
                evaluation_type = rcm_category['control_category']

        # 세션에 저장
        session['current_design_rcm_id'] = int(rcm_id)
        session['current_evaluation_type'] = evaluation_type
        if evaluation_session:
            session['current_evaluation_session'] = evaluation_session
    else:
        # GET 요청인 경우 URL 파라미터 우선, 없으면 세션에서 가져오기
        rcm_id = request.args.get('rcm_id')
        evaluation_session = request.args.get('session')

        if rcm_id:
            # URL 파라미터가 있으면 세션에 저장
            rcm_id = int(rcm_id)
            session['current_design_rcm_id'] = rcm_id
            if evaluation_session:
                session['current_evaluation_session'] = evaluation_session

            # RCM의 실제 control_category를 조회하여 evaluation_type 설정
            with get_db() as conn:
                rcm_category = conn.execute('''
                    SELECT control_category FROM sb_rcm WHERE rcm_id = %s
                ''', (rcm_id,)).fetchone()

                if rcm_category and rcm_category['control_category']:
                    evaluation_type = rcm_category['control_category']
                    session['current_evaluation_type'] = evaluation_type
                else:
                    evaluation_type = 'ITGC'  # 기본값
                    session['current_evaluation_type'] = evaluation_type
        else:
            # URL 파라미터가 없으면 세션에서 가져오기
            rcm_id = session.get('current_design_rcm_id')
            evaluation_session = session.get('current_evaluation_session')
            evaluation_type = session.get('current_evaluation_type', 'ITGC')

        if not rcm_id:
            flash('RCM 정보가 없습니다. 다시 선택해주세요.', 'error')
            return redirect(url_for('link6.user_design_evaluation'))

    # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
    if user_info.get('admin_flag') != 'Y':
        user_rcms = get_user_rcms(user_info['user_id'])
        rcm_ids = [rcm['rcm_id'] for rcm in user_rcms]

        if rcm_id not in rcm_ids:
            flash('해당 RCM에 대한 접근 권한이 없습니다.', 'error')
            return redirect(url_for('link6.user_design_evaluation'))
    else:
        # 관리자는 모든 RCM에 접근 가능
        user_rcms = get_user_rcms(user_info['user_id'])
    
    # RCM 정보 조회
    rcm_info = None
    if user_info.get('admin_flag') == 'Y':
        # 관리자인 경우 모든 RCM 정보를 조회
        with get_db() as conn:
            rcm_data = conn.execute('''
                SELECT r.rcm_id, r.rcm_name, r.description, r.upload_date, r.user_id,
                       u.user_name as owner_name, u.company_name, 'admin' as permission_type
                FROM sb_rcm r
                LEFT JOIN sb_user u ON r.user_id = u.user_id
                WHERE r.rcm_id = %s
            ''', (rcm_id,)).fetchone()
            if rcm_data:
                rcm_info = dict(rcm_data)
    else:
        # 일반 사용자인 경우 권한이 있는 RCM 중에서 조회
        for rcm in user_rcms:
            if rcm['rcm_id'] == rcm_id:
                rcm_info = rcm
                break
    
    # evaluation_type에 따라 자동으로 control_category 설정
    # ITGC, ELC, TLC는 해당 카테고리만 보여줌
    control_category = evaluation_type if evaluation_type in ['ITGC', 'ELC', 'TLC'] else None

    # RCM 세부 데이터 조회 (evaluation_session이 있으면 해당 세션의 데이터만 조회)
    rcm_details = get_rcm_details(rcm_id, control_category=control_category, evaluation_session=evaluation_session)

    # 매핑 정보 조회
    rcm_mappings = get_rcm_detail_mappings(rcm_id)

    # 통제 카테고리별 통계
    with get_db() as conn:
        category_stats = conn.execute('''
            SELECT control_category, COUNT(*) as count
            FROM sb_rcm_detail
            WHERE rcm_id = %s
            GROUP BY control_category
            ORDER BY control_category
        ''', (rcm_id,)).fetchall()
        category_stats = {row['control_category']: row['count'] for row in category_stats}

    # 운영평가 존재 여부 확인 (evaluation_session이 있고 status >= 3인 경우)
    # status 3: 운영평가 진행중, status 4: 운영평가 완료
    has_operation_evaluation = False
    if evaluation_session:
        with get_db() as conn:
            op_eval = conn.execute('''
                SELECT status FROM sb_evaluation_header
                WHERE rcm_id = %s AND evaluation_name = %s AND status >= 3
                AND (archived IS NULL OR archived = 0)
            ''', (rcm_id, evaluation_session)).fetchone()
            has_operation_evaluation = op_eval is not None

    log_user_activity(user_info, 'PAGE_ACCESS', 'RCM 디자인 평가', '/design-evaluation/rcm',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link6_design_rcm_detail.jsp',
                         rcm_id=rcm_id,
                         rcm_info=rcm_info,
                         rcm_details=rcm_details,
                         rcm_mappings=rcm_mappings,
                         control_category=control_category,
                         category_stats=category_stats,
                         evaluation_type=evaluation_type,
                         evaluation_session=evaluation_session,
                         has_operation_evaluation=has_operation_evaluation,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

@bp_link6.route('/api/design-evaluation/get', methods=['GET'])
@login_required
def get_design_evaluation_api():
    """설계평가 결과 조회 API"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    design_evaluation_session = request.args.get('design_evaluation_session')
    control_code = request.args.get('control_code')

    # control_code가 없으면 전체 설계평가 목록 조회
    if not control_code:
        if not all([rcm_id, design_evaluation_session]):
            return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'}), 400

        try:
            # 전체 설계평가 목록 조회
            with get_db() as conn:
                evaluations = conn.execute('''
                    SELECT
                        l.control_code,
                        rd.control_name,
                        rd.control_description,
                        rd.control_frequency,
                        rd.control_type,
                        rd.control_nature,
                        rd.key_control,
                        l.description_adequacy as design_adequacy,
                        l.improvement_suggestion,
                        l.overall_effectiveness as conclusion,
                        l.evaluation_evidence,
                        l.evaluation_rationale,
                        l.recommended_actions,
                        l.evaluation_date
                    FROM sb_evaluation_line l
                    JOIN sb_evaluation_header h ON l.header_id = h.header_id
                    JOIN sb_rcm_detail_v rd ON h.rcm_id = rd.rcm_id AND l.control_code = rd.control_code
                    WHERE h.rcm_id = %s AND h.evaluation_name = %s
                    ORDER BY l.control_code
                ''', (rcm_id, design_evaluation_session)).fetchall()

                return jsonify({
                    'success': True,
                    'evaluations': [dict(row) for row in evaluations]
                })

        except Exception as e:
            print(f"설계평가 목록 조회 오류: {str(e)}")
            return jsonify({'success': False, 'message': str(e)}), 500

    # control_code가 있으면 단일 설계평가 조회
    if not all([rcm_id, design_evaluation_session, control_code]):
        return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'}), 400

    try:
        # 단일 통제의 설계평가 데이터 조회
        with get_db() as conn:
            design_eval = conn.execute('''
                SELECT
                    l.line_id,
                    l.description_adequacy,
                    l.improvement_suggestion,
                    l.overall_effectiveness,
                    l.evaluation_evidence,
                    l.evaluation_rationale,
                    l.recommended_actions,
                    l.design_comment,
                    l.evaluation_date,
                    l.no_occurrence,
                    l.no_occurrence_reason,
                    h.header_id
                FROM sb_evaluation_line l
                JOIN sb_evaluation_header h ON l.header_id = h.header_id
                WHERE h.rcm_id = %s AND h.evaluation_name = %s AND l.control_code = %s
            ''', (rcm_id, design_evaluation_session, control_code)).fetchone()

            if design_eval:
                # 이미지 DB에서 조회
                images = conn.execute('''
                    SELECT image_id, file_path, file_name, file_size, uploaded_at
                    FROM sb_evaluation_image
                    WHERE evaluation_type = %s AND line_id = %s
                    ORDER BY uploaded_at
                ''', ('design', design_eval['line_id'])).fetchall()

                return jsonify({
                    'success': True,
                    'evaluation': dict(design_eval),
                    'images': [dict(img) for img in images],
                    'design_evaluation': {
                        'conclusion': design_eval['overall_effectiveness'],
                        'deficiency_details': design_eval['improvement_suggestion'],
                        'improvement_plan': design_eval['recommended_actions'],
                        'test_procedure': design_eval['evaluation_rationale'],
                        'evaluated_at': design_eval['evaluation_date']
                    }
                })
            else:
                return jsonify({'success': False, 'message': '설계평가 데이터를 찾을 수 없습니다.'}), 404

    except Exception as e:
        print(f"설계평가 조회 오류: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@bp_link6.route('/api/design-evaluation/save', methods=['POST'])
@login_required
def save_design_evaluation_api():
    """설계평가 결과 저장 API"""
    user_info = get_user_info()
    
    
    # FormData와 JSON 모두 처리
    if request.content_type and 'multipart/form-data' in request.content_type:
        # FormData 처리 (이미지 포함)
        rcm_id = request.form.get('rcm_id')
        control_code = request.form.get('control_code')
        evaluation_data_str = request.form.get('evaluation_data')
        evaluation_session = request.form.get('evaluation_session')

        # JSON 문자열을 파싱
        evaluation_data = json.loads(evaluation_data_str) if evaluation_data_str else None
        
        # 이미지 파일 처리
        uploaded_images = []
        for key in request.files:
            if key.startswith('evaluation_image_'):
                file = request.files[key]
                if file and file.filename:
                    uploaded_images.append(file)
    else:
        # 기존 JSON 처리
        data = request.get_json()
        rcm_id = data.get('rcm_id') if data else None
        control_code = data.get('control_code') if data else None
        evaluation_data = data.get('evaluation_data') if data else None
        evaluation_session = data.get('evaluation_session') if data else None
        uploaded_images = []

    # 필수 데이터 검증
    
    if not all([rcm_id, control_code, evaluation_data, evaluation_session]):
        missing_fields = []
        if not rcm_id: missing_fields.append('RCM ID')
        if not control_code: missing_fields.append('통제 코드')
        if not evaluation_data: missing_fields.append('평가 데이터')
        if not evaluation_session: missing_fields.append('세션명')
        
        error_msg = f'필수 데이터가 누락되었습니다: {", ".join(missing_fields)}'
        
        return jsonify({
            'success': False,
            'message': error_msg
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
        
        # 이미지 파일 저장
        saved_images = []
        if uploaded_images:
            # 헤더 ID 가져오기 (이미지 저장 폴더명으로 사용)
            with get_db() as conn:
                header_id = get_or_create_evaluation_header(conn, rcm_id, user_info['user_id'], evaluation_session)
            
            # 이미지 저장 디렉토리 생성 (header_id로 분리)
            upload_dir = os.path.join('static', 'uploads', 'design_evaluations', str(rcm_id), str(header_id), control_code)
            os.makedirs(upload_dir, exist_ok=True)
            
            for i, image_file in enumerate(uploaded_images):
                if image_file and image_file.filename:
                    # 디버깅 로그
                    
                    # 확장자를 먼저 분리
                    original_name, original_ext = os.path.splitext(image_file.filename)
                    
                    # 안전한 파일명 생성 (확장자 제외)
                    safe_name = secure_filename(original_name)
                    
                    # secure_filename이 빈 문자열을 반환하면 기본 이름 사용
                    if not safe_name or safe_name.strip() == '' or safe_name == '_':
                        safe_name = 'evaluation_image'
                    
                    # 중복 방지를 위해 타임스탬프 추가
                    import time
                    timestamp = str(int(time.time()))
                    
                    # 최종 파일명: 안전한이름_타임스탬프_인덱스.확장자
                    safe_filename = f"{safe_name}_{timestamp}_{i}{original_ext.lower()}"
                    
                    # 파일 저장
                    file_path = os.path.join(upload_dir, safe_filename)
                    image_file.save(file_path)

                    # 상대 경로로 저장 (DB 저장용)
                    relative_path = f"static/uploads/design_evaluations/{rcm_id}/{header_id}/{control_code}/{safe_filename}"

                    # 파일 크기
                    file_size = os.path.getsize(file_path)

                    saved_images.append({
                        'path': relative_path,
                        'name': safe_filename,
                        'size': file_size
                    })

        # 설계평가 결과 저장
        save_design_evaluation(rcm_id, control_code, user_info['user_id'], evaluation_data, evaluation_session)

        # 설계평가 완료(status=1) 상태에서 내용을 수정한 경우, status를 0(진행중)으로 되돌림
        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id, status FROM sb_evaluation_header
                WHERE rcm_id = %s AND evaluation_name = %s
            ''', (rcm_id, evaluation_session)).fetchone()

            if header and header['status'] == 1:
                from datetime import datetime
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                conn.execute('''
                    UPDATE sb_evaluation_header
                    SET status = 0, last_updated = %s
                    WHERE header_id = %s
                ''', (current_time, header['header_id']))
                conn.commit()

        # 이미지를 DB에 저장
        if saved_images:
            with get_db() as conn:
                # line_id 조회
                line_record = conn.execute('''
                    SELECT l.line_id FROM sb_evaluation_line l
                    JOIN sb_evaluation_header h ON l.header_id = h.header_id
                    WHERE h.rcm_id = %s AND h.evaluation_name = %s AND l.control_code = %s
                ''', (rcm_id, evaluation_session, control_code)).fetchone()

                if line_record:
                    line_id = line_record['line_id']

                    for img_info in saved_images:
                        # 중복 체크
                        existing = conn.execute('''
                            SELECT image_id FROM sb_evaluation_image
                            WHERE evaluation_type = %s AND line_id = %s AND file_path = %s
                        ''', ('design', line_id, img_info['path'])).fetchone()

                        if not existing:
                            conn.execute('''
                                INSERT INTO sb_evaluation_image
                                (evaluation_type, line_id, file_path, file_name, file_size, uploaded_at)
                                VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                            ''', ('design', line_id, img_info['path'], img_info['name'], img_info['size']))

                    conn.commit()
        
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION', f'설계평가 저장 - {control_code}', 
                         f'/api/design-evaluation/save', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': '새로운 설계평가 결과가 저장되었습니다.'
        })
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        return jsonify({
            'success': False,
            'message': f'저장 중 오류가 발생했습니다: {str(e)}'
        })

@bp_link6.route('/api/design-evaluation/reset', methods=['POST'])
@login_required
def reset_design_evaluations_api():
    """설계평가 결과 초기화 API"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    
    if not rcm_id:
        return jsonify({
            'success': False,
            'message': 'RCM ID가 누락되었습니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
            
            # 해당 RCM의 평가 결과 초기화 (Header-Line 구조 유지)
            # 레코드는 삭제하지 않고 평가 결과 필드만 NULL로 리셋

            # 1. sample 레코드들 삭제 (샘플은 재추출 필요)
            conn.execute('''
                DELETE FROM sb_evaluation_sample
                WHERE line_id IN (
                    SELECT l.line_id FROM sb_evaluation_line l
                    JOIN sb_evaluation_header h ON l.header_id = h.header_id
                    WHERE h.rcm_id = %s
                )
            ''', (rcm_id,))

            # 2. line 레코드들의 설계평가 결과 필드 초기화 (레코드 유지)
            cursor = conn.execute('''
                UPDATE sb_evaluation_line
                SET description_adequacy = NULL,
                    improvement_suggestion = NULL,
                    overall_effectiveness = NULL,
                    evaluation_rationale = NULL,
                    design_comment = NULL,
                    recommended_actions = NULL,
                    no_occurrence = NULL,
                    no_occurrence_reason = NULL,
                    evaluation_evidence = NULL,
                    evaluation_date = NULL,
                    conclusion = NULL,
                    operation_comment = NULL,
                    last_updated = CURRENT_TIMESTAMP
                WHERE header_id IN (
                    SELECT header_id FROM sb_evaluation_header
                    WHERE rcm_id = %s
                )
            ''', (rcm_id,))
            reset_count = cursor.rowcount

            # 3. header 레코드 상태 초기화 (레코드 유지, status=0으로 리셋)
            conn.execute('''
                UPDATE sb_evaluation_header
                SET status = 0, last_updated = CURRENT_TIMESTAMP
                WHERE rcm_id = %s
            ''', (rcm_id,))
            
            conn.commit()
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_RESET', f'설계평가 초기화 - RCM ID: {rcm_id}', 
                         f'/api/design-evaluation/reset', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': f'{reset_count}개의 통제 항목 평가 결과가 초기화되었습니다.',
            'reset_count': reset_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '초기화 중 오류가 발생했습니다.'
        })

# 임시 비활성화 - 테이블 구조 문제로 인해
# @bp_link6.route('/api/design-evaluation/versions/<int:rcm_id>/<control_code>')
# @login_required
# def get_evaluation_versions_api(rcm_id, control_code):
#     return jsonify({'success': False, 'message': '기능 준비 중입니다.'}), 503

@bp_link6.route('/api/design-evaluation/sessions/<int:rcm_id>')
@login_required
def get_evaluation_sessions_api(rcm_id):
    """사용자의 설계평가 세션 목록 조회 API"""
    user_info = get_user_info()
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403
        
        # 평가 세션 목록 조회
        sessions = get_user_evaluation_sessions(rcm_id, user_info['user_id'])
        
        return jsonify({
            'success': True,
            'sessions': sessions,
            'total_count': len(sessions)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': '세션 조회 중 오류가 발생했습니다.'}), 500

@bp_link6.route('/api/design-evaluation/load/<int:rcm_id>')
@login_required
def load_evaluation_data_api(rcm_id):
    """특정 평가 세션의 데이터 로드 API"""
    user_info = get_user_info()
    evaluation_session = request.args.get('session')
    header_id = request.args.get('header_id')
    
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403
        
        # 평가 데이터 로드
        if header_id:
            evaluations = get_design_evaluations_by_header_id(rcm_id, user_info['user_id'], int(header_id))
        else:
            evaluations = get_design_evaluations(rcm_id, user_info['user_id'], evaluation_session)
            
            # 세션명으로 로드할 때 실제 header_id를 찾아서 반환
            if evaluations and evaluation_session:
                # 첫 번째 평가 데이터에서 header_id 추출
                actual_header_id = evaluations[0].get('header_id')
        
        # 통제별로 정리
        evaluation_dict = {}
        for eval_data in evaluations:
            control_code = eval_data['control_code']
            evaluation_date = eval_data.get('evaluation_date')
            
            # 디버깅용 로그 추가
            
            # 해당 통제의 이미지 파일 찾기
            saved_images = []
            
            # header_id를 찾기 위해 현재 평가 데이터에서 확인
            current_header_id = header_id
            if not current_header_id and evaluation_session:
                # header_id가 없으면 세션명으로 찾기
                try:
                    with get_db() as conn:
                        result = conn.execute('''
                            SELECT header_id FROM sb_evaluation_header
                            WHERE rcm_id = %s AND evaluation_name = %s
                        ''', (rcm_id, evaluation_session)).fetchone()
                        if result:
                            current_header_id = result['header_id']
                except Exception as e:
                    pass

            if current_header_id:
                # DB에서 이미지 정보 조회
                images_found = False
                try:
                    with get_db() as conn:
                        # line_id 조회
                        line_result = conn.execute('''
                            SELECT line_id FROM sb_evaluation_line
                            WHERE header_id = %s AND control_code = %s
                        ''', (current_header_id, control_code)).fetchone()

                        if line_result:
                            images = conn.execute('''
                                SELECT image_id, file_path, file_name
                                FROM sb_evaluation_image
                                WHERE evaluation_type = %s AND line_id = %s
                                ORDER BY uploaded_at
                            ''', ('design', line_result['line_id'])).fetchall()

                            for img in images:
                                # file_path는 이미 "static/uploads/..."로 시작
                                # static/ 제거하여 상대 경로로 변환
                                relative_path = img['file_path'].replace('static/', '', 1) if img['file_path'].startswith('static/') else img['file_path']
                                saved_images.append({
                                    'image_id': img['image_id'],
                                    'filename': img['file_name'],
                                    'path': relative_path,
                                    'url': f"/static/{relative_path}"
                                })
                                images_found = True
                except Exception as e:
                    import traceback
                    traceback.print_exc()

                # DB에 이미지가 없으면 파일 시스템에서 폴백 검색
                if not images_found:
                    # 다른 header_id로 시도해보기
                    base_dir = os.path.join('static', 'uploads', 'design_evaluations', str(rcm_id))
                    if os.path.exists(base_dir):
                        for folder_name in os.listdir(base_dir):
                            test_dir = os.path.join(base_dir, folder_name, control_code)
                            if os.path.exists(test_dir):
                                for filename in os.listdir(test_dir):
                                    if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')):
                                        relative_path = f"uploads/design_evaluations/{rcm_id}/{folder_name}/{control_code}/{filename}"
                                        saved_images.append({
                                            'filename': filename,
                                            'path': relative_path,
                                            'url': f"/static/{relative_path}"
                                        })
            else:
                pass

            evaluation_dict[control_code] = {
                'description_adequacy': eval_data['description_adequacy'],
                'improvement_suggestion': eval_data['improvement_suggestion'],
                'evaluation_evidence': eval_data.get('evaluation_evidence'),
                'overall_effectiveness': eval_data['overall_effectiveness'],
                'evaluation_rationale': eval_data['evaluation_rationale'],
                'design_comment': eval_data.get('design_comment'),
                'recommended_actions': eval_data['recommended_actions'],
                'evaluation_date': evaluation_date,
                'images': saved_images
            }
        
        response_data = {
            'success': True,
            'evaluations': evaluation_dict,
            'session_name': evaluation_session
        }

        # 세션명으로 로드할 때 실제 header_id를 응답에 포함
        if evaluations and evaluation_session and not header_id:
            actual_header_id = evaluations[0].get('header_id')
            response_data['header_id'] = actual_header_id

        # header의 status 정보도 포함
        try:
            with get_db() as conn:
                # header_id가 있으면 해당 header의 status 조회
                if header_id:
                    result = conn.execute('''
                        SELECT status FROM sb_evaluation_header
                        WHERE header_id = %s
                    ''', (int(header_id),)).fetchone()
                elif evaluation_session:
                    result = conn.execute('''
                        SELECT status FROM sb_evaluation_header
                        WHERE rcm_id = %s AND evaluation_name = %s
                    ''', (rcm_id, evaluation_session)).fetchone()
                else:
                    result = None


                if result:
                    response_data['header_status'] = result['status']
                    # 하위 호환성을 위해 completed_date도 추가 (status >= 2면 완료로 간주)
                    response_data['header_completed_date'] = 'completed' if result['status'] >= 2 else None
                else:
                    response_data['header_status'] = 0
                    response_data['header_completed_date'] = None
        except Exception as e:
            response_data['header_completed_date'] = None

        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'데이터 로드 중 오류가 발생했습니다: {str(e)}'}), 500

@bp_link6.route('/api/design-evaluation/delete-session', methods=['POST'])
@login_required
def delete_evaluation_session_api():
    """평가 세션 삭제 API"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')
    
    if not all([rcm_id, evaluation_session]):
        return jsonify({
            'success': False,
            'message': '필수 데이터가 누락되었습니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

            # 해당 설계평가 세션을 기반으로 한 운영평가가 있는지 확인
            operation_check = conn.execute('''
                SELECT COUNT(*) as count
                FROM sb_evaluation_header
                WHERE rcm_id = %s AND design_evaluation_session = %s
            ''', (rcm_id, evaluation_session)).fetchone()

            if operation_check and operation_check['count'] > 0:
                return jsonify({
                    'success': False,
                    'message': '⛔ 이 설계평가를 기반으로 진행 중인 운영평가가 있습니다. 운영평가를 먼저 삭제해주세요.'
                })

        # 세션 삭제
        deleted_count = delete_evaluation_session(rcm_id, user_info['user_id'], evaluation_session)
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_DELETE', f'설계평가 세션 삭제 - {evaluation_session}', 
                         f'/api/design-evaluation/delete-session', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': f'평가 세션 "{evaluation_session}"이 삭제되었습니다.',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '삭제 중 오류가 발생했습니다.'
        })

@bp_link6.route('/api/design-evaluation/create-evaluation', methods=['POST'])
@login_required
def create_design_evaluation_api():
    """새로운 설계평가 생성 API (Header + 모든 Line 생성)"""
    user_info = get_user_info()
    data = request.get_json()
    
    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')
    
    if not all([rcm_id, evaluation_session]):
        return jsonify({
            'success': False,
            'message': '필수 데이터가 누락되었습니다.'
        })
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })
        
        # 중복 세션명 체크
        existing_sessions = get_user_evaluation_sessions(rcm_id, user_info['user_id'])
        session_names = [session['evaluation_session'] for session in existing_sessions]
        
        if evaluation_session in session_names:
            return jsonify({
                'success': False,
                'message': f'"{evaluation_session}" 세션이 이미 존재합니다. 다른 이름을 사용해주세요.',
                'exists': True
            })
        
        # 평가 구조 생성 (Header + 모든 빈 Line 생성)
        header_id = create_evaluation_structure(rcm_id, user_info['user_id'], evaluation_session)
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_CREATE', 
                         f'설계평가 생성 - {evaluation_session}', 
                         f'/api/design-evaluation/create-evaluation', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        return jsonify({
            'success': True,
            'message': f'설계평가 "{evaluation_session}"이 생성되었습니다.',
            'header_id': header_id,
            'evaluation_session': evaluation_session
        })
        
    except ValueError as ve:
        return jsonify({
            'success': False,
            'message': str(ve)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': '[EVAL-001] 평가 생성 중 오류가 발생했습니다.'
        })

@bp_link6.route('/api/design-evaluation/complete', methods=['POST'])
@login_required
def complete_design_evaluation_api():
    """설계평가 완료 처리 API - status를 1에서 2로 업데이트 (운영평가 시작)"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')

    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })

    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

        # 통합 header 테이블에서 해당 평가 세션의 status 업데이트
        with get_db() as conn:
            # 현재 평가 세션이 존재하는지 확인
            header = conn.execute('''
                SELECT header_id, status FROM sb_evaluation_header
                WHERE rcm_id = %s AND evaluation_name = %s
            ''', (rcm_id, evaluation_session)).fetchone()

            if not header:
                return jsonify({
                    'success': False,
                    'message': '해당 평가 세션을 찾을 수 없습니다.'
                })

            header_dict = dict(header)
            current_status = header_dict.get('status', 0)

            # status를 2로 업데이트 (설계평가 완료 -> 운영평가 시작)
            from datetime import datetime
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            conn.execute('''
                UPDATE sb_evaluation_header
                SET status = 2, last_updated = %s
                WHERE header_id = %s
            ''', (current_time, header_dict['header_id']))

            conn.commit()

        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_COMPLETE',
                         f'설계평가 완료 - {evaluation_session}',
                         f'/api/design-evaluation/complete',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({
            'success': True,
            'message': '설계평가가 완료되었습니다. 운영평가를 시작할 수 있습니다.',
            'completed_date': current_time,
            'new_status': 2
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'완료 처리 중 오류가 발생했습니다: {str(e)}'
        })

@bp_link6.route('/api/design-evaluation/cancel', methods=['POST'])
@login_required
def cancel_design_evaluation_api():
    """설계평가 완료 취소 API - status를 2에서 1로 변경 (운영평가 진행중 → 설계평가 완료)"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')

    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })

    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

        # header 테이블에서 해당 평가 세션의 status 확인 및 업데이트
        with get_db() as conn:
            # 현재 평가 세션이 존재하는지 확인
            header = conn.execute('''
                SELECT header_id, status FROM sb_evaluation_header
                WHERE rcm_id = %s AND evaluation_name = %s
            ''', (rcm_id, evaluation_session)).fetchone()

            if not header:
                return jsonify({
                    'success': False,
                    'message': '해당 평가 세션을 찾을 수 없습니다.'
                })

            header_dict = dict(header)
            current_status = header_dict.get('status', 0)

            from datetime import datetime
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            if current_status >= 1:
                # 설계평가 완료 이상의 모든 상태 → status를 0으로 변경 (설계평가 진행중 상태로)
                new_status = 0
                if current_status >= 2:
                    message = '운영평가가 취소되고 설계평가 진행중 상태로 되돌렸습니다.'
                else:
                    message = '설계평가 완료가 취소되었습니다. 설계평가 진행중 상태로 되돌렸습니다.'
            else:
                return jsonify({
                    'success': False,
                    'message': '취소할 수 있는 상태가 아닙니다.'
                })

            conn.execute('''
                UPDATE sb_evaluation_header
                SET status = %s, last_updated = %s
                WHERE header_id = %s
            ''', (new_status, current_time, header_dict['header_id']))

            conn.commit()

        # 활동 로그 기록
        log_activity_message = f'설계평가 완료 취소 - {evaluation_session}' if new_status == 0 else f'운영평가 시작 취소 - {evaluation_session}'
        log_user_activity(user_info, 'DESIGN_EVALUATION_CANCEL',
                         log_activity_message,
                         f'/api/design-evaluation/cancel',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'완료 취소 중 오류가 발생했습니다: {str(e)}'
        })


@bp_link6.route('/api/design-evaluation/download-excel/<int:rcm_id>')
@login_required
def download_evaluation_excel(rcm_id):
    """설계평가 엑셀 다운로드 - 원본 파일의 Template 시트를 통제별로 복사해서 생성"""
    # URL 파라미터에서 세션명 가져오기
    evaluation_session = request.args.get('evaluation_session') or request.args.get('session')
    
    # mimetypes 초기화 및 .mpo 오류 방지
    import mimetypes
    try:
        # 강제로 mimetypes 재초기화
        mimetypes.init()
        # 여러 방법으로 .mpo 확장자 등록 시도
        mimetypes.add_type('application/octet-stream', '.mpo')
        mimetypes.add_type('image/jpeg', '.mpo')  # MPO는 Multi Picture Object 포맷
        
        # types_map에 직접 추가
        if hasattr(mimetypes, 'types_map'):
            if True in mimetypes.types_map:
                mimetypes.types_map[True]['.mpo'] = 'image/jpeg'
            if False in mimetypes.types_map:
                mimetypes.types_map[False]['.mpo'] = 'image/jpeg'
        
    except Exception as mt_error:
        # 최후의 수단: monkey patch
        try:
            import openpyxl.packaging.manifest
            original_register = openpyxl.packaging.manifest.Manifest._register_mimetypes
            def patched_register(self, filenames):
                try:
                    return original_register(self, filenames)
                except KeyError as e:
                    if '.mpo' in str(e):
                        # .mpo 파일을 제외하고 다시 시도
                        filtered_filenames = [f for f in filenames if not f.lower().endswith('.mpo')]
                        return original_register(self, filtered_filenames)
                    raise
            openpyxl.packaging.manifest.Manifest._register_mimetypes = patched_register
        except Exception as patch_error:
            pass

    user_info = get_user_info()
    
    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()
                
                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    }), 403

            # RCM 정보 및 파일 경로 가져오기 (업로드한 사용자의 회사명, 최신 설계평가 세션명 포함)
            rcm_info = conn.execute('''
                SELECT r.rcm_name, r.original_filename, u.company_name
                FROM sb_rcm r
                LEFT JOIN sb_user u ON r.user_id = u.user_id
                WHERE r.rcm_id = %s
            ''', (rcm_id,)).fetchone()
            
            # 평가 세션명 조회 (통합 테이블) - 파라미터로 전달되지 않았을 경우 최신 세션 사용
            if not evaluation_session:
                session_info = conn.execute('''
                    SELECT evaluation_name
                    FROM sb_evaluation_header
                    WHERE rcm_id = %s
                    ORDER BY created_at DESC
                    LIMIT 1
                ''', (rcm_id,)).fetchone()

                if session_info:
                    evaluation_session = session_info['evaluation_name']
            
            if not rcm_info:
                return jsonify({
                    'success': False,
                    'message': 'RCM 정보를 찾을 수 없습니다.'
                }), 404
            
            # RCM 상세 정보 가져오기
            rcm_details = conn.execute('''
                SELECT detail_id, control_code, control_name, control_frequency, control_type, 
                       test_procedure, control_description
                FROM sb_rcm_detail 
                WHERE rcm_id = %s 
                ORDER BY detail_id
            ''', (rcm_id,)).fetchall()
            
            if not rcm_details:
                return jsonify({
                    'success': False,
                    'message': 'RCM 상세 정보를 찾을 수 없습니다.'
                }), 404
            
            # 설계평가 데이터 조회 (평가 근거 및 첫부 이미지)
            evaluation_data = {}
            if evaluation_session:
                try:
                    # 해당 세션의 설계평가 결과 조회 (N+1 문제 해결: line_id 포함)
                    eval_results = conn.execute('''
                        SELECT l.control_code, l.evaluation_evidence, l.evaluation_rationale,
                               l.overall_effectiveness, l.line_id, h.header_id
                        FROM sb_evaluation_line l
                        JOIN sb_evaluation_header h ON l.header_id = h.header_id
                        WHERE h.rcm_id = %s AND h.evaluation_name = %s
                        ORDER BY l.control_sequence, l.control_code
                    ''', (rcm_id, evaluation_session)).fetchall()

                    if eval_results:
                        # 모든 line_id 수집
                        line_ids = [eval['line_id'] for eval in eval_results]

                        # 한 번의 쿼리로 모든 이미지 조회
                        placeholders = ','.join(['%s'] * len(line_ids))
                        all_images = conn.execute(f'''
                            SELECT line_id, file_path, file_name
                            FROM sb_evaluation_image
                            WHERE evaluation_type = %s AND line_id IN ({placeholders})
                            ORDER BY line_id, uploaded_at
                        ''', ['design'] + line_ids).fetchall()

                        # line_id별로 이미지 그룹화
                        images_by_line = {}
                        for img in all_images:
                            line_id = img['line_id']
                            if line_id not in images_by_line:
                                images_by_line[line_id] = []
                            images_by_line[line_id].append({'file': img['file_name'], 'path': img['file_path']})

                        # 각 평가 결과에 이미지 매핑
                        for eval_result in eval_results:
                            control_code = eval_result['control_code']
                            evidence = eval_result['evaluation_evidence'] or ''
                            rationale = eval_result['evaluation_rationale'] or ''
                            effectiveness = eval_result['overall_effectiveness'] or ''
                            line_id = eval_result['line_id']

                            # 해당 line_id의 이미지 가져오기
                            images = images_by_line.get(line_id, [])
                            images_info = json.dumps(images) if images else ''

                            evaluation_data[control_code] = {
                                'evidence': evidence,
                                'rationale': rationale,
                                'effectiveness': effectiveness,
                                'images': images_info
                            }
                except Exception as e:
                    pass

        # 원본 파일 경로 결정
        original_file_path = None
        if rcm_info.get('original_filename'):
            # 회사별 폴더 구조 또는 루트 uploads 폴더 확인
            if os.path.sep in rcm_info['original_filename'] or '/' in rcm_info['original_filename']:
                potential_path = os.path.join('uploads', rcm_info['original_filename'])
            else:
                # 기본 rcm 업로드 경로도 확인 (auth.py의 get_rcm_file_path 참고)
                potential_path = os.path.join('static', 'uploads', 'rcm', str(rcm_id), rcm_info['original_filename'])
            
            if os.path.exists(potential_path):
                original_file_path = potential_path
            else:
                # uploads 루트에서 한 번 더 확인
                root_potential_path = os.path.join('uploads', rcm_info['original_filename'])
                if os.path.exists(root_potential_path):
                    original_file_path = root_potential_path

        # 워크북 로드 (로드 실패 시 Fallback)
        workbook = None
        if original_file_path:
            try:
                workbook = load_workbook(original_file_path)
            except Exception as e:
                pass

        # Fallback: 원본 파일이 없거나 로드 실패 시 시스템 템플릿 사용
        if not workbook:
            fallback_template = os.path.join(os.path.dirname(__file__), 'paper_templates', 'Design_Template.xlsx')
            if os.path.exists(fallback_template):
                try:
                    workbook = load_workbook(fallback_template)
                except Exception as e:
                    return jsonify({'success': False, 'message': f'템플릿 로드 실패: {str(e)}'}), 500
            else:
                return jsonify({'success': False, 'message': '원본 파일과 시스템 템플릿을 모두 찾을 수 없습니다.'}), 404
        
        # Template 시트가 있는지 확인
        if 'Template' not in workbook.sheetnames:
            return jsonify({
                'success': False,
                'message': '원본 파일에 Template 시트가 없습니다.'
            }), 404
            
        template_sheet = workbook['Template']
        
        # 각 통제별로 Template 시트를 복사해서 새 시트 생성
        for detail in rcm_details:
            control_code = detail['control_code']
            
            # Template 시트 복사
            new_sheet = workbook.copy_worksheet(template_sheet)
            new_sheet.title = control_code
            
            # B3, B5, C7~C11 셀에 정보 입력
            new_sheet['B3'] = user_info.get('company_name', '')  # 회사명
            new_sheet['B5'] = user_info.get('user_name', '')  # 사용자명
            new_sheet['C7'] = detail['control_code']  # 통제코드
            new_sheet['C8'] = detail['control_name']  # 통제명
            new_sheet['C9'] = detail['control_frequency']  # 통제주기
            new_sheet['C10'] = detail['control_type']  # 통제구분
            new_sheet['C11'] = detail['test_procedure'] or ''  # 테스트 절차
            
            # C12에 증빙, C13에 평가 근거, C14에 첨부 이미지, C15에 효과성 추가
            eval_info = evaluation_data.get(control_code, {})
            evidence_value = eval_info.get('evidence', '')
            rationale_value = eval_info.get('rationale', '')
            effectiveness_value = eval_info.get('effectiveness', '')

            new_sheet['C12'] = evidence_value  # 증빙
            new_sheet['C13'] = rationale_value  # 평가 근거

            # 여러 라인 텍스트가 있는 셀들의 높이 자동 조정
            # C11 (테스트 절차) 셀 설정
            c11_cell = new_sheet['C11']
            c11_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # C12 (증빙) 셀 설정
            c12_cell = new_sheet['C12']
            c12_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # C13 (평가 근거) 셀 설정
            c13_cell = new_sheet['C13']
            c13_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # 텍스트 길이에 따라 행 높이 자동 조정
            test_procedure_text = detail['test_procedure'] or ''
            evidence_text = evidence_value or ''
            rationale_text = rationale_value or ''

            # 줄바꿈 개수와 텍스트 길이를 기준으로 높이 계산
            c11_lines = max(1, len(test_procedure_text.split('\n'))) if test_procedure_text else 1
            c12_lines = max(1, len(evidence_text.split('\n'))) if evidence_text else 1
            c13_lines = max(1, len(rationale_text.split('\n'))) if rationale_text else 1
            
            # 긴 텍스트의 경우 줄바꿈이 없어도 자동 줄바꿈으로 인해 여러 줄이 될 수 있음
            # 약 50자당 1줄로 가정 (엑셀 C열 기준)
            if test_procedure_text:
                estimated_c11_lines = max(c11_lines, (len(test_procedure_text) // 50) + 1)
            else:
                estimated_c11_lines = c11_lines

            if evidence_text:
                estimated_c12_lines = max(c12_lines, (len(evidence_text) // 50) + 1)
            else:
                estimated_c12_lines = c12_lines

            if rationale_text:
                estimated_c13_lines = max(c13_lines, (len(rationale_text) // 50) + 1)
            else:
                estimated_c13_lines = c13_lines

            # 최소 높이는 25pt, 최대 높이는 150pt로 제한 (기본 20pt + 줄 수 * 18pt)
            c11_height = max(25, min(300, 20 + (estimated_c11_lines - 1) * 18))
            c12_height = max(25, min(150, 20 + (estimated_c12_lines - 1) * 18))
            c13_height = max(25, min(150, 20 + (estimated_c13_lines - 1) * 18))

            new_sheet.row_dimensions[11].height = c11_height
            new_sheet.row_dimensions[12].height = c12_height
            new_sheet.row_dimensions[13].height = c13_height

            # C15에 효과성 및 시트 탭 색상 설정
            if effectiveness_value == '효과적' or effectiveness_value.lower() == 'effective':
                new_sheet['C15'] = 'Effective' # 효과적인 경우
                new_sheet.sheet_properties.tabColor = Color(rgb="90EE90")
            else:
                new_sheet['C15'] = 'Ineffective'
                # 비효과적인 경우 시트 탭을 연한 빨간색으로 설정
                new_sheet.sheet_properties.tabColor = Color(rgb="FFA0A0")  # 연한 빨간색
            
            # 첨부 이미지 처리 (C14 셀에 이미지 삽입)
            images_info = eval_info.get('images', '')
            if images_info:
                try:
                    from openpyxl.drawing.image import Image as ExcelImage

                    # JSON 형태로 저장된 이미지 정보 파싱
                    if images_info.startswith('['):
                        image_list = json.loads(images_info)
                        if image_list and len(image_list) > 0:
                            # 최대 3개 이미지를 C14, D14, E14에 가로로 배치
                            target_cells = ['C14', 'D14', 'E14']
                            max_images = min(len(image_list), 3)

                            # C14 셀 크기 기준으로 통일된 이미지 크기 계산
                            col_width = new_sheet.column_dimensions['C'].width or 8.43
                            cell_width_px = int(col_width * 7)  # 문자 단위를 픽셀로 변환

                            row_height = new_sheet.row_dimensions[14].height or 15
                            cell_height_px = int(row_height * 1.33)  # 포인트를 픽셀로 변환

                            # 셀 크기의 90%로 최대 크기 설정 (모든 이미지에 동일 적용)
                            uniform_max_width = max(int(cell_width_px * 0.9), 150)
                            uniform_max_height = max(int(cell_height_px * 0.9), 150)

                            for idx in range(max_images):
                                image_path = image_list[idx].get('path', '')
                                if image_path and os.path.exists(image_path):
                                    try:
                                        # PIL을 사용해 이미지를 정리한 후 임시 파일로 저장
                                        from PIL import Image as PILImage
                                        import tempfile

                                        # 원본 이미지 로드 (메타데이터 제거)
                                        pil_image = PILImage.open(image_path)
                                        # RGB로 변환 (EXIF 데이터 제거)
                                        if pil_image.mode != 'RGB':
                                            pil_image = pil_image.convert('RGB')

                                        # 임시 파일로 저장 (메타데이터 없이)
                                        temp_image_file = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
                                        temp_image_file.close()
                                        pil_image.save(temp_image_file.name, 'JPEG', quality=85)
                                        pil_image.close()

                                        # 정리된 이미지를 엑셀에 삽입
                                        img = ExcelImage(temp_image_file.name)

                                        # 원본 이미지 크기 비율 계산
                                        original_width = img.width
                                        original_height = img.height

                                        # 가로세로 비율 유지하며 통일된 크기로 조정
                                        width_ratio = uniform_max_width / original_width
                                        height_ratio = uniform_max_height / original_height
                                        scale_ratio = min(width_ratio, height_ratio)

                                        img.width = int(original_width * scale_ratio)
                                        img.height = int(original_height * scale_ratio)

                                        # 해당 셀에 이미지 추가
                                        new_sheet.add_image(img, target_cells[idx])

                                        # 임시 파일은 나중에 정리
                                        if not hasattr(workbook, 'temp_files'):
                                            workbook.temp_files = []
                                        workbook.temp_files.append(temp_image_file.name)

                                    except Exception as img_error:
                                        # 오류 발생시 해당 셀에 메시지 표시
                                        new_sheet[target_cells[idx]] = f'이미지 {idx+1} 오류'
                    else:
                        # 단순 문자열 경우
                        new_sheet['C13'] = images_info
                except Exception as e:
                    new_sheet['C13'] = f'이미지 처리 오류: {str(e)}'
            else:
                new_sheet['C13'] = ''
        
        # 원본 Template 시트 삭제
        workbook.remove(template_sheet)
        
        # downloads 폴더에 파일 저장
        company_name = rcm_info['company_name'] or '회사명없음'
        evaluation_name = evaluation_session or '설계평가'
        safe_filename = f"{company_name}_{evaluation_name}.xlsx"
        downloads_path = os.path.join('downloads', safe_filename)
        
        # 기존 파일이 있으면 삭제
        if os.path.exists(downloads_path):
            os.remove(downloads_path)
        
        # .mpo 파일 문제 해결을 위한 대체 저장 방법
        try:
            workbook.save(downloads_path)
        except Exception as save_error:
            
            # 임시 디렉토리에서 작업
            import tempfile
            import shutil
            
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = os.path.join(temp_dir, "evaluation.xlsx")
                
                try:
                    # 임시 경로에 저장 시도
                    workbook.save(temp_path)
                    # 성공하면 최종 경로로 복사
                    shutil.copy2(temp_path, downloads_path)
                    
                except Exception as temp_save_error:
                    
                    # 최후의 수단: BytesIO를 사용한 메모리 저장
                    try:
                        from io import BytesIO
                        buffer = BytesIO()
                        workbook.save(buffer)
                        buffer.seek(0)
                        
                        with open(downloads_path, 'wb') as f:
                            f.write(buffer.read())
                        
                    except Exception as bytesio_error:
                        raise save_error  # 원래 오류를 다시 발생시킴
        
        # 임시 이미지 파일들 정리
        if hasattr(workbook, 'temp_files'):
            for temp_file in workbook.temp_files:
                try:
                    os.unlink(temp_file)
                except:
                    pass  # 임시 파일 삭제 실패 시 무시
        
        workbook.close()
        
        def remove_download_file():
            try:
                if os.path.exists(downloads_path):
                    os.unlink(downloads_path)
            except:
                pass
        
        # 활동 로그 기록
        log_user_activity(user_info, 'DESIGN_EVALUATION_DOWNLOAD', 
                         f'설계평가 엑셀 다운로드 - {rcm_info["rcm_name"]}', 
                         f'/api/design-evaluation/download-excel/{rcm_id}', 
                         request.remote_addr, request.headers.get('User-Agent'))
        
        # 다운로드 파일 전송
        response = send_file(
            downloads_path,
            as_attachment=True,
            download_name=safe_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 요청 완료 후 다운로드 파일 삭제
        response.call_on_close(remove_download_file)
        
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        
        # 구체적인 오류 메시지 반환
        error_msg = str(e)
        if "No such file or directory" in error_msg:
            error_msg = "원본 RCM 파일을 찾을 수 없습니다."
        elif "Template" in error_msg:
            error_msg = "원본 파일에 Template 시트가 없습니다."
        elif "Permission denied" in error_msg:
            error_msg = "파일 접근 권한이 없습니다."
        else:
            error_msg = f"엑셀 생성 중 오류 발생: {error_msg}"
            
        return jsonify({
            'success': False,
            'message': error_msg
        }), 500

@bp_link6.route('/api/design-evaluation/archive', methods=['POST'])
@login_required
def archive_design_evaluation_api():
    """설계평가 세션 Archive 처리 API"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')

    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })

    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

        # 세션 Archive 처리
        result = archive_design_evaluation_session(rcm_id, user_info['user_id'], evaluation_session)

        if result.get('success'):
            # 활동 로그 기록
            log_user_activity(user_info, 'DESIGN_EVALUATION_ARCHIVE',
                             f'설계평가 세션 Archive - {evaluation_session}',
                             f'/api/design-evaluation/archive',
                             request.remote_addr, request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'설계평가 세션 "{evaluation_session}"이 Archive 처리되었습니다.'
            })
        else:
            return jsonify({
                'success': False,
                'message': result.get('message', '해당 평가 세션을 찾을 수 없습니다.')
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Archive 처리 중 오류가 발생했습니다.'
        })

@bp_link6.route('/api/design-evaluation/unarchive', methods=['POST'])
@login_required
def unarchive_design_evaluation_api():
    """설계평가 세션 Unarchive 처리 API"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    evaluation_session = data.get('evaluation_session')

    if not rcm_id or not evaluation_session:
        return jsonify({
            'success': False,
            'message': 'RCM ID와 평가 세션이 필요합니다.'
        })

    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인 (관리자 권한 포함)
        with get_db() as conn:
            # 관리자인지 먼저 확인
            if user_info.get('admin_flag') == 'Y':
                pass  # 관리자는 모든 RCM에 접근 가능
            else:
                # 일반 사용자는 명시적 권한 확인
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], rcm_id)).fetchone()

                if not access_check:
                    return jsonify({
                        'success': False,
                        'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                    })

        # 세션 Unarchive 처리
        result = unarchive_design_evaluation_session(rcm_id, user_info['user_id'], evaluation_session)

        if result.get('success'):
            # 활동 로그 기록
            log_user_activity(user_info, 'DESIGN_EVALUATION_UNARCHIVE',
                             f'설계평가 세션 Unarchive - {evaluation_session}',
                             f'/api/design-evaluation/unarchive',
                             request.remote_addr, request.headers.get('User-Agent'))

            return jsonify({
                'success': True,
                'message': f'설계평가 세션 "{evaluation_session}"이 복원되었습니다.'
            })
        else:
            return jsonify({
                'success': False,
                'message': result.get('message', '해당 평가 세션을 찾을 수 없습니다.')
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Unarchive 처리 중 오류가 발생했습니다.'
        })
# ============================================================================
# ELC 설계평가 (수동통제만, 기준통제 매핑 없음)
# ============================================================================

@bp_link6.route('/elc/design-evaluation')
@login_required
def elc_design_evaluation():
    """ELC 평가 페이지 (설계평가 + 운영평가 통합)"""
    user_info = get_user_info()

    # ELC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    elc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'ELC']

    # 각 RCM에 대한 설계평가/운영평가 세션 정보 추가
    from evaluation_utils import get_evaluation_status

    for rcm in elc_rcms:
        with get_db() as conn:
            # 설계평가 현황: archived가 아닌 모든 평가
            design_sessions = conn.execute('''
                SELECT header_id, evaluation_name, created_at, last_updated
                FROM sb_evaluation_header
                WHERE rcm_id = ? AND (archived IS NULL OR archived = 0)
                ORDER BY last_updated DESC
            ''', (rcm['rcm_id'],)).fetchall()

            # 운영평가 현황은 설계평가 세션 중에서 status >= 2인 것만 필터링 (나중에 처리)

        # 설계평가 세션 정보 처리
        design_list = []
        operation_list = []

        for session in design_sessions:
            session_dict = dict(session)
            header_id = session_dict['header_id']

            # 실시간으로 status와 progress 계산
            with get_db() as conn:
                status_info = get_evaluation_status(conn, header_id)

            session_dict['status'] = status_info['status']
            session_dict['progress'] = status_info['design_progress']

            # 템플릿 호환성을 위해 evaluation_session 키 추가
            session_dict['evaluation_session'] = session_dict['evaluation_name']

            # 설계평가 완료 여부: status >= 1
            if session_dict["status"] >= 1:
                session_dict["is_completed"] = True
                session_dict["completed_date"] = session_dict.get("last_updated", "")
            else:
                session_dict["is_completed"] = False
                session_dict["completed_date"] = ""

            # 운영평가 가능한 통제 개수 (설계평가 완료된 경우만)
            if session_dict['is_completed']:
                session_dict['eligible_control_count'] = status_info['operation_total_count']
            else:
                session_dict['eligible_control_count'] = 0

            # 설계평가 목록: status >= 0 (모든 평가 세션을 표시)
            # status 0: 설계평가 진행중, 1: 설계평가 완료, 2~4: 운영평가 단계
            design_list.append(session_dict)

            # 운영평가 목록: status >= 2 (운영평가가 시작된 것)
            if session_dict['status'] >= 2:
                op_dict = session_dict.copy()
                op_dict['is_completed'] = session_dict['status'] == 4
                op_dict['progress'] = status_info['operation_progress']
                op_dict['operation_completed_count'] = status_info['operation_completed_count']
                op_dict['design_evaluation_name'] = session_dict['evaluation_name']
                op_dict['eligible_control_count'] = status_info['operation_total_count']
                # 설계평가 완료 정보 추가 (템플릿에서 "설계평가 기반" 칼럼에 표시)
                op_dict['design_completed_count'] = status_info['design_completed_count']
                op_dict['design_total_count'] = status_info['design_total_count']
                operation_list.append(op_dict)

        rcm['design_sessions'] = design_list
        rcm['operation_sessions'] = operation_list
        rcm['has_design_sessions'] = len(design_list) > 0
        rcm['has_operation_sessions'] = len(operation_list) > 0

        # 하위 호환성을 위해 기존 키도 유지
        # 운영평가 시작 가능한 세션: status = 1 (설계평가 완료)이면서 운영평가가 아직 시작되지 않은 것만
        # 운영평가가 시작되면 status >= 2가 되므로, status = 1인 것만 필터링하면 됨
        # 하지만 안전하게 operation_list에 없는 것만 포함
        operation_session_names = {s['evaluation_name'] for s in operation_list}
        rcm['completed_design_sessions'] = [
            s for s in design_list
            if s['status'] == 1 and s['evaluation_name'] not in operation_session_names
        ]
        rcm['design_evaluation_completed'] = len(rcm['completed_design_sessions']) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'ELC 평가', '/elc/design-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    # 세션에서 현재 선택된 RCM과 평가 세션 정보 가져오기
    from flask import session as flask_session
    current_rcm_id = flask_session.get('current_rcm_id')
    current_evaluation_session = flask_session.get('current_evaluation_session')

    return render_template('link6_evaluation.jsp',
                         category='ELC',
                         rcms=elc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         current_rcm_id=current_rcm_id,
                         current_evaluation_session=current_evaluation_session)

@bp_link6.route('/elc/design-evaluation/start', methods=['POST'])
@login_required
def elc_design_evaluation_start():
    """ELC 설계평가 시작 - 평가 세션 생성 및 평가 페이지로 이동"""
    user_info = get_user_info()
    rcm_id = request.form.get('rcm_id')
    evaluation_name = request.form.get('evaluation_name', '').strip()

    if not rcm_id or not evaluation_name:
        flash('RCM ID와 평가명은 필수입니다.', 'danger')
        return redirect(url_for('link6.elc_design_evaluation'))

    try:
        rcm_id = int(rcm_id)
    except ValueError:
        flash('유효하지 않은 RCM ID입니다.', 'danger')
        return redirect(url_for('link6.elc_design_evaluation'))

    # sb_evaluation_header에 새 레코드 생성 (status=0: 설계 시작)
    with get_db() as conn:
        # 중복 평가명 체크
        existing = conn.execute('''
            SELECT header_id FROM sb_evaluation_header
            WHERE rcm_id = %s AND evaluation_name = %s
        ''', (rcm_id, evaluation_name)).fetchone()

        if existing:
            flash(f'이미 "{evaluation_name}" 평가명이 존재합니다. 다른 이름을 사용해주세요.', 'warning')
            return redirect(url_for('link6.elc_design_evaluation'))

        # 새 평가 세션 생성
        # 평가 기간 날짜 가져오기
        evaluation_period_start = request.form.get('evaluation_period_start', '').strip() or None
        evaluation_period_end = request.form.get('evaluation_period_end', '').strip() or None

        cursor = conn.execute('''
            INSERT INTO sb_evaluation_header
            (rcm_id, evaluation_name, evaluation_period_start, evaluation_period_end, status, progress)
            VALUES (%s, %s, %s, %s, 0, 0)
        ''', (rcm_id, evaluation_name, evaluation_period_start, evaluation_period_end))
        header_id = cursor.lastrowid

        # RCM의 ELC 통제 목록을 sb_evaluation_line에 삽입
        conn.execute('''
            INSERT INTO sb_evaluation_line (header_id, control_code)
            SELECT %s, control_code
            FROM sb_rcm_detail
            WHERE rcm_id = %s AND control_category = 'ELC'
        ''', (header_id, rcm_id))

        conn.commit()

        log_user_activity(user_info, 'CREATE', 'ELC 설계평가 시작',
                         f'rcm_id={rcm_id}, evaluation_name={evaluation_name}',
                         request.remote_addr, request.headers.get('User-Agent'))

    # 세션에 정보 저장 후 설계평가 페이지로 리다이렉트
    session['current_design_rcm_id'] = rcm_id
    session['current_evaluation_type'] = 'ELC'
    session['current_evaluation_session'] = evaluation_name

    flash(f'"{evaluation_name}" 설계평가를 시작합니다.', 'success')
    return redirect(url_for('link6.user_design_evaluation_rcm'))


@bp_link6.route('/elc/evaluation/delete', methods=['POST'])
@login_required
def delete_elc_evaluation():
    """ELC 평가 삭제 (진행중인 평가만 삭제 가능, 운영평가 시작된 경우 불가)"""
    user_info = get_user_info()
    data = request.get_json()
    header_id = data.get('header_id')

    if not header_id:
        return jsonify({'success': False, 'error': 'header_id is required'}), 400

    try:
        with get_db() as conn:
            # 평가 정보 조회
            evaluation = conn.execute(
                'SELECT evaluation_name, status FROM sb_evaluation_header WHERE header_id = ?',
                (header_id,)
            ).fetchone()

            if not evaluation:
                return jsonify({'success': False, 'error': '평가를 찾을 수 없습니다.'}), 404

            eval_dict = dict(evaluation)
            current_status = eval_dict['status']

            # status 0 또는 1: 완전 삭제
            # status >= 2: status를 1로 롤백 (운영평가 데이터 삭제)
            if current_status <= 1:
                # 설계평가만 있는 경우: 완전 삭제
                conn.execute('DELETE FROM sb_evaluation_line WHERE header_id = ?', (header_id,))
                conn.execute('DELETE FROM sb_evaluation_header WHERE header_id = ?', (header_id,))
                action = '삭제'
            else:
                # 운영평가가 시작된 경우: status를 1로 롤백
                conn.execute('UPDATE sb_evaluation_header SET status = 1, progress = 100 WHERE header_id = ?', (header_id,))
                
                # 운영평가 관련 필드만 NULL로 초기화
                conn.execute('''
                    UPDATE sb_evaluation_line 
                    SET conclusion = NULL, 
                        improvement_plan = NULL,
                        sample_size = NULL,
                        exception_details = NULL
                    WHERE header_id = ?
                ''', (header_id,))
                action = 'status 롤백 (1로)'
            
            conn.commit()

            log_user_activity(user_info, 'DELETE', f'ELC 평가 {action}',
                             f'header_id={header_id}, evaluation_name={eval_dict["evaluation_name"]}, prev_status={current_status}',
                             request.remote_addr, request.headers.get('User-Agent'))

            return jsonify({'success': True, 'action': action})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



@bp_link6.route('/elc/evaluation/archive', methods=['POST'])
@login_required
def archive_elc_evaluation():
    """ELC 평가 아카이브 (status를 5로 변경)"""
    user_info = get_user_info()
    data = request.get_json()
    header_id = data.get('header_id')

    if not header_id:
        return jsonify({'success': False, 'error': 'header_id is required'}), 400

    try:
        with get_db() as conn:
            # 평가 정보 조회
            evaluation = conn.execute(
                'SELECT evaluation_name, status, rcm_id FROM sb_evaluation_header WHERE header_id = ?',
                (header_id,)
            ).fetchone()

            if not evaluation:
                return jsonify({'success': False, 'error': '평가를 찾을 수 없습니다.'}), 404

            # auth.py의 공통 함수 사용 (user_id는 rcm_id 권한 확인용으로 이미 상위에서 get_user_info() 되어 있음)
            result = archive_design_evaluation_session(evaluation['rcm_id'], user_info['user_id'], evaluation['evaluation_name'])

            if result.get('success'):
                log_user_activity(user_info, 'ARCHIVE', 'ELC 평가 아카이브',
                                 f'header_id={header_id}, evaluation_name={evaluation["evaluation_name"]}',
                                 request.remote_addr, request.headers.get('User-Agent'))
                return jsonify({'success': True, 'message': 'ELC 평가가 아카이브되었습니다.'})
            else:
                return jsonify({'success': False, 'error': result.get('message', '아카이브 중 오류가 발생했습니다.')})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# TLC 설계평가 (자동통제 포함, 기준통제 매핑 없음)
# ============================================================================

@bp_link6.route('/tlc/design-evaluation')
@login_required
def tlc_design_evaluation():
    """TLC 평가 페이지 (설계평가 + 운영평가 통합)"""
    user_info = get_user_info()

    # TLC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    tlc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'TLC']

    # 각 RCM에 대한 설계평가/운영평가 세션 정보 추가
    from evaluation_utils import get_evaluation_status

    for rcm in tlc_rcms:
        with get_db() as conn:
            # 설계평가 현황: archived가 아닌 모든 평가
            design_sessions = conn.execute('''
                SELECT header_id, evaluation_name, created_at, last_updated
                FROM sb_evaluation_header
                WHERE rcm_id = ? AND (archived IS NULL OR archived = 0)
                ORDER BY last_updated DESC
            ''', (rcm['rcm_id'],)).fetchall()

        # 설계평가 세션 정보 처리
        design_list = []
        operation_list = []

        for session in design_sessions:
            session_dict = dict(session)
            header_id = session_dict['header_id']

            # 실시간으로 status와 progress 계산
            with get_db() as conn:
                status_info = get_evaluation_status(conn, header_id)

            session_dict['status'] = status_info['status']
            session_dict['progress'] = status_info['design_progress']

            # 템플릿 호환성을 위해 evaluation_session 키 추가
            session_dict['evaluation_session'] = session_dict['evaluation_name']

            # 설계평가 완료 여부: status >= 1
            if session_dict["status"] >= 1:
                session_dict["is_completed"] = True
                session_dict["completed_date"] = session_dict.get("last_updated", "")
            else:
                session_dict["is_completed"] = False
                session_dict["completed_date"] = ""

            # 운영평가 가능한 통제 개수 (설계평가 완료된 경우만)
            if session_dict['is_completed']:
                session_dict['eligible_control_count'] = status_info['operation_total_count']
            else:
                session_dict['eligible_control_count'] = 0

            # 설계평가 목록: status >= 0 (모든 평가 세션을 표시)
            design_list.append(session_dict)

            # 운영평가 목록: status >= 2 (운영평가가 시작된 것)
            if session_dict['status'] >= 2:
                op_dict = session_dict.copy()
                op_dict['is_completed'] = session_dict['status'] == 4
                op_dict['progress'] = status_info['operation_progress']
                op_dict['operation_completed_count'] = status_info['operation_completed_count']
                op_dict['design_evaluation_name'] = session_dict['evaluation_name']
                op_dict['eligible_control_count'] = status_info['operation_total_count']
                op_dict['design_completed_count'] = status_info['design_completed_count']
                op_dict['design_total_count'] = status_info['design_total_count']
                operation_list.append(op_dict)

        rcm['design_sessions'] = design_list
        rcm['operation_sessions'] = operation_list
        rcm['has_design_sessions'] = len(design_list) > 0
        rcm['has_operation_sessions'] = len(operation_list) > 0

        # 하위 호환성을 위해 기존 키도 유지
        operation_session_names = {s['evaluation_name'] for s in operation_list}
        rcm['completed_design_sessions'] = [
            s for s in design_list
            if s['status'] == 1 and s['evaluation_name'] not in operation_session_names
        ]
        rcm['design_evaluation_completed'] = len(rcm['completed_design_sessions']) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'TLC 평가', '/tlc/design-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    # 세션에서 현재 선택된 RCM과 평가 세션 정보 가져오기
    from flask import session as flask_session
    current_rcm_id = flask_session.get('current_rcm_id')
    current_evaluation_session = flask_session.get('current_evaluation_session')

    return render_template('link6_evaluation.jsp',
                         category='TLC',
                         rcms=tlc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         current_rcm_id=current_rcm_id,
                         current_evaluation_session=current_evaluation_session)

@bp_link6.route('/tlc-evaluation')
@login_required
def tlc_evaluation():
    """TLC 평가 페이지 (설계평가 + 운영평가 통합)"""
    user_info = get_user_info()

    # TLC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    tlc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'TLC']

    # 각 RCM에 대한 설계평가/운영평가 세션 정보 추가
    from evaluation_utils import get_evaluation_status

    for rcm in tlc_rcms:
        with get_db() as conn:
            # 설계평가 현황: archived가 아닌 모든 평가 (통합 테이블 사용)
            design_sessions = conn.execute('''
                SELECT header_id, evaluation_name, created_at, last_updated
                FROM sb_evaluation_header
                WHERE rcm_id = ? AND (archived IS NULL OR archived = 0)
                ORDER BY last_updated DESC
            ''', (rcm['rcm_id'],)).fetchall()

        # 설계평가 세션 정보 처리
        design_list = []
        operation_list = []

        for session in design_sessions:
            session_dict = dict(session)
            header_id = session_dict['header_id']

            # 실시간으로 status와 progress 계산
            with get_db() as conn:
                status_info = get_evaluation_status(conn, header_id)

            session_dict['status'] = status_info['status']
            session_dict['progress'] = status_info['design_progress']

            # 템플릿 호환성을 위해 evaluation_session 키 추가
            session_dict['evaluation_session'] = session_dict['evaluation_name']

            # 설계평가 완료 여부: status >= 1
            if session_dict["status"] >= 1:
                session_dict["is_completed"] = True
                session_dict["completed_date"] = session_dict.get("last_updated", "")
            else:
                session_dict["is_completed"] = False
                session_dict["completed_date"] = ""

            # 운영평가 가능한 통제 개수
            if session_dict['is_completed']:
                session_dict['eligible_control_count'] = status_info['operation_total_count']
            else:
                session_dict['eligible_control_count'] = 0

            # 설계평가 목록: status >= 0 (모든 평가 세션을 표시)
            # status 0: 설계평가 진행중, 1: 설계평가 완료, 2~4: 운영평가 단계
            design_list.append(session_dict)

            # 운영평가 목록: status >= 2 (운영평가가 시작된 것)
            if session_dict['status'] >= 2:
                op_dict = session_dict.copy()
                op_dict['is_completed'] = session_dict['status'] == 4
                op_dict['progress'] = status_info['operation_progress']
                op_dict['operation_completed_count'] = status_info['operation_completed_count']
                op_dict['design_evaluation_name'] = session_dict['evaluation_name']
                op_dict['eligible_control_count'] = status_info['operation_total_count']
                op_dict['design_completed_count'] = status_info['design_completed_count']
                op_dict['design_total_count'] = status_info['design_total_count']
                operation_list.append(op_dict)

        rcm['design_sessions'] = design_list
        rcm['operation_sessions'] = operation_list
        rcm['has_design_sessions'] = len(design_list) > 0
        rcm['has_operation_sessions'] = len(operation_list) > 0

        # 운영평가 시작 가능한 세션: 설계평가 완료(is_completed=True)이고 운영평가 목록에 없는 것
        operation_session_names = {s['evaluation_name'] for s in operation_list}
        rcm['completed_design_sessions'] = [
            s for s in design_list
            if s['is_completed'] and s['evaluation_name'] not in operation_session_names
        ]
        rcm['design_evaluation_completed'] = len(rcm['completed_design_sessions']) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'TLC 통합 평가', '/tlc-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    # 세션에서 현재 선택된 RCM과 평가 세션 정보 가져오기
    from flask import session as flask_session
    current_rcm_id = flask_session.get('current_rcm_id')
    current_evaluation_session = flask_session.get('current_evaluation_session')

    return render_template('link6_evaluation.jsp',
                         category='TLC',
                         rcms=tlc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         current_rcm_id=current_rcm_id,
                         current_evaluation_session=current_evaluation_session)

@bp_link6.route('/tlc/design-evaluation/start', methods=['POST'])
@login_required
def tlc_design_evaluation_start():
    """TLC 설계평가 시작 - 평가 세션 생성 및 평가 페이지로 이동"""
    user_info = get_user_info()
    rcm_id = request.form.get('rcm_id')
    evaluation_name = request.form.get('evaluation_name', '').strip()

    if not rcm_id or not evaluation_name:
        flash('RCM ID와 평가명은 필수입니다.', 'danger')
        return redirect(url_for('link6.tlc_evaluation'))

    try:
        rcm_id = int(rcm_id)
    except ValueError:
        flash('유효하지 않은 RCM ID입니다.', 'danger')
        return redirect(url_for('link6.tlc_evaluation'))

    # sb_evaluation_header에 새 레코드 생성 (status=0: 설계 시작)
    with get_db() as conn:
        # 중복 평가명 체크
        existing = conn.execute('''
            SELECT header_id FROM sb_evaluation_header
            WHERE rcm_id = %s AND evaluation_name = %s
        ''', (rcm_id, evaluation_name)).fetchone()

        if existing:
            flash(f'이미 "{evaluation_name}" 평가명이 존재합니다. 다른 이름을 사용해주세요.', 'warning')
            return redirect(url_for('link6.tlc_evaluation'))

        # 새 평가 세션 생성
        evaluation_period_start = request.form.get('evaluation_period_start', '').strip() or None
        evaluation_period_end = request.form.get('evaluation_period_end', '').strip() or None

        cursor = conn.execute('''
            INSERT INTO sb_evaluation_header
            (rcm_id, evaluation_name, evaluation_period_start, evaluation_period_end, status, progress)
            VALUES (%s, %s, %s, %s, 0, 0)
        ''', (rcm_id, evaluation_name, evaluation_period_start, evaluation_period_end))
        header_id = cursor.lastrowid

        # RCM의 TLC 통제 목록을 sb_evaluation_line에 삽입
        conn.execute('''
            INSERT INTO sb_evaluation_line (header_id, control_code)
            SELECT %s, control_code
            FROM sb_rcm_detail
            WHERE rcm_id = %s AND control_category = 'TLC'
        ''', (header_id, rcm_id))

        conn.commit()

        log_user_activity(user_info, 'CREATE', 'TLC 설계평가 시작',
                         f'rcm_id={rcm_id}, evaluation_name={evaluation_name}',
                         request.remote_addr, request.headers.get('User-Agent'))

    # 세션에 정보 저장 후 설계평가 페이지로 리다이렉트
    session['current_design_rcm_id'] = rcm_id
    session['current_evaluation_type'] = 'TLC'
    session['current_evaluation_session'] = evaluation_name

    flash(f'"{evaluation_name}" 설계평가를 시작합니다.', 'success')
    return redirect(url_for('link6.user_design_evaluation_rcm'))

@bp_link6.route('/itgc-evaluation')
@login_required
def itgc_evaluation():
    """ITGC 평가 페이지 (설계평가 + 운영평가 통합)"""
    user_info = get_user_info()

    # ITGC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    itgc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'ITGC']

    # 각 RCM에 대한 설계평가/운영평가 세션 정보 추가
    from evaluation_utils import get_evaluation_status

    for rcm in itgc_rcms:
        with get_db() as conn:
            # 설계평가 현황: archived가 아닌 모든 평가 (통합 테이블 사용)
            design_sessions = conn.execute('''
                SELECT header_id, evaluation_name, created_at, last_updated
                FROM sb_evaluation_header
                WHERE rcm_id = ? AND (archived IS NULL OR archived = 0)
                ORDER BY last_updated DESC
            ''', (rcm['rcm_id'],)).fetchall()

        # 설계평가 세션 정보 처리
        design_list = []
        operation_list = []

        for session in design_sessions:
            session_dict = dict(session)
            header_id = session_dict['header_id']

            # 실시간으로 status와 progress 계산
            with get_db() as conn:
                status_info = get_evaluation_status(conn, header_id)

            session_dict['status'] = status_info['status']
            session_dict['progress'] = status_info['design_progress']

            # 템플릿 호환성을 위해 evaluation_session 키 추가
            session_dict['evaluation_session'] = session_dict['evaluation_name']

            # 설계평가 완료 여부: status >= 1
            if session_dict["status"] >= 1:
                session_dict["is_completed"] = True
                session_dict["completed_date"] = session_dict.get("last_updated", "")
            else:
                session_dict["is_completed"] = False
                session_dict["completed_date"] = ""

            # 운영평가 가능한 통제 개수
            if session_dict['is_completed']:
                session_dict['eligible_control_count'] = status_info['operation_total_count']
            else:
                session_dict['eligible_control_count'] = 0

            # 설계평가 목록: status >= 0 (모든 평가 세션을 표시)
            # status 0: 설계평가 진행중, 1: 설계평가 완료, 2~4: 운영평가 단계
            design_list.append(session_dict)

            # 운영평가 목록: status >= 2 (운영평가가 시작된 것)
            if session_dict['status'] >= 2:
                op_dict = session_dict.copy()
                op_dict['is_completed'] = session_dict['status'] == 4
                op_dict['progress'] = status_info['operation_progress']
                op_dict['operation_completed_count'] = status_info['operation_completed_count']
                op_dict['design_evaluation_name'] = session_dict['evaluation_name']
                op_dict['eligible_control_count'] = status_info['operation_total_count']
                op_dict['design_completed_count'] = status_info['design_completed_count']
                op_dict['design_total_count'] = status_info['design_total_count']
                operation_list.append(op_dict)

        rcm['design_sessions'] = design_list
        rcm['operation_sessions'] = operation_list
        rcm['has_design_sessions'] = len(design_list) > 0
        rcm['has_operation_sessions'] = len(operation_list) > 0

        # 운영평가 시작 가능한 세션: 설계평가 완료(is_completed=True)이고 운영평가 목록에 없는 것
        operation_session_names = {s['evaluation_name'] for s in operation_list}
        rcm['completed_design_sessions'] = [
            s for s in design_list
            if s['is_completed'] and s['evaluation_name'] not in operation_session_names
        ]
        rcm['design_evaluation_completed'] = len(rcm['completed_design_sessions']) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'ITGC 통합 평가', '/itgc-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    # 세션에서 현재 선택된 RCM과 평가 세션 정보 가져오기
    from flask import session as flask_session
    current_rcm_id = flask_session.get('current_rcm_id')
    current_evaluation_session = flask_session.get('current_evaluation_session')

    # start_design 파라미터가 있으면 해당 RCM으로 설계평가 시작 모달 자동 표시
    start_design_rcm_id = request.args.get('start_design')
    start_design_rcm_name = None
    if start_design_rcm_id:
        # RCM 이름 조회
        for rcm in itgc_rcms:
            if str(rcm['rcm_id']) == str(start_design_rcm_id):
                start_design_rcm_name = rcm['rcm_name']
                break

    return render_template('link6_evaluation.jsp',
                         category='ITGC',
                         rcms=itgc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         current_rcm_id=current_rcm_id,
                         current_evaluation_session=current_evaluation_session,
                         start_design_rcm_id=start_design_rcm_id,
                         start_design_rcm_name=start_design_rcm_name)

@bp_link6.route('/itgc/design-evaluation/start', methods=['POST'])
@login_required
def itgc_design_evaluation_start():
    """ITGC 설계평가 시작 - 평가 세션 생성 및 평가 페이지로 이동"""
    user_info = get_user_info()
    rcm_id = request.form.get('rcm_id')
    evaluation_name = request.form.get('evaluation_name', '').strip()

    if not rcm_id or not evaluation_name:
        flash('RCM ID와 평가명은 필수입니다.', 'danger')
        return redirect(url_for('link6.itgc_evaluation'))

    try:
        rcm_id = int(rcm_id)
    except ValueError:
        flash('유효하지 않은 RCM ID입니다.', 'danger')
        return redirect(url_for('link6.itgc_evaluation'))

    # sb_evaluation_header에 새 레코드 생성 (status=0: 설계 시작)
    with get_db() as conn:
        # 중복 평가명 체크
        existing = conn.execute('''
            SELECT header_id FROM sb_evaluation_header
            WHERE rcm_id = ? AND evaluation_name = ?
        ''', (rcm_id, evaluation_name)).fetchone()

        if existing:
            flash(f'이미 "{evaluation_name}" 평가명이 존재합니다. 다른 이름을 사용해주세요.', 'warning')
            return redirect(url_for('link6.itgc_evaluation'))

        # 새 평가 세션 생성
        # 평가 기간 날짜 가져오기
        evaluation_period_start = request.form.get('evaluation_period_start', '').strip() or None
        evaluation_period_end = request.form.get('evaluation_period_end', '').strip() or None

        cursor = conn.execute('''
            INSERT INTO sb_evaluation_header
            (rcm_id, evaluation_name, evaluation_period_start, evaluation_period_end, status, progress)
            VALUES (?, ?, ?, ?, 0, 0)
        ''', (rcm_id, evaluation_name, evaluation_period_start, evaluation_period_end))
        header_id = cursor.lastrowid

        # RCM의 ITGC 통제 목록을 sb_evaluation_line에 삽입
        conn.execute('''
            INSERT INTO sb_evaluation_line (header_id, control_code)
            SELECT ?, control_code
            FROM sb_rcm_detail
            WHERE rcm_id = ? AND control_category = 'ITGC'
        ''', (header_id, rcm_id))

        conn.commit()

        log_user_activity(user_info, 'CREATE', 'ITGC 설계평가 시작',
                         f'rcm_id={rcm_id}, evaluation_name={evaluation_name}',
                         request.remote_addr, request.headers.get('User-Agent'))

    # 세션에 정보 저장 후 설계평가 페이지로 리다이렉트
    session['current_design_rcm_id'] = rcm_id
    session['current_evaluation_type'] = 'ITGC'
    session['current_evaluation_session'] = evaluation_name

    flash(f'"{evaluation_name}" 설계평가를 시작합니다.', 'success')
    return redirect(url_for('link6.user_design_evaluation_rcm'))

@bp_link6.route('/itgc/evaluation/delete', methods=['POST'])
@login_required
def delete_itgc_evaluation():
    """
    ITGC 평가 삭제
    - 설계평가 삭제 (status <= 1): header, line, sample 완전 삭제
    - 운영평가 삭제 (status >= 2): status를 1로 되돌리고 운영평가 필드만 초기화
    """
    data = request.get_json()
    header_id = data.get('header_id')

    if not header_id:
        return jsonify({'success': False, 'error': 'header_id is required'})

    try:
        with get_db() as conn:
            # header 존재 확인 및 status, rcm_id 조회
            header = conn.execute('SELECT rcm_id, status FROM sb_evaluation_header WHERE header_id = ?', (header_id,)).fetchone()
            if not header:
                return jsonify({'success': False, 'error': 'Evaluation not found'})

            header_dict = dict(header)
            rcm_id = header_dict['rcm_id']
            current_status = header_dict.get('status', 0)

            import shutil

            if current_status <= 1:
                # 설계평가 상태 (status <= 1): 세션 완전 삭제

                # 설계평가 파일 삭제 (파일시스템)
                design_upload_dir = os.path.join('static', 'uploads', 'design_evaluations', str(rcm_id), str(header_id))
                if os.path.exists(design_upload_dir):
                    try:
                        shutil.rmtree(design_upload_dir)
                    except Exception as file_error:
                        print(f"설계평가 파일 삭제 중 오류 발생: {file_error}")

                # sample 삭제
                conn.execute('''
                    DELETE FROM sb_evaluation_sample
                    WHERE line_id IN (
                        SELECT line_id FROM sb_evaluation_line WHERE header_id = ?
                    )
                ''', (header_id,))

                # line 삭제
                conn.execute('DELETE FROM sb_evaluation_line WHERE header_id = ?', (header_id,))

                # header 삭제
                conn.execute('DELETE FROM sb_evaluation_header WHERE header_id = ?', (header_id,))

            else:
                # 운영평가 상태 (status >= 2): 설계평가 완료 상태로 되돌림

                # 운영평가 파일 삭제 (파일시스템)
                operation_upload_dir = os.path.join('static', 'uploads', 'operation_evaluations', str(rcm_id), str(header_id))
                if os.path.exists(operation_upload_dir):
                    try:
                        shutil.rmtree(operation_upload_dir)
                    except Exception as file_error:
                        print(f"운영평가 파일 삭제 중 오류 발생: {file_error}")

                # status를 1(설계평가 완료)로 되돌림
                conn.execute('UPDATE sb_evaluation_header SET status = 1 WHERE header_id = ?', (header_id,))

                # 운영평가 관련 필드만 초기화 (설계평가 필드는 유지)
                conn.execute('''
                    UPDATE sb_evaluation_line
                    SET conclusion = NULL,
                        operation_comment = NULL,
                        exception_details = NULL,
                        sample_size = NULL
                    WHERE header_id = ?
                ''', (header_id,))

                # 운영평가 sample만 삭제 (evaluation_type='operation'인 것만)
                # 설계평가 sample(evaluation_type='design')은 유지
                conn.execute('''
                    DELETE FROM sb_evaluation_sample
                    WHERE line_id IN (
                        SELECT line_id FROM sb_evaluation_line WHERE header_id = ?
                    )
                    AND (evaluation_type = 'operation' OR evaluation_type IS NULL)
                ''', (header_id,))

            conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@bp_link6.route('/itgc/evaluation/archive', methods=['POST'])
@login_required
def archive_itgc_evaluation():
    """ITGC 평가 아카이브"""
    data = request.get_json()
    header_id = data.get('header_id')

    if not header_id:
        return jsonify({'success': False, 'error': 'header_id is required'})

    try:
        with get_db() as conn:
            # header 존재 확인
            header = conn.execute('SELECT evaluation_name, rcm_id FROM sb_evaluation_header WHERE header_id = ?', (header_id,)).fetchone()
            if not header:
                return jsonify({'success': False, 'error': 'Evaluation not found'})

            # auth.py의 공통 함수 사용
            result = archive_design_evaluation_session(header['rcm_id'], user_info['user_id'], header['evaluation_name'])

            if result.get('success'):
                # 활동 로그 기록
                log_user_activity(user_info, 'ARCHIVE_ITGC', 
                                f'ITGC 설계평가 아카이브 - {header["evaluation_name"]}',
                                f'/itgc/evaluation/archive', 
                                request.remote_addr, request.headers.get('User-Agent'))
                return jsonify({'success': True, 'message': 'ITGC 설계평가가 아카이브되었습니다.'})
            else:
                return jsonify({'success': False, 'error': result.get('message', '아카이브 중 오류가 발생했습니다.')})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ============================================================================
# 설계평가 다운로드 기능
# ============================================================================

def format_evidence_for_excel(evaluation_evidence_json, attributes_dict, population_count):
    """
    증빙 데이터를 "컬럼명: 증빙값" 형식으로 포맷팅

    Args:
        evaluation_evidence_json (str): JSON 형식의 증빙 데이터
        attributes_dict (dict): {'attribute0': '거래일자', 'attribute1': '거래금액', ...}
        population_count (int): 모집단 항목 개수 (이 값 이후부터 증빙으로 간주)

    Returns:
        tuple: (formatted_evidence, line_count)
            - formatted_evidence (str): 줄바꿈으로 구분된 증빙 내용
            - line_count (int): 줄 수 (셀 높이 계산용)
    """
    # 빈 값 처리
    if not evaluation_evidence_json or not evaluation_evidence_json.strip():
        return '', 0

    try:
        # JSON 파싱 시도
        evidence_data = json.loads(evaluation_evidence_json)

        # JSON 객체가 아니면 원본 반환
        if not isinstance(evidence_data, dict):
            # 줄 수 계산
            line_count = evaluation_evidence_json.count('\n') + 1
            return evaluation_evidence_json, line_count

        # 증빙 항목만 추출 (모집단 제외)
        evidence_lines = []
        for i in range(population_count, 10):  # attribute{population_count}부터 attribute9까지
            attr_key = f'attribute{i}'
            attr_name = attributes_dict.get(attr_key)
            attr_value = evidence_data.get(attr_key, '')

            # 속성명과 값이 모두 있을 때만 추가
            if attr_name and attr_value:
                evidence_lines.append(f'{attr_name}: {attr_value}')

        # 줄바꿈으로 연결
        formatted_evidence = '\n'.join(evidence_lines)
        line_count = len(evidence_lines)

        return formatted_evidence, line_count

    except (json.JSONDecodeError, ValueError, TypeError):
        # JSON 파싱 실패 시 (레거시 데이터) 원본 텍스트 반환
        line_count = evaluation_evidence_json.count('\n') + 1
        return evaluation_evidence_json, line_count


@bp_link6.route('/design-evaluation/download')
@login_required
def download_design_evaluation():
    """설계평가 결과 엑셀 다운로드 (레거시 경로 - 새로운 API로 리다이렉트)"""
    rcm_id = request.args.get('rcm_id')
    evaluation_session = request.args.get('evaluation_session')

    if not rcm_id:
        flash('RCM ID가 필요합니다.', 'error')
        return redirect(url_for('link6.user_design_evaluation'))

    return redirect(url_for('link6.download_evaluation_excel', rcm_id=rcm_id, evaluation_session=evaluation_session))
@bp_link6.route('/api/client-log', methods=['POST'])
def client_log():
    """클라이언트 JavaScript 로그를 Flask 콘솔에 출력"""
    try:
        data = request.get_json()
        message = data.get('message', 'No message')
        log_data = data.get('data', {})
        
        print(f'[CLIENT LOG] {message}')
        if log_data:
            import json
            print(f'[CLIENT DATA] {json.dumps(log_data, ensure_ascii=False, indent=2)}')
        
        return jsonify({'success': True})
    except Exception as e:
        print(f'[CLIENT LOG ERROR] {e}')
        return jsonify({'success': False})


@bp_link6.route('/api/design-evaluation/delete-image/<int:image_id>', methods=['POST'])
@login_required
def delete_design_evaluation_image(image_id):
    """설계평가 이미지 삭제 API"""
    import os

    try:
        user_info = get_user_info()

        with get_db() as conn:
            # 이미지 정보 조회 (권한 확인용)
            image_info = conn.execute('''
                SELECT ei.file_path, ei.line_id, h.rcm_id
                FROM sb_evaluation_image ei
                JOIN sb_evaluation_line l ON ei.line_id = l.line_id
                JOIN sb_evaluation_header h ON l.header_id = h.header_id
                WHERE ei.image_id = %s AND ei.evaluation_type = 'design'
            ''', (image_id,)).fetchone()

            if not image_info:
                return jsonify({'success': False, 'message': '이미지를 찾을 수 없습니다.'})

            # 권한 확인 (관리자이거나 해당 RCM에 접근 권한이 있는지)
            if user_info.get('admin_flag') != 'Y':
                access_check = conn.execute('''
                    SELECT permission_type FROM sb_user_rcm
                    WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
                ''', (user_info['user_id'], image_info['rcm_id'])).fetchone()

                if not access_check:
                    return jsonify({'success': False, 'message': '삭제 권한이 없습니다.'})

            # 파일 삭제
            file_path = image_info['file_path']
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    # 파일 삭제 실패해도 DB 레코드는 삭제 진행
                    pass

            # DB 레코드 삭제
            conn.execute('''
                DELETE FROM sb_evaluation_image
                WHERE image_id = %s
            ''', (image_id,))
            conn.commit()

            return jsonify({'success': True, 'message': '이미지가 삭제되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'이미지 삭제 중 오류가 발생했습니다: {str(e)}'})
