from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, session, send_file
from auth import login_required, get_current_user, get_user_rcms, get_rcm_details, get_key_rcm_details, save_operation_evaluation, get_operation_evaluations, get_operation_evaluation_samples, get_design_evaluation_sample, log_user_activity, get_db, is_design_evaluation_completed, get_completed_design_evaluation_sessions
from snowball_link5 import get_user_info, is_logged_in
import json
import os
import tempfile
from openpyxl import load_workbook

bp_link7 = Blueprint('link7', __name__)

# ============================================================================
# 운영평가 통제 설정 (control_config.py 통합)
# 새로운 통제 추가 시 이 섹션의 설정만 추가하면 자동으로 기능이 구현됩니다.
# ============================================================================

# 수동통제 설정 (APD01, APD07, APD09, APD12 등)
MANUAL_CONTROLS = {
    'APD01': {
        'name': '사용자 권한 부여 승인',
        'description': '사용자 권한 부여 승인 테스트',

        # 모집단 필드 매핑
        'population_fields': ['user_id', 'user_name', 'department', 'permission', 'grant_date'],
        'population_headers': ['사용자ID', '사용자명', '부서명', '권한명', '권한부여일'],

        # 엑셀 헤더 (한글)
        'excel_headers': {
            'population': ['사용자ID', '사용자명', '부서명', '권한명', '권한부여일'],
            'testing': ['No', '사용자ID', '사용자명', '부서', '권한', '부여일',
                       '요청번호', '요청자명', '요청자부서', '승인자', '승인자부서', '승인일자', '예외', '비고']
        },

        # 필드 매핑 UI 라벨
        'field_labels': ['사용자ID', '사용자명', '부서명', '권한명', '권한부여일'],

        # 필수 컬럼 (업로드 시 안내)
        'required_fields': ['사용자ID', '사용자명', '부서명', '권한명', '권한부여일'],

        # Sticky 컬럼 설정 (좌측 고정)
        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 120, 'left': 50},    # 사용자ID
            {'width': 100, 'left': 170},   # 사용자명
            {'width': 120, 'left': 270},   # 부서
            {'width': 150, 'left': 390},   # 권한
            {'width': 120, 'left': 540}    # 부여일
        ],

        # 일반 컬럼 설정
        'normal_columns': [
            {'label': '요청번호', 'width': 180},
            {'label': '요청자명', 'width': 150},
            {'label': '요청자부서', 'width': 180},
            {'label': '승인자', 'width': 150},
            {'label': '승인자부서', 'width': 180},
            {'label': '승인일자', 'width': 150},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'APD07': {
        'name': 'DB 변경 승인',
        'description': 'DB 변경 승인 테스트',

        'population_fields': ['change_id', 'change_date'],
        'population_headers': ['쿼리(변경내역)', '변경일자'],

        'excel_headers': {
            'population': ['쿼리(변경내역)', '변경일자', '변경유형', '테이블명', '변경자', '승인일자'],
            'testing': ['No', '쿼리', '실행일자',
                       '변경 요청서 번호', '요청자명', '요청부서', '승인자명', '승인자부서', '승인일자', '승인여부', '사전승인여부', '결론', '비고']
        },

        'field_labels': ['쿼리(변경내역)', '변경일자'],
        'required_fields': ['쿼리', '실행일자'],

        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 200, 'left': 50},    # 쿼리
            {'width': 150, 'left': 250}    # 실행일자
        ],

        'normal_columns': [
            {'label': '변경 요청서 번호', 'width': 180},
            {'label': '요청자명', 'width': 150},
            {'label': '요청부서', 'width': 180},
            {'label': '승인자명', 'width': 150},
            {'label': '승인자부서', 'width': 180},
            {'label': '승인일자', 'width': 150},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'APD09': {
        'name': 'OS 접근권한 부여 승인',
        'description': 'OS 접근권한 부여 승인 테스트',

        'population_fields': ['account', 'grant_date'],
        'population_headers': ['OS 접근권한명', '접근권한부여일'],

        'excel_headers': {
            'population': ['OS 접근권한명', '접근권한부여일'],
            'testing': ['No', 'OS 접근권한명', '접근권한부여일',
                       '권한 요청서 번호', '요청자명', '요청부서', '승인자명', '승인자부서', '승인일자', '승인여부', '사전승인여부', '결론', '비고']
        },

        'field_labels': ['접근권한 부여 계정', '권한부여일'],
        'required_fields': ['접근권한 부여 계정', '권한부여일'],

        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 200, 'left': 50},    # 접근권한명
            {'width': 150, 'left': 250}    # 권한부여일
        ],

        'normal_columns': [
            {'label': '요청번호', 'width': 180},
            {'label': '요청자명', 'width': 150},
            {'label': '요청자부서', 'width': 180},
            {'label': '승인자', 'width': 150},
            {'label': '승인자부서', 'width': 180},
            {'label': '승인일자', 'width': 150},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'APD12': {
        'name': 'DB 접근권한 부여 승인',
        'description': 'DB 접근권한 부여 승인 테스트',

        'population_fields': ['account', 'grant_date'],
        'population_headers': ['DB 접근권한명', '접근권한부여일'],

        'excel_headers': {
            'population': ['DB 접근권한명', '접근권한부여일'],
            'testing': ['No', 'DB 접근권한명', '접근권한부여일',
                       '권한 요청서 번호', '요청자명', '요청부서', '승인자명', '승인자부서', '승인일자', '승인여부', '사전승인여부', '결론', '비고']
        },

        'field_labels': ['접근권한 부여 계정', '권한부여일'],
        'required_fields': ['접근권한 부여 계정', '권한부여일'],

        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 200, 'left': 50},    # 접근권한명
            {'width': 150, 'left': 250}    # 권한부여일
        ],

        'normal_columns': [
            {'label': '요청번호', 'width': 180},
            {'label': '요청자명', 'width': 150},
            {'label': '요청자부서', 'width': 180},
            {'label': '승인자', 'width': 150},
            {'label': '승인자부서', 'width': 180},
            {'label': '승인일자', 'width': 150},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'PC01': {
        'name': '프로그램 변경 승인',
        'description': '프로그램 변경 승인 테스트',

        # 모집단 필드 매핑 (필수: 프로그램명, 반영일자)
        'population_fields': ['program_name', 'deploy_date'],
        'population_headers': ['프로그램명', '반영일자'],

        # 엑셀 헤더
        'excel_headers': {
            'population': ['프로그램명', '반영일자'],
            'testing': ['No', '프로그램명', '반영일자',
                       '요청번호', '요청자', '요청자부서', '승인자', '승인자부서', '승인일자',
                       '개발담당자', '배포담당자', '예외', '비고']
        },

        # 필드 매핑 UI 라벨
        'field_labels': ['프로그램명', '반영일자'],

        # 필수 컬럼 (업로드 시 안내)
        'required_fields': ['프로그램명', '반영일자'],

        # Sticky 컬럼 설정 (좌측 고정)
        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 200, 'left': 50},    # 프로그램명
            {'width': 120, 'left': 250}    # 반영일자
        ],

        # 일반 컬럼 설정
        'normal_columns': [
            {'label': '변경 요청서 번호', 'width': 180},
            {'label': '요청자', 'width': 120},
            {'label': '요청자부서', 'width': 150},
            {'label': '승인자', 'width': 120},
            {'label': '승인자부서', 'width': 150},
            {'label': '승인일자', 'width': 120},
            {'label': '개발담당자', 'width': 120},
            {'label': '배포담당자', 'width': 120},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'PC02': {
        'name': '사용자 테스트',
        'description': '사용자 테스트 수행 확인',

        # PC01과 동일한 모집단 필드 (PC01 데이터 참조)
        'population_fields': ['program_name', 'deploy_date'],
        'population_headers': ['프로그램명', '반영일자'],

        # 엑셀 헤더
        'excel_headers': {
            'population': ['프로그램명', '반영일자'],
            'testing': ['No', '프로그램명', '반영일자',
                       '사용자테스트 유무', '사용자테스트담당자', '사용자테스트일자', '예외', '비고']
        },

        # PC01 참조용 플래그
        'depends_on': 'PC01',
        'skip_upload': True,  # 모집단 업로드 스킵 (PC01 데이터 사용)

        # 필드 매핑 UI는 사용 안함
        'field_labels': [],
        'required_fields': [],

        # Sticky 컬럼 설정 (PC01과 동일)
        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 200, 'left': 50},    # 프로그램명
            {'width': 120, 'left': 250}    # 반영일자
        ],

        # 일반 컬럼 설정 (PC02 전용)
        'normal_columns': [
            {'label': '변경 요청서 번호', 'width': 180},
            {'label': '사용자테스트 유무', 'width': 150},
            {'label': '사용자테스트담당자', 'width': 150},
            {'label': '사용자테스트일자', 'width': 150},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'PC03': {
        'name': '배포 승인',
        'description': '배포 승인 확인',

        # PC01과 동일한 모집단 필드 (PC01 데이터 참조)
        'population_fields': ['program_name', 'deploy_date'],
        'population_headers': ['프로그램명', '반영일자'],

        # 엑셀 헤더
        'excel_headers': {
            'population': ['프로그램명', '반영일자'],
            'testing': ['No', '프로그램명', '반영일자',
                       '변경 요청서 번호', '배포요청자', '배포요청자부서', '배포승인자', '배포승인일자', '예외', '비고']
        },

        # PC01 참조용 플래그
        'depends_on': 'PC01',
        'skip_upload': True,  # 모집단 업로드 스킵 (PC01 데이터 사용)

        # 필드 매핑 UI는 사용 안함
        'field_labels': [],
        'required_fields': [],

        # Sticky 컬럼 설정 (PC01과 동일)
        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 200, 'left': 50},    # 프로그램명
            {'width': 120, 'left': 250}    # 반영일자
        ],

        # 일반 컬럼 설정 (PC03 전용)
        'normal_columns': [
            {'label': '변경 요청서 번호', 'width': 180},
            {'label': '배포요청자', 'width': 150},
            {'label': '배포요청자부서', 'width': 180},
            {'label': '배포승인자', 'width': 150},
            {'label': '배포승인자부서', 'width': 180},
            {'label': '배포승인일자', 'width': 150},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'CO01': {
        'name': '배치 스케줄 승인',
        'description': '배치 스케줄 등록 및 승인 확인',

        # 모집단 필드
        'population_fields': ['batch_schedule_name', 'register_date'],
        'population_headers': ['배치스케줄이름', '등록일자'],

        # 엑셀 헤더
        'excel_headers': {
            'population': ['배치스케줄이름', '등록일자'],
            'testing': ['No', '배치스케줄이름', '등록일자',
                       '요청번호', '요청자', '요청자부서', '승인자', '승인자부서', '승인일자', '예외', '비고']
        },

        # 필드 라벨 및 필수 필드
        'field_labels': {
            'batch_schedule_name': '배치스케줄이름',
            'register_date': '등록일자'
        },
        'required_fields': ['batch_schedule_name', 'register_date'],

        # Sticky 컬럼 설정
        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 250, 'left': 50},    # 배치스케줄이름
            {'width': 120, 'left': 300}    # 등록일자
        ],

        # 일반 컬럼 설정
        'normal_columns': [
            {'label': '요청번호', 'width': 180},
            {'label': '요청자', 'width': 120},
            {'label': '요청자부서', 'width': 150},
            {'label': '승인자', 'width': 120},
            {'label': '승인자부서', 'width': 150},
            {'label': '승인일자', 'width': 120},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    },

    'GENERIC': {
        'name': '일반 수동통제',
        'description': '범용 수동통제 운영평가',

        # 모집단 필드 (2개 필드만 요구 - 컬럼명은 사용자가 지정)
        'population_fields': ['field1', 'field2'],
        'population_headers': ['필드1', '필드2'],

        # 엑셀 헤더
        'excel_headers': {
            'population': ['필드1', '필드2'],
            'testing': ['No', '필드1', '필드2',
                       '증빙1', '증빙2', '증빙3', '증빙4', '증빙5', '예외', '비고']
        },

        # 필드 라벨 (사용자 정의)
        'field_labels': ['주요 필드1', '주요 필드2'],
        'required_fields': ['주요 필드1', '주요 필드2'],

        # Sticky 컬럼 설정
        'sticky_columns': [
            {'width': 50, 'left': 0},      # No
            {'width': 200, 'left': 50},    # 필드1
            {'width': 200, 'left': 250}    # 필드2
        ],

        # 일반 컬럼 설정 (범용 증빙 컬럼)
        'normal_columns': [
            {'label': '증빙1', 'width': 200},
            {'label': '증빙2', 'width': 200},
            {'label': '증빙3', 'width': 200},
            {'label': '증빙4', 'width': 200},
            {'label': '증빙5', 'width': 200},
            {'label': '예외', 'width': 80},
            {'label': '비고', 'width': 300}
        ]
    }
}


def get_control_config(control_code):
    """통제 코드로 설정 조회"""
    return MANUAL_CONTROLS.get(control_code)


def get_all_manual_controls():
    """모든 수동통제 목록 조회"""
    return list(MANUAL_CONTROLS.keys())


def is_manual_control(control_code):
    """수동통제 여부 확인"""
    return control_code in MANUAL_CONTROLS

# ============================================================================
# 운영평가 관련 기능들
# ============================================================================

@bp_link7.route('/operation-evaluation')
@login_required
def user_operation_evaluation():
    """운영평가 페이지"""
    user_info = get_user_info()

    # 사용자가 접근 가능한 RCM 목록 조회 (ITGC만)
    user_rcms = get_user_rcms(user_info['user_id'], control_category='ITGC')

    # 각 RCM에 대해 모든 설계평가 세션 조회 (진행중 + 완료)
    from auth import get_all_design_evaluation_sessions
    for rcm in user_rcms:
        all_sessions = get_all_design_evaluation_sessions(rcm['rcm_id'], user_info['user_id'])
        completed_sessions = [s for s in all_sessions if s['completed_date'] is not None]
        in_progress_sessions = [s for s in all_sessions if s['completed_date'] is None]

        # 완료된 세션에 대해서만 운영평가 진행상황 조회
        for session in completed_sessions:
            operation_evaluation_session = f"OP_{session['evaluation_session']}"

            # 운영평가 진행 통제 수 조회
            from auth import count_completed_operation_evaluations
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm['rcm_id'], user_info['user_id'], operation_evaluation_session, session['evaluation_session'])).fetchone()

            if header:
                completed_count = count_completed_operation_evaluations(header['header_id'])
                session['operation_completed_count'] = completed_count
            else:
                session['operation_completed_count'] = 0

            # 운영평가 가능한 통제 개수 추가
            eligible_controls = get_key_rcm_details(rcm['rcm_id'], user_info['user_id'], session['evaluation_session'])
            session['eligible_control_count'] = len(eligible_controls)

        rcm['all_design_sessions'] = all_sessions
        rcm['completed_design_sessions'] = completed_sessions
        rcm['in_progress_design_sessions'] = in_progress_sessions
        rcm['design_evaluation_completed'] = len(completed_sessions) > 0

        # 핵심통제 개수 조회 (모든 핵심통제)
        key_controls = get_key_rcm_details(rcm['rcm_id'])
        rcm['key_control_count'] = len(key_controls)
        rcm['has_key_controls'] = len(key_controls) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', '운영평가', '/user/operation-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link7_evaluation.jsp',
                         evaluation_type='ITGC',
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         user_rcms=user_rcms,
                         remote_addr=request.remote_addr)

@bp_link7.route('/operation-evaluation/rcm', methods=['GET', 'POST'])
@login_required
def user_operation_evaluation_rcm():
    """RCM별 운영평가 페이지 (설계평가 세션 기반)"""
    user_info = get_user_info()

    # POST로 전달된 RCM ID와 설계평가 세션 정보 받기
    if request.method == 'POST':
        rcm_id = request.form.get('rcm_id')
        design_evaluation_session = request.form.get('design_evaluation_session')
        new_operation_session = request.form.get('new_operation_session')  # 신규 운영평가 세션명
        action = request.form.get('action')  # 'start', 'continue' 등

        if not rcm_id:
            flash('RCM 정보가 없습니다.', 'error')
            return redirect(url_for('link7.user_operation_evaluation'))
        if not design_evaluation_session:
            flash('설계평가 세션 정보가 없습니다.', 'error')
            return redirect(url_for('link7.user_operation_evaluation'))

        # 세션에 저장
        session['current_operation_rcm_id'] = int(rcm_id)
        session['current_design_evaluation_session'] = design_evaluation_session

        # 신규 운영평가 세션인 경우
        if new_operation_session:
            session['new_operation_session_name'] = new_operation_session
            flash(f'새로운 운영평가 세션 "{new_operation_session}"을 시작합니다.', 'success')
        else:
            # 기존 세션 제거
            session.pop('new_operation_session_name', None)

    # POST든 GET이든 세션에서 정수형 rcm_id를 가져옴
    rcm_id = session.get('current_operation_rcm_id')
    design_evaluation_session = session.get('current_design_evaluation_session')

    if not rcm_id:
        flash('RCM 정보가 없습니다. 다시 선택해주세요.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))
    if not design_evaluation_session:
        flash('설계평가 세션 정보가 없습니다. 다시 선택해주세요.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 사용자가 해당 RCM에 접근 권한이 있는지 확인
    user_rcms = get_user_rcms(user_info['user_id'])
    rcm_ids = [rcm['rcm_id'] for rcm in user_rcms]

    if rcm_id not in rcm_ids:
        flash('해당 RCM에 대한 접근 권한이 없습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))


    # 해당 설계평가 세션이 완료되었는지 확인

    # 통합 테이블 사용 (sb_evaluation_header) - ELC, TLC, ITGC 모두
    # status >= 2이면 운영평가 가능
    with get_db() as conn:
        evaluation_header = conn.execute('''
            SELECT header_id, status, progress, evaluation_name
            FROM sb_evaluation_header
            WHERE rcm_id = ? AND evaluation_name = ?
        ''', (rcm_id, design_evaluation_session)).fetchone()

    if not evaluation_header:
        # 통합 테이블에 없으면 기존 방식으로 확인 (구 ITGC용)
        completed_sessions = get_completed_design_evaluation_sessions(rcm_id, user_info['user_id'])
        session_found = False
        for session_item in completed_sessions:
            if session_item['evaluation_session'] == design_evaluation_session:
                session_found = True
                break

        if not session_found:
            flash(f'설계평가 세션 "{design_evaluation_session}"이 완료되지 않아 운영평가를 수행할 수 없습니다.', 'warning')
            return redirect(url_for('link7.user_operation_evaluation'))
    else:
        # 통합 테이블 사용: status 확인 및 업데이트
        evaluation_dict = dict(evaluation_header)

        # action='start'이고 status=1(설계평가 완료)이면 status를 2(운영평가 시작)로 변경
        if request.method == 'POST' and action == 'start' and evaluation_dict['status'] == 1:
            with get_db() as conn:
                conn.execute('''
                    UPDATE sb_evaluation_header
                    SET status = 2, last_updated = CURRENT_TIMESTAMP
                    WHERE header_id = ?
                ''', (evaluation_dict['header_id'],))
                conn.commit()
            evaluation_dict['status'] = 2  # 메모리상 값도 업데이트

        # status < 1 (설계평가 미완료)이면 리다이렉트
        if evaluation_dict['status'] < 1:
            flash(f'설계평가 세션 "{design_evaluation_session}"이 완료되지 않아 운영평가를 수행할 수 없습니다.', 'warning')
            # RCM 정보 조회하여 동적 리다이렉트
            user_rcms_temp = get_user_rcms(user_info['user_id'])
            rcm_info_temp = None
            for rcm in user_rcms_temp:
                if rcm['rcm_id'] == rcm_id:
                    rcm_info_temp = rcm
                    break

            if rcm_info_temp:
                control_category = rcm_info_temp.get('control_category', 'ELC')
                if control_category == 'ITGC':
                    return redirect(url_for('link6.itgc_evaluation'))
                elif control_category == 'TLC':
                    return redirect(url_for('link6.tlc_evaluation'))
                else:
                    return redirect(url_for('link6.elc_design_evaluation'))
            else:
                return redirect(url_for('link6.elc_design_evaluation'))

        # status가 2이면 3으로 변경 (운영평가 시작 → 진행중)
        if evaluation_dict['status'] == 2:
            with get_db() as conn:
                conn.execute('''
                    UPDATE sb_evaluation_header
                    SET status = 3, last_updated = CURRENT_TIMESTAMP
                    WHERE header_id = ?
                ''', (evaluation_dict['header_id'],))
                conn.commit()

    # RCM 정보 조회
    rcm_info = None
    for rcm in user_rcms:
        if rcm['rcm_id'] == rcm_id:
            rcm_info = rcm
            break
    
    # RCM 핵심통제 데이터 조회 (운영평가는 핵심통제이면서 설계평가가 '적정'인 통제만 대상)
    rcm_details = get_key_rcm_details(rcm_id, user_info['user_id'], design_evaluation_session)
    
    # 매핑 정보 조회
    from auth import get_rcm_detail_mappings
    rcm_mappings_list = get_rcm_detail_mappings(rcm_id)
    # control_code를 키로 하는 딕셔너리로 변환
    rcm_mappings = {m['control_code']: m for m in rcm_mappings_list}

    # 핵심통제이면서 설계평가가 '적정'인 통제가 없는 경우 안내 메시지 표시
    if not rcm_details:
        flash('해당 RCM에 설계평가 결과가 "적정"인 핵심통제가 없어 운영평가를 수행할 수 없습니다.', 'warning')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 각 통제 코드에 대한 config 정보 미리 로드
    control_configs = {}
    for detail in rcm_details:
        control_configs[detail['control_code']] = get_control_config(detail['control_code'])

    # 운영평가 세션명 생성 (설계평가 세션 기반)
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    # 운영평가 Header/Line 데이터 동기화 (설계평가 결과 변경 반영)
    sync_messages = []
    operation_header = None

    # ELC 평가인지 확인
    is_elc = rcm_info and rcm_info.get('control_category') == 'ELC'

    try:
        # 모든 평가 유형(ELC, ITGC, TLC): 통합 테이블 + get_evaluation_status() 함수 사용
        from evaluation_utils import get_evaluation_status

        with get_db() as conn:
            # 헤더 정보 조회
            evaluation_header = conn.execute('''
                SELECT header_id, status, evaluation_name
                FROM sb_evaluation_header
                WHERE rcm_id = ? AND evaluation_name = ?
            ''', (rcm_id, design_evaluation_session)).fetchone()

            if evaluation_header:
                eval_dict = dict(evaluation_header)
                header_id = eval_dict['header_id']

                # get_evaluation_status() 함수로 진행률 계산 (통일된 방식)
                status_info = get_evaluation_status(conn, header_id)

                # operation_header 형식으로 변환 (템플릿 호환성 유지)
                operation_header = {
                    'header_id': header_id,
                    'evaluated_controls': status_info['operation_completed_count'],
                    'total_controls': status_info['operation_total_count'],
                    'progress_percentage': status_info['operation_progress'],
                    'evaluation_status': 'COMPLETED' if eval_dict['status'] == 4 else 'IN_PROGRESS'
                }
            else:
                # 헤더가 없으면 생성할 수 없음 (설계평가가 선행되어야 함)
                flash('설계평가 정보를 찾을 수 없습니다.', 'error')
                return redirect(url_for('link7.user_operation_evaluation'))
    except Exception as e:
        flash(f"운영평가 데이터 동기화 중 오류 발생: {str(e)}", 'error')

    # 기존 운영평가 내역 불러오기 (Header-Line 구조)
    try:
        evaluations = get_operation_evaluations(rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)

        # 평가가 완료된 통제(conclusion 값이 있는 경우) 또는 샘플이 업로드된 통제를 control_code를 키로 하는 딕셔너리로 변환
        # 중복이 있는 경우 가장 최신(last_updated 또는 evaluation_date 기준) 레코드만 사용
        evaluated_controls = {}
        for eval in evaluations:
            # line_id가 있거나, conclusion이 있거나, 샘플이 있으면 포함
            if eval.get('line_id') or eval.get('conclusion') or (eval.get('sample_lines') and len(eval.get('sample_lines', [])) > 0):
                control_code = eval['control_code']

                # 기존에 없거나, 더 최신 데이터인 경우만 업데이트
                if control_code not in evaluated_controls:
                    evaluated_controls[control_code] = eval
                else:
                    # last_updated 또는 evaluation_date로 최신 판단
                    existing_date = evaluated_controls[control_code].get('last_updated') or evaluated_controls[control_code].get('evaluation_date')
                    new_date = eval.get('last_updated') or eval.get('evaluation_date')
                    if new_date and existing_date and new_date > existing_date:
                        evaluated_controls[control_code] = eval

    except Exception as e:
        evaluated_controls = {}

    log_user_activity(user_info, 'PAGE_ACCESS', 'RCM 운영평가', '/operation-evaluation/rcm',
                      request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link7_detail.jsp',
                         rcm_id=rcm_id,
                         design_evaluation_session=design_evaluation_session,
                         evaluation_session=design_evaluation_session,  # 템플릿 호환성
                         operation_evaluation_session=operation_evaluation_session,
                         operation_header=operation_header,  # 진행률 표시용
                         rcm_info=rcm_info,
                         rcm_details=rcm_details,
                         rcm_mappings=rcm_mappings,
                         evaluated_controls=evaluated_controls,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr,
                         control_configs=control_configs)

@bp_link7.route('/api/operation-evaluation/save', methods=['POST'])
@login_required
def save_operation_evaluation_api():
    """운영평가 결과 저장 API"""
    user_info = get_user_info()

    # JSON과 FormData 모두 처리
    if request.content_type and 'multipart/form-data' in request.content_type:
        # FormData로 전송된 경우
        data = request.form.to_dict()
        rcm_id = data.get('rcm_id')
        design_evaluation_session = data.get('design_evaluation_session')
        control_code = data.get('control_code')

        # evaluation_data는 JSON 문자열로 전송되므로 파싱
        import json
        evaluation_data_str = data.get('evaluation_data')
        if evaluation_data_str:
            try:
                evaluation_data = json.loads(evaluation_data_str)
            except json.JSONDecodeError:
                return jsonify({
                    'success': False,
                    'message': 'evaluation_data 파싱 오류'
                })
        else:
            # 개별 필드로 전송된 경우 (구버전 호환)
            evaluation_data = {
                'sample_size': int(data.get('sample_size', 0)),
                'exception_count': int(data.get('exception_count', 0)),
                'exception_details': data.get('exception_details', ''),
                'conclusion': data.get('conclusion'),
                'improvement_plan': data.get('improvement_plan', '')
            }
    else:
        # JSON으로 전송된 경우
        data = request.get_json()
        rcm_id = data.get('rcm_id')
        design_evaluation_session = data.get('design_evaluation_session')
        control_code = data.get('control_code')
        evaluation_data = data.get('evaluation_data')

    if not all([rcm_id, design_evaluation_session, control_code, evaluation_data]):
        return jsonify({
            'success': False,
            'message': '필수 데이터가 누락되었습니다.'
        })

    # 운영평가 세션명 생성
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    try:
        # 사용자가 해당 RCM에 접근 권한이 있는지 확인
        with get_db() as conn:
            access_check = conn.execute('''
                SELECT permission_type FROM sb_user_rcm
                WHERE user_id = %s AND rcm_id = %s AND is_active = 'Y'
            ''', (user_info['user_id'], rcm_id)).fetchone()

            if not access_check:
                return jsonify({
                    'success': False,
                    'message': '해당 RCM에 대한 접근 권한이 없습니다.'
                })

        # 해당 설계평가 세션이 완료되었는지 확인
        completed_sessions = get_completed_design_evaluation_sessions(rcm_id, user_info['user_id'])
        session_found = False
        for session in completed_sessions:
            if session['evaluation_session'] == design_evaluation_session:
                session_found = True
                break

        if not session_found:
            return jsonify({
                'success': False,
                'message': f'설계평가 세션 "{design_evaluation_session}"이 완료되지 않아 운영평가를 수행할 수 없습니다.'
            })

        # RCM에 설정된 권장 표본수 확인
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail and rcm_detail['recommended_sample_size'] is not None else 0

        # 표본 크기 유효성 검사 (no_occurrence가 아니고, 설계평가 대체가 아닌 경우에만)
        is_no_occurrence = evaluation_data.get('no_occurrence', False)
        use_design_evaluation = evaluation_data.get('use_design_evaluation', False)
        if not is_no_occurrence and not use_design_evaluation and recommended_size > 0:
            submitted_sample_size = evaluation_data.get('sample_size')
            if submitted_sample_size is not None:
                submitted_sample_size = int(submitted_sample_size)
                if submitted_sample_size < recommended_size:
                    return jsonify({
                        'success': False,
                        'message': f'표본 크기({submitted_sample_size})는 권장 표본수({recommended_size})보다 작을 수 없습니다.'
                    })

        # 운영평가 결과 저장 (Header-Line 구조)
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'], operation_evaluation_session, design_evaluation_session, evaluation_data)

        # 활동 로그 기록
        log_user_activity(user_info, 'OPERATION_EVALUATION', f'운영평가 저장 - {control_code}',
                         f'/api/operation-evaluation/save',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({
            'success': True,
            'message': '운영평가 결과가 저장되었습니다.'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'저장 중 오류가 발생했습니다: {str(e)}'
        })

@bp_link7.route('/api/operation-evaluation/load/<int:rcm_id>/<design_evaluation_session>')
@login_required
def load_operation_evaluation(rcm_id, design_evaluation_session):
    """운영평가 데이터 로드 API (설계평가 세션별)"""
    user_info = get_user_info()

    try:
        # 권한 체크
        user_rcms = get_user_rcms(user_info['user_id'])
        rcm_ids = [rcm['rcm_id'] for rcm in user_rcms]

        if rcm_id not in rcm_ids:
            return jsonify({'success': False, 'message': '접근 권한이 없습니다.'}), 403

        # 운영평가 세션명 생성
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        evaluations = get_operation_evaluations(rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)

        evaluation_dict = {}
        for eval_data in evaluations:
            control_code = eval_data['control_code']
            evaluation_dict[control_code] = {
                'sample_size': eval_data['sample_size'],
                'exception_count': eval_data['exception_count'],
                'mitigating_factors': eval_data.get('mitigating_factors'),
                'exception_details': eval_data['exception_details'],
                'conclusion': eval_data['conclusion'],
                'improvement_plan': eval_data['improvement_plan']
            }

        return jsonify({'success': True, 'evaluations': evaluation_dict})

    except Exception as e:
        return jsonify({'success': False, 'message': '데이터 로드 중 오류가 발생했습니다.'}), 500

@bp_link7.route('/api/operation-evaluation/samples/<int:line_id>')
@login_required
def load_operation_evaluation_samples(line_id):
    """평가 버튼 클릭 시 특정 line_id의 샘플 데이터 조회 API
    통합 테이블(sb_evaluation_line) 사용
    """
    user_info = get_user_info()

    try:
        # line_id에 해당하는 통제 정보 조회
        with get_db() as conn:
            line_info = conn.execute('''
                SELECT h.rcm_id, r.control_category, l.control_code
                FROM sb_evaluation_line l
                JOIN sb_evaluation_header h ON l.header_id = h.header_id
                JOIN sb_rcm r ON h.rcm_id = r.rcm_id
                WHERE l.line_id = ?
            ''', (line_id,)).fetchone()

            if not line_info:
                return jsonify({'success': False, 'message': '평가 데이터를 찾을 수 없습니다.'}), 404

        # 샘플 데이터 조회
        sample_lines = get_operation_evaluation_samples(line_id)

        # 설계평가 샘플 조회
        design_sample = get_design_evaluation_sample(line_id)

        # attributes 정의 조회 (RCM detail에서)
        attributes = []
        population_attribute_count = 0

        with get_db() as conn:
            # line_info에서 이미 control_code와 rcm_id를 가져옴
            # RCM detail에서 attribute 정의 조회
            rcm_detail = conn.execute('''
                SELECT population_attribute_count,
                       attribute0, attribute1, attribute2, attribute3, attribute4,
                       attribute5, attribute6, attribute7, attribute8, attribute9
                FROM sb_rcm_detail
                WHERE rcm_id = ? AND control_code = ?
            ''', (line_info['rcm_id'], line_info['control_code'])).fetchone()

            if rcm_detail:
                population_attribute_count = rcm_detail['population_attribute_count'] or 0

                # attribute 정의 생성 (RCM detail에 정의된 모든 attributes 반환)
                for i in range(10):
                    # RCM detail에서 attribute 이름 가져오기
                    attr_name = rcm_detail[f'attribute{i}'] if rcm_detail[f'attribute{i}'] else None

                    # 이름이 정의되지 않은 attribute는 skip
                    if not attr_name:
                        continue

                    # population_attribute_count를 기준으로 모집단/증빙 구분
                    if i < population_attribute_count:
                        attr_type = 'population'
                    else:
                        attr_type = 'evidence'

                    attributes.append({
                        'attribute': f'attribute{i}',
                        'name': attr_name,
                        'type': attr_type
                    })

        return jsonify({
            'success': True,
            'samples': sample_lines,
            'design_sample': design_sample,
            'attributes': attributes,
            'population_attribute_count': population_attribute_count
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'샘플 데이터 로드 중 오류가 발생했습니다: {str(e)}'}), 500

# REMOVED: Duplicate reset API that was deleting entire header
# The correct reset API is at line 589: operation_evaluation_reset()
# That one only deletes specific control's line and files, not the header
# ===================================================================
# APD01 표준통제 테스트 API
# ===================================================================

@bp_link7.route('/api/operation-evaluation/apd01/upload-population', methods=['POST'])
@login_required
def apd01_upload_population():
    """APD01 모집단 업로드 및 파싱"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    import json
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"
        from auth import get_db

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (recommended_sample_size 전달)
        result = file_manager.parse_apd01_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd01_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'파일 처리 오류: {str(e)}'})


# The following routes are deprecated and replaced by the generic implementation
# in operation_evaluation_generic.py. They are kept here for reference but can be removed.
# - apd01_save_test_results
# - user_operation_evaluation_apd01


@bp_link7.route('/api/operation-evaluation/reset', methods=['POST'])
@login_required
def operation_evaluation_reset():
    """운영평가 파일 삭제 및 리셋 (모든 통제 공통)"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    control_code = data.get('control_code')
    design_evaluation_session = data.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 데이터가 누락되었습니다.'})

    try:
        import os
        operation_evaluation_session = f"OP_{design_evaluation_session}"


        # DB에서 operation_header_id 조회 (있으면)
        from auth import get_db
        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()


            if header:
                operation_header_id = header['header_id']


                # DB 라인 데이터 삭제 (해당 통제만)
                deleted_rows = conn.execute('''
                    DELETE FROM sb_evaluation_line
                    WHERE header_id = %s AND control_code = %s
                ''', (operation_header_id, control_code))
                conn.commit()


                # 파일 삭제 (절대 경로 사용, control_code 폴더 제거)
                app_root = os.path.dirname(os.path.abspath(__file__))
                excel_path = os.path.join(app_root, 'static', 'uploads', 'operation_evaluations',
                                        str(rcm_id), str(operation_header_id), f'{control_code}_evaluation.xlsx')


                if os.path.exists(excel_path):
                    os.remove(excel_path)

        # 세션 정리 (통제별로 다른 키 사용)
        if control_code == 'APD01':
            session_key = f'apd01_test_{rcm_id}_{control_code}'
        elif control_code == 'APD07':
            session_key = f'apd07_test_{rcm_id}_{control_code}'
        else:
            session_key = f'{control_code.lower()}_test_{rcm_id}_{control_code}'

        session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'{control_code} 리셋',
                         '/api/operation-evaluation/reset',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': '초기화되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'리셋 오류: {str(e)}'})


# The following routes are deprecated and replaced by the generic implementation
# in operation_evaluation_generic.py. They are kept here for reference but can be removed.
# - apd01_save_test_results
# - user_operation_evaluation_apd01


@bp_link7.route('/api/design-evaluation/get', methods=['GET'])
@login_required
def get_design_evaluation_data():
    """설계평가 데이터 조회 (운영평가에서 보기용)"""
    try:
        user_info = get_user_info()
        rcm_id_param = request.args.get('rcm_id')
        evaluation_session = request.args.get('evaluation_session')


        if not rcm_id_param or not evaluation_session:
            return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'})

        rcm_id = int(rcm_id_param)

        # 설계평가 데이터 조회
        from auth import get_design_evaluations
        evaluations = get_design_evaluations(rcm_id, user_info['user_id'], evaluation_session)

        # RCM 상세 정보와 조인하여 통제 정보 추가
        rcm_details = get_rcm_details(rcm_id)
        rcm_dict = {detail['control_code']: detail for detail in rcm_details}

        # 매핑 정보 조회
        from auth import get_rcm_detail_mappings
        rcm_mappings_list = get_rcm_detail_mappings(rcm_id)
        rcm_mappings = {m['control_code']: m for m in rcm_mappings_list}

        # 평가 데이터에 통제 정보 추가
        result = []
        for eval_data in evaluations:
            control_code = eval_data['control_code']
            if control_code in rcm_dict:
                detail = rcm_dict[control_code]
                mapping = rcm_mappings.get(control_code, {})
                result.append({
                    'control_code': control_code,
                    'control_name': detail['control_name'],
                    'control_description': detail['control_description'],
                    'control_frequency': detail['control_frequency'],
                    'control_frequency_name': detail.get('control_frequency_name'),
                    'control_nature': detail['control_nature'],
                    'control_nature_name': detail.get('control_nature_name'),
                    'key_control': detail.get('key_control'),
                    'std_control_code': mapping.get('std_control_code'),
                    'std_control_name': mapping.get('std_control_name'),
                    'design_adequacy': eval_data.get('overall_effectiveness'),
                    'improvement_plan': eval_data.get('recommended_actions'),
                    'evaluated_date': eval_data.get('evaluation_date')
                })

        return jsonify({'success': True, 'evaluations': result})

    except Exception as e:
        return jsonify({'success': False, 'message': f'조회 오류: {str(e)}'})

@bp_link7.route('/operation-evaluation/apd07')
@login_required
def user_operation_evaluation_apd07():
    """APD07 운영평가 페이지"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    control_code = request.args.get('control_code')
    control_name = request.args.get('control_name')
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 기존 운영평가 데이터 조회
    existing_data = None
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    try:
        with get_db() as conn:
            # 운영평가 헤더 조회 (있으면)
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 저장된 파일에서 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code
                )

                if loaded_data and loaded_data['samples_data']:
                    existing_data = {
                        'samples': loaded_data['samples_data'].get('samples', []),
                        'population_count': loaded_data['samples_data'].get('population_count', 0),
                        'sample_size': loaded_data['samples_data'].get('sample_size', 0),
                        'test_results': loaded_data['samples_data'].get('test_results', {}),
                        'operation_header_id': operation_header_id
                    }

                    # 세션에 operation_header_id 저장 (저장 시 필요)
                    session_key = f'apd07_test_{rcm_id}_{control_code}'
                    session[session_key] = {
                        'operation_header_id': operation_header_id,
                        'population_count': existing_data['population_count'],
                        'sample_size': existing_data['sample_size']
                    }

    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        # 오류가 발생해도 페이지는 정상적으로 표시
        pass

    log_user_activity(user_info, 'PAGE_ACCESS', 'APD07 운영평가', '/operation-evaluation/apd07',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('user_operation_evaluation_apd07.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         existing_data=existing_data,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)

# ===================================================================
# APD07 표준통제 테스트 API
# ===================================================================

@bp_link7.route('/api/operation-evaluation/apd07/upload-population', methods=['POST'])
@login_required
def apd07_upload_population():
    """APD07 모집단 업로드 및 파싱 (데이터 직접변경 승인)"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    import json
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"
        from auth import get_db

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (APD07용, recommended_sample_size 전달)
        result = file_manager.parse_apd07_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd07_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'파일 처리 오류: {str(e)}'})


@bp_link7.route('/api/operation-evaluation/apd07/save-test-results', methods=['POST'])
@login_required
def apd07_save_test_results():
    """APD07 테스트 결과 저장 (데이터 직접변경 승인)"""
    user_info = get_user_info()
    data = request.get_json()

    rcm_id = data.get('rcm_id')
    control_code = data.get('control_code')
    design_evaluation_session = data.get('design_evaluation_session')
    test_results = data.get('test_results')  # 표본별 테스트 결과

    if not all([rcm_id, control_code, design_evaluation_session, test_results]):
        return jsonify({'success': False, 'message': '필수 데이터가 누락되었습니다.'})

    try:
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 세션에서 파일 경로 정보 가져오기
        session_key = f'apd07_test_{rcm_id}_{control_code}'
        test_data = session.get(session_key)

        if not test_data:
            return jsonify({'success': False, 'message': '테스트 데이터를 찾을 수 없습니다. 모집단을 다시 업로드해주세요.'})

        # 세션에서 operation_header_id 가져오기
        operation_header_id = test_data.get('operation_header_id')
        if not operation_header_id:
            return jsonify({'success': False, 'message': '운영평가 헤더 ID를 찾을 수 없습니다.'})

        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # 템플릿 기반 엑셀 파일 업데이트 (테스트 결과 추가)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=samples_data.get('field_mapping', {}),
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': 'APD07'
            }
        )

        # 평가 데이터 구성 (메타데이터만 DB에 저장)
        evaluation_data = {
            'test_type': 'APD07',
            'population_count': samples_data['population_count'],
            'sample_size': samples_data['sample_size'],
            'population_path': None,  # 템플릿 방식에서는 엑셀에 통합
            'samples_path': file_paths.get('samples_path'),
            'test_results_path': file_paths.get('excel_path'),  # 엑셀 파일 경로
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception'
        }

        # 운영평가 저장
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        # 세션 정리 제거 - 다시 저장할 수 있도록 세션 유지
        # session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'APD07 테스트 저장 - {control_code}',
                         '/api/operation-evaluation/apd07/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': 'APD07 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})

# ===================================================================
# APD09 운영평가 라우트
# ===================================================================

@bp_link7.route('/operation-evaluation/apd09')
@login_required
def user_operation_evaluation_apd09():
    """APD09 운영평가 페이지"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    control_code = request.args.get('control_code')
    control_name = request.args.get('control_name')
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 기존 운영평가 데이터 조회
    existing_data = None
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    try:
        with get_db() as conn:
            # 운영평가 헤더 조회 (있으면)
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 저장된 파일에서 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code
                )

                if loaded_data and loaded_data['samples_data']:
                    existing_data = {
                        'samples': loaded_data['samples_data'].get('samples', []),
                        'population_count': loaded_data['samples_data'].get('population_count', 0),
                        'sample_size': loaded_data['samples_data'].get('sample_size', 0),
                        'test_results': loaded_data['samples_data'].get('test_results', {}),
                        'operation_header_id': operation_header_id
                    }

                    # 세션에 operation_header_id 저장 (저장 시 필요)
                    session_key = f'apd09_test_{rcm_id}_{control_code}'
                    session[session_key] = {
                        'operation_header_id': operation_header_id,
                        'population_count': existing_data['population_count'],
                        'sample_size': existing_data['sample_size']
                    }

    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        # 오류가 발생해도 페이지는 정상적으로 표시
        pass

    log_user_activity(user_info, 'PAGE_ACCESS', 'APD09 운영평가', '/operation-evaluation/apd09',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('user_operation_evaluation_apd09.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         existing_data=existing_data,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)


@bp_link7.route('/api/operation-evaluation/apd09/upload-population', methods=['POST'])
@login_required
def upload_apd09_population():
    """APD09 모집단 업로드 및 파싱 (OS 접근권한 부여 승인)"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (APD09용, recommended_sample_size 전달)
        result = file_manager.parse_apd09_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd09_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id,
            'field_mapping': field_mapping,
            'population_count': result['count'],
            'sample_size': result['sample_size']
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@bp_link7.route('/api/operation-evaluation/apd09/save-test-results', methods=['POST'])
@login_required
def save_apd09_test_results():
    """APD09 테스트 결과 저장"""
    try:
        user_info = get_user_info()
        data = request.json
        rcm_id = data.get('rcm_id')
        control_code = data.get('control_code')
        design_evaluation_session = data.get('design_evaluation_session')
        test_results = data.get('test_results', [])

        # 세션에서 operation_header_id 가져오기
        session_key = f'apd09_test_{rcm_id}_{control_code}'
        session_data = session.get(session_key)

        if not session_data:
            return jsonify({'success': False, 'message': '세션 정보가 없습니다. 모집단을 다시 업로드해주세요.'})

        operation_header_id = session_data['operation_header_id']
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # 엑셀 파일에 테스트 결과 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=session_data.get('field_mapping', {}),  # 세션에서 가져오기
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': 'APD09'
            }
        )

        # 운영평가 데이터 저장
        evaluation_data = {
            'sample_size': session_data['sample_size'],
            'population_path': file_paths.get('population_file'),
            'samples_path': file_paths.get('excel_path'),
            'test_results_path': file_paths.get('excel_path'),  # 엑셀 파일 경로
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception'
        }

        # 운영평가 저장
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        # 세션 정리 제거 - 다시 저장할 수 있도록 세션 유지
        # session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'APD09 테스트 저장 - {control_code}',
                         '/api/operation-evaluation/apd09/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': 'APD09 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})


# ===================================================================
# 운영평가 리셋 API
# ===================================================================

# ===================================================================
# APD12 운영평가 라우트
# ===================================================================

@bp_link7.route('/operation-evaluation/apd12')
@login_required
def user_operation_evaluation_apd12():
    """APD12 운영평가 페이지"""
    user_info = get_user_info()

    rcm_id = request.args.get('rcm_id')
    control_code = request.args.get('control_code')
    control_name = request.args.get('control_name')
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation'))

    # 기존 운영평가 데이터 조회
    existing_data = None
    operation_evaluation_session = f"OP_{design_evaluation_session}"

    try:
        with get_db() as conn:
            # 운영평가 헤더 조회 (있으면)
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 저장된 파일에서 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code
                )

                if loaded_data and loaded_data['samples_data']:
                    existing_data = {
                        'samples': loaded_data['samples_data'].get('samples', []),
                        'population_count': loaded_data['samples_data'].get('population_count', 0),
                        'sample_size': loaded_data['samples_data'].get('sample_size', 0),
                        'test_results': loaded_data['samples_data'].get('test_results', {}),
                        'operation_header_id': operation_header_id
                    }

                    # 세션에 operation_header_id 저장 (저장 시 필요)
                    session_key = f'apd12_test_{rcm_id}_{control_code}'
                    session[session_key] = {
                        'operation_header_id': operation_header_id,
                        'population_count': existing_data['population_count'],
                        'sample_size': existing_data['sample_size']
                    }

    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")
        import traceback
        traceback.print_exc()
        # 오류가 발생해도 페이지는 정상적으로 표시
        pass

    log_user_activity(user_info, 'PAGE_ACCESS', 'APD12 운영평가', '/operation-evaluation/apd12',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('user_operation_evaluation_apd12.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         existing_data=existing_data,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         remote_addr=request.remote_addr)


@bp_link7.route('/api/operation-evaluation/apd12/upload-population', methods=['POST'])
@login_required
def upload_apd12_population():
    """APD12 모집단 업로드 및 파싱 (DB 접근권한 부여 승인)"""
    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기 (JSON)
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    control_code = request.form.get('control_code')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회 (RCM 페이지에서 이미 생성되어 있어야 함)
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        import tempfile
        import os
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (APD12용, recommended_sample_size 전달)
        result = file_manager.parse_apd12_population(temp_file.name, field_mapping, recommended_size)

        # 표본 선택
        samples = file_manager.select_random_samples(result['population'], result['sample_size'])

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장 (운영평가 헤더 ID 사용)
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=result['population'],
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None  # 아직 테스트 결과 없음
        )

        # 세션에 파일 경로만 저장 (나중에 저장할 때 사용)
        session_key = f'apd12_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id,
            'field_mapping': field_mapping,
            'population_count': result['count'],
            'sample_size': result['sample_size']
        }

        return jsonify({
            'success': True,
            'population_count': result['count'],
            'sample_size': result['sample_size'],
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


@bp_link7.route('/api/operation-evaluation/apd12/save-test-results', methods=['POST'])
@login_required
def save_apd12_test_results():
    """APD12 테스트 결과 저장"""
    try:
        user_info = get_user_info()
        data = request.json
        rcm_id = data.get('rcm_id')
        control_code = data.get('control_code')
        design_evaluation_session = data.get('design_evaluation_session')
        test_results = data.get('test_results', [])

        # 세션에서 operation_header_id 가져오기
        session_key = f'apd12_test_{rcm_id}_{control_code}'
        session_data = session.get(session_key)

        if not session_data:
            return jsonify({'success': False, 'message': '세션 정보가 없습니다. 모집단을 다시 업로드해주세요.'})

        operation_header_id = session_data['operation_header_id']
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 엑셀 파일에 테스트 결과 ��장
        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # 엑셀 파일에 테스트 결과 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=session_data.get('field_mapping', {}),  # 세션에서 가져오기
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': 'APD12'
            }
        )

        # 운영평가 데이터 저장
        evaluation_data = {
            'test_type': 'APD12',
            'population_count': samples_data['population_count'],
            'sample_size': samples_data['sample_size'],
            'population_path': None,  # 템플릿 방식에서는 엑셀에 통합
            'samples_path': file_paths.get('samples_path'),
            'test_results_path': file_paths.get('excel_path'),  # 엑셀 파일 경로
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception'
        }

        # 운영평가 저장
        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        # 세션 정리 제거 - 다시 저장할 수 있도록 세션 유지
        # session.pop(session_key, None)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'APD12 테스트 저장 - {control_code}',
                         '/api/operation-evaluation/apd12/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': 'APD12 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})

# ============================================================================
# ELC 운영평가 (수동통제만)
# ============================================================================

@bp_link7.route('/elc/operation-evaluation')
@login_required
def elc_operation_evaluation():
    """ELC 운영평가 페이지"""
    user_info = get_user_info()

    # ELC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    elc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'ELC']

    # 각 RCM에 대해 모든 설계평가 세션 조회 (통합 테이블에서)
    from evaluation_utils import get_evaluation_status
    for rcm in elc_rcms:
        with get_db() as conn:
            # 설계평가 세션 조회 (archived가 아닌 것만)
            sessions = conn.execute('''
                SELECT header_id, evaluation_name, created_at, last_updated
                FROM sb_evaluation_header
                WHERE rcm_id = ? AND user_id = ? AND (archived IS NULL OR archived = 0)
                ORDER BY last_updated DESC
            ''', (rcm['rcm_id'], user_info['user_id'])).fetchall()

        all_sessions = []
        for session_row in sessions:
            session = dict(session_row)
            header_id = session['header_id']

            # 실시간으로 status 계산
            with get_db() as conn:
                status_info = get_evaluation_status(conn, header_id)

            session['evaluation_session'] = session['evaluation_name']
            session['status'] = status_info['status']
            session['completed_date'] = session['last_updated'] if status_info['status'] >= 1 else None
            all_sessions.append(session)

        completed_sessions = [s for s in all_sessions if s['completed_date'] is not None]
        in_progress_sessions = [s for s in all_sessions if s['completed_date'] is None]

        # 완료된 세션에 대해서만 운영평가 진행상황 조회
        for session in completed_sessions:
            operation_evaluation_session = f"OP_{session['evaluation_session']}"

            from auth import count_completed_operation_evaluations
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm['rcm_id'], user_info['user_id'], operation_evaluation_session, session['evaluation_session'])).fetchone()

            if header:
                completed_count = count_completed_operation_evaluations(header['header_id'])
                session['operation_completed_count'] = completed_count
            else:
                session['operation_completed_count'] = 0

            # 운영평가 가능한 통제 개수 추가
            eligible_controls = get_key_rcm_details(rcm['rcm_id'], user_info['user_id'], session['evaluation_session'], control_category='ELC')
            session['eligible_control_count'] = len(eligible_controls)

        rcm['all_design_sessions'] = all_sessions
        rcm['completed_design_sessions'] = completed_sessions
        rcm['in_progress_design_sessions'] = in_progress_sessions
        rcm['design_evaluation_completed'] = len(completed_sessions) > 0

        # 핵심통제 개수 조회
        key_controls = get_key_rcm_details(rcm['rcm_id'], control_category='ELC')
        rcm['key_control_count'] = len(key_controls)
        rcm['has_key_controls'] = len(key_controls) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'ELC 운영평가', '/elc/operation-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    # 세션에서 현재 선택된 RCM과 평가 세션 정보 가져오기
    from flask import session as flask_session
    current_rcm_id = flask_session.get('current_rcm_id')
    current_evaluation_session = flask_session.get('current_evaluation_session')

    return render_template('link7_evaluation.jsp',
                         evaluation_type='ELC',
                         user_rcms=elc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info,
                         current_rcm_id=current_rcm_id,
                         current_evaluation_session=current_evaluation_session)

# ============================================================================
# TLC 운영평가 (자동통제 포함)
# ============================================================================

@bp_link7.route('/tlc/operation-evaluation')
@login_required
def tlc_operation_evaluation():
    """TLC 운영평가 페이지"""
    user_info = get_user_info()

    # TLC RCM 목록만 필터링
    all_rcms = get_user_rcms(user_info['user_id'])
    tlc_rcms = [rcm for rcm in all_rcms if rcm.get('control_category') == 'TLC']

    # 각 RCM에 대해 모든 설계평가 세션 조회 (진행중 + 완료)
    from auth import get_all_design_evaluation_sessions
    for rcm in tlc_rcms:
        all_sessions = get_all_design_evaluation_sessions(rcm['rcm_id'], user_info['user_id'])
        completed_sessions = [s for s in all_sessions if s['completed_date'] is not None]
        in_progress_sessions = [s for s in all_sessions if s['completed_date'] is None]

        # 완료된 세션에 대해서만 운영평가 진행상황 조회
        for session in completed_sessions:
            operation_evaluation_session = f"OP_{session['evaluation_session']}"

            from auth import count_completed_operation_evaluations
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm['rcm_id'], user_info['user_id'], operation_evaluation_session, session['evaluation_session'])).fetchone()

            if header:
                completed_count = count_completed_operation_evaluations(header['header_id'])
                session['operation_completed_count'] = completed_count
            else:
                session['operation_completed_count'] = 0

            # 운영평가 가능한 통제 개수 추가
            eligible_controls = get_key_rcm_details(rcm['rcm_id'], user_info['user_id'], session['evaluation_session'], control_category='TLC')
            session['eligible_control_count'] = len(eligible_controls)

        rcm['all_design_sessions'] = all_sessions
        rcm['completed_design_sessions'] = completed_sessions
        rcm['in_progress_design_sessions'] = in_progress_sessions
        rcm['design_evaluation_completed'] = len(completed_sessions) > 0

        # 핵심통제 개수 조회
        key_controls = get_key_rcm_details(rcm['rcm_id'], control_category='TLC')
        rcm['key_control_count'] = len(key_controls)
        rcm['has_key_controls'] = len(key_controls) > 0

    log_user_activity(user_info, 'PAGE_ACCESS', 'TLC 운영평가', '/tlc/operation-evaluation',
                     request.remote_addr, request.headers.get('User-Agent'))

    return render_template('link7_evaluation.jsp',
                         evaluation_type='TLC',
                         user_rcms=tlc_rcms,
                         is_logged_in=is_logged_in(),
                         user_info=user_info)


# ===================================================================
# 일반 통제 모집단 업로드 API (표본수 0인 경우)
# ===================================================================

@bp_link7.route('/api/operation-evaluation/upload-population', methods=['POST'])
@login_required
def upload_general_population():
    """일반 통제 모집단 업로드 및 표본 추출 (표본수 0인 경우)"""
    import os
    from openpyxl import load_workbook
    from werkzeug.utils import secure_filename

    user_info = get_user_info()

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'error': '파일을 선택해주세요.'})

    # 파라미터 받기
    control_code = request.form.get('control_code')
    rcm_id = request.form.get('rcm_id')
    design_evaluation_session = request.form.get('design_evaluation_session')
    field_mapping_str = request.form.get('field_mapping')

    if not all([control_code, rcm_id, design_evaluation_session, field_mapping_str]):
        return jsonify({'success': False, 'error': '필수 파라미터가 누락되었습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'error': '필드 매핑 파싱 실패'})

    try:
        # 파일 저장
        upload_folder = os.path.join('uploads', 'populations')
        os.makedirs(upload_folder, exist_ok=True)

        # 원본 파일명에서 확장자 추출
        original_filename = file.filename
        file_ext = os.path.splitext(original_filename)[1]  # .xlsx

        # secure_filename으로 안전한 이름 생성
        filename = secure_filename(file.filename)

        # secure_filename이 파일명을 완전히 제거한 경우 (한글 등)
        if not filename or filename == file_ext.replace('.', ''):
            filename = f"population{file_ext}"

        # 확장자가 없으면 원본에서 가져온 확장자 추가
        if not os.path.splitext(filename)[1]:
            filename = filename + file_ext

        # 파일 확장자 확인
        if not filename.lower().endswith(('.xlsx', '.xlsm')):
            return jsonify({'success': False, 'error': '.xlsx 또는 .xlsm 형식의 파일만 지원됩니다. (.xls 파일은 Excel에서 .xlsx로 변환 후 업로드해주세요)'})

        filepath = os.path.join(upload_folder, f"{user_info['user_id']}_{control_code}_{filename}")
        file.save(filepath)

        # 엑셀 파일 읽기 (openpyxl 사용)
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
        except Exception as excel_error:
            return jsonify({'success': False, 'error': f'엑셀 파일을 읽을 수 없습니다. 파일이 손상되었거나 암호로 보호되어 있을 수 있습니다. ({str(excel_error)})'})


        # 헤더 읽기 (첫 번째 행)
        headers = [cell.value for cell in ws[1]]

        # 필드 매핑 적용
        number_col_idx = field_mapping['number']
        description_col_idx = field_mapping['description']

        # 모집단 데이터 파싱
        population = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # 헤더 제외
            if row[number_col_idx] is not None:  # 빈 행 건너뛰기
                population.append({
                    'number': str(row[number_col_idx]),
                    'description': str(row[description_col_idx]) if row[description_col_idx] else ''
                })

        wb.close()

        population_count = len(population)

        # 표본 크기 자동 계산
        sample_size = file_manager.calculate_sample_size(population_count)

        # 무작위 표본 추출
        import random
        sample_indices = random.sample(range(population_count), min(sample_size, population_count))
        samples = [population[i] for i in sorted(sample_indices)]

        # 운영평가 세션 확인/생성
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            # Header 확인/생성
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                conn.execute('''
                    INSERT INTO sb_evaluation_header (rcm_id, user_id, evaluation_session, design_evaluation_session)
                    VALUES (%s, %s, %s, %s)
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session))
                conn.commit()

                header = conn.execute('''
                    SELECT header_id FROM sb_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            header_id = header['header_id']

            # 기존 Line 확인
            existing_line = conn.execute('''
                SELECT line_id FROM sb_evaluation_line
                WHERE header_id = %s AND control_code = %s
            ''', (header_id, control_code)).fetchone()

            if existing_line:
                line_id = existing_line['line_id']

                # 기존 샘플 삭제
                conn.execute('DELETE FROM sb_evaluation_sample WHERE line_id = %s', (line_id,))

                # Line 업데이트 (sample_size만)
                conn.execute('''
                    UPDATE sb_evaluation_line
                    SET sample_size = %s
                    WHERE line_id = %s
                ''', (sample_size, line_id))
            else:
                # 새 Line 생성
                conn.execute('''
                    INSERT INTO sb_evaluation_line
                    (header_id, control_code, sample_size)
                    VALUES (%s, %s, %s)
                ''', (header_id, control_code, sample_size))

                # SQLite용 last_insert_rowid() 사용
                line_id = conn.execute('SELECT last_insert_rowid() as id').fetchone()['id']

            # 샘플 데이터 저장 (attribute0에 번호, attribute1에 설명 저장)
            for idx, sample in enumerate(samples, 1):
                conn.execute('''
                    INSERT INTO sb_evaluation_sample
                    (line_id, sample_number, evaluation_type, attribute0, attribute1)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (line_id, idx, 'operation', sample['number'], sample['description']))

            conn.commit()

            # 저장된 샘플 데이터 조회하여 sample_lines 형식으로 반환
            sample_lines = []
            saved_samples = conn.execute('''
                SELECT sample_number, evidence, has_exception, mitigation,
                       attribute0, attribute1, attribute2, attribute3, attribute4,
                       attribute5, attribute6, attribute7, attribute8, attribute9
                FROM sb_evaluation_sample
                WHERE line_id = %s
                ORDER BY sample_number
            ''', (line_id,)).fetchall()

            for sample in saved_samples:
                # attribute 데이터 수집
                attributes = {}
                for i in range(10):
                    attr_val = sample[f'attribute{i}']
                    if attr_val is not None:
                        attributes[f'attribute{i}'] = attr_val

                sample_lines.append({
                    'sample_number': sample['sample_number'],
                    'evidence': sample['evidence'] or '',
                    'result': 'exception' if sample['has_exception'] else 'no_exception',
                    'mitigation': sample['mitigation'] or '',
                    'attributes': attributes if attributes else None
                })

            # RCM detail에서 attribute 정의 조회
            rcm_detail = conn.execute('''
                SELECT population_attribute_count,
                       attribute0, attribute1, attribute2, attribute3, attribute4,
                       attribute5, attribute6, attribute7, attribute8, attribute9
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

            population_attr_count = rcm_detail['population_attribute_count'] if rcm_detail and rcm_detail['population_attribute_count'] else 2

            # 샘플 데이터를 확인하여 실제 사용된 attribute 찾기
            used_attributes = set()
            for sample in saved_samples:
                for i in range(10):
                    if sample[f'attribute{i}'] is not None:
                        used_attributes.add(i)

            # attribute 정의 생성 (RCM detail에 정의된 모든 attributes 반환)
            attributes = []
            for i in range(10):
                # RCM detail에서 attribute 이름 가져오기
                attr_name = rcm_detail[f'attribute{i}'] if rcm_detail else None

                # 이름이 정의되지 않은 attribute는 skip
                if not attr_name:
                    continue

                # population_attr_count를 기준으로 모집단/증빙 구분
                if i < population_attr_count:
                    attr_type = 'population'
                else:
                    attr_type = 'evidence'

                attributes.append({
                    'attribute': f'attribute{i}',
                    'name': attr_name,
                    'type': attr_type
                })

        return jsonify({
            'success': True,
            'population_count': population_count,
            'sample_size': sample_size,
            'line_id': line_id,
            'sample_lines': sample_lines,
            'attributes': attributes,
            'population_attribute_count': population_attr_count
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@bp_link7.route('/api/operation-evaluation/save-attributes', methods=['POST'])
@login_required
def save_attributes():
    """Attribute 필드 설정 저장"""
    user_info = get_user_info()
    data = request.get_json()

    line_id = data.get('line_id')
    attributes = data.get('attributes', [])

    if not line_id or not attributes:
        return jsonify({'success': False, 'error': '필수 데이터가 누락되었습니다.'})

    try:
        # 성공 응답
        return jsonify({
            'success': True,
            'message': 'Attribute 설정이 저장되었습니다.'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@bp_link7.route('/api/operation-evaluation/reset-population', methods=['POST'])
@login_required
def reset_population_upload():
    """모집단 업로드 초기화 (파일 삭제 + DB 데이터 삭제)"""
    import os
    user_info = get_user_info()
    data = request.get_json()

    control_code = data.get('control_code')
    line_id = data.get('line_id')

    if not control_code:
        return jsonify({'success': False, 'message': '통제 코드가 필요합니다.'})

    try:
        # 1. 업로드된 모집단 파일 삭제
        upload_folder = os.path.join('uploads', 'populations')
        if os.path.exists(upload_folder):
            # 파일명 패턴: {user_id}_{control_code}_*.xlsx
            file_pattern = f"{user_info['user_id']}_{control_code}_"
            for filename in os.listdir(upload_folder):
                if filename.startswith(file_pattern):
                    filepath = os.path.join(upload_folder, filename)
                    try:
                        os.remove(filepath)
                    except Exception:
                        pass

        # 2. DB에서 표본 데이터 삭제
        if line_id:
            with get_db() as conn:
                # 표본 데이터 삭제
                conn.execute('DELETE FROM sb_evaluation_sample WHERE line_id = %s', (line_id,))

                # 라인 데이터 삭제
                conn.execute('DELETE FROM sb_evaluation_line WHERE line_id = %s', (line_id,))

                conn.commit()

        log_user_activity(user_info, 'DATA_DELETE', '모집단 업로드 초기화',
                         f'/api/operation-evaluation/reset-population (control: {control_code})',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({
            'success': True,
            'message': '모집단 업로드가 초기화되었습니다.'
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


# ============================================================================
# 운영평가 다운로드 기능
# ============================================================================

@bp_link7.route('/operation-evaluation/download')
@login_required
def download_operation_evaluation():
    """운영평가 결과를 Template_Manual.xlsx 양식으로 다운로드 (통제별)"""
    from flask import make_response
    import urllib.parse

    user_info = get_user_info()

    # URL 파라미터 받기
    rcm_id = request.args.get('rcm_id')
    evaluation_session = request.args.get('evaluation_session')
    design_evaluation_session = request.args.get('design_evaluation_session')
    control_code = request.args.get('control_code')

    # 필수 파라미터 검증
    if not all([rcm_id, evaluation_session, design_evaluation_session, control_code]):
        flash('RCM ID, 운영평가 세션, 설계평가 세션, 통제번호가 필요합니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation_rcm'))

    # 자동통제 여부 확인
    is_auto_control = (evaluation_session == 'AUTO')

    try:
        # 템플릿 파일 경로
        template_path = os.path.join(os.path.dirname(__file__), 'paper_templates', 'Template_Manual.xlsx')

        if not os.path.exists(template_path):
            flash('템플릿 파일을 찾을 수 없습니다.', 'error')
            return redirect(url_for('link7.user_operation_evaluation_rcm'))

        # 템플릿 로드 (외부 링크 제거)
        wb = load_workbook(template_path, keep_links=False)

        # RCM 정보 조회
        with get_db() as conn:
            rcm_info = conn.execute("""
                SELECT rcm_name, description
                FROM sb_rcm
                WHERE rcm_id = %s
            """, (rcm_id,)).fetchone()

            if not rcm_info:
                flash('RCM 정보를 찾을 수 없습니다.', 'error')
                return redirect(url_for('link7.user_operation_evaluation_rcm'))

            # 운영평가 결과 조회 (해당 통제 1개만)
            # 자동통제인 경우에도 저장된 운영평가 데이터가 있으면 사용
            if is_auto_control:
                # 먼저 저장된 운영평가 데이터 찾기 (design_evaluation_session으로 연결)
                evaluation = conn.execute("""
                    SELECT
                        l.line_id,
                        l.control_code,
                        rd.control_name,
                        rd.control_description,
                        rd.control_frequency,
                        rd.control_type,
                        rd.control_nature,
                        l.sample_size,
                        l.exception_count,
                        l.exception_details,
                        l.conclusion,
                        l.improvement_plan,
                        l.review_comment,
                        l.evaluation_date,
                        d.attribute0, d.attribute1, d.attribute2, d.attribute3, d.attribute4,
                        d.attribute5, d.attribute6, d.attribute7, d.attribute8, d.attribute9,
                        d.population_attribute_count,
                        d.recommended_sample_size
                    FROM sb_evaluation_line l
                    JOIN sb_evaluation_header h ON l.header_id = h.header_id
                    JOIN sb_rcm_detail_v rd ON h.rcm_id = rd.rcm_id AND l.control_code = rd.control_code
                    JOIN sb_rcm_detail d ON d.rcm_id = h.rcm_id AND d.control_code = l.control_code
                    WHERE h.rcm_id = %s
                      AND h.evaluation_name = %s
                      AND h.status >= 2
                      AND l.control_code = %s
                      AND l.conclusion IS NOT NULL
                """, (rcm_id, design_evaluation_session, control_code)).fetchone()

                # 저장된 데이터가 없으면 RCM 정보만 가져오기
                if not evaluation:
                    evaluation = conn.execute("""
                        SELECT
                            NULL as line_id,
                            d.control_code,
                            rd.control_name,
                            rd.control_description,
                            rd.control_frequency,
                            rd.control_type,
                            rd.control_nature,
                            0 as sample_size,
                            0 as exception_count,
                            NULL as exception_details,
                            NULL as conclusion,
                            NULL as improvement_plan,
                            NULL as review_comment,
                            NULL as evaluation_date,
                            d.attribute0, d.attribute1, d.attribute2, d.attribute3, d.attribute4,
                            d.attribute5, d.attribute6, d.attribute7, d.attribute8, d.attribute9,
                            d.population_attribute_count,
                            d.recommended_sample_size
                        FROM sb_rcm_detail d
                        JOIN sb_rcm_detail_v rd ON d.rcm_id = rd.rcm_id AND d.control_code = rd.control_code
                        WHERE d.rcm_id = %s AND d.control_code = %s
                    """, (rcm_id, control_code)).fetchone()
            else:
                evaluation = conn.execute("""
                    SELECT
                        l.line_id,
                        l.control_code,
                        rd.control_name,
                        rd.control_description,
                        rd.control_frequency,
                        rd.control_type,
                        rd.control_nature,
                        l.sample_size,
                        l.exception_count,
                        l.exception_details,
                        l.conclusion,
                        l.improvement_plan,
                        l.review_comment,
                        l.evaluation_date,
                        d.attribute0, d.attribute1, d.attribute2, d.attribute3, d.attribute4,
                        d.attribute5, d.attribute6, d.attribute7, d.attribute8, d.attribute9,
                        d.population_attribute_count,
                        d.recommended_sample_size
                    FROM sb_evaluation_line l
                    JOIN sb_evaluation_header h ON l.header_id = h.header_id
                    JOIN sb_rcm_detail_v rd ON h.rcm_id = rd.rcm_id AND l.control_code = rd.control_code
                    JOIN sb_rcm_detail d ON d.rcm_id = h.rcm_id AND d.control_code = l.control_code
                    WHERE h.rcm_id = %s
                      AND h.evaluation_name = %s
                      AND h.status >= 2
                      AND l.control_code = %s
                      AND l.conclusion IS NOT NULL
                """, (rcm_id, design_evaluation_session, control_code)).fetchone()

            # 설계평가 결과 조회 (design_comment, evaluation_evidence 및 line_id 가져오기)
            design_evaluation = conn.execute("""
                SELECT l.design_comment, l.line_id, l.evaluation_evidence
                FROM sb_evaluation_line l
                JOIN sb_evaluation_header h ON l.header_id = h.header_id
                WHERE h.rcm_id = %s
                  AND h.evaluation_name = %s
                  AND l.control_code = %s
            """, (rcm_id, design_evaluation_session, control_code)).fetchone()

            # 설계평가 이미지 조회 (DB에서)
            design_image_files = []
            if design_evaluation:
                # line_id 조회
                line_result = conn.execute("""
                    SELECT line_id
                    FROM sb_evaluation_line
                    WHERE header_id = (
                        SELECT header_id
                        FROM sb_evaluation_header
                        WHERE rcm_id = %s AND evaluation_name = %s
                    ) AND control_code = %s
                """, (rcm_id, design_evaluation_session, control_code)).fetchone()

                if line_result:
                    images = conn.execute("""
                        SELECT file_path
                        FROM sb_evaluation_image
                        WHERE evaluation_type = %s AND line_id = %s
                        ORDER BY uploaded_at
                    """, ('design', line_result['line_id'])).fetchall()

                    design_image_files = [img['file_path'] for img in images]

            # 운영평가 이미지 조회 (DB에서) - 자동통제는 건너뜀
            operation_image_files = []
            if evaluation and not is_auto_control:
                # line_id 조회
                op_line_result = conn.execute("""
                    SELECT line_id
                    FROM sb_evaluation_line
                    WHERE header_id = (
                        SELECT header_id
                        FROM sb_evaluation_header
                        WHERE rcm_id = %s AND evaluation_name = %s AND status >= 2
                    ) AND control_code = %s
                """, (rcm_id, design_evaluation_session, control_code)).fetchone()

                if op_line_result:
                    op_images = conn.execute("""
                        SELECT file_path
                        FROM sb_evaluation_image
                        WHERE evaluation_type = %s AND line_id = %s
                        ORDER BY uploaded_at
                    """, ('operation', op_line_result['line_id'])).fetchall()

                    operation_image_files = [img['file_path'] for img in op_images]

        if not evaluation:
            flash('다운로드할 운영평가 결과가 없습니다.', 'warning')
            return redirect(url_for('link7.user_operation_evaluation_rcm'))

        eval_dict = dict(evaluation)
        design_eval_dict = dict(design_evaluation) if design_evaluation else {}

        # Template 시트에 직접 내용 작성
        template_sheet = wb['Template']

        # Client 정보 (C2)
        template_sheet['C2'] = user_info.get('company_name', '')

        # Prepared by (C4)
        template_sheet['C4'] = user_info.get('user_name', '')

        # 통제번호 (C7)
        template_sheet['C7'] = control_code

        # 통제명 (C8)
        template_sheet['C8'] = eval_dict.get('control_name', '')

        # 주기 (C9)
        template_sheet['C9'] = eval_dict.get('control_frequency', '')

        # 구분 (C10)
        template_sheet['C10'] = eval_dict.get('control_type', '')

        # 통제 설명 (C11)
        control_description = eval_dict.get('control_description', '')
        template_sheet['C11'] = control_description

        # C11 셀의 행 높이 자동 조정 (텍스트 길이에 따라)
        if control_description:
            # 줄바꿈 개수 계산
            line_count = control_description.count('\n') + 1
            # 기본 행 높이(15) + 각 줄당 추가 높이(15)
            row_height = 15 + (line_count * 15)
            # 최대 높이 제한 (300)
            row_height = min(row_height, 300)
            template_sheet.row_dimensions[11].height = row_height

        # 설계평가 검토 결과 (C12)
        design_comment = design_eval_dict.get('design_comment', '')
        template_sheet['C12'] = design_comment

        # C12 셀의 행 높이 자동 조정 (텍스트 길이에 따라)
        if design_comment:
            # 줄바꿈 개수 계산
            line_count = design_comment.count('\n') + 1
            # 기본 행 높이(15) + 각 줄당 추가 높이(15)
            row_height = 15 + (line_count * 15)
            # 최대 높이 제한 (300)
            row_height = min(row_height, 300)
            template_sheet.row_dimensions[12].height = row_height

        # 운영평가 의견 작성 (C13)
        operation_review_comment = eval_dict.get('review_comment', '')
        template_sheet['C13'] = operation_review_comment

        # C13 셀의 행 높이 자동 조정 (텍스트 길이에 따라)
        if operation_review_comment:
            # 줄바꿈 개수 계산
            line_count = operation_review_comment.count('\n') + 1
            # 기본 행 높이(15) + 각 줄당 추가 높이(15)
            row_height = 15 + (line_count * 15)
            # 최대 높이 제한 (300)
            row_height = min(row_height, 300)
            template_sheet.row_dimensions[13].height = row_height

        # 운영평가 결론 작성 (C14) - Effective, Ineffective 등
        operation_conclusion = eval_dict.get('conclusion', '')
        template_sheet['C14'] = operation_conclusion

        # Template 시트명을 통제코드로 변경
        template_sheet.title = control_code[:31]  # Excel 시트명 31자 제한

        # Testing Table 시트에 샘플 데이터 작성
        testing_table = wb['Testing Table']
        line_id = eval_dict.get('line_id')
        population_count = eval_dict.get('population_attribute_count', 2)
        sample_size = eval_dict.get('sample_size', 0)

        # 자동통제는 표본수가 0이지만 설계평가 데이터 1건을 표시해야 함
        if is_auto_control and sample_size == 0:
            sample_size = 1

        # 모집단 attribute 개수와 증빙 attribute 개수 계산
        evidence_attributes = []
        for i in range(population_count, 10):
            attr_key = f'attribute{i}'
            attr_name = eval_dict.get(attr_key)
            if attr_name:
                evidence_attributes.append((i, attr_name))

        evidence_count = len(evidence_attributes)

        # 템플릿 구조:
        # - C~L(3~12): Attribute0~9 (10개)
        # - M(13): 결론
        # - N(14): 천고사항/비고
        # 작업: 사용하는 attribute만 헤더 작성하고, 사용안하는 attribute 컬럼 삭제

        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # C열(3번)부터 시작
        current_col = 3

        # 모집단 항목 헤더 작성 (노란색 배경)
        for i in range(population_count):
            attr_key = f'attribute{i}'
            attr_name = eval_dict.get(attr_key, f'모집단{i+1}')
            if attr_name:
                cell = testing_table.cell(row=4, column=current_col, value=attr_name)
                cell.fill = yellow_fill
                current_col += 1

        # 증빙 항목 헤더 작성 (초록색 배경)
        evidence_col_start = current_col
        for i, attr_name in evidence_attributes:
            cell = testing_table.cell(row=4, column=current_col, value=attr_name)
            cell.fill = green_fill
            current_col += 1

        # 사용하지 않는 attribute 컬럼 삭제 (current_col부터 12번까지)
        # 13번(M)은 결론, 14번(N)은 천고사항/비고 - 템플릿 그대로 유지
        if current_col <= 12:
            cols_to_delete = 12 - current_col + 1
            testing_table.delete_cols(current_col, cols_to_delete)

            # 컬럼 삭제 후 3번 행의 "Detailed Testing Table" 병합 복원
            # 삭제 후 마지막 컬럼은 current_col + 1 (결론 + 비고)
            last_col = current_col + 1
            from openpyxl.utils import get_column_letter
            merge_range = f'B3:{get_column_letter(last_col)}3'
            testing_table.merge_cells(merge_range)

            # 삭제 후 결론 컬럼 위치는 current_col이 됨 (M열이 앞으로 당겨짐)
            conclusion_col = current_col
        else:
            # attribute가 10개 모두 사용되는 경우
            conclusion_col = 13  # M열

        # 자동통제인 경우 처리
        if is_auto_control:
            # attribute가 하나도 없으면 Testing Table 전체 삭제
            if population_count == 0 and len(evidence_attributes) == 0:
                # Testing Table 시트 전체를 삭제하는 대신, 필요한 행만 남기고 나머지 삭제
                # 1~2행(설명), 3행(제목) 유지, 4행 이후 삭제
                testing_table.delete_rows(3, 100)  # 3번 행부터 100개 행 삭제
            else:
                # 설계평가 샘플 데이터 조회 (sb_evaluation_sample에서)
                evidence_data = {}
                if design_eval_dict.get('line_id'):
                    design_line_id = design_eval_dict['line_id']

                    with get_db() as conn:
                        design_sample = conn.execute("""
                            SELECT attribute0, attribute1, attribute2, attribute3, attribute4,
                                   attribute5, attribute6, attribute7, attribute8, attribute9
                            FROM sb_evaluation_sample
                            WHERE line_id = %s AND evaluation_type = 'design'
                            LIMIT 1
                        """, (design_line_id,)).fetchone()

                        if design_sample:
                            evidence_data = dict(design_sample)

                # 5번 행에 설계평가 증빙 값 표시
                testing_table.cell(row=5, column=2, value=1)  # No.

                # 자동통제는 모집단과 증빙을 구분하여 표시 (수동통제와 동일한 구조)
                col = 3

                # 모집단 데이터 (C열부터, 빈 값으로 두거나 설계평가 데이터에서 가져오기)
                for i in range(population_count):
                    attr_key = f'attribute{i}'
                    # 자동통제의 경우 모집단은 빈 값 또는 설계평가 데이터에서 가져오기
                    attr_value = evidence_data.get(attr_key, '')
                    testing_table.cell(row=5, column=col, value=attr_value)
                    col += 1

                # 증빙 데이터 (모집단 다음 컬럼부터)
                for i, attr_name in evidence_attributes:
                    attr_key = f'attribute{i}'
                    attr_value = evidence_data.get(attr_key, '')
                    testing_table.cell(row=5, column=col, value=attr_value)
                    col += 1

                # 결론 컬럼
                conclusion_value = 'Effective' if eval_dict.get('conclusion') == 'effective' else 'See Comments'
                testing_table.cell(row=5, column=conclusion_col, value=conclusion_value)

        # B열에 순번 작성 (1, 2, 3, ...)
        if sample_size > 0:
            for i in range(sample_size):
                testing_table.cell(row=5 + i, column=2, value=i + 1)  # B열 = column 2

        # 샘플 데이터 입력 (5행부터) - 자동통제는 건너뜀
        if line_id and not is_auto_control:
            samples = get_operation_evaluation_samples(line_id)
            if samples:
                for row_idx, sample in enumerate(samples, start=5):
                    sample_attributes = sample.get('attributes', {})

                    # 모집단 데이터 (C열부터)
                    col = 3
                    for i in range(population_count):
                        attr_key = f'attribute{i}'
                        attr_value = sample_attributes.get(attr_key, '')
                        testing_table.cell(row=row_idx, column=col, value=attr_value)
                        col += 1

                    # 증빙 데이터 (모집단 다음 컬럼부터)
                    for i, attr_name in evidence_attributes:
                        attr_key = f'attribute{i}'
                        attr_value = sample_attributes.get(attr_key, '')
                        testing_table.cell(row=row_idx, column=col, value=attr_value)
                        col += 1

                    # 결과(결론) 데이터 - 저장된 결론 컬럼 위치에 작성
                    # result: 'exception' -> 'Exception', 'no_exception' -> 'No Exception'
                    result = sample.get('result', 'no_exception')
                    result_text = 'Exception' if result == 'exception' else 'No Exception'
                    testing_table.cell(row=row_idx, column=conclusion_col, value=result_text)

                    # 경감요소(비고) 데이터 - 결론 다음 컬럼에 작성
                    mitigation = sample.get('mitigation', '')
                    if mitigation:
                        testing_table.cell(row=row_idx, column=conclusion_col + 1, value=mitigation)

        # 66번 행("Testing Table")의 색상을 행 전체에 미리 적용 (행 삭제 전)
        from copy import copy
        from openpyxl.styles import PatternFill
        source_cell_66 = testing_table.cell(row=66, column=2)
        if source_cell_66.fill:
            for col in range(2, 16385):
                cell = testing_table.cell(row=66, column=col)
                cell.fill = copy(source_cell_66.fill)

        # 사용하지 않는 행 삭제 (5~64행까지 60개 준비되어 있음)
        if sample_size < 60:
            first_row_to_delete = 5 + sample_size
            rows_to_delete = 64 - first_row_to_delete + 1
            if rows_to_delete > 0:
                testing_table.delete_rows(first_row_to_delete, rows_to_delete)

        # 행 삭제 후 "Testing Table" 구분자 행 위치 계산
        # 원래 66번 행이 (66 - rows_to_delete)번으로 이동
        if sample_size < 60 and rows_to_delete > 0:
            testing_table_separator_row = 66 - rows_to_delete
        else:
            testing_table_separator_row = 66

        # 이미지 삽입 시작 위치 초기화 (구분자 2칸 아래)
        current_row = testing_table_separator_row + 2

        # 설계평가 이미지를 Testing Table 구분자 다음에 삽입
        if design_image_files:
            from openpyxl.drawing.image import Image as XLImage

            for image_path in design_image_files:
                if os.path.exists(image_path):
                    try:
                        # 이미지 객체 생성
                        xl_img = XLImage(image_path)

                        # 이미지 크기 조정 (최대 너비 400px)
                        max_width = 400
                        if xl_img.width > max_width:
                            ratio = max_width / xl_img.width
                            xl_img.width = max_width
                            xl_img.height = int(xl_img.height * ratio)

                        # 이미지 삽입 (모든 설계평가 이미지는 같은 행에 삽입)
                        design_img_row = testing_table_separator_row + 2
                        xl_img.anchor = f'B{design_img_row}'
                        testing_table.add_image(xl_img)

                        # 행 높이 조정 (가장 큰 이미지 높이로 설정)
                        current_height = testing_table.row_dimensions[design_img_row].height or 0
                        new_height = (xl_img.height * 0.75) + 5
                        if new_height > current_height:
                            testing_table.row_dimensions[design_img_row].height = new_height

                    except Exception as e:
                        print(f"설계평가 이미지 삽입 실패 ({image_path}): {e}")

        # 운영평가 이미지 삽입 (설계평가 바로 다음)
        if operation_image_files:
            from openpyxl.drawing.image import Image as XLImage

            # 운영평가 이미지는 설계평가 이미지 다음 행 (설계평가 +0, 빈칸 +1, 운영평가 +2)
            # 설계평가가 (separator + 2)에 있으므로, 운영평가는 (separator + 2 + 2) = separator + 4
            if design_image_files:
                operation_img_row = testing_table_separator_row + 4
            else:
                operation_img_row = testing_table_separator_row + 2

            for image_path in operation_image_files:
                if os.path.exists(image_path):
                    try:
                        # 이미지 객체 생성
                        xl_img = XLImage(image_path)

                        # 이미지 크기 조정 (최대 너비 400px)
                        max_width = 400
                        if xl_img.width > max_width:
                            ratio = max_width / xl_img.width
                            xl_img.width = max_width
                            xl_img.height = int(xl_img.height * ratio)

                        # 이미지 삽입 (모든 운영평가 이미지는 같은 행에 삽입)
                        xl_img.anchor = f'B{operation_img_row}'
                        testing_table.add_image(xl_img)

                        # 행 높이 조정 (가장 큰 이미지 높이로 설정)
                        current_height = testing_table.row_dimensions[operation_img_row].height or 0
                        new_height = (xl_img.height * 0.75) + 5
                        if new_height > current_height:
                            testing_table.row_dimensions[operation_img_row].height = new_height

                    except Exception as e:
                        print(f"운영평가 이미지 삽입 실패 ({image_path}): {e}")

        # Population 시트 처리
        recommended_sample_size = eval_dict.get('recommended_sample_size', 0)
        if recommended_sample_size == 0 and 'Population' in wb.sheetnames:
            # 표본수가 0인 경우: 업로드한 모집단 데이터를 Population 시트에 채움
            population_sheet = wb['Population']

            # 업로드한 모집단 파일 경로
            upload_folder = os.path.join('uploads', 'populations')
            population_file_pattern = f"{user_info['user_id']}_{control_code}_*"
            population_files = []

            if os.path.exists(upload_folder):
                import glob
                population_files = glob.glob(os.path.join(upload_folder, population_file_pattern))

            if population_files:
                # 가장 최근 파일 사용
                population_file = max(population_files, key=os.path.getmtime)

                try:
                    # 모집단 파일 읽기
                    pop_wb = load_workbook(population_file, read_only=True)
                    pop_ws = pop_wb.active

                    # 헤더 복사 (1행)
                    for col_idx, cell in enumerate(pop_ws[1], start=1):
                        if cell.value:
                            population_sheet.cell(row=1, column=col_idx, value=cell.value)

                    # 데이터 복사 (2행부터)
                    row_idx = 2
                    for row in pop_ws.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):  # 빈 행이 아니면
                            for col_idx, value in enumerate(row, start=1):
                                population_sheet.cell(row=row_idx, column=col_idx, value=value)
                            row_idx += 1

                    pop_wb.close()
                except Exception:
                    pass
        elif recommended_sample_size != 0 and 'Population' in wb.sheetnames:
            # 표본수가 0이 아닌 경우: Population 시트 삭제
            wb.remove(wb['Population'])

        # 시트 순서 조정: 통제명 시트를 가장 앞에, Testing Table을 두 번째로
        control_sheet_index = wb.index(template_sheet)
        testing_table_index = wb.index(wb['Testing Table'])

        # 통제명 시트를 맨 앞으로 이동
        wb.move_sheet(template_sheet, offset=-control_sheet_index)
        # Testing Table을 두 번째로 이동 (통제명 시트 다음)
        wb.move_sheet(wb['Testing Table'], offset=-testing_table_index + 1)

        # Population 시트가 있으면 세 번째로 이동
        if 'Population' in wb.sheetnames:
            population_index = wb.index(wb['Population'])
            wb.move_sheet(wb['Population'], offset=-population_index + 2)

        # 명명된 범위(defined names) 제거 (깨진 참조 방지)
        if hasattr(wb, 'defined_names'):
            try:
                # openpyxl 3.x 버전
                names_to_remove = list(wb.defined_names.definedName)
                for name in names_to_remove:
                    del wb.defined_names[name.name]
            except AttributeError:
                # openpyxl 2.x 버전 또는 다른 구조
                try:
                    names_to_remove = [name for name in wb.defined_names]
                    for name in names_to_remove:
                        try:
                            del wb.defined_names[name]
                        except:
                            pass
                except:
                    pass

        # 외부 링크(external links) 제거
        if hasattr(wb, '_external_links'):
            wb._external_links = []

        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        # 다운로드 파일명 생성 (통제번호_통제명)
        control_name = eval_dict.get('control_name', '')
        filename = f"{control_code}_{control_name}.xlsx"
        unsafe_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
        for char in unsafe_chars:
            filename = filename.replace(char, '_')

        # 파일 전송 (UTF-8 인코딩)
        response = make_response(send_file(
            temp_file.name,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ))

        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'다운로드 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('link7.user_operation_evaluation_rcm'))


@bp_link7.route('/api/operation-evaluation/upload-image', methods=['POST'])
@login_required
def upload_operation_image():
    """운영평가 이미지 업로드"""
    try:
        rcm_id = request.form.get('rcm_id')
        header_id = request.form.get('header_id')
        control_code = request.form.get('control_code')

        if not all([rcm_id, header_id, control_code]):
            return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'}), 400

        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다.'}), 400

        # 파일 확장자 검증
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'message': '허용되지 않는 파일 형식입니다.'}), 400

        # 저장 경로 생성
        upload_dir = os.path.join('static', 'uploads', 'operation_evaluations', str(rcm_id), str(header_id), control_code)
        os.makedirs(upload_dir, exist_ok=True)

        # 파일명 생성 (타임스탬프 포함)
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{file.filename}"
        filepath = os.path.join(upload_dir, filename)

        # 파일 저장
        file.save(filepath)

        # DB에 이미지 정보 저장
        from db_config import get_db
        with get_db() as conn:
            # line_id 조회
            line_record = conn.execute('''
                SELECT l.line_id FROM sb_evaluation_line l
                JOIN sb_evaluation_header h ON l.header_id = h.header_id
                WHERE h.header_id = %s AND l.control_code = %s
            ''', (header_id, control_code)).fetchone()

            if line_record:
                line_id = line_record['line_id']
                relative_path = filepath.replace('\\', '/')
                file_size = os.path.getsize(filepath)

                # 중복 체크
                existing = conn.execute('''
                    SELECT image_id FROM sb_evaluation_image
                    WHERE evaluation_type = %s AND line_id = %s AND file_path = %s
                ''', ('operation', line_id, relative_path)).fetchone()

                if not existing:
                    conn.execute('''
                        INSERT INTO sb_evaluation_image
                        (evaluation_type, line_id, file_path, file_name, file_size, uploaded_at)
                        VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                    ''', ('operation', line_id, relative_path, filename, file_size))
                    conn.commit()

        return jsonify({
            'success': True,
            'message': '이미지가 업로드되었습니다.',
            'filepath': filepath.replace('\\', '/')
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'업로드 실패: {str(e)}'}), 500


@bp_link7.route('/api/operation-evaluation/images/<int:rcm_id>/<int:header_id>/<control_code>')
@login_required
def get_operation_images(rcm_id, header_id, control_code):
    """운영평가 이미지 목록 조회"""
    try:
        from db_config import get_db
        with get_db() as conn:
            # line_id 조회
            line_record = conn.execute('''
                SELECT l.line_id FROM sb_evaluation_line l
                JOIN sb_evaluation_header h ON l.header_id = h.header_id
                WHERE h.header_id = %s AND l.control_code = %s
            ''', (header_id, control_code)).fetchone()

            if not line_record:
                return jsonify({'success': True, 'images': []})

            line_id = line_record['line_id']

            # DB에서 이미지 조회
            image_records = conn.execute('''
                SELECT image_id, file_path, file_name, file_size, uploaded_at
                FROM sb_evaluation_image
                WHERE evaluation_type = %s AND line_id = %s
                ORDER BY uploaded_at
            ''', ('operation', line_id)).fetchall()

            images = []
            for img in image_records:
                images.append({
                    'filename': img['file_name'],
                    'url': f"/{img['file_path']}"
                })

            return jsonify({'success': True, 'images': images})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link7.route('/api/operation-evaluation/delete-image', methods=['POST'])
@login_required
def delete_operation_image():
    """운영평가 이미지 삭제"""
    try:
        data = request.get_json()
        filepath = data.get('filepath')

        if not filepath:
            return jsonify({'success': False, 'message': '파일 경로가 없습니다.'}), 400

        # DB에서 이미지 삭제
        from db_config import get_db
        with get_db() as conn:
            conn.execute('''
                DELETE FROM sb_evaluation_image
                WHERE evaluation_type = %s AND file_path = %s
            ''', ('operation', filepath))
            conn.commit()

        # 파일 존재 확인 및 삭제
        if os.path.exists(filepath):
            os.remove(filepath)

        return jsonify({'success': True, 'message': '이미지가 삭제되었습니다.'})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@bp_link7.route('/api/design-evaluation/samples', methods=['GET'])
@login_required
def get_design_evaluation_samples():
    """설계평가 표본 데이터 조회 (운영평가에서 참고용)"""
    try:
        user_info = get_user_info()
        rcm_id = request.args.get('rcm_id')
        evaluation_session = request.args.get('evaluation_session')
        control_code = request.args.get('control_code')

        if not all([rcm_id, evaluation_session, control_code]):
            return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'})

        rcm_id = int(rcm_id)

        # 통합 테이블에서 설계평가 line_id 조회
        with get_db() as conn:
            # 헤더 조회
            header = conn.execute('''
                SELECT header_id
                FROM sb_evaluation_header
                WHERE rcm_id = ? AND evaluation_name = ?
            ''', (rcm_id, evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '설계평가 세션을 찾을 수 없습니다.'})

            header_id = dict(header)['header_id']

            # 해당 통제의 라인 조회
            line = conn.execute('''
                SELECT line_id
                FROM sb_evaluation_line
                WHERE header_id = ? AND control_code = ?
            ''', (header_id, control_code)).fetchone()

            if not line:
                return jsonify({'success': False, 'message': '해당 통제의 평가 데이터를 찾을 수 없습니다.'})

            line_id = dict(line)['line_id']

            # 표본 데이터 조회
            samples = conn.execute('''
                SELECT sample_number, evidence, result, mitigation
                FROM sb_evaluation_sample
                WHERE line_id = ?
                ORDER BY sample_number
            ''', (line_id,)).fetchall()

            # 딕셔너리로 변환
            samples_list = [dict(sample) for sample in samples]

            return jsonify({
                'success': True,
                'samples': samples_list,
                'count': len(samples_list)
            })

    except Exception as e:
        print(f"[get_design_evaluation_samples] 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'조회 오류: {str(e)}'})


# =============================================================================
# 통합 운영평가 API (Generic Implementation)
# 모든 수동통제(APD01, APD07, APD09, APD12 등)에 대해 공통으로 사용
# =============================================================================

@bp_link7.route('/operation-evaluation/manual/<control_code>')
@login_required
def manual_control_evaluation(control_code):
    """
    통합 운영평가 페이지
    모든 수동통제에 대해 공통으로 사용
    """
    user_info = get_user_info()

    # 통제 설정 확인
    config = get_control_config(control_code)
    if not config:
        flash(f'지원하지 않는 통제 코드입니다: {control_code}', 'error')
        return redirect(url_for('link7.user_operation_evaluation_index'))

    # URL 파라미터
    rcm_id = request.args.get('rcm_id')
    control_code_param = control_code
    control_name = request.args.get('control_name', config['name'])
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, design_evaluation_session]):
        flash('필수 정보가 누락되었습니다.', 'error')
        return redirect(url_for('link7.user_operation_evaluation_index'))

    # 기존 운영평가 데이터 조회
    operation_evaluation_session = f"OP_{design_evaluation_session}"
    existing_data = None
    pc01_data = None  # PC02용 PC01 데이터

    try:
        with get_db() as conn:
            # 운영평가 헤더 조회
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if header:
                operation_header_id = header['header_id']

                # 기존 데이터 로드
                loaded_data = file_manager.load_operation_test_data(
                    rcm_id=rcm_id,
                    operation_header_id=operation_header_id,
                    control_code=control_code_param
                )

                if loaded_data and loaded_data.get('samples_data'):
                    existing_data = loaded_data

                # 데이터베이스에서 no_occurrence 정보 추가 로드
                if existing_data:
                    eval_line = conn.execute('''
                        SELECT conclusion, no_occurrence, no_occurrence_reason
                        FROM sb_evaluation_line
                        WHERE header_id = %s AND control_code = %s
                        ORDER BY line_id DESC
                        LIMIT 1
                    ''', (operation_header_id, control_code_param)).fetchone()

                    if eval_line:
                        existing_data['no_occurrence'] = eval_line['no_occurrence'] if eval_line['no_occurrence'] else False
                        existing_data['no_occurrence_reason'] = eval_line['no_occurrence_reason'] or ''
                        existing_data['conclusion'] = eval_line['conclusion']
                elif not loaded_data:
                    # 엑셀 파일은 없지만 DB에 no_occurrence 데이터가 있을 수 있음
                    eval_line = conn.execute('''
                        SELECT conclusion, no_occurrence, no_occurrence_reason
                        FROM sb_evaluation_line
                        WHERE header_id = %s AND control_code = %s
                        ORDER BY line_id DESC
                        LIMIT 1
                    ''', (operation_header_id, control_code_param)).fetchone()

                    if eval_line and eval_line['no_occurrence']:
                        existing_data = {
                            'no_occurrence': True,
                            'no_occurrence_reason': eval_line['no_occurrence_reason'] or '',
                            'conclusion': eval_line['conclusion'],
                            'samples_data': None,
                            'population_data': []
                        }

                # PC02, PC03인 경우 PC01 데이터도 로드
                if control_code in ['PC02', 'PC03']:
                    pc01_loaded = file_manager.load_operation_test_data(
                        rcm_id=rcm_id,
                        operation_header_id=operation_header_id,
                        control_code='PC01'
                    )
                    if pc01_loaded and pc01_loaded.get('samples_data'):
                        pc01_data = pc01_loaded

                        # PC02/PC03 파일이 없으면 PC01 데이터로 파일 생성
                        if not existing_data:
                            # PC01의 표본 데이터로 엑셀 파일 생성
                            pc01_samples = pc01_data['samples_data']['samples']
                            pc01_population = pc01_data.get('population_data', [])

                            # PC02/PC03 파일 생성
                            file_manager.save_operation_test_data(
                                rcm_id=rcm_id,
                                operation_header_id=operation_header_id,
                                control_code=control_code,
                                population_data=pc01_population,
                                field_mapping={'program_name': '0', 'deploy_date': '1'},
                                samples=pc01_samples,
                                test_results_data=None  # 테스트 결과는 아직 없음
                            )

                            # 생성된 파일 로드
                            existing_data = file_manager.load_operation_test_data(
                                rcm_id=rcm_id,
                                operation_header_id=operation_header_id,
                                control_code=control_code
                            )
                    else:
                        flash('PC01 모집단 데이터가 없습니다. PC01을 먼저 수행해주세요.', 'error')
                        return redirect(url_for('link7.user_operation_evaluation_rcm',
                                              rcm_id=rcm_id,
                                              evaluation_session=design_evaluation_session))

        # 속성 메타데이터 조회
        attribute_meta = conn.execute('''
            SELECT attribute_name, attribute_label, display_order
            FROM sb_control_attribute_meta
            WHERE control_code = %s
            ORDER BY display_order
        ''', (control_code,)).fetchall()

    except Exception as e:
        print(f"기존 데이터 로드 오류: {e}")
        attribute_meta = []

    return render_template('link7_manual_generic.jsp',
                         rcm_id=rcm_id,
                         control_code=control_code_param,
                         control_name=control_name,
                         design_evaluation_session=design_evaluation_session,
                         config=config,
                         existing_data=existing_data,
                         pc01_data=pc01_data,
                         attribute_meta=[dict(row) for row in attribute_meta])


@bp_link7.route('/api/operation-evaluation/manual/<control_code>/upload-population', methods=['POST'])
@login_required
def upload_manual_population(control_code):
    """
    통합 모집단 업로드 API
    모든 수동통제에 대해 공통으로 사용
    """
    user_info = get_user_info()

    # 통제 설정 확인
    config = get_control_config(control_code)
    if not config:
        return jsonify({'success': False, 'message': f'지원하지 않는 통제 코드입니다: {control_code}'})

    # 파일 받기
    if 'population_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'})

    file = request.files['population_file']
    if not file.filename:
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'})

    # 필드 매핑 정보 받기
    field_mapping_str = request.form.get('field_mapping')
    if not field_mapping_str:
        return jsonify({'success': False, 'message': '필드 매핑 정보가 없습니다.'})

    try:
        field_mapping = json.loads(field_mapping_str)
    except:
        return jsonify({'success': False, 'message': '필드 매핑 형식이 올바르지 않습니다.'})

    # RCM 정보
    rcm_id = request.form.get('rcm_id')
    design_evaluation_session = request.form.get('design_evaluation_session')

    if not all([rcm_id, control_code, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

    try:
        # 운영평가 헤더 조회
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다. RCM 페이지에서 다시 시작해주세요.'})

            operation_header_id = header['header_id']

        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()

        # RCM detail에서 recommended_sample_size 가져오기
        with get_db() as conn:
            rcm_detail = conn.execute('''
                SELECT recommended_sample_size
                FROM sb_rcm_detail
                WHERE rcm_id = %s AND control_code = %s
            ''', (rcm_id, control_code)).fetchone()

        recommended_size = rcm_detail['recommended_sample_size'] if rcm_detail else None

        # 모집단 파싱 (통합 함수 사용)
        population_data = file_manager.parse_population_excel(temp_file.name, field_mapping)
        count = len(population_data)

        # recommended_sample_size가 있으면 사용, 없으면 자동 계산
        if recommended_size is not None and recommended_size > 0:
            sample_size = min(recommended_size, count)
        else:
            sample_size = file_manager.calculate_sample_size(count)

        # 표본 선택
        samples = file_manager.select_random_samples(population_data, sample_size)

        # 임시 파일 삭제 (Windows에서 파일 핸들 문제로 실패할 수 있으므로 무시)
        try:
            os.unlink(temp_file.name)
        except Exception as e:
            print(f"임시 파일 삭제 실패 (무시됨): {e}")

        # 템플릿 기반 엑셀 파일 생성 및 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=population_data,
            field_mapping=field_mapping,
            samples=samples,
            test_results_data=None
        )

        # 세션에 정보 저장
        session_key = f'manual_test_{rcm_id}_{control_code}'
        session[session_key] = {
            'file_paths': file_paths,
            'rcm_id': rcm_id,
            'control_code': control_code,
            'design_evaluation_session': design_evaluation_session,
            'operation_header_id': operation_header_id,
            'field_mapping': field_mapping,
            'population_count': count,
            'sample_size': sample_size
        }

        return jsonify({
            'success': True,
            'population_count': count,
            'sample_size': sample_size,
            'samples': samples
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@bp_link7.route('/api/operation-evaluation/manual/<control_code>/get-design-data', methods=['GET'])
@login_required
def get_design_data_for_manual_control(control_code):
    """
    수동 통제 운영평가에서 사용할 설계평가 데이터 조회
    """
    user_info = get_user_info()
    rcm_id = request.args.get('rcm_id')
    design_evaluation_session = request.args.get('design_evaluation_session')

    if not all([rcm_id, design_evaluation_session]):
        return jsonify({'success': False, 'message': '필수 파라미터가 누락되었습니다.'}), 400

    try:
        with get_db() as conn:
            # 설계평가 결과 조회
            design_eval = conn.execute('''
                SELECT l.evaluation_evidence, l.evaluation_rationale
                FROM sb_evaluation_line l
                JOIN sb_evaluation_header h ON l.header_id = h.header_id
                WHERE h.rcm_id = %s AND h.user_id = %s AND h.evaluation_name = %s AND l.control_code = %s
            ''', (rcm_id, user_info['user_id'], design_evaluation_session, control_code)).fetchone()

            if design_eval:
                return jsonify({
                    'success': True,
                    'design_data': dict(design_eval)
                })
            else:
                return jsonify({'success': False, 'message': '설계평가 데이터를 찾을 수 없습니다.'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': f'설계평가 데이터 조회 중 오류 발생: {str(e)}'}), 500


@bp_link7.route('/api/operation-evaluation/manual/<control_code>/save-test-results', methods=['POST'])
@login_required
def save_manual_test_results(control_code):
    """
    통합 테스트 결과 저장 API
    모든 수동통제에 대해 공통으로 사용
    """
    user_info = get_user_info()

    # 통제 설정 확인
    config = get_control_config(control_code)
    if not config:
        return jsonify({'success': False, 'message': f'지원하지 않는 통제 코드입니다: {control_code}'})

    try:
        data = request.json
        rcm_id = data.get('rcm_id')
        design_evaluation_session = data.get('design_evaluation_session')
        test_results = data.get('test_results', [])

        # 세션에서 정보 가져오기
        session_key = f'manual_test_{rcm_id}_{control_code}'
        session_data = session.get(session_key)
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        # 세션이 없으면 DB에서 operation_header_id 조회
        if not session_data:
            with get_db() as conn:
                header = conn.execute('''
                    SELECT header_id FROM sb_evaluation_header
                    WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

                if not header:
                    return jsonify({'success': False, 'message': '운영평가 세션을 찾을 수 없습니다.'})

                operation_header_id = header['header_id']
        else:
            operation_header_id = session_data['operation_header_id']

        # 저장된 파일에서 표본 데이터 로드
        loaded_data = file_manager.load_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code
        )

        if not loaded_data or not loaded_data['samples_data']:
            return jsonify({'success': False, 'message': '저장된 표본 데이터를 찾을 수 없습니다.'})

        samples_data = loaded_data['samples_data']

        # field_mapping 가져오기 (세션 또는 기본 매핑)
        if session_data and 'field_mapping' in session_data:
            field_mapping = session_data['field_mapping']
        else:
            # 기본 field_mapping 생성 (인덱스 기반)
            config = get_control_config(control_code)
            field_mapping = {field: str(idx) for idx, field in enumerate(config['population_fields'])}

        # 엑셀 파일에 테스트 결과 저장
        file_paths = file_manager.save_operation_test_data(
            rcm_id=rcm_id,
            operation_header_id=operation_header_id,
            control_code=control_code,
            population_data=loaded_data.get('population_data', []),
            field_mapping=field_mapping,
            samples=samples_data['samples'],
            test_results_data={
                'test_results': test_results,
                'exceptions': [r for r in test_results if r.get('has_exception')],
                'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
                'test_type': control_code
            }
        )

        # 운영평가 데이터 저장

        # sample_lines 데이터 구성
        sample_lines = []
        for i, res in enumerate(test_results):
            sample_lines.append({
                'sample_number': i + 1,
                'evidence': res.get('notes', ''),
                'result': 'exception' if res.get('has_exception') else 'no_exception',
                'mitigation': '', # 수동통제 UI에서는 개별 경감요소 없음
                'attributes': res.get('attributes', {})
            })

        evaluation_data = {
            'test_type': control_code,
            'population_count': samples_data['population_count'],
            'sample_size': samples_data['sample_size'],
            'test_results_path': file_paths.get('excel_path'),
            'conclusion': 'effective' if not any(r.get('has_exception') for r in test_results) else 'exception',
            'sample_lines': sample_lines
        }

        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'{control_code} 테스트 저장',
                         f'/api/operation-evaluation/manual/{control_code}/save-test-results',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': f'{control_code} 테스트 결과가 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})


@bp_link7.route('/api/operation-evaluation/save-no-occurrence', methods=['POST'])
@login_required
def save_no_occurrence():
    """
    당기 발생사실 없음 저장 API
    모든 수동통제에 대해 공통으로 사용
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    user_info = get_user_info()

    try:
        data = request.get_json()
        rcm_id = data.get('rcm_id')
        control_code = data.get('control_code')
        design_evaluation_session = data.get('design_evaluation_session')
        no_occurrence_reason = data.get('no_occurrence_reason', '')

        if not all([rcm_id, control_code, design_evaluation_session]):
            return jsonify({'success': False, 'message': '필수 정보가 누락되었습니다.'})

        # 운영평가 헤더 조회 또는 생성
        operation_evaluation_session = f"OP_{design_evaluation_session}"

        with get_db() as conn:
            header = conn.execute('''
                SELECT header_id FROM sb_evaluation_header
                WHERE rcm_id = %s AND user_id = %s AND evaluation_session = %s AND design_evaluation_session = %s
            ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session)).fetchone()

            if not header:
                # 헤더 생성
                cursor = conn.execute('''
                    INSERT INTO sb_evaluation_header
                    (rcm_id, user_id, evaluation_session, design_evaluation_session, created_date)
                    VALUES (%s, %s, %s, %s, datetime('now'))
                ''', (rcm_id, user_info['user_id'], operation_evaluation_session, design_evaluation_session))
                operation_header_id = cursor.lastrowid
                conn.commit()
            else:
                operation_header_id = header['header_id']

        # 통제 설정 조회
        config = get_control_config(control_code)
        if not config:
            config = get_control_config('GENERIC')  # GENERIC 설정 사용

        # "당기 발생사실 없음" 엑셀 파일 생성
        wb = Workbook()

        # Population 시트
        ws_pop = wb.active
        ws_pop.title = 'Population'

        # 헤더 스타일
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Population 시트에 "당기 발생사실 없음" 메시지
        ws_pop['A1'] = '당기 발생사실 없음'
        ws_pop['A1'].font = Font(bold=True, size=14, color='FF0000')
        ws_pop['A1'].alignment = Alignment(horizontal='center', vertical='center')

        if no_occurrence_reason:
            ws_pop['A2'] = f'사유: {no_occurrence_reason}'
            ws_pop['A2'].font = Font(size=11)

        ws_pop.column_dimensions['A'].width = 50

        # Test Table 시트
        ws_test = wb.create_sheet('Test Table')

        # 헤더 행 작성
        excel_headers = config['excel_headers']['testing']
        for col_idx, header in enumerate(excel_headers, 1):
            cell = ws_test.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 한 줄의 빈 데이터 행 추가 (No = 1)
        ws_test.cell(row=2, column=1, value=1)  # No
        for col_idx in range(2, len(excel_headers) + 1):
            ws_test.cell(row=2, column=col_idx, value='')

        # 비고 컬럼에 "당기 발생사실 없음" 표시
        remark_col_idx = len(excel_headers)  # 마지막 컬럼이 비고
        ws_test.cell(row=2, column=remark_col_idx, value='당기 발생사실 없음')

        # 컬럼 너비 자동 조정
        for col_idx, header in enumerate(excel_headers, 1):
            ws_test.column_dimensions[ws_test.cell(row=1, column=col_idx).column_letter].width = 15

        # 엑셀 파일 저장
        excel_path = file_manager.get_test_results_path(rcm_id, operation_header_id, control_code)
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb.save(excel_path)

        # 운영평가 데이터 저장 (당기 발생사실 없음)
        evaluation_data = {
            'sample_size': 0,
            'exception_count': 0,
            'exception_details': '',
            'conclusion': 'effective',
            'improvement_plan': '',
            'no_occurrence': True,
            'no_occurrence_reason': no_occurrence_reason,
            'population_count': 0,
            'population_path': None,
            'samples_path': None,
            'test_results_path': excel_path
        }

        save_operation_evaluation(rcm_id, control_code, user_info['user_id'],
                                 operation_evaluation_session, design_evaluation_session, evaluation_data)

        log_user_activity(user_info, 'OPERATION_EVALUATION', f'{control_code} 당기 발생사실 없음 저장',
                         '/api/operation-evaluation/save-no-occurrence',
                         request.remote_addr, request.headers.get('User-Agent'))

        return jsonify({'success': True, 'message': '당기 발생사실 없음이 저장되었습니다.'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'저장 오류: {str(e)}'})
